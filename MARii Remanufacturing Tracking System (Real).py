###########################################################################
## Python code generated with wxFormBuilder (version Jun 17 2015)
## http://www.wxformbuilder.org/
##
## PLEASE DO "NOT" EDIT THIS FILE!
###########################################################################

import wx
import wx.xrc
import wx.adv
import openpyxl

###########################################################################
## Load The Workbook(Experimental and Will be updated in GUI)
###########################################################################

class LoadWorkbook(object):
    def __init__ (self, parent):
        self.wb = openpyxl.load_workbook('Audit Report Score.xlsx')
        self.wb.get_sheet_names()

class LoadWorkbook1(object):
    def __init__(self, parent):
        self.wb1 = openpyxl.load_workbook('Checklist Old.xlsx')
        self.wb1.get_sheet_names()

class LoadWorkbook2(object):
    def __init__ (self, parent):
        self.wb2 = openpyxl.load_workbook('Checklist New.xlsx')
        self.wb2.get_sheet_names()

###########################################################################
## Class MainSystem
###########################################################################

class MainSystem ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Main System", pos = wx.DefaultPosition, size = wx.Size( 722,456 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

    	self.SetSizeHintsSz( wx.Size(722,456),wx.Size(722,456))
    	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_BTNHIGHLIGHT ) )

    	bSizer1 = wx.BoxSizer( wx.VERTICAL )

    	bSizer258 = wx.BoxSizer( wx.HORIZONTAL )


    	bSizer258.Add( ( 5, 0), 0, 0, 5 )

    	self.bitmap1 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"MARii logo(Small).jpg", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.Size( 300,80 ), 0 )
    	bSizer258.Add( self.bitmap1, 0, 0, 5 )

    	self.staticText226 = wx.StaticText( self, wx.ID_ANY, u"REMANUFACTURING \nTRACKING SYSTEM", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText226.Wrap( -1 )
    	self.staticText226.SetFont( wx.Font( 18, 74, 93, 92, False, "Arial" ) )

    	bSizer258.Add( self.staticText226, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )


    	bSizer1.Add( bSizer258, 1, wx.EXPAND, 5 )

    	bSizer257 = wx.BoxSizer( wx.VERTICAL )

    	bSizer2 = wx.BoxSizer( wx.HORIZONTAL )

    	bSizer4 = wx.BoxSizer( wx.VERTICAL )


    	bSizer4.Add( ( 0, 20), 0, 0, 5 )

    	self.button1 = wx.Button( self, wx.ID_ANY, u"Details \nand Report", wx.Point( -1,-1 ), wx.Size( 180,100 ), 0 )
    	self.button1.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

    	bSizer4.Add( self.button1, 0, wx.ALIGN_BOTTOM|wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


    	bSizer2.Add( bSizer4, 1, wx.EXPAND, 5 )

    	bSizer6 = wx.BoxSizer( wx.VERTICAL )


    	bSizer6.Add( ( 0, 20), 0, 0, 5 )

    	self.button2 = wx.Button( self, wx.ID_ANY, u"Checklist", wx.Point( 0,0 ), wx.Size( 180,100 ), 0 )
    	self.button2.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

    	bSizer6.Add( self.button2, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


    	bSizer2.Add( bSizer6, 1, wx.EXPAND, 5 )


    	bSizer257.Add( bSizer2, 1, wx.EXPAND, 5 )

    	bSizer5 = wx.BoxSizer( wx.HORIZONTAL )

    	bSizer7 = wx.BoxSizer( wx.VERTICAL )


    	bSizer7.Add( ( 0, 20), 0, 0, 5 )

    	self.button3 = wx.Button( self, wx.ID_ANY, u"Audit \nScore", wx.DefaultPosition, wx.Size( 180,100 ), 0 )
    	self.button3.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

    	bSizer7.Add( self.button3, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


    	bSizer5.Add( bSizer7, 1, wx.EXPAND, 5 )

    	bSizer10 = wx.BoxSizer( wx.VERTICAL )


    	bSizer10.Add( ( 0, 20), 0, 0, 5 )

    	self.button4 = wx.Button( self, wx.ID_ANY, u"Executive\nSummary", wx.DefaultPosition, wx.Size( 180,100 ), 0 )
    	self.button4.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

    	bSizer10.Add( self.button4, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


    	bSizer5.Add( bSizer10, 1, wx.EXPAND, 5 )


    	bSizer257.Add( bSizer5, 1, wx.EXPAND, 5 )


    	bSizer257.Add( ( 0, 0), 1, wx.EXPAND, 5 )

    	bSizer91 = wx.BoxSizer( wx.HORIZONTAL )


    	bSizer91.Add( ( 0, 0), 1, wx.EXPAND, 5 )

    	self.button23 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button23.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer91.Add( self.button23, 0, wx.ALL, 5 )

    	self.button26 = wx.Button( self, wx.ID_ANY, u"Generate", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button26.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer91.Add( self.button26, 0, wx.ALL, 5 )

    	self.button24 = wx.Button( self, wx.ID_ANY, u"Check Report", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button24.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer91.Add( self.button24, 0, wx.ALL, 5 )

    	self.button25 = wx.Button( self, wx.ID_ANY, u"Upload", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button25.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer91.Add( self.button25, 0, wx.ALL, 5 )


    	bSizer257.Add( bSizer91, 0, wx.EXPAND, 5 )


    	bSizer1.Add( bSizer257, 1, wx.EXPAND, 5 )


    	self.SetSizer( bSizer1 )
    	self.Layout()

    	self.Centre( wx.BOTH )

    	# Connect Events
    	self.button1.Bind( wx.EVT_BUTTON, self.DetailnReportWindows )
    	self.button2.Bind( wx.EVT_BUTTON, self.ChecklistWindows )
    	self.button3.Bind( wx.EVT_BUTTON, self.AuditWindows )
    	self.button4.Bind( wx.EVT_BUTTON, self.ExecutiveWindows )
    	self.button23.Bind( wx.EVT_BUTTON, self.Reset )
    	self.button26.Bind( wx.EVT_BUTTON, self.Generate )
    	self.button24.Bind( wx.EVT_BUTTON, self.Check )
    	self.button25.Bind( wx.EVT_BUTTON, self.Upload )



def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Exterminate( self,event):
        event.Skip()

def DetailnReportWindows( self, event ):
        Detail = DetailsnReport(parent=self.button1)
        Detail.Show()
        #event.Skip()

def ChecklistWindows( self, event ):
        #event.Skip()
        Checklist = ChecklistMain(parent=self.button2)
        Checklist.Show()

def AuditWindows( self, event ):
        #event.Skip()
        Audit = AuditScore(parent=self.button3)
        Audit.Show()

def ExecutiveWindows( self, event ):
        #event.Skip()
        Executive = ExecutiveSummary(parent=self.button4)
        Executive.Show()

def Reset( self, event ):
	event.Skip()

def Generate( self, event ):
	event.Skip()

def Check( self, event ):
	event.Skip()

def Upload( self, event ):
	event.Skip()


###########################################################################
## Class DetailsnReport
###########################################################################

class DetailsnReport ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Detail And Report", pos = wx.DefaultPosition, size = wx.Size( 608,345 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

    	self.SetSizeHintsSz( wx.Size(608,345), wx.Size(608.345) )
    	self.SetForegroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_WINDOW ) )
    	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

    	bSizer8 = wx.BoxSizer( wx.VERTICAL )

    	bSizer8.SetMinSize( wx.Size( -1,500 ) )

    	bSizer8.Add( ( 0, 50), 0, 0, 5 )

    	bSizer9 = wx.BoxSizer( wx.HORIZONTAL )

    	self.text1 = wx.StaticText( self, wx.ID_ANY, u"Company Name ", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.text1.Wrap( -1 )
    	self.text1.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer9.Add( self.text1, 0, wx.ALL, 5 )

    	bSizer14 = wx.BoxSizer( wx.VERTICAL )

    	self.input1 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 200,-1 ), 0 )
    	bSizer14.Add( self.input1, 0, wx.ALL|wx.EXPAND, 5 )


    	bSizer9.Add( bSizer14, 1, 0, 5 )


    	bSizer8.Add( bSizer9, 0, wx.EXPAND, 5 )

    	bSizer11 = wx.BoxSizer( wx.HORIZONTAL )

    	self.text3 = wx.StaticText( self, wx.ID_ANY, u"Address \t\t\t\t ", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.text3.Wrap( -1 )
    	self.text3.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer11.Add( self.text3, 0, wx.ALL, 5 )

    	bSizer15 = wx.BoxSizer( wx.VERTICAL )

    	bSizer15.SetMinSize( wx.Size( 100,-1 ) )
    	self.input2 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 200,-1 ), 0 )
    	bSizer15.Add( self.input2, 0, wx.ALL|wx.EXPAND, 5 )


    	bSizer11.Add( bSizer15, 1, wx.EXPAND, 5 )


    	bSizer8.Add( bSizer11, 0, wx.EXPAND, 5 )

    	bSizer16 = wx.BoxSizer( wx.HORIZONTAL )


    	bSizer16.Add( ( 100, 0), 0, 0, 5 )

    	bSizer17 = wx.BoxSizer( wx.VERTICAL )

    	self.text5 = wx.StaticText( self, wx.ID_ANY, u"Contact No", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.text5.Wrap( -1 )
    	self.text5.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer17.Add( self.text5, 0, wx.ALL, 5 )

    	bSizer20 = wx.BoxSizer( wx.VERTICAL )

    	self.input3 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 200,-1 ), 0 )
    	bSizer20.Add( self.input3, 0, wx.ALL, 5 )

    	self.m_textCtrl4 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 200,-1 ), 0 )
    	bSizer20.Add( self.m_textCtrl4, 0, wx.ALL, 5 )

    	self.m_textCtrl5 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 200,-1 ), 0 )
    	bSizer20.Add( self.m_textCtrl5, 0, wx.ALL, 5 )


    	bSizer17.Add( bSizer20, 1, wx.EXPAND, 5 )


    	bSizer16.Add( bSizer17, 1, wx.EXPAND, 5 )


    	bSizer8.Add( bSizer16, 0, wx.EXPAND, 5 )

    	bSizer13 = wx.BoxSizer( wx.VERTICAL )


    	bSizer13.Add( ( 0, 10), 0, 0, 5 )


    	bSizer8.Add( bSizer13, 0, 0, 5 )

    	bSizer12 = wx.BoxSizer( wx.HORIZONTAL )


    	bSizer12.Add( ( 0, 0), 1, wx.EXPAND, 5 )

    	self.button6 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button6.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer12.Add( self.button6, 0, wx.ALL, 5 )

    	self.button5 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button5.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer12.Add( self.button5, 0, wx.ALL|wx.ALIGN_RIGHT, 5 )


    	bSizer8.Add( bSizer12, 1, wx.EXPAND|wx.ALIGN_RIGHT, 5 )


    	self.SetSizer( bSizer8 )
    	self.Layout()

    	self.Centre( wx.BOTH )

    	# Connect Events
    	self.input1.Bind( wx.EVT_TEXT, self.CompanyName )
    	self.input2.Bind( wx.EVT_TEXT, self.Address )
    	self.input3.Bind( wx.EVT_TEXT, self.Contact1 )
    	self.m_textCtrl4.Bind( wx.EVT_TEXT, self.Contact2 )
    	self.m_textCtrl5.Bind( wx.EVT_TEXT, self.Contact3 )
    	self.button6.Bind( wx.EVT_BUTTON, self.Reset )
    	self.button5.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def CompanyName( self, event ):
	event.Skip()

def Address( self, event ):
	event.Skip()

def Contact1( self, event ):
	event.Skip()

def Contact2( self, event ):
	event.Skip()

def Contact3( self, event ):
	event.Skip()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
	event.Skip()


###########################################################################
## Class ChecklistMain
###########################################################################

class ChecklistMain ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Type Of Checklist", pos = wx.DefaultPosition, size = wx.Size( 550,219 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

    	self.SetSizeHintsSz( wx.Size(550,219), wx.Size(550,219) )
    	self.SetForegroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_WINDOW ) )
    	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

    	bSizer24 = wx.BoxSizer( wx.VERTICAL )


    	bSizer24.Add( ( 0, 30), 0, 0, 5 )

    	self.text5 = wx.StaticText( self, wx.ID_ANY, u"Please Select The Format", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.text5.Wrap( -1 )
    	self.text5.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

    	bSizer24.Add( self.text5, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


    	bSizer24.Add( ( 0, 0), 0, 0, 5 )

    	bSizer25 = wx.BoxSizer( wx.HORIZONTAL )


    	bSizer25.Add( ( 70, 0), 0, 0, 5 )

    	self.button13 = wx.Button( self, wx.ID_ANY, u"Old", wx.DefaultPosition, wx.Size( 180,100 ), 0 )
    	self.button13.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

    	bSizer25.Add( self.button13, 0, wx.ALL, 5 )


    	bSizer25.Add( ( 20, 0), 0, 0, 5 )

    	self.button14 = wx.Button( self, wx.ID_ANY, u"New", wx.DefaultPosition, wx.Size( 180,100 ), 0 )
    	self.button14.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

    	bSizer25.Add( self.button14, 0, wx.ALL, 5 )


    	bSizer24.Add( bSizer25, 1, wx.EXPAND, 5 )


    	self.SetSizer( bSizer24 )
    	self.Layout()

    	self.Centre( wx.BOTH )

    	# Connect Events
    	self.button13.Bind( wx.EVT_BUTTON, self.Old )
    	self.button14.Bind( wx.EVT_BUTTON, self.New )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Old( self, event ):
        #event.Skip()
        OldFormat = ChecklistOld(parent=self.button13)
        OldFormat.Show()

def New( self, event ):
        #event.Skip()
        NewFormat = ChecklistNew(parent=self.button14)
        NewFormat.Show()

###########################################################################
## Class ChecklistOld
###########################################################################

class ChecklistOld ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Checklist Old", pos = wx.DefaultPosition, size = wx.Size( 371,480 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

    	self.SetSizeHintsSz( wx.Size(371,480), wx.Size(371,480) )
    	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

    	bSizer20 = wx.BoxSizer( wx.VERTICAL )


    	bSizer20.Add( ( 0, 20), 0, 0, 5 )

    	bSizer253 = wx.BoxSizer( wx.HORIZONTAL )


    	bSizer253.Add( ( 80, 0), 0, 0, 5 )

    	self.staticText222 = wx.StaticText( self, wx.ID_ANY, u"State Date", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText222.Wrap( -1 )
    	self.staticText222.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

    	bSizer253.Add( self.staticText222, 0, wx.ALL, 5 )

    	self.datePicker1 = wx.adv.DatePickerCtrl( self, wx.ID_ANY, wx.DefaultDateTime, wx.DefaultPosition, wx.DefaultSize, wx.adv.DP_DEFAULT )
    	self.datePicker1.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

    	bSizer253.Add( self.datePicker1, 0, wx.ALL, 5 )


    	bSizer20.Add( bSizer253, 1, wx.EXPAND, 5 )

    	bSizer21 = wx.BoxSizer( wx.VERTICAL )

    	self.button9 = wx.Button( self, wx.ID_ANY, u"Licences and Permit", wx.DefaultPosition, wx.Size( 200,60 ), 0 )
    	self.button9.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

    	bSizer21.Add( self.button9, 1, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


    	bSizer21.Add( ( 0, 5), 0, 0, 5 )

    	self.button10 = wx.Button( self, wx.ID_ANY, u"Best industries Practices", wx.DefaultPosition, wx.Size( 200,60 ), 0 )
    	self.button10.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

    	bSizer21.Add( self.button10, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


    	bSizer21.Add( ( 0, 5), 0, 0, 5 )

    	self.button11 = wx.Button( self, wx.ID_ANY, u"Remanufacturing Work Flow\nSystem and S.O.P", wx.DefaultPosition, wx.Size( 200,60 ), 0 )
    	self.button11.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

    	bSizer21.Add( self.button11, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


    	bSizer21.Add( ( 0, 5), 0, 0, 5 )

    	self.button12 = wx.Button( self, wx.ID_ANY, u"Standard Compliance To\nVehicle Type Approval (VTA)\n-Safety Parts-", wx.DefaultPosition, wx.Size( 200,60 ), 0 )
    	self.button12.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

    	bSizer21.Add( self.button12, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


    	bSizer20.Add( bSizer21, 1, wx.EXPAND, 5 )

    	bSizer90 = wx.BoxSizer( wx.HORIZONTAL )


    	bSizer90.Add( ( 0, 0), 1, wx.EXPAND, 5 )

    	self.button21 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button21.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

    	bSizer90.Add( self.button21, 0, wx.ALL, 5 )


    	bSizer20.Add( bSizer90, 1, wx.EXPAND, 5 )


    	self.SetSizer( bSizer20 )
    	self.Layout()

    	self.Centre( wx.BOTH )

    	# Connect Events
    	self.datePicker1.Bind( wx.adv.EVT_DATE_CHANGED, self.Date1 )
    	self.button9.Bind( wx.EVT_BUTTON, self.License )
    	self.button10.Bind( wx.EVT_BUTTON, self.Best )
    	self.button11.Bind( wx.EVT_BUTTON, self.Remanufacturing )
    	self.button12.Bind( wx.EVT_BUTTON, self.Standard )
    	self.button21.Bind( wx.EVT_BUTTON, self.Reset )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Date1( self, event ):
	event.Skip()

def License( self, event ):
        #event.Skip()
        LicenseForm = PartA(parent = self.button9)
        LicenseForm.Show()

def Best( self, event ):
        #event.Skip()
        BestForm = PartB(parent = self.button10)
        BestForm.Show()

def Remanufacturing( self, event ):
        #event.Skip()
        RemanufacturingForm = PartC(parent = self.button11)
        RemanufacturingForm.Show()

def Standard( self, event ):
        #event.Skip()
        StandardForm = PartD(parent = self.button12)
        StandardForm.Show()

def Reset( self, event ):
	event.Skip()


###########################################################################
## Class PartA
###########################################################################

class PartA ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Licences And Permit", pos = wx.DefaultPosition, size = wx.Size( 479,355 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(479,355), wx.Size(479,355) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )

        bSizer23 = wx.BoxSizer( wx.VERTICAL )

        self.staticText52 = wx.StaticText( self, wx.ID_ANY, u"AVAILABILITY OF LICENCES AND PERMIT", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText52.Wrap( -1 )
        self.staticText52.SetFont( wx.Font( 16, 74, 90, 92, False, "Arial" ) )

        bSizer23.Add( self.staticText52, 0, wx.ALL, 5 )


        bSizer23.Add( ( 0, 10), 0, 0, 5 )

        bSizer30 = wx.BoxSizer( wx.VERTICAL )

        bSizer31 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer31.Add( ( 50, 0), 0, wx.ALIGN_CENTER_VERTICAL, 5 )

        bSizer24 = wx.BoxSizer( wx.VERTICAL )

        self.text5 = wx.StaticText( self, wx.ID_ANY, u"Number of AP allocated by MITI", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.text5.Wrap( -1 )
        self.text5.SetFont( wx.Font( 14, 74, 90, 92, False, "Arial" ) )

        bSizer24.Add( self.text5, 0, wx.ALL, 5 )

        bSizer25 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer25.Add( ( 20, 0), 0, 0, 5 )

        self.checkbox1 = wx.CheckBox( self, wx.ID_ANY, u"The number of AP used to import vehicle", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkbox1.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

        bSizer25.Add( self.checkbox1, 0, wx.ALL, 5 )


        bSizer24.Add( bSizer25, 0, 0, 5 )

        self.Text6 = wx.StaticText( self, wx.ID_ANY, u"Manufacturing Licence (ML) from MITI", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.Text6.Wrap( -1 )
        self.Text6.SetFont( wx.Font( 14, 74, 90, 92, False, "Arial" ) )

        bSizer24.Add( self.Text6, 0, wx.ALL, 5 )

        bSizer27 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer27.Add( ( 20, 0), 0, 0, 5 )

        self.checkBox4 = wx.CheckBox( self, wx.ID_ANY, u"The business activities match the ML", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox4.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

        bSizer27.Add( self.checkBox4, 0, wx.ALL, 5 )


        bSizer24.Add( bSizer27, 0, 0, 5 )

        self.m_staticText7 = wx.StaticText( self, wx.ID_ANY, u"Licenses from related local authorities (1,3,4,5,6)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText7.Wrap( -1 )
        self.m_staticText7.SetFont( wx.Font( 14, 74, 90, 92, False, "Arial" ) )

        bSizer24.Add( self.m_staticText7, 0, wx.ALL, 5 )

        bSizer28 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer28.Add( ( 20, 0), 0, 0, 5 )

        bSizer29 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox6 = wx.CheckBox( self, wx.ID_ANY, u"Business License from PBT & SSM registration", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox6.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

        bSizer29.Add( self.checkBox6, 0, wx.ALL, 5 )

        self.checkBox7 = wx.CheckBox( self, wx.ID_ANY, u"License/Certificate from Customs & GST", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox7.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

        bSizer29.Add( self.checkBox7, 0, wx.ALL, 5 )


        bSizer28.Add( bSizer29, 1, wx.EXPAND, 5 )


        bSizer24.Add( bSizer28, 0, 0, 5 )


        bSizer31.Add( bSizer24, 0, 0, 5 )


        bSizer30.Add( bSizer31, 1, wx.EXPAND, 5 )


        bSizer23.Add( bSizer30, 1, wx.EXPAND, 5 )

        bSizer84 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer84.Add( ( 0, 0), 1, 0, 5 )

        self.button13 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button13.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer84.Add( self.button13, 0, wx.ALL, 5 )

        self.button14 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button14.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer84.Add( self.button14, 0, wx.ALL, 5 )


        bSizer23.Add( bSizer84, 0, wx.EXPAND, 5 )


        self.SetSizer( bSizer23 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.checkbox1.Bind( wx.EVT_CHECKBOX, self.Check1 )
        self.checkBox4.Bind( wx.EVT_CHECKBOX, self.Check2 )
        self.checkBox6.Bind( wx.EVT_CHECKBOX, self.Check3 )
        self.checkBox7.Bind( wx.EVT_CHECKBOX, self.Check4 )
        self.button13.Bind( wx.EVT_BUTTON, self.Reset )
        self.button14.Bind( wx.EVT_BUTTON, self.Save )
def __del__( self ):
		pass



##	def Check1( self, event ):
def Check1( self, event):
        #event.Skip()
        tick1 = LoadWorkbook1(parent=self.checkbox1)
        mtick1 = tick1.wb1.get_sheet_by_name('Audit Checklist Q2')
        put1 = mtick1['C7']
        box1 = self.checkbox1.GetValue()

        if box1 == True:
            put1.value = "/"
        else:
            put1.value = ""

def Check2( self, event ):
        #event.Skip()
        tick2 = LoadWorkbook1(parent=self.checkBox4)
        mtick2 = tick2.wb1.get_sheet_by_name('Audit Checklist Q2')
        put2 = mtick2['C10']
        box2 = self.checkBox4.GetValue()

        if box2 == True:
            put2.value = "/"
        else:
            put2.value = ""

def Check3( self, event ):
        #event.Skip()
        tick3 = LoadWorkbook1(parent=self.checkBox6)
        mtick3 = tick3.wb1.get_sheet_by_name('Audit Checklist Q2')
        put3 = mtick3['C13']
        box3 = self.checkBox6.GetValue()

        if box3 == True:
            put3.value = "/"
        else:
            put3.value = ""

def Check4( self, event ):
        #event.Skip()
        tick4 = LoadWorkbook1(parent=self.checkBox7)
        mtick4 = tick4.wb1.get_sheet_by_name('Audit Checklist Q2')
        put4 = mtick4['C14']
        box4 = self.checkBox7.GetValue()

        if box4 == True:
            put4.value = "/"
        else:
            put4.value = ""

def Reset( self, event ):
        #event.Skip()
        self.checkbox1.SetValue(False)
        self.checkBox4.SetValue(False)
        self.checkBox6.SetValue(False)
        self.checkBox7.SetValue(False)


def Save( self, event):
        #event.Skip()
        check1 = self.checkbox1.GetValue()
        check2 = self.checkBox4.GetValue()
        check3 = self.checkBox6.GetValue()
        check4 = self.checkBox7.GetValue()

        if check1 == True:
            a1 = 1
        else:
            a1 = 0

        if check2 == True:
            a2 = 1
        else:
            a2 = 0

        if check3 == True:
            a3 = 1
        else:
            a3 = 0

        if check4 == True:
            a4 = 1
        else:
            a4 = 0


###########################################################################
## Class PartB
###########################################################################

class PartB ( wx.Frame ):

    def __init__( self, parent):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Best Industries Practices", pos = wx.DefaultPosition, size = wx.Size( 1180,772 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
    
    	self.SetSizeHintsSz( wx.Size(1180,772), wx.Size(1180,772) )
    	self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
    
    	bSizer32 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer89 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"COMPLIANCES OF BEST INDUSTRY PRACTICES", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText51.Wrap( -1 )
    	self.staticText51.SetFont( wx.Font( 18, 70, 90, 92, False, "Arial" ) )
    
    	bSizer89.Add( self.staticText51, 0, wx.ALL, 5 )
    
    
    	bSizer32.Add( bSizer89, 1, wx.EXPAND, 5 )
    
    	bSizer85 = wx.BoxSizer( wx.HORIZONTAL )
    
    	Left = wx.BoxSizer( wx.VERTICAL )
    
    	Man = wx.BoxSizer( wx.VERTICAL )
    
    	self.text8 = wx.StaticText( self, wx.ID_ANY, u"Man", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.text8.Wrap( -1 )
    	self.text8.SetFont( wx.Font( 16, 70, 90, 92, False, "Arial" ) )
    
    	Man.Add( self.text8, 0, wx.ALL, 5 )
    
    	bSizer48 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer48.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer47 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox9 = wx.CheckBox( self, wx.ID_ANY, u"Manpower Planning, *list of employee & Organisation Chart\n*to refer EPF forms (A KWSP) & SOCSO(8A -PERKESO) and document for foreign worker", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox9.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer47.Add( self.checkBox9, 0, wx.ALL, 5 )
    
    	self.checkBox10 = wx.CheckBox( self, wx.ID_ANY, u"Manpower qualifications", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox10.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer47.Add( self.checkBox10, 0, wx.ALL, 5 )
    
    	self.checkBox11 = wx.CheckBox( self, wx.ID_ANY, u"Manpower training and record", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox11.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer47.Add( self.checkBox11, 0, wx.ALL, 5 )
    
    	self.checkBox12 = wx.CheckBox( self, wx.ID_ANY, u"Automotive Engineer (In-house)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox12.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer47.Add( self.checkBox12, 0, wx.ALL, 5 )
    
    	self.checkBox13 = wx.CheckBox( self, wx.ID_ANY, u"Professional Engineer (PE) status", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox13.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer47.Add( self.checkBox13, 0, wx.ALL, 5 )
    
    
    	bSizer48.Add( bSizer47, 1, wx.EXPAND, 5 )
    
    
    	Man.Add( bSizer48, 1, wx.EXPAND, 5 )
    
    
    	Left.Add( Man, 1, wx.EXPAND, 5 )
    
    	Machine = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText9 = wx.StaticText( self, wx.ID_ANY, u"Machine", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText9.Wrap( -1 )
    	self.staticText9.SetFont( wx.Font( 16, 70, 90, 92, False, "Arial" ) )
    
    	Machine.Add( self.staticText9, 0, wx.ALL, 5 )
    
    	bSizer49 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer49.Add( ( 20, 0), 0, 0, 5 )
    
    	self.checkBox14 = wx.CheckBox( self, wx.ID_ANY, u"Equipment and Testing equiptment based on JPJ & SIRIM Audit result", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox14.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer49.Add( self.checkBox14, 0, wx.ALL, 5 )
    
    
    	Machine.Add( bSizer49, 1, wx.EXPAND, 5 )
    
    
    	Left.Add( Machine, 1, wx.EXPAND, 5 )
    
    
    	bSizer85.Add( Left, 1, wx.EXPAND, 5 )
    
    	Right = wx.BoxSizer( wx.VERTICAL )
    
    	Material = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText10 = wx.StaticText( self, wx.ID_ANY, u"Material", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText10.Wrap( -1 )
    	self.staticText10.SetFont( wx.Font( 16, 70, 90, 92, False, "Arial" ) )
    
    	Material.Add( self.staticText10, 0, wx.ALL, 5 )
    
    	bSizer51 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer51.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer50 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox15 = wx.CheckBox( self, wx.ID_ANY, u"Incoming parts control", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox15.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer50.Add( self.checkBox15, 0, wx.ALL, 5 )
    
    	self.checkBox16 = wx.CheckBox( self, wx.ID_ANY, u"Reject parts S.O.P", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox16.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer50.Add( self.checkBox16, 0, wx.ALL, 5 )
    
    	self.checkBox17 = wx.CheckBox( self, wx.ID_ANY, u"Storage Management", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox17.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer50.Add( self.checkBox17, 0, wx.ALL, 5 )
    
    	self.checkBox18 = wx.CheckBox( self, wx.ID_ANY, u"Regulatory & Statutory requirements", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox18.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer50.Add( self.checkBox18, 0, wx.ALL, 5 )
    
    
    	bSizer51.Add( bSizer50, 1, wx.EXPAND, 5 )
    
    
    	Material.Add( bSizer51, 1, wx.EXPAND, 5 )
    
    
    	Right.Add( Material, 1, wx.EXPAND, 5 )
    
    	Method = wx.BoxSizer( wx.VERTICAL )
    
    	self.m_staticText11 = wx.StaticText( self, wx.ID_ANY, u"Method", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.m_staticText11.Wrap( -1 )
    	self.m_staticText11.SetFont( wx.Font( 16, 70, 90, 92, False, "Arial" ) )
    
    	Method.Add( self.m_staticText11, 0, wx.ALL, 5 )
    
    	bSizer52 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer52.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer39 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer40 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText12 = wx.StaticText( self, wx.ID_ANY, u"Process Flow & Control Plan", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText12.Wrap( -1 )
    	self.staticText12.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer40.Add( self.staticText12, 0, wx.ALL, 5 )
    
    	bSizer54 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer54.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer53 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox31 = wx.CheckBox( self, wx.ID_ANY, u"SOP/Work Instruction/Checksheet /Record", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox31.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer53.Add( self.checkBox31, 0, wx.ALL, 5 )
    
    
    	bSizer54.Add( bSizer53, 1, wx.EXPAND, 5 )
    
    
    	bSizer40.Add( bSizer54, 1, wx.EXPAND, 5 )
    
    
    	bSizer39.Add( bSizer40, 0, 0, 5 )
    
    	bSizer41 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText13 = wx.StaticText( self, wx.ID_ANY, u"Safety, Health & Environment", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText13.Wrap( -1 )
    	self.staticText13.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer41.Add( self.staticText13, 0, wx.ALL, 5 )
    
    	bSizer55 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer55.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer56 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox32 = wx.CheckBox( self, wx.ID_ANY, u"Manpower", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox32.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer56.Add( self.checkBox32, 0, wx.ALL, 5 )
    
    	self.checkBox33 = wx.CheckBox( self, wx.ID_ANY, u"Workplace", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox33.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer56.Add( self.checkBox33, 0, wx.ALL, 5 )
    
    	self.checkBox34 = wx.CheckBox( self, wx.ID_ANY, u"Environment", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox34.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer56.Add( self.checkBox34, 0, wx.ALL, 5 )
    
    
    	bSizer55.Add( bSizer56, 1, wx.EXPAND, 5 )
    
    
    	bSizer41.Add( bSizer55, 1, wx.EXPAND, 5 )
    
    
    	bSizer39.Add( bSizer41, 0, 0, 5 )
    
    	bSizer46 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText20 = wx.StaticText( self, wx.ID_ANY, u"Quality Management System - ISO 9001:2015", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText20.Wrap( -1 )
    	self.staticText20.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer46.Add( self.staticText20, 0, wx.ALL, 5 )
    
    	bSizer561 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer561.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer58 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox36 = wx.CheckBox( self, wx.ID_ANY, u"Registration", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox36.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer58.Add( self.checkBox36, 0, wx.ALL, 5 )
    
    	self.checkBox37 = wx.CheckBox( self, wx.ID_ANY, u"Training", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox37.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer58.Add( self.checkBox37, 0, wx.ALL, 5 )
    
    	self.checkBox38 = wx.CheckBox( self, wx.ID_ANY, u"Certification", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox38.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer58.Add( self.checkBox38, 0, wx.ALL, 5 )
    
    
    	bSizer561.Add( bSizer58, 1, wx.EXPAND, 5 )
    
    
    	bSizer46.Add( bSizer561, 1, wx.EXPAND, 5 )
    
    
    	bSizer39.Add( bSizer46, 1, wx.EXPAND, 5 )
    
    	bSizer42 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText15 = wx.StaticText( self, wx.ID_ANY, u"MAI 4R2S Industry Standards", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText15.Wrap( -1 )
    	self.staticText15.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer42.Add( self.staticText15, 0, wx.ALL, 5 )
    
    	bSizer57 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer57.Add( ( 20, 0), 0, 0, 5 )
    
    	self.checkBox35 = wx.CheckBox( self, wx.ID_ANY, u"Awareness", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox35.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer57.Add( self.checkBox35, 0, wx.ALL, 5 )
    
    
    	bSizer42.Add( bSizer57, 1, wx.EXPAND, 5 )
    
    
    	bSizer39.Add( bSizer42, 0, 0, 5 )
    
    	bSizer43 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText16 = wx.StaticText( self, wx.ID_ANY, u"Vehicle Testing Report on Road Worthiness Test (In House & PUSPAKOM)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText16.Wrap( -1 )
    	self.staticText16.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer43.Add( self.staticText16, 0, wx.ALL, 5 )
    
    	bSizer45 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer59 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText17 = wx.StaticText( self, wx.ID_ANY, u"Brake Test", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText17.Wrap( -1 )
    	self.staticText17.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer59.Add( self.staticText17, 0, wx.ALL, 5 )
    
    
    	bSizer59.Add( ( 100, 0), 1, 0, 5 )
    
    	bSizer62 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox183 = wx.CheckBox( self, wx.ID_ANY, u"InHouse", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox183.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer62.Add( self.checkBox183, 0, wx.ALL, 5 )
    
    	self.checkBox184 = wx.CheckBox( self, wx.ID_ANY, u"External (PUSPAKOM)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox184.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer62.Add( self.checkBox184, 0, wx.ALL, 5 )
    
    
    	bSizer59.Add( bSizer62, 1, wx.EXPAND, 5 )
    
    
    	bSizer45.Add( bSizer59, 1, wx.EXPAND, 5 )
    
    	bSizer60 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText18 = wx.StaticText( self, wx.ID_ANY, u"Side Slip Test", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText18.Wrap( -1 )
    	self.staticText18.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer60.Add( self.staticText18, 0, wx.ALL, 5 )
    
    
    	bSizer60.Add( ( 50, 0), 1, 0, 5 )
    
    	bSizer61 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox185 = wx.CheckBox( self, wx.ID_ANY, u"InHouse", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox185.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer61.Add( self.checkBox185, 0, wx.ALL, 5 )
    
    	self.checkBox186 = wx.CheckBox( self, wx.ID_ANY, u"External (PUSPAKOM)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox186.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer61.Add( self.checkBox186, 0, wx.ALL, 5 )
    
    
    	bSizer60.Add( bSizer61, 1, wx.EXPAND, 5 )
    
    
    	bSizer45.Add( bSizer60, 1, wx.EXPAND, 5 )
    
    	bSizer63 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText19 = wx.StaticText( self, wx.ID_ANY, u"Smoke Emission Test", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText19.Wrap( -1 )
    	self.staticText19.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer63.Add( self.staticText19, 0, wx.ALL, 5 )
    
    	bSizer64 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer64.Add( ( 0, 0), 1, wx.EXPAND, 5 )
    
    	self.checkBox187 = wx.CheckBox( self, wx.ID_ANY, u"InHouse", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox187.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer64.Add( self.checkBox187, 0, wx.ALL, 5 )
    
    	self.checkBox188 = wx.CheckBox( self, wx.ID_ANY, u"External (PUSPAKOM)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox188.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer64.Add( self.checkBox188, 0, wx.ALL, 5 )
    
    
    	bSizer63.Add( bSizer64, 1, wx.EXPAND, 5 )
    
    
    	bSizer45.Add( bSizer63, 1, wx.EXPAND, 5 )
    
    
    	bSizer43.Add( bSizer45, 1, wx.EXPAND, 5 )
    
    
    	bSizer39.Add( bSizer43, 1, wx.EXPAND, 5 )
    
    
    	bSizer52.Add( bSizer39, 1, wx.EXPAND, 5 )
    
    
    	Method.Add( bSizer52, 1, wx.EXPAND, 5 )
    
    
    	Right.Add( Method, 1, wx.EXPAND, 5 )
    
    
    	bSizer85.Add( Right, 1, wx.EXPAND, 5 )
    
    
    	bSizer32.Add( bSizer85, 1, wx.EXPAND, 5 )
    
    	bSizer86 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer86.Add( ( 0, 0), 1, 0, 5 )
    
    	self.button15 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button15.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.button15, 0, wx.ALL, 5 )
    
    	self.button16 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button16.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.button16, 0, wx.ALL, 5 )
    
    
    	bSizer32.Add( bSizer86, 0, wx.EXPAND, 5 )
    
    
    	self.SetSizer( bSizer32 )
    	self.Layout()
    
    	self.Centre( wx.BOTH )
    
    	# Connect Events
    	self.checkBox9.Bind( wx.EVT_CHECKBOX, self.Check1 )
    	self.checkBox10.Bind( wx.EVT_CHECKBOX, self.Check2 )
    	self.checkBox11.Bind( wx.EVT_CHECKBOX, self.Check3 )
    	self.checkBox12.Bind( wx.EVT_CHECKBOX, self.Check4 )
    	self.checkBox13.Bind( wx.EVT_CHECKBOX, self.Check5 )
    	self.checkBox14.Bind( wx.EVT_CHECKBOX, self.Check6 )
    	self.checkBox15.Bind( wx.EVT_CHECKBOX, self.Check7 )
    	self.checkBox16.Bind( wx.EVT_CHECKBOX, self.Check8 )
    	self.checkBox17.Bind( wx.EVT_CHECKBOX, self.Check9 )
    	self.checkBox18.Bind( wx.EVT_CHECKBOX, self.Check10 )
    	self.checkBox31.Bind( wx.EVT_CHECKBOX, self.Check11 )
    	self.checkBox32.Bind( wx.EVT_CHECKBOX, self.Check12 )
    	self.checkBox33.Bind( wx.EVT_CHECKBOX, self.Check13 )
    	self.checkBox34.Bind( wx.EVT_CHECKBOX, self.Check14 )
    	self.checkBox36.Bind( wx.EVT_CHECKBOX, self.Check15 )
    	self.checkBox37.Bind( wx.EVT_CHECKBOX, self.Check16 )
    	self.checkBox38.Bind( wx.EVT_CHECKBOX, self.Check17 )
    	self.checkBox35.Bind( wx.EVT_CHECKBOX, self.Check18 )
    	self.checkBox183.Bind( wx.EVT_CHECKBOX, self.Check19 )
    	self.checkBox184.Bind( wx.EVT_CHECKBOX, self.Check20 )
    	self.checkBox185.Bind( wx.EVT_CHECKBOX, self.Check21 )
    	self.checkBox186.Bind( wx.EVT_CHECKBOX, self.Check22 )
    	self.checkBox187.Bind( wx.EVT_CHECKBOX, self.Check23 )
    	self.checkBox188.Bind( wx.EVT_CHECKBOX, self.Check24 )
    	self.button15.Bind( wx.EVT_BUTTON, self.Reset )
    	self.button16.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
    	pass
    
    
    # Virtual event handlers, overide them in your derived class
    def Check1( self, event ):
            #event.Skip()
            tick5 = LoadWorkbook1(parent=self.checkBox9)
            mtick5 = tick5.wb1.get_sheet_by_name('Audit Checklist Q2')
            put5 = mtick5['C19']
            box5 = self.checkBox9.GetValue()
    
            if box5 == True:
                put5.value = "/"
            else:
                put5.value = ""
    
    def Check2( self, event ):
            #event.Skip()
            tick6 = LoadWorkbook1(parent=self.checkBox10)
            mtick6 = tick6.wb1.get_sheet_by_name('Audit Checklist Q2')
            put6 = mtick6['C21']
            box6 = self.checkBox10.GetValue()
    
            if box6 == True:
                put6.value = "/"
            else:
                put6.value = ""
    
    def Check3( self, event ):
            #event.Skip()
            tick7 = LoadWorkbook1(parent=self.checkBox11)
            mtick7 = tick7.wb1.get_sheet_by_name('Audit Checklist Q2')
            put7 = mtick7['C22']
            box7 = self.checkBox11.GetValue()
    
            if box7 == True:
                put7.value = "/"
            else:
                put7.value = ""
    
    
    def Check4( self, event ):
            #event.Skip()
            tick8 = LoadWorkbook1(parent=self.checkBox12)
            mtick8 = tick8.wb1.get_sheet_by_name('Audit Checklist Q2')
            put8 = mtick8['C23']
    
            box8 = self.checkBox12.GetValue()
    
            if box8 == True:
                put8.value = "/"
            else:
                put8.value = ""
    
    def Check5( self, event ):
            #event.Skip()
            tick9 = LoadWorkbook1(parent=self.checkBox13)
            mtick9 = tick9.wb1.get_sheet_by_name('Audit Checklist Q2')
            put9 = mtick9['C24']
            box9 = self.checkBox13.GetValue()
    
            if box9 == True:
                put9.value = "/"
            else:
                put9.value = ""
    
    
    def Check6( self, event ):
            #event.Skip()
            tick10 = LoadWorkbook1(parent=self.checkBox14)
            mtick10 = tick10.wb1.get_sheet_by_name('Audit Checklist Q2')
            put10 = mtick10['C27']
            box10 = self.checkBox14.GetValue()
    
            if box10 == True:
                put10.value = "/"
            else:
                put10.value = ""
    
    def Check7( self, event ):
            #event.Skip()
            tick11 = LoadWorkbook1(parent=self.checkBox15)
            mtick11 = tick11.wb1.get_sheet_by_name('Audit Checklist Q2')
            put11 = mtick11['C30']
            box11 = self.checkBox15.GetValue()
    
            if box11 == True:
                put11.value = "/"
            else:
                put11.value = ""
    
    def Check8( self, event ):
            #event.Skip()
            tick12 = LoadWorkbook1(parent=self.checkBox16)
            mtick12 = tick12.wb1.get_sheet_by_name('Audit Checklist Q2')
            put12 = mtick12['C31']
            box12 = self.checkBox16.GetValue()
    
            if box12 == True:
                put12.value = "/"
            else:
                put12.value = ""
    
    def Check9( self, event ):
            #event.Skip()
            global team13
            tick13 = LoadWorkbook1(parent=self.checkBox17)
            mtick13 = tick13.wb1.get_sheet_by_name('Audit Checklist Q2')
            put13 = mtick13['C32']
            box13 = self.checkBox17.GetValue()
    
            if box13 == True:
                put13.value = "/"
            else:
                put13.value = ""
    
    def Check10( self, event ):
            #event.Skip()
            tick14 = LoadWorkbook1(parent=self.checkBox18)
            mtick14 = tick14.wb1.get_sheet_by_name('Audit Checklist Q2')
            put14 = mtick14['C33']
            box14 = self.checkBox18.GetValue()
    
            if box14 == True:
                put14.value = "/"
            else:
                put14.value = ""
    
    def Check11( self, event ):
            #event.Skip()
            global team15
            tick15 = LoadWorkbook1(parent=self.checkBox31)
            mtick15 = tick15.wb1.get_sheet_by_name('Audit Checklist Q2')
            put15 = mtick15['C36']
            box15 = self.checkBox31.GetValue()
    
            if box15 == True:
                put15.value = "/"
            else:
                put15.value = ""
    
    def Check12( self, event ):
            #event.Skip()
            tick16 = LoadWorkbook1(parent=self.checkBox32)
            mtick16 = tick16.wb1.get_sheet_by_name('Audit Checklist Q2')
            put16 = mtick16['C40']
            box16 = self.checkBox32
    
            if box16 == True:
                put16.value = "/"
            else:
                put16.value = ""
    
    def Check13( self, event ):
            #event.Skip()
            tick17 = LoadWorkbook1(parent=self.checkBox33)
            mtick17 = tick17.wb1.get_sheet_by_name('Audit Checklist Q2')
            put17 = mtick17['C41']
            box17 = self.checkBox33.GetValue()
    
            if box17 == True:
                put17.value = "/"
            else:
                put17.value = ""
    
    
    def Check14( self, event ):
            #event.Skip()
            tick18 = LoadWorkbook1(parent=self.checkBox34)
            mtick18 = tick18.wb1.get_sheet_by_name('Audit Checklist Q2')
            put18 = mtick18['C42']
            box18 = self.checkBox34.GetValue()
    
            if box18 == True:
                put18.value = "/"
            else:
                put18.value = ""
    
    def Check15( self, event ):
            #event.Skip()
            tick19 = LoadWorkbook1(parent=self.checkBox36)
            mtick19 = tick19.wb1.get_sheet_by_name('Audit Checklist Q2')
            put19 = mtick19['C45']
            box19 = self.checkBox36.GetValue()
    
            if box19 == True:
                put19.value = "/"
            else:
                put19.value = ""
    
    def Check16( self, event ):
            #event.Skip()
            tick20 = LoadWorkbook1(parent=self.checkBox37)
            mtick20 = tick20.wb1.get_sheet_by_name('Audit Checklist Q2')
            put20 = mtick20['C46']
            box20 = self.checkBox37.GetValue()
    
            if box20 == True:
                put20.value = "/"
            else:
                put20.value = ""
    
    def Check17( self, event ):
            #event.Skip()
            tick21 = LoadWorkbook1(parent=self.checkBox38)
            mtick21 = tick21.wb1.get_sheet_by_name('Audit Checklist Q2')
            put21 = mtick21['C47']
            box21 = self.checkBox38.GetValue()
    
            if box21 == True:
                put21.value = "/"
            else:
                put21.value = ""
    
    def Check18( self, event ):
            #event.Skip()
            tick22 = LoadWorkbook1(parent=self.checkBox35)
            mtick22 = tick22.wb1.get_sheet_by_name('Audit Checklist Q2')
            put22 = mtick22['C51']
            box22 = self.checkBox35.GetValue()
    
            if box22 == True:
                put22.value = "/"
            else:
                put22.value = ""
    
    def Check19( self, event ):
            #event.Skip()
            tick23 = LoadWorkbook1(parent=self.checkBox183)
            mtick23 = tick23.wb1.get_sheet_by_name('Audit Checklist Q2')
            put23 = mtick23['C54']
            box23 = self.checkBox183.GetValue()
    
            if box23 == True:
                put23.value = "/"
            else:
                put23.value = ""
    
    
    def Check20( self, event ):
            #event.Skip()
            tick24 = LoadWorkbook1(parent=self.checkBox184)
            mtick24 = tick24.wb1.get_sheet_by_name('Audit Checklist Q2')
            put24 = mtick24['D54']
            box24 = self.checkBox184.GetValue()
    
            if box24 == True:
                put24.value = "/"
            else:
                put24.value = ""
    
    def Check21( self, event ):
            #event.Skip()
            tick25 = LoadWorkbook1(parent=self.checkBox185)
            mtick25 = tick25.wb1.get_sheet_by_name('Audit Checklist Q2')
            put25 = mtick25['C55']
            box25 = self.checkBox185.GetValue()
    
            if box25 == True:
                put25.value = "/"
            else:
                put25.value = ""
    
    def Check22( self, event ):
            #event.Skip()
            tick26 = LoadWorkbook1(parent=self.checkBox186)
            mtick26 = tick26.wb1.get_sheet_by_name('Audit Checklist Q2')
            put26 = mtick26['D55']
            box26 = self.checkBox186.GetValue()
    
            if box26 == True:
                put26.value = "/"
            else:
                put26.value = ""
    
    def Check23( self, event ):
            #event.Skip()
            tick27 = LoadWorkbook1(parent=self.checkBox187)
            mtick27 = tick27.wb1.get_sheet_by_name('Audit Checklist Q2')
            put27 = mtick27['C56']
            box27 = self.checkBox187.GetValue()
    
            if box27 == True:
                put27.value = "/"
            else:
                put27.value = ""
    
    def Check24( self, event ):
            #event.Skip()
            tick28 = LoadWorkbook1(parent=self.checkBox188)
            mtick28 = tick28.wb1.get_sheet_by_name('Audit Checklist Q2')
            put28 = mtick28['D56']
            box28 = self.checkBox188.GetValue()
    
            if box28 == True:
                put28.value = "/"
            else:
                put28.value = ""
    
    
    def Reset( self, event ):
            #event.Skip()
            self.checkBox9.SetValue(False)
            self.checkBox10.SetValue(False)
            self.checkBox11.SetValue(False)
            self.checkBox12.SetValue(False)
            self.checkBox13.SetValue(False)
            self.checkBox14.SetValue(False)
            self.checkBox15.SetValue(False)
            self.checkBox16.SetValue(False)
            self.checkBox17.SetValue(False)
            self.checkBox18.SetValue(False)
            self.checkBox31.SetValue(False)
            self.checkBox32.SetValue(False)
            self.checkBox33.SetValue(False)
            self.checkBox34.SetValue(False)
            self.checkBox36.SetValue(False)
            self.checkBox37.SetValue(False)
            self.checkBox38.SetValue(False)
            self.checkBox35.SetValue(False)
            self.checkBox183.SetValue(False)
            self.checkBox184.SetValue(False)
            self.checkBox185.SetValue(False)
            self.checkBox186.SetValue(False)
            self.checkBox187.SetValue(False)
            self.checkBox188.SetValue(False)
    
    def Save( self, event ):
            #event.Skip()
            check5 = self.checkBox9.GetValue()
            check6 = self.checkBox10.GetValue()
            check7 = self.checkBox11.GetValue()
            check8 = self.checkBox12.GetValue()
            check9 = self.checkBox13.GetValue()
            check10 = self.checkBox14.GetValue()
            check11 = self.checkBox15.GetValue()
            check12 = self.checkBox16.GetValue()
            check13 = self.checkBox17.GetValue()
            check14 = self.checkBox18.Getvalue()
            check15 = self.checkBox31.Getvalue()
            check16 = self.checkBox32.GetValue()
            check17 = self.checkBox33.GetValue()
            check18 = self.checkBox34.GetValue()
            check19 = self.checkBox36.GetValue()
            check20 = self.checkBox37.GetValue()
            check21 = self.checkBox38.GetValue()
            check22 = self.checkBox35.GetValue()
            check23 = self.checkBox183.GetValue()
            check24 = self.checkBox184.GetValue()
            check25 = self.checkBox185.GetValue()
            check26 = self.checkBox186.GetValue()
            check27 = self.checkBox187.GetValue()
            check28 = self.checkBox188.GetValue()
    
            if check5 == True:
                a5 = 1
            else:
                a5 = 0
    
            if check6 == True:
                a6 = 1
            else:
                a6 = 0
    
            if check7 == True:
                a7 = 1
            else:
                a7 = 0
    
            if check8 == True:
                a8 = 1
            else:
                a8 = 0
    
            if check9 == True:
                a9 = 1
            else:
                a9 = 0
    
            if check10 == True:
                a10 = 1
            else:
                a10 = 0
    
            if check11 == True:
                a11 = 1
            else:
                a11 = 0
    
            if check12 == True:
                a12 = 1
            else:
                a12 = 0
    
            if check13 == True:
                a13 = 1
            else:
                a13 = 0
    
            if check14 == True:
                a14 = 1
            else:
                a14 = 0
    
            if check15 == True:
                a15 = 1
            else:
                a15 = 0
    
            if check16 == True:
                a16 = 1
            else:
                a16 = 0
    
            if check17 == True:
                a17 = 1
            else:
                a17 = 0
    
            if check18 == True:
                a18 = 1
            else:
                a18 = 0
    
            if check19 == True:
                a19 = 1
            else:
                a19 = 0
    
            if check20 == True:
                a20 = 1
            else:
                a20 = 0
    
            if check21 == True:
                a21 = 1
            else:
                a21 = 0
    
            if check22 == True:
                a22 = 1
            else:
                a22 = 0
    
            if check23 == True:
                a23 = 1
            else:
                a23 = 0
    
            if check24 == True:
                a24 = 1
            else:
                a24 = 0
    
            if check25 == True:
                a25 = 1
            else:
                a25 = 0
    
            if check26 == True:
                a26 = 1
            else:
                a26 = 0
    
            if check27 == True:
                a27 = 1
            else:
                a27 = 0
    
            if check28 == True:
                a28 = 1
            else:
                a28 = 0





###########################################################################
## Class PartC
###########################################################################

class PartC ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Reamanufacturing Work Flow System And SOP", pos = wx.DefaultPosition, size = wx.Size( 764,591 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
    
    	self.SetSizeHintsSz( wx.Size(764,561), wx.Size(764,561) )
    	self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
    
    	bSizer65 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer88 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText50 = wx.StaticText( self, wx.ID_ANY, u"COMPLIANCES OF REMANUFACTURING WORK FLOW SYSTEM & S.O.P", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText50.Wrap( -1 )
    	self.staticText50.SetFont( wx.Font( 14, 70, 90, 92, False, "Arial" ) )
    
    	bSizer88.Add( self.staticText50, 0, wx.ALL, 5 )
    
    
    	bSizer65.Add( bSizer88, 1, wx.EXPAND, 5 )
    
    	bSizer68 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer68.Add( ( 270, 0), 1, 0, 5 )
    
    	self.staticText32 = wx.StaticText( self, wx.ID_ANY, u"Brake System", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText32.Wrap( -1 )
    	self.staticText32.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer68.Add( self.staticText32, 0, wx.ALL, 5 )
    
    
    	bSizer68.Add( ( 0, 0), 1, 0, 5 )
    
    	self.staticText33 = wx.StaticText( self, wx.ID_ANY, u"Intercooler", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText33.Wrap( -1 )
    	self.staticText33.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer68.Add( self.staticText33, 0, wx.ALL, 5 )
    
    
    	bSizer68.Add( ( 0, 0), 1, 0, 5 )
    
    	self.staticText34 = wx.StaticText( self, wx.ID_ANY, u"Turbocharger", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText34.Wrap( -1 )
    	self.staticText34.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer68.Add( self.staticText34, 0, wx.ALL, 5 )
    
    
    	bSizer68.Add( ( 0, 0), 1, 0, 5 )
    
    	self.staticText35 = wx.StaticText( self, wx.ID_ANY, u"Stater Motor", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText35.Wrap( -1 )
    	self.staticText35.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer68.Add( self.staticText35, 0, wx.ALL, 5 )
    
    
    	bSizer68.Add( ( 0, 0), 1, 0, 5 )
    
    	self.staticText36 = wx.StaticText( self, wx.ID_ANY, u"Alternator", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText36.Wrap( -1 )
    	self.staticText36.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer68.Add( self.staticText36, 0, wx.ALL, 5 )
    
    
    	bSizer68.Add( ( 0, 0), 1, 0, 5 )
    
    	self.staticText37 = wx.StaticText( self, wx.ID_ANY, u"Radiator", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText37.Wrap( -1 )
    	self.staticText37.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer68.Add( self.staticText37, 0, wx.ALL, 5 )
    
    
    	bSizer68.Add( ( 0, 0), 1, 0, 5 )
    
    
    	bSizer65.Add( bSizer68, 0, wx.EXPAND, 5 )
    
    	bSizer66 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer74 = wx.BoxSizer( wx.HORIZONTAL )
    
    	bSizer73 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText21 = wx.StaticText( self, wx.ID_ANY, u"Core Management", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText21.Wrap( -1 )
    	self.staticText21.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText21, 0, wx.ALL, 5 )
    
    	self.staticText22 = wx.StaticText( self, wx.ID_ANY, u"Sorting", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText22.Wrap( -1 )
    	self.staticText22.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText22, 0, wx.ALL, 5 )
    
    	self.staticText23 = wx.StaticText( self, wx.ID_ANY, u"Disassembly of parts and components", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText23.Wrap( -1 )
    	self.staticText23.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText23, 0, wx.ALL, 5 )
    
    	self.staticText24 = wx.StaticText( self, wx.ID_ANY, u"Cleaning of parts and components", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText24.Wrap( -1 )
    	self.staticText24.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText24, 0, wx.ALL, 5 )
    
    	self.staticText25 = wx.StaticText( self, wx.ID_ANY, u"Machining and/or polishing", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText25.Wrap( -1 )
    	self.staticText25.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText25, 0, wx.ALL, 5 )
    
    	self.staticText26 = wx.StaticText( self, wx.ID_ANY, u"Inspection and Testing", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText26.Wrap( -1 )
    	self.staticText26.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText26, 0, wx.ALL, 5 )
    
    	self.staticText27 = wx.StaticText( self, wx.ID_ANY, u"Assembly", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText27.Wrap( -1 )
    	self.staticText27.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText27, 0, wx.ALL, 5 )
    
    	self.staticText28 = wx.StaticText( self, wx.ID_ANY, u"Painting/Polishing", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText28.Wrap( -1 )
    	self.staticText28.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText28, 0, wx.ALL, 5 )
    
    	self.staticText29 = wx.StaticText( self, wx.ID_ANY, u"Labelling", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText29.Wrap( -1 )
    	self.staticText29.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText29, 0, wx.ALL, 5 )
    
    	self.staticText31 = wx.StaticText( self, wx.ID_ANY, u"Warranty of parts (1 year waranty)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText31.Wrap( -1 )
    	self.staticText31.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer73.Add( self.staticText31, 0, wx.ALL, 5 )
    
    
    	bSizer74.Add( bSizer73, 0, wx.EXPAND, 5 )
    
    
    	bSizer74.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer72 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox34 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox34, 1, wx.ALL, 5 )
    
    	self.checkBox35 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox35, 1, wx.ALL, 5 )
    
    	self.checkBox36 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox36, 1, wx.ALL, 5 )
    
    	self.checkBox37 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox37, 1, wx.ALL, 5 )
    
    	self.checkBox38 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox38, 1, wx.ALL, 5 )
    
    	self.checkBox39 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox39, 1, wx.ALL, 5 )
    
    	self.checkBox40 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox40, 1, wx.ALL, 5 )
    
    	self.checkBox41 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox41, 1, wx.ALL, 5 )
    
    	self.checkBox42 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox42, 1, wx.ALL, 5 )
    
    	self.checkBox43 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer72.Add( self.checkBox43, 1, wx.ALL, 5 )
    
    
    	bSizer74.Add( bSizer72, 0, wx.EXPAND, 5 )
    
    
    	bSizer74.Add( ( 55, 0), 0, 0, 5 )
    
    	bSizer75 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox44 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox44, 1, wx.ALL, 5 )
    
    	self.checkBox45 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox45, 1, wx.ALL, 5 )
    
    	self.checkBox46 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox46, 1, wx.ALL, 5 )
    
    	self.checkBox47 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox47, 1, wx.ALL, 5 )
    
    	self.checkBox48 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox48, 1, wx.ALL, 5 )
    
    	self.checkBox49 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox49, 1, wx.ALL, 5 )
    
    	self.checkBox50 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox50, 1, wx.ALL, 5 )
    
    	self.checkBox51 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox51, 1, wx.ALL, 5 )
    
    	self.checkBox52 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox52, 1, wx.ALL, 5 )
    
    	self.checkBox53 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer75.Add( self.checkBox53, 1, wx.ALL, 5 )
    
    
    	bSizer74.Add( bSizer75, 0, wx.EXPAND, 5 )
    
    
    	bSizer74.Add( ( 55, 0), 0, 0, 5 )
    
    	bSizer77 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox54 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox54, 1, wx.ALL, 5 )
    
    	self.checkBox55 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox55, 1, wx.ALL, 5 )
    
    	self.checkBox56 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox56, 1, wx.ALL, 5 )
    
    	self.checkBox57 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox57, 1, wx.ALL, 5 )
    
    	self.checkBox58 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox58, 1, wx.ALL, 5 )
    
    	self.checkBox59 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox59, 1, wx.ALL, 5 )
    
    	self.checkBox60 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox60, 1, wx.ALL, 5 )
    
    	self.checkBox61 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox61, 1, wx.ALL, 5 )
    
    	self.checkBox62 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox62, 1, wx.ALL, 5 )
    
    	self.checkBox63 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer77.Add( self.checkBox63, 1, wx.ALL, 5 )
    
    
    	bSizer74.Add( bSizer77, 0, wx.EXPAND, 5 )
    
    
    	bSizer74.Add( ( 70, 0), 0, 0, 5 )
    
    	bSizer78 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox64 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox64, 1, wx.ALL, 5 )
    
    	self.checkBox65 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox65, 1, wx.ALL, 5 )
    
    	self.checkBox66 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox66, 1, wx.ALL, 5 )
    
    	self.checkBox67 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox67, 1, wx.ALL, 5 )
    
    	self.checkBox68 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox68, 1, wx.ALL, 5 )
    
    	self.checkBox69 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox69, 1, wx.ALL, 5 )
    
    	self.checkBox70 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox70, 1, wx.ALL, 5 )
    
    	self.checkBox71 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox71, 1, wx.ALL, 5 )
    
    	self.checkBox72 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox72, 1, wx.ALL, 5 )
    
    	self.checkBox73 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer78.Add( self.checkBox73, 1, wx.ALL, 5 )
    
    
    	bSizer74.Add( bSizer78, 0, wx.EXPAND, 5 )
    
    
    	bSizer74.Add( ( 55, 0), 0, 0, 5 )
    
    	bSizer79 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox74 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox74, 1, wx.ALL, 5 )
    
    	self.checkBox75 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox75, 1, wx.ALL, 5 )
    
    	self.checkBox76 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox76, 1, wx.ALL, 5 )
    
    	self.checkBox77 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox77, 1, wx.ALL, 5 )
    
    	self.checkBox78 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox78, 1, wx.ALL, 5 )
    
    	self.checkBox79 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox79, 1, wx.ALL, 5 )
    
    	self.checkBox80 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox80, 1, wx.ALL, 5 )
    
    	self.checkBox81 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox81, 1, wx.ALL, 5 )
    
    	self.checkBox82 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox82, 1, wx.ALL, 5 )
    
    	self.checkBox83 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer79.Add( self.checkBox83, 1, wx.ALL, 5 )
    
    
    	bSizer74.Add( bSizer79, 0, wx.EXPAND, 5 )
    
    
    	bSizer74.Add( ( 45, 0), 0, 0, 5 )
    
    	bSizer80 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox84 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox84, 1, wx.ALL, 5 )
    
    	self.checkBox85 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox85, 1, wx.ALL, 5 )
    
    	self.checkBox86 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox86, 1, wx.ALL, 5 )
    
    	self.checkBox87 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox87, 1, wx.ALL, 5 )
    
    	self.checkBox88 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox88, 1, wx.ALL, 5 )
    
    	self.checkBox89 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox89, 1, wx.ALL, 5 )
    
    	self.checkBox90 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox90, 1, wx.ALL, 5 )
    
    	self.checkBox91 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox91, 1, wx.ALL, 5 )
    
    	self.checkBox92 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox92, 1, wx.ALL, 5 )
    
    	self.checkBox93 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	bSizer80.Add( self.checkBox93, 1, wx.ALL, 5 )
    
    
    	bSizer74.Add( bSizer80, 0, wx.EXPAND, 5 )
    
    
    	bSizer66.Add( bSizer74, 1, wx.EXPAND, 5 )
    
    
    	bSizer65.Add( bSizer66, 1, wx.EXPAND, 5 )
    
    	bSizer81 = wx.BoxSizer( wx.VERTICAL )
    
    
    	bSizer81.Add( ( 0, 20), 0, 0, 5 )
    
    	self.staticText38 = wx.StaticText( self, wx.ID_ANY, u"Final Inspection and Testing on Road Worthiness", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText38.Wrap( -1 )
    	self.staticText38.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer81.Add( self.staticText38, 0, wx.ALL, 5 )
    
    	bSizer83 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer83.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer82 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox194 = wx.CheckBox( self, wx.ID_ANY, u"Brake Test", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox194.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer82.Add( self.checkBox194, 0, wx.ALL, 5 )
    
    	self.checkBox195 = wx.CheckBox( self, wx.ID_ANY, u"Side Slip Test", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox195.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer82.Add( self.checkBox195, 0, wx.ALL, 5 )
    
    	self.checkBox196 = wx.CheckBox( self, wx.ID_ANY, u"Smoke Emission Test", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox196.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer82.Add( self.checkBox196, 0, wx.ALL, 5 )
    
    
    	bSizer83.Add( bSizer82, 1, wx.EXPAND, 5 )
    
    
    	bSizer81.Add( bSizer83, 1, wx.EXPAND, 5 )
    
    
    	bSizer65.Add( bSizer81, 1, wx.EXPAND, 5 )
    
    	bSizer801 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer801.Add( ( 0, 0), 1, wx.EXPAND, 5 )
    
    	self.button17 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button17.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer801.Add( self.button17, 0, wx.ALL, 5 )
    
    	self.button18 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button18.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer801.Add( self.button18, 0, wx.ALL, 5 )
    
    
    	bSizer65.Add( bSizer801, 1, wx.EXPAND, 5 )
    
    
    	self.SetSizer( bSizer65 )
    	self.Layout()
    
    	self.Centre( wx.BOTH )
    
    	# Connect Events
    	self.checkBox34.Bind( wx.EVT_CHECKBOX, self.Check1 )
    	self.checkBox35.Bind( wx.EVT_CHECKBOX, self.Check2 )
    	self.checkBox36.Bind( wx.EVT_CHECKBOX, self.Check3 )
    	self.checkBox37.Bind( wx.EVT_CHECKBOX, self.Check4 )
    	self.checkBox38.Bind( wx.EVT_CHECKBOX, self.Check5 )
    	self.checkBox39.Bind( wx.EVT_CHECKBOX, self.Check6 )
    	self.checkBox40.Bind( wx.EVT_CHECKBOX, self.Check7 )
    	self.checkBox41.Bind( wx.EVT_CHECKBOX, self.Check8 )
    	self.checkBox42.Bind( wx.EVT_CHECKBOX, self.Check9 )
    	self.checkBox43.Bind( wx.EVT_CHECKBOX, self.Check10 )
    	self.checkBox44.Bind( wx.EVT_CHECKBOX, self.Check11 )
    	self.checkBox45.Bind( wx.EVT_CHECKBOX, self.Check12 )
    	self.checkBox46.Bind( wx.EVT_CHECKBOX, self.Check13 )
    	self.checkBox47.Bind( wx.EVT_CHECKBOX, self.Check14 )
    	self.checkBox48.Bind( wx.EVT_CHECKBOX, self.Check15 )
    	self.checkBox49.Bind( wx.EVT_CHECKBOX, self.Check16 )
    	self.checkBox50.Bind( wx.EVT_CHECKBOX, self.Check17 )
    	self.checkBox51.Bind( wx.EVT_CHECKBOX, self.Check18 )
    	self.checkBox52.Bind( wx.EVT_CHECKBOX, self.Check19 )
    	self.checkBox53.Bind( wx.EVT_CHECKBOX, self.Check20 )
    	self.checkBox54.Bind( wx.EVT_CHECKBOX, self.Check21 )
    	self.checkBox55.Bind( wx.EVT_CHECKBOX, self.Check22 )
    	self.checkBox56.Bind( wx.EVT_CHECKBOX, self.Check23 )
    	self.checkBox57.Bind( wx.EVT_CHECKBOX, self.Check24 )
    	self.checkBox58.Bind( wx.EVT_CHECKBOX, self.Check25 )
    	self.checkBox59.Bind( wx.EVT_CHECKBOX, self.Check26 )
    	self.checkBox60.Bind( wx.EVT_CHECKBOX, self.Check27 )
    	self.checkBox61.Bind( wx.EVT_CHECKBOX, self.Check28 )
    	self.checkBox62.Bind( wx.EVT_CHECKBOX, self.Check29 )
    	self.checkBox63.Bind( wx.EVT_CHECKBOX, self.Check30 )
    	self.checkBox64.Bind( wx.EVT_CHECKBOX, self.Check31 )
    	self.checkBox65.Bind( wx.EVT_CHECKBOX, self.Check32 )
    	self.checkBox66.Bind( wx.EVT_CHECKBOX, self.Check33 )
    	self.checkBox67.Bind( wx.EVT_CHECKBOX, self.Check34 )
    	self.checkBox68.Bind( wx.EVT_CHECKBOX, self.Check35 )
    	self.checkBox69.Bind( wx.EVT_CHECKBOX, self.Check36 )
    	self.checkBox70.Bind( wx.EVT_CHECKBOX, self.Check37 )
    	self.checkBox71.Bind( wx.EVT_CHECKBOX, self.Check38 )
    	self.checkBox72.Bind( wx.EVT_CHECKBOX, self.Check39 )
    	self.checkBox73.Bind( wx.EVT_CHECKBOX, self.Check40 )
    	self.checkBox74.Bind( wx.EVT_CHECKBOX, self.Check41 )
    	self.checkBox75.Bind( wx.EVT_CHECKBOX, self.Check42 )
    	self.checkBox76.Bind( wx.EVT_CHECKBOX, self.Check43 )
    	self.checkBox77.Bind( wx.EVT_CHECKBOX, self.Check44 )
    	self.checkBox78.Bind( wx.EVT_CHECKBOX, self.Check45 )
    	self.checkBox79.Bind( wx.EVT_CHECKBOX, self.Check46 )
    	self.checkBox80.Bind( wx.EVT_CHECKBOX, self.Check47 )
    	self.checkBox81.Bind( wx.EVT_CHECKBOX, self.Check48 )
    	self.checkBox82.Bind( wx.EVT_CHECKBOX, self.Check49 )
    	self.checkBox83.Bind( wx.EVT_CHECKBOX, self.Check50 )
    	self.checkBox84.Bind( wx.EVT_CHECKBOX, self.Check51 )
    	self.checkBox85.Bind( wx.EVT_CHECKBOX, self.Check52 )
    	self.checkBox86.Bind( wx.EVT_CHECKBOX, self.Check53 )
    	self.checkBox87.Bind( wx.EVT_CHECKBOX, self.Check54 )
    	self.checkBox88.Bind( wx.EVT_CHECKBOX, self.Check55 )
    	self.checkBox89.Bind( wx.EVT_CHECKBOX, self.Check56 )
    	self.checkBox90.Bind( wx.EVT_CHECKBOX, self.Check57 )
    	self.checkBox91.Bind( wx.EVT_CHECKBOX, self.Check58 )
    	self.checkBox92.Bind( wx.EVT_CHECKBOX, self.Check59 )
    	self.checkBox93.Bind( wx.EVT_CHECKBOX, self.Check60 )
    	self.checkBox194.Bind( wx.EVT_CHECKBOX, self.Check61 )
    	self.checkBox195.Bind( wx.EVT_CHECKBOX, self.Check62 )
    	self.checkBox196.Bind( wx.EVT_CHECKBOX, self.Check63 )
    	self.button17.Bind( wx.EVT_BUTTON, self.Reset )
    	self.button18.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
    	pass
    
    
    # Virtual event handlers, overide them in your derived class
    def Check1( self, event ):
            #event.Skip()
            tick29 = LoadWorkbook1(parent=self.checkBox34)
            mtick29 = tick29.wb1.get_sheet_by_name('Audit Checklist Q2')
            put29 = mtick29['C63']
            box29 = self.checkBox34.GetValue()
    
            if box29 == True:
                put29.value = "/"
            else:
                put29.value = ""
    
    def Check2( self, event ):
            #event.Skip()
            tick30 = LoadWorkbook1(parent=self.checkBox35)
            mtick30 = tick30.wb1.get_sheet_by_name('Audit Checklist Q2')
            put30 = mtick30['C64']
            box30 = self.checkBox35.GetValue()
    
            if box30 == True:
                put30.value = "/"
            else:
                put30.value = ""
    
    def Check3( self, event ):
            #event.Skip()
            tick31 = LoadWorkbook1(parent=self.checkBox36)
            mtick31 = tick31.wb1.get_sheet_by_name('Audit Checklist Q2')
            put31 = mtick31['C65']
            box31 = self.checkBox36.GetValue()
    
            if box31 == True:
                put31.value = "/"
            else:
                put31.value = ""
    
    
    def Check4( self, event ):
            #event.Skip()
            tick32 = LoadWorkbook1(parent=self.checkBox37)
            mtick32 = tick32.wb1.get_sheet_by_name('Audit Checklist Q2')
            put32 = mtick32['C66']
            box32 = self.checkBox37.GetValue()
    
            if box32 == True:
                put32.value = "/"
            else:
                put32.value = ""
    
    def Check5( self, event ):
            #event.Skip()
            tick33 = LoadWorkbook1(parent=self.checkBox38)
            mtick33 = tick33.wb1.get_sheet_by_name('Audit Checklist Q2')
            put33 = mtick33['C67']
            box33 = self.checkBox38.GetValue()
    
            if box33 == True:
                put33.value = "/"
            else:
                put33.value = ""
    
    def Check6( self, event ):
            #event.Skip()
            tick34 = LoadWorkbook1(parent=self.checkBox39)
            mtick34 = tick34.wb1.get_sheet_by_name('Audit Checklist Q2')
            put34 = mtick34['C68']
            box34 = self.checkBox39.GetValue()
    
            if box34 == True:
                put34.value = "/"
            else:
                put34.value = ""
    
    def Check7( self, event ):
            #event.Skip()
            tick35 = LoadWorkbook1(parent=self.checkBox40)
            mtick35 = tick35.wb1.get_sheet_by_name('Audit Checklist Q2')
            put35 = mtick35['C69']
            box35 = self.checkBox40.GetValue()
    
            if box35 == True:
                put35.value = "/"
            else:
                put35.value = ""
    
    def Check8( self, event ):
            #event.Skip()
            tick36 = LoadWorkbook1(parent=self.checkBox41)
            mtick36 = tick36.wb1.get_sheet_by_name('Audit Checklist Q2')
            put36 = mtick36['C70']
            box36 = self.checkBox41.GetValue()
    
            if box36 == True:
                put36.value = "/"
            else:
                put36.value = ""
    
    def Check9( self, event ):
            #event.Skip()
            tick37 = LoadWorkbook1(parent=self.checkBox42)
            mtick37 = tick37.wb1.get_sheet_by_name('Audit Checklist Q2')
            put37 = mtick37['C71']
            box37 = self.checkBox42.GetValue()
    
            if box37 == True:
                put37.value = "/"
            else:
                put37.value = ""
    
    def Check10( self, event ):
            #event.Skip()
            tick38 = LoadWorkbook1(parent=self.checkBox43)
            mtick38 = tick38.wb1.get_sheet_by_name('Audit Checklist Q2')
            put38 = mtick38['C72']
            box38 = self.checkBox43.GetValue()
    
            if box38 == True:
                put38.value = "/"
            else:
                put38.value = ""
    
    def Check11( self, event ):
            #event.Skip()
            tick39 = LoadWorkbook1(parent=self.checkBox44)
            mtick39 = tick39.wb1.get_sheet_by_name('Audit Checklist Q2')
            put39 = mtick39['E63']
            box39 = self.checkBox44.GetValue()
    
            if box39 == True:
                put39.value = "/"
            else:
                put39.value = ""
    
    def Check12( self, event ):
            #event.Skip()
            global team40
            tick40 = LoadWorkbook1(parent=self.checkBox45)
            mtick40 = tick40.wb1.get_sheet_by_name('Audit Checklist Q2')
            put40 = mtick40['E64']
            box40 = self.checkBox45.GetValue()
    
            if box40 == True:
                put40.value = "/"
            else:
                put40.value = ""
    
    def Check13( self, event ):
            #event.Skip()
            tick41 = LoadWorkbook1(parent=self.checkBox46)
            mtick41 = tick41.wb1.get_sheet_by_name('Audit Checklist Q2')
            put41 = mtick41['E65']
            box41 = self.checkBox46.GetValue()
    
            if box41 == True:
                put41.value = "/"
            else:
                put41.value = ""
    
    def Check14( self, event ):
            #event.Skip()
            tick42 = LoadWorkbook1(parent=self.checkBox47)
            mtick42 = tick42.wb1.get_sheet_by_name('Audit Checklist Q2')
            put42 = mtick42['E66']
            box42 = self.checkBox47.GetValue()
    
            if box42 == True:
                put42.value = "/"
            else:
                put42.value = ""
    
    def Check15( self, event ):
            #event.Skip()
            tick43 = LoadWorkbook1(parent=self.checkBox48)
            mtick43 = tick43.wb1.get_sheet_by_name('Audit Checklist Q2')
            put43 = mtick43['E67']
            box43 = self.checkBox48.GetValue()
    
            if box43 == True:
                put43.value = "/"
            else:
                put43.value = ""
    
    def Check16( self, event ):
            #event.Skip()
            tick44 = LoadWorkbook1(parent=self.checkBox49)
            mtick44 = tick44.wb1.get_sheet_by_name('Audit Checklist Q2')
            put44 = mtick44['E68']
            box44 = self.checkBox48.GetValue()
    
            if box44 == True:
                put44.value = "/"
            else:
                put44.value = ""
    
    def Check17( self, event ):
            #event.Skip()
            tick45 = LoadWorkbook1(parent=self.checkBox50)
            mtick45 = tick45.wb1.get_sheet_by_name('Audit Checklist Q2')
            put45 = mtick45['E69']
            box45 = self.checkBox50.GetValue()
    
            if box45 == True:
                put45.value = "/"
            else:
                put45.value = ""
    
    def Check18( self, event ):
            #event.Skip()
            tick46 = LoadWorkbook1(parent=self.checkBox51)
            mtick46 = tick46.wb1.get_sheet_by_name('Audit Checklist Q2')
            put46 = mtick46['E70']
            box46 = self.checkBox51.GetValue()
    
            if box46 == True:
                put46.value = "/"
            else:
                put46.value = ""
    
    def Check19( self, event ):
            #event.Skip()
            tick47 = LoadWorkbook1(parent=self.checkBox52)
            mtick47 = tick47.wb1.get_sheet_by_name('Audit Checklist Q2')
            put47 = mtick47['E71']
            box47 = self.checkBox52.GetValue()
    
            if box47 == True:
                put47.value = "/"
            else:
                put47.value = ""
    
    def Check20( self, event ):
            #event.Skip()
            tick48 = LoadWorkbook1(parent=self.checkBox53)
            mtick48 = tick48.wb1.get_sheet_by_name('Audit Checklist Q2')
            put48 = mtick48['E72']
            box48 = self.checkBox53.GetValue()
    
            if box48 == True:
                put48.value = "/"
            else:
                put48.value = ""
    
    def Check21( self, event ):
            #event.Skip()
            tick49 = LoadWorkbook1(parent=self.checkBox54)
            mtick49 = tick49.wb1.get_sheet_by_name('Audit Checklist Q2')
            put49 = mtick49['G63']
            box49 = self.checkBox54.GetValue()
    
            if box49 == True:
                put49.value = "/"
            else:
                put49.value = ""
    
    def Check22( self, event ):
            #event.Skip()
            tick50 = LoadWorkbook1(parent=self.checkBox55)
            mtick50 = tick50.wb1.get_sheet_by_name('Audit Checklist Q2')
            put50 = mtick50['G64']
            box50 = self.checkBox55.GetValue()
    
            if box50 == True:
                put50.value = "/"
            else:
                put50.value = ""
    
    def Check23( self, event ):
            #event.Skip()
            tick51 = LoadWorkbook1(parent=self.checkBox56)
            mtick51 = tick51.wb1.get_sheet_by_name('Audit Checklist Q2')
            put51 = mtick51['G65']
            box51 = self.checkBox56.GetValue()
    
            if box51 == True:
                put51.value = "/"
            else:
                put51.value = ""
    
    
    def Check24( self, event ):
            #event.Skip()
            global team52
            tick52 = LoadWorkbook1(parent=self.checkBox57)
            mtick52 = tick52.wb1.get_sheet_by_name('Audit Checklist Q2')
            put52 = mtick52['G66']
            box52 = self.checkBox57.GetValue()
    
            if box52 == True:
                put52.value = "/"
    
            else:
                put52.value = ""
    
    
    def Check25( self, event ):
            #event.Skip()
            tick53 = LoadWorkbook1(parent=self.checkBox58)
            mtick53 = tick53.wb1.get_sheet_by_name('Audit Checklist Q2')
            put53 = mtick53['G67']
            box53 = self.checkBox58.GetValue()
    
            if box53 == True:
                put53.value = "/"
    
            else:
                put53.value = ""
    
    def Check26( self, event ):
            #event.Skip()
            tick54 = LoadWorkbook1(parent=self.checkBox59)
            mtick54 = tick54.wb1.get_sheet_by_name('Audit Checklist Q2')
            put54 = mtick54['G68']
            box54 = self.checkBox59.GetValue()
    
            if box54 == True:
                put54.value = "/"
    
            else:
                put54.value = ""
    
    def Check27( self, event ):
            #event.Skip()
            tick55 = LoadWorkbook1(parent=self.checkBox60)
            mtick55 = tick55.wb1.get_sheet_by_name('Audit Checklist Q2')
            put55 = mtick55['G69']
            box55 = self.checkBox60.GetValue()
    
            if box55 == True:
                put55.value = "/"
    
            else:
                put55.value = ""
    
    def Check28( self, event ):
            #event.Skip()
            tick56 = LoadWorkbook1(parent=self.checkBox61)
            mtick56 = tick56.wb1.get_sheet_by_name('Audit Checklist Q2')
            put56 = mtick56['G70']
            box56 = self.checkBox61.GetValue()
    
            if box56 == True:
                put56.value = "/"
    
            else:
                put56.value = ""
    
    def Check29( self, event ):
            #event.Skip()
            tick57 = LoadWorkbook1(parent=self.checkBox62)
            mtick57 = tick57.wb1.get_sheet_by_name('Audit Checklist Q2')
            put57 = mtick57['G71']
            box57 = self.checkBox62.GetValue()
    
            if box57 == True:
                put57.value = "/"
    
            else:
                put57.value = ""
    
    def Check30( self, event ):
            #event.Skip()
            tick58 = LoadWorkbook1(parent=self.checkBox63)
            mtick58 = tick58.wb1.get_sheet_by_name('Audit Checklist Q2')
            put58 = mtick58['G72']
            box58 = self.checkBox63.GetValue()
    
            if box58 == True:
                put58.value = "/"
    
            else:
                put58.value = ""
    
    def Check31( self, event ):
            #event.Skip()
            tick59 = LoadWorkbook1(parent=self.checkBox64)
            mtick59 = tick59.wb1.get_sheet_by_name('Audit Checklist Q2')
            put59 = mtick59['I63']
            box59 = self.checkBox64.GetValue()
    
            if box59 == True:
                put59.value = "/"
    
            else:
                put59.value = ""
    
    def Check32( self, event ):
            #event.Skip()
            tick60 = LoadWorkbook1(parent=self.checkBox65)
            mtick60 = tick60.wb1.get_sheet_by_name('Audit Checklist Q2')
            put60 = mtick60['I64']
            box60 = self.checkBox65.GetValue()
    
            if box60 == True:
                put60.value = "/"
    
            else:
                put60.value = ""
    
    def Check33( self, event ):
            #event.Skip()
            tick61 = LoadWorkbook1(parent=self.checkBox66)
            mtick61 = tick61.wb1.get_sheet_by_name('Audit Checklist Q2')
            put61 = mtick61['I65']
            box61 = self.checkBox66.GetValue()
    
            if box61 == True:
                put61.value = "/"
    
            else:
                put61.value = ""
    
    def Check34( self, event ):
            #event.Skip()
            tick62 = LoadWorkbook1(parent=self.checkBox67)
            mtick62 = tick62.wb1.get_sheet_by_name('Audit Checklist Q2')
            put62 = mtick62['I66']
            box62 = self.checkBox67.GetValue()
    
            if box62 == True:
                put62.value = "/"
    
            else:
                put62.value = ""
    
    def Check35( self, event ):
            #event.Skip()
            tick63 = LoadWorkbook1(parent=self.checkBox68)
            mtick63 = tick63.wb1.get_sheet_by_name('Audit Checklist Q2')
            put63 = mtick63['I67']
            box63 = self.checkBox68.GetValue()
    
            if box63 == True:
                put63.value = "/"
    
            else:
                put63.value = ""
    
    def Check36( self, event ):
            #event.Skip()
            tick64 = LoadWorkbook1(parent=self.checkBox69)
            mtick64 = tick64.wb1.get_sheet_by_name('Audit Checklist Q2')
            put64 = mtick64['I68']
            box64 = self.checkBox69.GetValue()
    
            if box64 == True:
                put64.value = "/"
    
            else:
                put64.value = ""
    
    def Check37( self, event ):
            #event.Skip()
            tick65 = LoadWorkbook1(parent=self.checkBox70)
            mtick65 = tick65.wb1.get_sheet_by_name('Audit Checklist Q2')
            put65 = mtick65['I69']
            box65 = self.checkBox70.GetValue()
    
            if box65 == True:
                put65.value = "/"
    
            else:
                put65.value = ""
    
    def Check38( self, event ):
            #event.Skip()
            tick66 = LoadWorkbook1(parent=self.checkBox71)
            mtick66 = tick66.wb1.get_sheet_by_name('Audit Checklist Q2')
            put66 = mtick66['I70']
            box66 = self.checkBox71.GetValue()
    
            if box66 == True:
                put66.value = "/"
    
            else:
                put66.value = ""
    
    def Check39( self, event ):
            #event.Skip()
            tick67 = LoadWorkbook1(parent=self.checkBox72)
            mtick67 = tick67.wb1.get_sheet_by_name('Audit Checklist Q2')
            put67 = mtick67['I71']
            box67 = self.checkBox72.GetValue()
    
            if box67 == True:
                put67.value = "/"
    
            else:
                put67.value = ""
    
    def Check40( self, event ):
            #event.Skip()
            tick68 = LoadWorkbook1(parent=self.checkBox73)
            mtick68 = tick68.wb1.get_sheet_by_name('Audit Checklist Q2')
            put68 = mtick68['I72']
            box68 = self.checkBox73.GetValue()
    
            if box68 == True:
                put68.value = "/"
    
            else:
                put68.value = ""
    
    def Check41( self, event ):
            #event.Skip()
            tick69 = LoadWorkbook1(parent=self.checkBox74)
            mtick69 = tick69.wb1.get_sheet_by_name('Audit Checklist Q2')
            put69 = mtick69['K63']
            box69 = self.checkBox74.GetValue()
    
            if box69 == True:
                put69.value = "/"
    
            else:
                put69.value = ""
    
    def Check42( self, event ):
            #event.Skip()
            tick70 = LoadWorkbook1(parent=self.checkBox75)
            mtick70 = tick70.wb1.get_sheet_by_name('Audit Checklist Q2')
            put70 = mtick70['K64']
            box70 = self.checkBox75.GetValue()
    
            if box70 == True:
                put70.value = "/"
    
            else:
                put70.value = ""
    
    def Check43( self, event ):
            #event.Skip()
            tick71 = LoadWorkbook1(parent=self.checkBox76)
            mtick71 = tick71.wb1.get_sheet_by_name('Audit Checklist Q2')
            put71 = mtick71['K65']
            box71 = self.checkBox76.GetValue()
    
            if box71 == True:
                put71.value = "/"
    
            else:
                put71.value = ""
    
    def Check44( self, event ):
            #event.Skip()
            tick72 = LoadWorkbook1(parent=self.checkBox77)
            mtick72 = tick72.wb1.get_sheet_by_name('Audit Checklist Q2')
            put72 = mtick72['K66']
            box72 = self.checkBox77.GetValue()
    
            if box72 == True:
                put72.value = "/"
    
            else:
                put72.value = ""
    
    def Check45( self, event ):
            #event.Skip()
            tick73 = LoadWorkbook1(parent=self.checkBox78)
            mtick73 = tick73.wb1.get_sheet_by_name('Audit Checklist Q2')
            put73 = mtick73['K67']
            box73 = self.checkBox78.GetValue()
    
            if box73 == True:
                put73.value = "/"
    
            else:
                put73.value = ""
    
    def Check46( self, event ):
            #event.Skip()
            tick74 = LoadWorkbook1(parent=self.checkBox79)
            mtick74 = tick74.wb1.get_sheet_by_name('Audit Checklist Q2')
            put74 = mtick74['K68']
            box74 = self.checkBox79.GetValue()
    
            if box74 == True:
                put74.value = "/"
    
            else:
                put74.value = ""
    
    def Check47( self, event ):
            #event.Skip()
            tick75 = LoadWorkbook1(parent=self.checkBox80)
            mtick75 = tick75.wb1.get_sheet_by_name('Audit Checklist Q2')
            put75 = mtick75['K69']
            box75 = self.checkBox80.GetValue()
    
            if box75 == True:
                put75.value = "/"
    
            else:
                put75.value = ""
    
    def Check48( self, event ):
            #event.Skip()
            tick76 = LoadWorkbook1(parent=self.checkBox81)
            mtick76 = tick76.wb1.get_sheet_by_name('Audit Checklist Q2')
            put76 = mtick76['K70']
            box76 = self.checkBox81.GetValue()
    
            if box76 == True:
                put76.value = "/"
    
            else:
                put76.value = ""
    
    def Check49( self, event ):
            #event.Skip()
            tick77 = LoadWorkbook1(parent=self.checkBox82)
            mtick77 = tick77.wb1.get_sheet_by_name('Audit Checklist Q2')
            put77 = mtick77['K71']
            box77 = self.checkBox82.GetValue()
    
            if box77 == True:
                put77.value = "/"
    
            else:
                put77.value = ""
    
    
    def Check50( self, event ):
            #event.Skip()
            tick78 = LoadWorkbook1(parent=self.checkBox83)
            mtick78 = tick78.wb1.get_sheet_by_name('Audit Checklist Q2')
            put78 = mtick78['K72']
            box78 = self.checkBox83.GetValue()
    
            if box78 == True:
                put78.value = "/"
    
            else:
                put78.value = ""
    
    def Check51( self, event ):
            #event.Skip()
            global team79
            tick79 = LoadWorkbook1(parent=self.checkBox84)
            mtick79 = tick79.wb1.get_sheet_by_name('Audit Checklist Q2')
            put79 = mtick79['M63']
            box79 = self.checkBox84.GetValue()
    
            if box79 == True:
                put79.value = "/"
    
            else:
                put79.value = ""
    
    def Check52( self, event ):
            #event.Skip()
            tick80 = LoadWorkbook1(parent=self.checkBox85)
            mtick80 = tick80.wb1.get_sheet_by_name('Audit Checklist Q2')
            put80 = mtick80['M64']
            box80 = self.checkBox85.GetValue()
    
            if box80 == True:
                put80.value = "/"
    
            else:
                put80.value = ""
    
    def Check53( self, event ):
            #event.Skip()
            tick81 = LoadWorkbook1(parent=self.checkBox86)
            mtick81 = tick81.wb1.get_sheet_by_name('Audit Checklist Q2')
            put81 = mtick81['M65']
            box81 = self.checkBox86.GetValue()
    
            if box81 == True:
                put81.value = "/"
    
            else:
                put81.value = ""
    
    def Check54( self, event ):
            #event.Skip()
            tick82 = LoadWorkbook1(parent=self.checkBox87)
            mtick82 = tick82.wb1.get_sheet_by_name('Audit Checklist Q2')
            put82 = mtick82['M66']
            box82 = self.checkBox87.GetValue()
    
            if box82 == True:
                put82.value = "/"
    
            else:
                put82.value = ""
    
    def Check55( self, event ):
            #event.Skip()
            tick83 = LoadWorkbook1(parent=self.checkBox88)
            mtick83 = tick83.wb1.get_sheet_by_name('Audit Checklist Q2')
            put83 = mtick83['M67']
            box83 = self.checkBox88.GetValue()
    
    
            if box83 == True:
                put83.value = "/"
    
            else:
                put83.value = ""
    
    def Check56( self, event ):
            #event.Skip()
            tick84 = LoadWorkbook1(parent=self.checkBox89)
            mtick84 = tick84.wb1.get_sheet_by_name('Audit Checklist Q2')
            put84 = mtick84['M68']
            box84 = self.checkBox89.GetValue()
    
            if box84 == True:
                put84.value = "/"
    
            else:
                put84.value = ""
    
    def Check57( self, event ):
            #event.Skip()
            tick85 = LoadWorkbook1(parent=self.checkBox90)
            mtick85 = tick85.wb1.get_sheet_by_name('Audit Checklist Q2')
            put85 = mtick85['M69']
            box85 = self.checkBox90.GetValue()
    
            if box85 == True:
                put85.value = "/"
    
            else:
                put85.value = ""
    
    def Check58( self, event ):
            #event.Skip()
            tick86 = LoadWorkbook1(parent=self.checkBox91)
            mtick86 = tick86.wb1.get_sheet_by_name('Audit Checklist Q2')
            put86 = mtick86['M70']
            box86 = self.checkBox91.GetValue()
    
            if box86 == True:
                put86.value = "/"
    
            else:
                put86.value = ""
    
    def Check59( self, event ):
            #event.Skip()
            tick87 = LoadWorkbook1(parent=self.checkBox92)
            mtick87 = tick87.wb1.get_sheet_by_name('Audit Checklist Q2')
            put87 = mtick87['M71']
            box87 = self.checkBox92.GetValue()
    
            if box87 == True:
                put87.value = "/"
    
            else:
                put87.value = ""
    
    def Check60( self, event ):
            #event.Skip()
            tick88 = LoadWorkbook1(parent=self.checkBox93)
            mtick88 = tick88.wb1.get_sheet_by_name('Audit Checklist Q2')
            put88 = mtick88['M72']
            box88 = self.checkBox93.GetValue()
    
            if box88 == True:
                put88.value = "/"
    
            else:
                put88.value = ""
    
    def Check61( self, event ):
            #event.Skip()
            tick89 = LoadWorkbook1(parent=self.checkBox194)
            mtick89 = tick89.wb1.get_sheet_by_name('Audit Checklist Q2')
            put89 = mtick89['C74']
            box89 = self.checkBox194.GetValue()
    
            if box89 == True:
                put89.value = "/"
    
            else:
                put89.value = ""
    
    def Check62( self, event ):
            #event.Skip()
            tick90 = LoadWorkbook1(parent=self.checkBox195)
            mtick90 = tick90.wb1.get_sheet_by_name('Audit Checklist Q2')
            put90 = mtick90['C75']
            box90 = self.checkBox195.GetValue()
    
            if box90 == True:
                put90.value = "/"
    
            else:
                put90.value = ""
    
    def Check63( self, event ):
            #event.Skip()
            tick91 = LoadWorkbook1(parent=self.checkBox196)
            mtick91 = tick91.wb1.get_sheet_by_name('Audit Checklist Q2')
            put91 = mtick91['C76']
            box91 = self.checkBox196.GetValue()
    
            if box91 == True:
                put91.value = "/"
    
            else:
                put91.value = ""
    
    def Reset( self, event ):
            #event.Skip()
            self.checkBox34.SetValue(False)
            self.checkBox35.SetValue(False)
            self.checkBox36.SetValue(False)
            self.checkBox37.SetValue(False)
            self.checkBox38.SetValue(False)
            self.checkBox39.SetValue(False)
            self.checkBox40.SetValue(False)
            self.checkBox41.SetValue(False)
            self.checkBox42.SetValue(False)
            self.checkBox43.SetValue(False)
            self.checkBox44.SetValue(False)
            self.checkBox45.SetValue(False)
            self.checkBox46.SetValue(False)
            self.checkBox47.SetValue(False)
            self.checkBox48.SetValue(False)
            self.checkBox49.SetValue(False)
            self.checkBox50.SetValue(False)
            self.checkBox51.SetValue(False)
            self.checkBox52.SetValue(False)
            self.checkBox53.SetValue(False)
            self.checkBox54.SetValue(False)
            self.checkBox55.SetValue(False)
            self.checkBox56.SetValue(False)
            self.checkBox57.SetValue(False)
            self.checkBox58.SetValue(False)
            self.checkBox59.SetValue(False)
            self.checkBox60.SetValue(False)
            self.checkBox61.SetValue(False)
            self.checkBox62.SetValue(False)
            self.checkBox63.SetValue(False)
            self.checkBox64.SetValue(False)
            self.checkBox65.SetValue(False)
            self.checkBox66.SetValue(False)
            self.checkBox67.SetValue(False)
            self.checkBox68.SetValue(False)
            self.checkBox69.SetValue(False)
            self.checkBox70.SetValue(False)
            self.checkBox71.SetValue(False)
            self.checkBox72.SetValue(False)
            self.checkBox73.SetValue(False)
            self.checkBox74.SetValue(False)
            self.checkBox75.SetValue(False)
            self.checkBox76.SetValue(False)
            self.checkBox77.SetValue(False)
            self.checkBox78.SetValue(False)
            self.checkBox79.SetValue(False)
            self.checkBox80.SetValue(False)
            self.checkBox81.SetValue(False)
            self.checkBox82.SetValue(False)
            self.checkBox83.SetValue(False)
            self.checkBox84.SetValue(False)
            self.checkBox85.SetValue(False)
            self.checkBox86.SetValue(False)
            self.checkBox87.SetValue(False)
            self.checkBox88.SetValue(False)
            self.checkBox89.SetValue(False)
            self.checkBox90.SetValue(False)
            self.checkBox91.SetValue(False)
            self.checkBox92.SetValue(False)
            self.checkBox93.SetValue(False)
            self.checkBox194.SetValue(False)
            self.checkBox195.SetValue(False)
            self.checkBox196.SetValue(False)
    
    
    
    def Save( self, event ):
            #event.Skip()
            check29 = self.checkBox34.GetValue()
            check30 = self.checkBox35.GetValue()
            check31 = self.checkBox36.GetValue()
            check32 = self.checkBox37.GetValue()
            check33 = self.checkBox38.GetValue()
            check34 = self.checkBox39.GetValue()
            check35 = self.checkBox40.GetValue()
            check36 = self.checkBox41.GetValue()
            check37 = self.checkBox42.GetValue()
            check38 = self.checkBox43.GetValue()
            check39 = self.checkBox44.GetValue()
            check40 = self.checkBox45.GetValue()
            check41 = self.checkBox46.GetValue()
            check42 = self.checkBox47.GetValue()
            check43 = self.checkBox48.GetValue()
            check44 = self.checkBox49.GetValue()
            check45 = self.checkBox50.GetValue()
            check46 = self.checkBox51.GetValue()
            check47 = self.checkBox52.GetValue()
            check48 = self.checkBox53.GetValue()
            check49 = self.checkBox54.GetValue()
            check50 = self.checkBox55.GetValue()
            check51 = self.checkBox56.GetValue()
            check52 = self.checkBox57.GetValue()
            check53 = self.checkBox58.GetValue()
            check54 = self.checkBox59.GetValue()
            check55 = self.checkBox60.GetValue()
            check56 = self.checkBox61.GetValue()
            check57 = self.checkBox62.GetValue()
            check58 = self.checkBox63.GetValue()
            check59 = self.checkBox64.GetValue()
            check60 = self.checkBox65.GetValue()
            check61 = self.checkBox66.GetValue()
            check62 = self.checkBox67.GetValue()
            check63 = self.checkBox68.GetValue()
            check64 = self.checkBox69.GetValue()
            check65 = self.checkBox70.GetValue()
            check66 = self.checkBox71.GetValue()
            check67 = self.checkBox72.GetValue()
            check68 = self.checkBox73.GetValue()
            check69 = self.checkBox74.GetValue()
            check70 = self.checkBox75.GetValue()
            check71 = self.checkBox76.GetValue()
            check72 = self.checkBox77.GetValue()
            check73 = self.checkBox78.GetValue()
            check74 = self.checkBox79.GetValue()
            check75 = self.checkBox80.GetValue()
            check76 = self.checkBox81.GetValue()
            check77 = self.checkBox82.GetValue()
            check78 = self.checkBox83.GetValue()
            check79 = self.checkBox84.GetValue()
            check80 = self.checkBox85.GetValue()
            check81 = self.checkBox86.GetValue()
            check82 = self.checkBox87.GetValue()
            check83 = self.checkBox88.GetValue()
            check84 = self.checkBox89.GetValue()
            check85 = self.checkBox90.GetValue()
            check86 = self.checkBox91.GetValue()
            check87 = self.checkBox92.GetValue()
            check88 = self.checkBox93.GetValue()
            check89 = self.checkBox194.GetValue()
            check90 = self.checkBox195.GetValue()
            check91 = self.checkBox196.GetValue()
    
            if check29 == True:
                a29 = 1
            else:
                a29 = 0
    
            if check30 == True:
                a30 = 1
            else:
                a30 = 0
    
            if check31 == True:
                a31 = 1
            else:
                a31 = 0
    
            if check32 == True:
                a32 = 1
            else:
                a32 = 0
    
            if check33 == True:
                a33 = 1
            else:
                a33 = 0
    
            if check34 == True:
                a34 = 1
            else:
                a34 = 0
    
            if check35 == True:
                a35 = 1
            else:
                a35 = 0
    
            if check36 == True:
                a36 = 1
            else:
                a36 = 0
    
            if check37 == True:
                a37 = 1
            else:
                a37 = 0
    
            if check38 == True:
                a38 = 1
            else:
                a38 = 0
    
            if check39 == True:
                a39 = 1
            else:
                a39 = 0
    
            if check40 == True:
                a40 = 1
            else:
                a40 = 0
    
            if check41 == True:
                a41 = 1
            else:
                a41 = 0
    
            if check42 == True:
                a42 = 1
            else:
                a42 = 0
    
            if check43 == True:
                a43 = 1
            else:
                a43 = 0
    
            if check44 == True:
                a44 = 1
            else:
                a44 = 0
    
            if check45 == True:
                a45 = 1
            else:
                a45 = 0
    
            if check46 == True:
                a46 = 1
            else:
                a46 = 0
    
            if check47 == True:
                a47 = 1
            else:
                a47 = 0
    
            if check48 == True:
                a48 = 1
            else:
                a48 = 0
    
            if check49 == True:
                a49 = 1
            else:
                a49 = 0
    
            if check50 == True:
                a50 = 1
            else:
                a50 = 0
    
            if check51 == True:
                a51 = 1
            else:
                a51 = 0
    
            if check52 == True:
                a52 = 1
            else:
                a52 = 0
    
            if check53 == True:
                a53 = 1
            else:
                a53 = 0
    
            if check54 == True:
                a54 = 1
            else:
                a54 = 0
    
            if check55 == True:
                a55 = 1
            else:
                a55 = 0
    
            if check56 == True:
                a56 = 1
            else:
                a56 = 0
    
            if check57 == True:
                a57 = 1
            else:
                a57 = 0
    
            if check58 == True:
                a58 = 1
            else:
                a58 = 0
    
            if check59 == True:
                a59 = 1
            else:
                a59 = 0
    
            if check60 == True:
                a60 = 1
            else:
                a60 = 0
    
            if check60 == True:
                a60 = 1
            else:
                a60 = 0
    
            if check61 == True:
                a61 = 1
            else:
                a61 = 0
    
            if check62 == True:
                a62 = 1
            else:
                a62 = 0
    
            if check63 == True:
                a63 = 1
            else:
                a63 = 0
    
            if check64 == True:
                a64 = 1
            else:
                a64 = 0
    
            if check65 == True:
                a65 = 1
            else:
                a65 = 0
    
            if check66 == True:
                a66 = 1
            else:
                a66 = 0
    
            if check67 == True:
                a67 = 1
            else:
                a67 = 0
    
            if check68 == True:
                a68 = 1
            else:
                a68 = 0
    
            if check69 == True:
                a69 = 1
            else:
                a69 = 0
    
            if check70 == True:
                a70 = 1
            else:
                a70 = 0
    
            if check71 == True:
                a71 = 1
            else:
                a71 = 0
    
            if check72 == True:
                a72 = 1
            else:
                a72 = 0
    
            if check73 == True:
                a73 = 1
            else:
                a73 = 0
    
            if check74 == True:
                a74 = 1
            else:
                a74 = 0
    
            if check75 == True:
                a75 = 1
            else:
                a75 = 0
    
            if check76 == True:
                a76 = 1
            else:
                a76 = 0
    
            if check77 == True:
                a77 = 1
            else:
                a77 = 0
    
            if check78 == True:
                a78 = 1
            else:
                a78 = 0
    
            if check79 == True:
                a79 = 1
            else:
                a79 = 0
    
            if check80 == True:
                a80 = 1
            else:
                a80 = 0
    
            if check81 == True:
                a81 = 1
            else:
                a81 = 0
    
            if check82 == True:
                a82 = 1
            else:
                a82 = 0
    
            if check83 == True:
                a83 = 1
            else:
                a83 = 0
    
            if check84 == True:
                a84 = 1
            else:
                a84 = 0
    
            if check85 == True:
                a85 = 1
            else:
                a85 = 0
    
            if check86 == True:
                a86 = 1
            else:
                a86 = 0
    
            if check87 == True:
                a87 = 1
            else:
                a87 = 0
    
            if check88 == True:
                a88 = 1
            else:
                a88 = 0
    
            if check89 == True:
                a89 = 1
            else:
                a89 = 0
    
            if check90 == True:
                a90 = 1
            else:
                a90 = 0
    
            if check91 == True:
                a91 = 1
            else:
                a91 = 0



###########################################################################
## Class PartD
###########################################################################

class PartD ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Standard Compliance To\nVehicle Type Approval (VTA)\n-Safety Parts-", pos = wx.DefaultPosition, size = wx.Size( 791,320 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
    
    	self.SetSizeHintsSz( wx.Size(791,320), wx.Size(791,320) )
    	self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
    
    	bSizer81 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer82 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText37 = wx.StaticText( self, wx.ID_ANY, u"Compliance to E-marking that follow condition MS/ UN", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText37.Wrap( -1 )
    	self.staticText37.SetFont( wx.Font( 14, 70, 90, 92, False, "Arial" ) )
    
    	bSizer82.Add( self.staticText37, 0, wx.ALL, 5 )
    
    
    	bSizer82.Add( ( 50, 0), 0, 0, 5 )
    
    	self.staticText38 = wx.StaticText( self, wx.ID_ANY, u"Standards", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText38.Wrap( -1 )
    	self.staticText38.SetFont( wx.Font( 14, 70, 90, 92, False, "Arial" ) )
    
    	bSizer82.Add( self.staticText38, 0, wx.ALL, 5 )
    
    
    	bSizer81.Add( bSizer82, 0, wx.EXPAND, 5 )
    
    	bSizer84 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer84.Add( ( 20, 0), 0, 0, 5 )
    
    	bSizer83 = wx.BoxSizer( wx.VERTICAL )
    
    	self.checkBox87 = wx.CheckBox( self, wx.ID_ANY, u"Safety Glass", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox87.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer83.Add( self.checkBox87, 0, wx.ALL, 5 )
    
    	self.checkBox88 = wx.CheckBox( self, wx.ID_ANY, u"Brake Lamp Performance", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox88.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer83.Add( self.checkBox88, 0, wx.ALL, 5 )
    
    	self.checkBox89 = wx.CheckBox( self, wx.ID_ANY, u"Indicator Performance", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox89.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer83.Add( self.checkBox89, 0, wx.ALL, 5 )
    
    	self.checkBox90 = wx.CheckBox( self, wx.ID_ANY, u"Headlamp Performance", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox90.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer83.Add( self.checkBox90, 0, wx.ALL, 5 )
    
    	self.checkBox91 = wx.CheckBox( self, wx.ID_ANY, u"Tire and Wheel", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox91.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer83.Add( self.checkBox91, 0, wx.ALL, 5 )
    
    	self.checkBox92 = wx.CheckBox( self, wx.ID_ANY, u"Rear Marking Plates", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox92.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer83.Add( self.checkBox92, 0, wx.ALL, 5 )
    
    	self.m_checkBox93 = wx.CheckBox( self, wx.ID_ANY, u"Front Underrun Protection", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.m_checkBox93.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer83.Add( self.m_checkBox93, 0, wx.ALL, 5 )
    
    
    	bSizer84.Add( bSizer83, 1, wx.EXPAND, 5 )
    
    	bSizer85 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer85.Add( ( 30, 0), 0, 0, 5 )
    
    	bSizer86 = wx.BoxSizer( wx.VERTICAL )
    
    	self.staticText43 = wx.StaticText( self, wx.ID_ANY, u"MS 595/UN R43", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText43.Wrap( -1 )
    	self.staticText43.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.staticText43, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )
    
    	self.staticText44 = wx.StaticText( self, wx.ID_ANY, u"MS 1776/UN R7", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText44.Wrap( -1 )
    	self.staticText44.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.staticText44, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )
    
    	self.staticText45 = wx.StaticText( self, wx.ID_ANY, u"MS 1851/UN R6", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText45.Wrap( -1 )
    	self.staticText45.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.staticText45, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )
    
    	self.staticText46 = wx.StaticText( self, wx.ID_ANY, u"UN R112", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText46.Wrap( -1 )
    	self.staticText46.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.staticText46, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )
    
    	self.staticText47 = wx.StaticText( self, wx.ID_ANY, u"UN R54", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText47.Wrap( -1 )
    	self.staticText47.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.staticText47, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )
    
    	self.staticText48 = wx.StaticText( self, wx.ID_ANY, u"UN R70", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText48.Wrap( -1 )
    	self.staticText48.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.staticText48, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )
    
    	self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"UN R 93", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText49.Wrap( -1 )
    	self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer86.Add( self.staticText49, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )
    
    
    	bSizer85.Add( bSizer86, 1, wx.EXPAND, 5 )
    
    
    	bSizer84.Add( bSizer85, 1, wx.EXPAND, 5 )
    
    
    	bSizer81.Add( bSizer84, 1, wx.EXPAND, 5 )
    
    	bSizer87 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer87.Add( ( 0, 0), 1, wx.EXPAND, 5 )
    
    	self.button19 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button19.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer87.Add( self.button19, 0, wx.ALL, 5 )
    
    	self.button20 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button20.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer87.Add( self.button20, 0, wx.ALL, 5 )
    
    
    	bSizer81.Add( bSizer87, 1, wx.EXPAND, 5 )
    
    
    	self.SetSizer( bSizer81 )
    	self.Layout()
    
    	self.Centre( wx.BOTH )
    	# Connect Events
    	self.checkBox87.Bind( wx.EVT_CHECKBOX, self.Check1 )
    	self.checkBox88.Bind( wx.EVT_CHECKBOX, self.Check2 )
    	self.checkBox89.Bind( wx.EVT_CHECKBOX, self.Check3 )
    	self.checkBox90.Bind( wx.EVT_CHECKBOX, self.Check4 )
    	self.checkBox91.Bind( wx.EVT_CHECKBOX, self.Check5 )
    	self.checkBox92.Bind( wx.EVT_CHECKBOX, self.Check6 )
    	self.m_checkBox93.Bind( wx.EVT_CHECKBOX, self.Check7 )
    	self.button19.Bind( wx.EVT_BUTTON, self.Reset )
    	self.button20.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
    	pass
    
    
    # Virtual event handlers, overide them in your derived class
    def Check1( self, event ):
            #event.Skip()
            tick92 = LoadWorkbook1(parent=self.checkBox87)
            mtick92 = tick92.wb1.get_sheet_by_name('Audit Checklist Q2')
            put92 = mtick92['H81']
            box92 = self.checkBox87.GetValue()
    
            if box92 == True:
                put92.value = "/"
    
            else:
                put92.value = ""
    
    def Check2( self, event ):
            #event.Skip()
            tick93 = LoadWorkbook1(parent=self.checkBox88)
            mtick93 = tick93.wb1.get_sheet_by_name('Audit Checklist Q2')
            put93 = mtick93['H82']
            box93 = self.checkBox88.GetValue()
    
            if box93 == True:
                put93.value = "/"
    
            else:
                put93.value = ""
    
    def Check3( self, event ):
            #event.Skip()
            tick94 = LoadWorkbook1(parent=self.checkBox89)
            mtick94 = tick94.wb1.get_sheet_by_name('Audit Checklist Q2')
            put94 = mtick94['H83']
            box94 = self.checkBox89.GetValue()
    
            if box94 == True:
                put94.value = "/"
    
            else:
                put94.value = ""
    
    def Check4( self, event ):
            #event.Skip()
            tick95 = LoadWorkbook1(parent=self.checkBox90)
            mtick95 = tick95.wb1.get_sheet_by_name('Audit Checklist Q2')
            put95 = mtick95['H84']
            box95 = self.checkBox90.GetValue()
    
            if box95 == True:
                put95.value = "/"
    
            else:
                put95.value = ""
    
    def Check5( self, event ):
            #event.Skip()
            tick96 = LoadWorkbook1(parent=self.checkBox91)
            mtick96 = tick96.wb1.get_sheet_by_name('Audit Checklist Q2')
            put96 = mtick96['H85']
            box96 = self.checkBox91.GetValue()
    
            if box96 == True:
                put96.value = "/"
    
            else:
                put96.value = ""
    
    def Check6( self, event ):
            #event.Skip()
            tick97 = LoadWorkbook1(parent=self.checkBox92)
            mtick97 = tick97.wb1.get_sheet_by_name('Audit Checklist Q2')
            put97 = mtick97['H86']
            box97 = self.checkBox92.GetValue()
    
            if box97 == True:
                put97.value = "/"
    
            else:
                put97.value = ""
    
    def Check7( self, event ):
            #event.Skip()
            tick98 = LoadWorkbook1(parent=self.m_checkBox93)
            mtick98 = tick98.wb1.get_sheet_by_name('Audit Checklist Q2')
            put98 = mtick98['C87']
            box98 = self.m_checkBox93.GetValue()
    
            if box98 == True:
                put98.value = "/"
    
            else:
                put98.value = ""
    
    def Reset( self, event ):
    		#event.Skip()
            self.checkBox87.SetValue(False)
            self.checkBox88.SetValue(False)
            self.checkBox89.SetValue(False)
            self.checkBox90.SetValue(False)
            self.checkBox91.SetValue(False)
            self.checkBox92.SetValue(False)
            self.m_checkBox93.SetValue(False)
    
    def Save( self, event ):
            #event.Skip()
            check92 = self.checkBox87.GetValue()
            check93 = self.checkBox88.GetValue()
            check94 = self.checkBox89.GetValue()
            check95 = self.checkBox90.GetValue()
            check96 = self.checkBox91.GetValue()
            check97 = self.checkBox92.GetValue()
            check98 = self.m_checkBox93.GetValue()
    
            if check92 == True:
                a92 = 1
            else:
                a92 = 0
    
            if check93 == True:
                a93 = 1
            else:
                a93 = 0
    
            if check94 == True:
                a94 = 1
            else:
                a94 = 0
    
            if check95 == True:
                a95 = 1
            else:
                a95 = 0
    
            if check96 == True:
                a96 = 1
            else:
                a96 = 0
    
            if check97 == True:
                a97 = 1
            else:
                a97 = 0
    
            if check98 == True:
                a98 = 1
            else:
                a98 = 0



###########################################################################
## Class ChecklistNew
###########################################################################

class ChecklistNew ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Checklist New", pos = wx.DefaultPosition, size = wx.Size( 500,441 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
    
    	self.SetSizeHintsSz( wx.Size(500,441), wx.Size(500,441))
    	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )
    
    	bSizer92 = wx.BoxSizer( wx.VERTICAL )
    
    
    	bSizer92.Add( ( 0, 15), 0, 0, 5 )
    
    	bSizer254 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer254.Add( ( 170, 0), 0, 0, 5 )
    
    	self.staticText223 = wx.StaticText( self, wx.ID_ANY, u"State Date", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText223.Wrap( -1 )
    	self.staticText223.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )
    
    	bSizer254.Add( self.staticText223, 0, wx.ALL, 5 )
    
    	self.datePicker2 = wx.adv.DatePickerCtrl( self, wx.ID_ANY, wx.DefaultDateTime, wx.DefaultPosition, wx.DefaultSize, wx.adv.DP_DEFAULT )
    	self.datePicker2.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )
    
    	bSizer254.Add( self.datePicker2, 0, wx.ALL, 5 )
    
    
    	bSizer92.Add( bSizer254, 1, wx.EXPAND, 5 )
    
    	bSizer93 = wx.BoxSizer( wx.VERTICAL )
    
    	self.button26 = wx.Button( self, wx.ID_ANY, u"Clause 7.1 \nGeneral Requirement", wx.DefaultPosition, wx.Size( 160,60 ), 0 )
    	self.button26.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )
    
    	bSizer93.Add( self.button26, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )
    
    	self.button27 = wx.Button( self, wx.ID_ANY, u"Clause 7.3\nRemanufacturing SOP", wx.DefaultPosition, wx.Size( 160,-1 ), 0 )
    	self.button27.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )
    
    	bSizer93.Add( self.button27, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )
    
    	self.button28 = wx.Button( self, wx.ID_ANY, u"Clause 9.1 \nLabelling and packaging", wx.DefaultPosition, wx.Size( 160,-1 ), 0 )
    	self.button28.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )
    
    	bSizer93.Add( self.button28, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )
    
    	self.button29 = wx.Button( self, wx.ID_ANY, u"Clause 10\nWarranty", wx.DefaultPosition, wx.Size( 160,-1 ), 0 )
    	self.button29.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )
    
    	bSizer93.Add( self.button29, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )
    
    	self.button30 = wx.Button( self, wx.ID_ANY, u"Clause 11\nSupplier mark", wx.DefaultPosition, wx.Size( 160,-1 ), 0 )
    	self.button30.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )
    
    	bSizer93.Add( self.button30, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )
    
    
    	bSizer92.Add( bSizer93, 1, wx.EXPAND, 5 )
    
    	bSizer94 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer94.Add( ( 0, 0), 1, wx.EXPAND, 5 )
    
    	self.button31 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button31.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer94.Add( self.button31, 0, wx.ALL, 5 )
    
    
    	bSizer92.Add( bSizer94, 0, wx.EXPAND, 5 )
    
    
    	self.SetSizer( bSizer92 )
    	self.Layout()
    
    	self.Centre( wx.BOTH )
    
    	# Connect Events
    	self.datePicker2.Bind( wx.adv.EVT_DATE_CHANGED, self.Date2 )
    	self.button26.Bind( wx.EVT_BUTTON, self.Clause1 )
    	self.button27.Bind( wx.EVT_BUTTON, self.Clause2 )
    	self.button28.Bind( wx.EVT_BUTTON, self.Clause3 )
    	self.button29.Bind( wx.EVT_BUTTON, self.Clause4 )
    	self.button30.Bind( wx.EVT_BUTTON, self.Clause5 )
    	self.button31.Bind( wx.EVT_BUTTON, self.Reset )
    
    def __del__( self ):
    	pass
    
    
    # Virtual event handlers, overide them in your derived class
    def Date2( self, event ):
    	event.Skip()
    
    def Clause1( self, event ):
            #event.Skip()
            Clause1Form = Clause71(parent = self.button26)
            Clause1Form.Show()
    
    def Clause2( self, event ):
            #event.Skip()
            Clause2Form = Clause73(parent = self.button27)
            Clause2Form.Show()
    
    def Clause3( self, event ):
            #event.Skip()
            Clause3Form = Clause91a(parent = self.button28)
            Clause3Form.Show()
    
    def Clause4( self, event ):
            #event.Skip()
            Clause4Form = Clause10(parent = self.button29)
            Clause4Form.Show()
    
    def Clause5( self, event ):
            #event.Skip()
            Clause5Form = Clause11(parent = self.button30)
            Clause5Form.Show()
    
    def Reset( self, event ):
    	event.Skip()


###########################################################################
## Class Clause71
###########################################################################
class Clause71 ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"General Requirement", pos = wx.DefaultPosition, size = wx.Size( 1055,591 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
    
    	self.SetSizeHintsSz( wx.Size(1055,591), wx.Size(1055,591) )
    	self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
    
    	bSizer95 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer117 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText64.Wrap( -1 )
    	self.staticText64.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
    
    	bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )
    
    
    	bSizer117.Add( ( 370, 0), 0, 0, 5 )
    
    	self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText65.Wrap( -1 )
    	self.staticText65.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
    
    	bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )
    
    
    	bSizer95.Add( bSizer117, 1, wx.EXPAND, 5 )
    
    	bSizer116 = wx.BoxSizer( wx.HORIZONTAL )
    
    	bSizer104 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer96 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.1", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText49.Wrap( -1 )
    	self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )
    
    
    	bSizer96.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"Handling remanufacturing of the\nused parts and components", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText103.Wrap( -1 )
    	self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )
    
    
    	bSizer96.Add( ( 90, 0), 1, 0, 5 )
    
    	self.staticText50 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing\nof the used parts and components shall \nbe to  demonstrate that  it has the legal \nright to transfer their ownership to \nanother party", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText50.Wrap( -1 )
    	self.staticText50.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer96.Add( self.staticText50, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer96, 0, 0, 5 )
    
    	bSizer97 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.2", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText51.Wrap( -1 )
    	self.staticText51.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer97.Add( self.staticText51, 0, wx.ALL, 5 )
    
    
    	bSizer97.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText105 = wx.StaticText( self, wx.ID_ANY, u"Provide specified information \nof remanufactured parts\nand components", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText105.Wrap( -1 )
    	self.staticText105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer97.Add( self.staticText105, 0, wx.ALL, 5 )
    
    
    	bSizer97.Add( ( 105, 0), 0, 0, 5 )
    
    	self.staticText52 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing\nof the used parts and components shall \nbe to  demonstrate that  it has the legal \nright to transfer their ownership to \nanother party", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText52.Wrap( -1 )
    	self.staticText52.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer97.Add( self.staticText52, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer97, 1, wx.EXPAND, 5 )
    
    	bSizer98 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText57 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.3", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText57.Wrap( -1 )
    	self.staticText57.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer98.Add( self.staticText57, 0, wx.ALL, 5 )
    
    
    	bSizer98.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText106 = wx.StaticText( self, wx.ID_ANY, u"Manpower certification ", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText106.Wrap( -1 )
    	self.staticText106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer98.Add( self.staticText106, 0, wx.ALL, 5 )
    
    
    	bSizer98.Add( ( 147, 0), 0, 0, 5 )
    
    	self.staticText58 = wx.StaticText( self, wx.ID_ANY, u"Person dismantling, cleaning, examining,\nremediating, re-assembling, testing and \nhandling of the core and remanufactured\nparts or components shall be of \ntechnical specialist", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText58.Wrap( -1 )
    	self.staticText58.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer98.Add( self.staticText58, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer98, 0, wx.EXPAND, 5 )
    
    	bSizer100 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText59 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.4", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText59.Wrap( -1 )
    	self.staticText59.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText59, 0, wx.ALL, 5 )
    
    
    	bSizer100.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText107 = wx.StaticText( self, wx.ID_ANY, u"Tool and Equipment", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText107.Wrap( -1 )
    	self.staticText107.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText107, 0, wx.ALL, 5 )
    
    
    	bSizer100.Add( ( 165, 0), 0, 0, 5 )
    
    	self.staticText60 = wx.StaticText( self, wx.ID_ANY, u"The facilities of the company involved\nin the remanufacturing process shall have\nappropriate tools and equipment", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText60.Wrap( -1 )
    	self.staticText60.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText60, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer100, 1, wx.EXPAND, 5 )
    
    	bSizer101 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText61 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.5", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText61.Wrap( -1 )
    	self.staticText61.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer101.Add( self.staticText61, 0, wx.ALL, 5 )
    
    
    	bSizer101.Add( ( 60, 0), 0, 0, 5 )
    
    	self.m_staticText108 = wx.StaticText( self, wx.ID_ANY, u"Statutory and regulatory\nrequirements and \nindustry best practices", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.m_staticText108.Wrap( -1 )
    	self.m_staticText108.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer101.Add( self.m_staticText108, 0, wx.ALL, 5 )
    
    
    	bSizer101.Add( ( 140, 0), 0, 0, 5 )
    
    	self.staticText62 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing shall be performed in     \naccordance to the relevant statutory\nand regulatory requirements and\nindustry best practices.", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText62.Wrap( -1 )
    	self.staticText62.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer101.Add( self.staticText62, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer101, 1, wx.EXPAND, 5 )
    
    
    	bSizer116.Add( bSizer104, 1, wx.EXPAND, 5 )
    
    	bSizer105 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer106 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox93 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox93.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox93, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox94 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox94.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox94, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox95 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox95.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox95, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    
    	bSizer105.Add( bSizer106, 1, wx.EXPAND, 5 )
    
    	bSizer107 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox96 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox96.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox96, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox97 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox97.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox97, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox98 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox98.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox98, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    
    	bSizer105.Add( bSizer107, 1, wx.EXPAND, 5 )
    
    	bSizer108 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox99 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox99.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox99, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox100 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox100.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox100, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox101 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox101.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox101, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    
    	bSizer105.Add( bSizer108, 1, wx.EXPAND, 5 )
    
    	bSizer114 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox102 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox102.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox102, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox103 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox103, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox104 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox104.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox104, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    
    	bSizer105.Add( bSizer114, 1, wx.EXPAND, 5 )
    
    	bSizer115 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox105 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox105, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox106 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox106, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    	self.checkBox107 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox107.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox107, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
    
    
    	bSizer105.Add( bSizer115, 1, wx.EXPAND, 5 )
    
    
    	bSizer116.Add( bSizer105, 1, wx.EXPAND, 5 )
    
    
    	bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )
    
    	bSizer103 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )
    
    	self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer103.Add( self.button33, 0, wx.ALL, 5 )
    
    	self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer103.Add( self.button34, 0, wx.ALL, 5 )
    
    
    	bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )
    
    
    	self.SetSizer( bSizer95 )
    	self.Layout()
    
    	self.Centre( wx.BOTH )
    
    	# Connect Events
    	self.checkBox93.Bind( wx.EVT_CHECKBOX, self.check1 )
    	self.checkBox94.Bind( wx.EVT_CHECKBOX, self.check2 )
    	self.checkBox95.Bind( wx.EVT_CHECKBOX, self.check3 )
    	self.checkBox96.Bind( wx.EVT_CHECKBOX, self.check4 )
    	self.checkBox97.Bind( wx.EVT_CHECKBOX, self.check5 )
    	self.checkBox98.Bind( wx.EVT_CHECKBOX, self.check6 )
    	self.checkBox99.Bind( wx.EVT_CHECKBOX, self.check7 )
    	self.checkBox100.Bind( wx.EVT_CHECKBOX, self.check8 )
    	self.checkBox101.Bind( wx.EVT_CHECKBOX, self.check9 )
    	self.checkBox102.Bind( wx.EVT_CHECKBOX, self.check10 )
    	self.checkBox103.Bind( wx.EVT_CHECKBOX, self.check11 )
    	self.checkBox104.Bind( wx.EVT_CHECKBOX, self.check12 )
    	self.checkBox105.Bind( wx.EVT_CHECKBOX, self.check13 )
    	self.checkBox106.Bind( wx.EVT_CHECKBOX, self.check14 )
    	self.checkBox107.Bind( wx.EVT_CHECKBOX, self.check15 )
    	self.button33.Bind( wx.EVT_BUTTON, self.Reset )
    	self.button34.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
    	pass
    
    
    # Virtual event handlers, overide them in your derived class
    def check1( self, event ):
            #event.Skip()
            tick99 = LoadWorkbook2(parent=self.checkBox93)
            mtick99 = tick99.wb2.get_sheet_by_name('Audit Checklist')
            put99 = mtick99['E11']
            box99 = self.checkBox93.GetValue()
    
            if box99 == True:
                put99.value = "/"
            else:
                put99.value = ""
    
    def check2( self, event ):
            #event.Skip()
            tick100 = LoadWorkbook2(parent=self.checkBox94)
            mtick100 = tick100.wb2.get_sheet_by_name('Audit Checklist')
            put100 = mtick100['F11']
            box100 = self.checkBox94.GetValue()
    
            if box100 == True:
                put100.value = "/"
            else:
                put100.value = ""
    
    def check3( self, event ):
            #event.Skip()
            tick101 = LoadWorkbook2(parent=self.checkBox95)
            mtick101 = tick101.wb2.get_sheet_by_name('Audit Checklist')
            put101 = mtick101['G11']
            box101 = self.checkBox95.GetValue()
    
            if box101 == True:
                put101.value = "/"
            else:
                put101.value = ""
    
    def check4( self, event ):
            #event.Skip()
            tick102 = LoadWorkbook2(parent=self.checkBox96)
            mtick102 = tick102.wb2.get_sheet_by_name('Audit Checklist')
            put102 = mtick102['E13']
            box102 = self.checkBox96.GetValue()
    
            if box102 == True:
                put102.value = "/"
            else:
                put102.value = ""
    
    def check5( self, event ):
            #event.Skip()
            tick103 = LoadWorkbook2(parent=self.checkBox97)
            mtick103 = tick103.wb2.get_sheet_by_name('Audit Checklist')
            put103 = mtick103['F13']
            box103 = self.checkBox97.GetValue()
    
            if box103 == True:
                put103.value = "/"
            else:
                put103.value = ""
    
    def check6( self, event ):
            #event.Skip()
            tick104 = LoadWorkbook2(parent=self.checkBox98)
            mtick104 = tick104.wb2.get_sheet_by_name('Audit Checklist')
            put104 = mtick104['G13']
            box104 = self.checkBox104.GetValue()
    
            if box104 == True:
                put104.value = "/"
            else:
                put104.value = ""
    
    def check7( self, event ):
            #event.Skip()
            tick105 = LoadWorkbook2(parent=self.checkBox99)
            mtick105 = tick105.wb2.get_sheet_by_name('Audit Checklist')
            put105 = mtick105['E15']
            box105 = self.checkBox99.GetValue()
    
            if box105 == True:
                put105.value = "/"
            else:
                put105.value = ""
    
    def check8( self, event ):
            #event.Skip()
            tick106 = LoadWorkbook2(parent=self.checkBox100)
            mtick106 = tick106.wb2.get_sheet_by_name('Audit Checklist')
            put106 = mtick106['F15']
            box106 = self.checkBox106.GetValue()
    
            if box106 == True:
                put106.value = "/"
            else:
                put106.value = ""
    
    def check9( self, event ):
            #event.Skip()
            tick107 = LoadWorkbook2(parent=self.checkBox101)
            mtick107 = tick107.wb2.get_sheet_by_name('Audit Checklist')
            put107 = mtick107['G15']
            box107 = self.checkBox107.GetValue()
    
            if box107 == True:
                put107.value = "/"
            else:
                put107.value = ""
    
    def check10( self, event ):
            #event.Skip()
            tick108 = LoadWorkbook2(parent=self.checkBox102)
            mtick108 = tick108.wb2.get_sheet_by_name('Audit Checklist')
            put108 = mtick108['E17']
            box108 = self.checkBox102.GetValue()
    
            if box108 == True:
                put108.value = "/"
            else:
                put108.value = ""
    
    def check11( self, event ):
            #event.Skip()
            tick109 = LoadWorkbook2(parent=self.checkBox103)
            mtick109 = tick109.wb2.get_sheet_by_name('Audit Checklist')
            put109 = mtick109['F17']
            box109 = self.checkBox103.GetValue()
    
            if box109 == True:
                put109.value = "/"
            else:
                put109.value = ""
    
    def check12( self, event ):
            #event.Skip()
            tick110 = LoadWorkbook2(parent=self.checkBox104)
            mtick110 = tick110.wb2.get_sheet_by_name('Audit Checklist')
            put110 = mtick110['G17']
            box110 = self.checkBox104.GetValue()
    
            if box110 == True:
                put110.value = "/"
            else:
                put110.value = ""
    
    def check13( self, event ):
            #event.Skip()
            tick111 = LoadWorkbook2(parent=self.checkBox105)
            mtick111 = tick111.wb2.get_sheet_by_name('Audit Checklist')
            put111 = mtick111['E19']
            box111 = self.checkBox105.GetValue()
    
            if box111 == True:
                put111.value = "/"
            else:
                put111.value = ""
    
    def check14( self, event ):
            #event.Skip()
            tick112 = LoadWorkbook2(parent=self.checkBox106)
            mtick112 = tick112.wb2.get_sheet_by_name('Audit Checklist')
            put112 = mtick112['F19']
            box112 = self.checkBox106.GetValue()
    
            if box112 == True:
                put112.value = "/"
            else:
                put112.value = ""
    
    def check15( self, event ):
            #event.Skip()
            tick113 = LoadWorkbook2(parent=self.checkBox107)
            mtick113 = tick113.wb2.get_sheet_by_name('Audit Checklist')
            put113 = mtick113['G19']
            box113 = self.checkBox107.GetValue()
    
            if box113 == True:
                put113.value = "/"
            else:
                put113.value = ""
    
    def Reset( self, event ):
    		#event.Skip()
            self.checkBox93.SetValue(False)
            self.checkBox94.SetValue(False)
            self.checkBox95.SetValue(False)
            self.checkBox96.SetValue(False)
            self.checkBox97.SetValue(False)
            self.checkBox98.SetValue(False)
            self.checkBox99.SetValue(False)
            self.checkBox100.SetValue(False)
            self.checkBox101.SetValue(False)
            self.checkBox102.SetValue(False)
            self.checkBox103.SetValue(False)
            self.checkBox104.SetValue(False)
            self.checkBox105.SetValue(False)
            self.checkBox106.SetValue(False)
            self.checkBox107.SetValue(False)
    
    def Save( self, event ):
            #event.Skip()
            check99 = self.checkBox93.GetValue()
            check100 = self.checkBox94.GetValue()
            check101 = self.checkBox95.GetValue()
            check102 = self.checkBox96.GetValue()
            check103 = self.checkBox97.GetValue()
            check104 = self.checkBox98.GetValue()
            check105 = self.checkBox99.GetValue()
            check106 = self.checkBox100.GetValue()
            check107 = self.checkBox101.GetValue()
            check108 = self.checkBox102.GetValue()
            check109 = self.checkBox103.GetValue()
            check110 = self.checkBox104.GetValue()
            check111 = self.checkBox105.GetValue()
            check112 = self.checkBox106.GetValue()
            check113 = self.checkBox107.GetValue()
    
            if check99 == True:
                b1 = 1
            else:
                b1 = 0
    
            if check100 == True:
                b2 = 1
            else:
                b3 = 0
    
            if check101 == True:
                b4 = 1
            else:
                b4 = 0
    
            if check102 == True:
                b5 = 1
            else:
                b5 = 0
    
            if check103 == True:
                b6 = 1
            else:
                b6 = 0
    
            if check104 == True:
                b7 = 1
            else:
                b7 = 0
    
            if check105 == True:
                b8 = 1
            else:
                b8 = 0
    
            if check106 == True:
                b9 = 1
            else:
                b9 = 0
    
            if check107 == True:
                b10 = 1
            else:
                b10 = 0
    
            if check108 == True:
                b11 = 1
            else:
                b11 = 0
    
            if check109 == True:
                b12 = 1
            else:
                b12 = 0
    
            if check110 == True:
                b13 = 1
            else:
                b13 = 0
    
            if check111 == True:
                b14 = 1
            else:
                b14 = 0
    
            if check112 == True:
                b15 = 1
            else:
                b15 = 0
    
            if check113 == True:
                b16 = 1
            else:
                b16 = 0


###########################################################################
## Class Clause73
###########################################################################

class Clause73 ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Remanufacturing SOP", pos = wx.DefaultPosition, size = wx.Size( 1148,1005 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
    
    	self.SetSizeHintsSz( wx.Size(1148,1005), wx.Size(1148,1005) )
    	self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
    
    	bSizer95 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer117 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText64.Wrap( -1 )
    	self.staticText64.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
    
    	bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )
    
    
    	bSizer117.Add( ( 360, 0), 0, 0, 5 )
    
    	self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText65.Wrap( -1 )
    	self.staticText65.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
    
    	bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )
    
    
    	bSizer95.Add( bSizer117, 1, wx.EXPAND, 5 )
    
    	bSizer116 = wx.BoxSizer( wx.HORIZONTAL )
    
    	bSizer104 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer96 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText49.Wrap( -1 )
    	self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )
    
    
    	bSizer96.Add( ( 78, 0), 0, 0, 5 )
    
    	self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"The remanufacturing process\ninvolves:", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText103.Wrap( -1 )
    	self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )
    
    
    	bSizer96.Add( ( 90, 0), 1, 0, 5 )
    
    	self.staticText50 = wx.StaticText( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText50.Wrap( -1 )
    	bSizer96.Add( self.staticText50, 1, wx.ALL|wx.EXPAND, 5 )
    
    
    	bSizer104.Add( bSizer96, 0, 0, 5 )
    
    	bSizer97 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(a)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText51.Wrap( -1 )
    	self.staticText51.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer97.Add( self.staticText51, 0, wx.ALL, 5 )
    
    
    	bSizer97.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText105 = wx.StaticText( self, wx.ID_ANY, u"Core management, core \nsorting, dismantling", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText105.Wrap( -1 )
    	self.staticText105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer97.Add( self.staticText105, 0, wx.ALL, 5 )
    
    
    	bSizer97.Add( ( 104, 0), 0, 0, 5 )
    
    	self.staticText52 = wx.StaticText( self, wx.ID_ANY, u"The core management process is fully \nimplemented in terms of traceability\nrecord of the parts and components.", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText52.Wrap( -1 )
    	self.staticText52.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer97.Add( self.staticText52, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer97, 1, wx.EXPAND, 5 )
    
    	bSizer98 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText57 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(b)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText57.Wrap( -1 )
    	self.staticText57.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer98.Add( self.staticText57, 0, wx.ALL, 5 )
    
    
    	bSizer98.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText106 = wx.StaticText( self, wx.ID_ANY, u"Cleaning of all internal \nand external components", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText106.Wrap( -1 )
    	self.staticText106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer98.Add( self.staticText106, 0, wx.ALL, 5 )
    
    
    	bSizer98.Add( ( 104, 0), 0, 0, 5 )
    
    	self.staticText58 = wx.StaticText( self, wx.ID_ANY, u"Cleaning is required for used\nparts and components to ensure\nthat they are free from any impurities\nand hazardous materials.\n", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText58.Wrap( -1 )
    	self.staticText58.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer98.Add( self.staticText58, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer98, 0, wx.EXPAND, 5 )
    
    	bSizer100 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText59 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(c)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText59.Wrap( -1 )
    	self.staticText59.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText59, 0, wx.ALL, 5 )
    
    
    	bSizer100.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText107 = wx.StaticText( self, wx.ID_ANY, u"Replacement of parts and\ncomponent", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText107.Wrap( -1 )
    	self.staticText107.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText107, 0, wx.ALL, 5 )
    
    
    	bSizer100.Add( ( 100, 0), 0, 0, 5 )
    
    	self.staticText60 = wx.StaticText( self, wx.ID_ANY, u"All missing parts, restoration of all\nimpaired, defective or substantially\nworn parts to a sound condition or\nreplacement thereof", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText60.Wrap( -1 )
    	self.staticText60.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText60, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer100, 1, wx.EXPAND, 5 )
    
    	bSizer101 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText61 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(d)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText61.Wrap( -1 )
    	self.staticText61.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer101.Add( self.staticText61, 0, wx.ALL, 5 )
    
    
    	bSizer101.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText108 = wx.StaticText( self, wx.ID_ANY, u"Reworking and machining ", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText108.Wrap( -1 )
    	self.staticText108.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer101.Add( self.staticText108, 0, wx.ALL, 5 )
    
    
    	bSizer101.Add( ( 95, 0), 0, 0, 5 )
    
    	self.staticText62 = wx.StaticText( self, wx.ID_ANY, u"Performing such other operations as\nare necessary to put the\npart in original\nworking condition or better", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText62.Wrap( -1 )
    	self.staticText62.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer101.Add( self.staticText62, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer101, 1, wx.EXPAND, 5 )
    
    	bSizer218 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText160 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(e)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText160.Wrap( -1 )
    	self.staticText160.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer218.Add( self.staticText160, 0, wx.ALL, 5 )
    
    
    	bSizer218.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText162 = wx.StaticText( self, wx.ID_ANY, u"Component assembly\n", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText162.Wrap( -1 )
    	self.staticText162.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer218.Add( self.staticText162, 0, wx.ALL, 5 )
    
    
    	bSizer218.Add( ( 125, 0), 0, 0, 5 )
    
    	self.staticText161 = wx.StaticText( self, wx.ID_ANY, u"Assembly done according to\nmanufacturer specification\n", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText161.Wrap( -1 )
    	self.staticText161.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer218.Add( self.staticText161, 0, wx.ALL|wx.EXPAND, 5 )
    
    
    	bSizer104.Add( bSizer218, 1, wx.EXPAND, 5 )
    
    	bSizer1011 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText611 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(f) \nand\nClause 7.3.6", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText611.Wrap( -1 )
    	self.staticText611.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1011.Add( self.staticText611, 0, wx.ALL, 5 )
    
    
    	bSizer1011.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText1081 = wx.StaticText( self, wx.ID_ANY, u"Final testing of each\nremanufactured part\n", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText1081.Wrap( -1 )
    	self.staticText1081.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1011.Add( self.staticText1081, 0, wx.ALL, 5 )
    
    
    	bSizer1011.Add( ( 135, 0), 0, 0, 5 )
    
    	self.staticText621 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing shall be performed in     \naccordance to the relevant statutory\nand regulatory requirements and\nindustry best practices.", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText621.Wrap( -1 )
    	self.staticText621.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1011.Add( self.staticText621, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer1011, 1, wx.EXPAND, 5 )
    
    	bSizer2181 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText1601 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.2", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText1601.Wrap( -1 )
    	self.staticText1601.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer2181.Add( self.staticText1601, 0, wx.ALL, 5 )
    
    
    	bSizer2181.Add( ( 75, 0), 0, 0, 5 )
    
    	self.staticText1621 = wx.StaticText( self, wx.ID_ANY, u"Identification mark", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText1621.Wrap( -1 )
    	self.staticText1621.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer2181.Add( self.staticText1621, 0, wx.ALL, 5 )
    
    
    	bSizer2181.Add( ( 153, 0), 0, 0, 5 )
    
    	self.staticText1611 = wx.StaticText( self, wx.ID_ANY, u"Any remanufactured parts and components that\nare regulated by relevant authorities shall be\naffixed with an identification mark. This\nidentification mark shall identify the source\nof the used parts and components. A detailed\nrecord of the identification mark shall be maintained", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText1611.Wrap( -1 )
    	self.staticText1611.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer2181.Add( self.staticText1611, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer2181, 1, wx.EXPAND, 5 )
    
    	bSizer971 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText511 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.3", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText511.Wrap( -1 )
    	self.staticText511.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer971.Add( self.staticText511, 0, wx.ALL, 5 )
    
    
    	bSizer971.Add( ( 75, 0), 0, 0, 5 )
    
    	self.staticText1051 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing permanent\nmark", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText1051.Wrap( -1 )
    	self.staticText1051.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer971.Add( self.staticText1051, 0, wx.ALL, 5 )
    
    
    	bSizer971.Add( ( 85, 0), 0, 0, 5 )
    
    	self.staticText521 = wx.StaticText( self, wx.ID_ANY, u"All remanufactured part shall\nhave a non-removable and\npermanent marking to identify\nit as a remanufactured product", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText521.Wrap( -1 )
    	self.staticText521.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer971.Add( self.staticText521, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer971, 1, wx.EXPAND, 5 )
    
    	bSizer972 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText512 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.4", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText512.Wrap( -1 )
    	self.staticText512.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer972.Add( self.staticText512, 0, wx.ALL, 5 )
    
    
    	bSizer972.Add( ( 73, 0), 0, 0, 5 )
    
    	self.staticText1052 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing records", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText1052.Wrap( -1 )
    	self.staticText1052.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer972.Add( self.staticText1052, 0, wx.ALL, 5 )
    
    
    	bSizer972.Add( ( 107, 0), 0, 0, 5 )
    
    	self.staticText522 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing process shall be conducted\nin accordance to an established industrialised\nprocess, guided by industry best practices\nand shall be fully documented and recorded", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText522.Wrap( -1 )
    	self.staticText522.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer972.Add( self.staticText522, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer972, 1, wx.EXPAND, 5 )
    
    	bSizer973 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText513 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.5", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText513.Wrap( -1 )
    	self.staticText513.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer973.Add( self.staticText513, 0, wx.ALL, 5 )
    
    
    	bSizer973.Add( ( 75, 0), 0, 0, 5 )
    
    	self.staticText1053 = wx.StaticText( self, wx.ID_ANY, u"Authorised from person\nin charge", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText1053.Wrap( -1 )
    	self.staticText1053.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer973.Add( self.staticText1053, 0, wx.ALL, 5 )
    
    
    	bSizer973.Add( ( 117, 0), 0, 0, 5 )
    
    	self.staticText523 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing of the\nused parts and components shall be to \ndemonstrate that  it has the legal right to\ntransfer their ownership to another party", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText523.Wrap( -1 )
    	self.staticText523.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer973.Add( self.staticText523, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer973, 1, wx.EXPAND, 5 )
    
    
    	bSizer116.Add( bSizer104, 1, wx.EXPAND, 5 )
    
    	bSizer105 = wx.BoxSizer( wx.VERTICAL )
    
    
    	bSizer105.Add( ( 0, 50), 0, 0, 5 )
    
    	bSizer106 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox108 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox108.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox108, 0, wx.ALL, 5 )
    
    	self.checkBox109 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox109.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox109, 0, wx.ALL, 5 )
    
    	self.checkBox110 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox110.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox110, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer106, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 60), 0, 0, 5 )
    
    	bSizer107 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox111 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox111.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox111, 0, wx.ALL, 5 )
    
    	self.checkBox112 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox112.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox112, 0, wx.ALL, 5 )
    
    	self.checkBox113 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox113.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox113, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer107, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 60), 0, 0, 5 )
    
    	bSizer108 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox114 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox114.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox114, 0, wx.ALL, 5 )
    
    	self.checkBox115 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox115.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox115, 0, wx.ALL, 5 )
    
    	self.checkBox116 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox116.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox116, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer108, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 50), 0, 0, 5 )
    
    	bSizer114 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox117 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox117.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox117, 0, wx.ALL, 5 )
    
    	self.checkBox118 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox118.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox118, 0, wx.ALL, 5 )
    
    	self.checkBox119 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox119.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox119, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer114, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 50), 0, 0, 5 )
    
    	bSizer115 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox120 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox120.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox120, 0, wx.ALL, 5 )
    
    	self.checkBox121 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox121.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox121, 0, wx.ALL, 5 )
    
    	self.checkBox122 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox122.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox122, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer115, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 50), 0, 0, 5 )
    
    	bSizer1061 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox123 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox123.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1061.Add( self.checkBox123, 0, wx.ALL, 5 )
    
    	self.checkBox124 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox124.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1061.Add( self.checkBox124, 0, wx.ALL, 5 )
    
    	self.checkBox125 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox125.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1061.Add( self.checkBox125, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer1061, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 60), 0, 0, 5 )
    
    	bSizer1062 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox126 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox126.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1062.Add( self.checkBox126, 0, wx.ALL, 5 )
    
    	self.checkBox127 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox127.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1062.Add( self.checkBox127, 0, wx.ALL, 5 )
    
    	self.checkBox128 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox128.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1062.Add( self.checkBox128, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer1062, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 80), 0, 0, 5 )
    
    	bSizer1063 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox129 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox129.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1063.Add( self.checkBox129, 0, wx.ALL, 5 )
    
    	self.checkBox130 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox130.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1063.Add( self.checkBox130, 0, wx.ALL, 5 )
    
    	self.checkBox131 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox131.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1063.Add( self.checkBox131, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer1063, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 50), 0, 0, 5 )
    
    	bSizer1064 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox132 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox132.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1064.Add( self.checkBox132, 0, wx.ALL, 5 )
    
    	self.checkBox133 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox133.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1064.Add( self.checkBox133, 0, wx.ALL, 5 )
    
    	self.checkBox134 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox134.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1064.Add( self.checkBox134, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer1064, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 60), 0, 0, 5 )
    
    	bSizer1065 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox135 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox135.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1065.Add( self.checkBox135, 0, wx.ALL, 5 )
    
    	self.checkBox136 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox136.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1065.Add( self.checkBox136, 0, wx.ALL, 5 )
    
    	self.checkBox137 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox137.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1065.Add( self.checkBox137, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer1065, 1, wx.EXPAND, 5 )
    
    
    	bSizer116.Add( bSizer105, 1, 0, 5 )
    
    
    	bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )
    
    	bSizer103 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )
    
    	self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer103.Add( self.button33, 0, wx.ALL, 5 )
    
    	self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer103.Add( self.button34, 0, wx.ALL, 5 )
    
    
    	bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )
    
    
    	self.SetSizer( bSizer95 )
    	self.Layout()
    
    	self.Centre( wx.BOTH )
    
    	# Connect Events
    	self.checkBox108.Bind( wx.EVT_CHECKBOX, self.check1 )
    	self.checkBox109.Bind( wx.EVT_CHECKBOX, self.check2 )
    	self.checkBox110.Bind( wx.EVT_CHECKBOX, self.check3 )
    	self.checkBox111.Bind( wx.EVT_CHECKBOX, self.check4 )
    	self.checkBox112.Bind( wx.EVT_CHECKBOX, self.check5 )
    	self.checkBox113.Bind( wx.EVT_CHECKBOX, self.check6 )
    	self.checkBox114.Bind( wx.EVT_CHECKBOX, self.check7 )
    	self.checkBox115.Bind( wx.EVT_CHECKBOX, self.check8 )
    	self.checkBox116.Bind( wx.EVT_CHECKBOX, self.check9 )
    	self.checkBox117.Bind( wx.EVT_CHECKBOX, self.check10 )
    	self.checkBox118.Bind( wx.EVT_CHECKBOX, self.check11 )
    	self.checkBox119.Bind( wx.EVT_CHECKBOX, self.check12 )
    	self.checkBox120.Bind( wx.EVT_CHECKBOX, self.check13 )
    	self.checkBox121.Bind( wx.EVT_CHECKBOX, self.check14 )
    	self.checkBox122.Bind( wx.EVT_CHECKBOX, self.check15 )
    	self.checkBox123.Bind( wx.EVT_CHECKBOX, self.check16 )
    	self.checkBox124.Bind( wx.EVT_CHECKBOX, self.check17 )
    	self.checkBox125.Bind( wx.EVT_CHECKBOX, self.check18 )
    	self.checkBox126.Bind( wx.EVT_CHECKBOX, self.check19 )
    	self.checkBox127.Bind( wx.EVT_CHECKBOX, self.check20 )
    	self.checkBox128.Bind( wx.EVT_CHECKBOX, self.check21 )
    	self.checkBox129.Bind( wx.EVT_CHECKBOX, self.check22 )
    	self.checkBox130.Bind( wx.EVT_CHECKBOX, self.check23 )
    	self.checkBox131.Bind( wx.EVT_CHECKBOX, self.check24 )
    	self.checkBox132.Bind( wx.EVT_CHECKBOX, self.check25 )
    	self.checkBox133.Bind( wx.EVT_CHECKBOX, self.check26 )
    	self.checkBox134.Bind( wx.EVT_CHECKBOX, self.check27 )
    	self.checkBox135.Bind( wx.EVT_CHECKBOX, self.check28 )
    	self.checkBox136.Bind( wx.EVT_CHECKBOX, self.check29 )
    	self.checkBox137.Bind( wx.EVT_CHECKBOX, self.check30 )
    	self.button33.Bind( wx.EVT_BUTTON, self.Reset )
    	self.button34.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
    	pass
    
    
    # Virtual event handlers, overide them in your derived class
    def check1( self, event ):
            #event.Skip()
            tick114 = LoadWorkbook2(parent=self.checkBox108)
            mtick114 = tick114.wb2.get_sheet_by_name('Audit Checklist')
            put114 = mtick114['E21']
            box114 = self.checkBox108.GetValue()
    
    
            if box114 == True:
                put114.value = "/"
            else:
                put114.value = ""
    
    def check2( self, event ):
            #event.Skip()
            tick115 = LoadWorkbook2(parent=self.checkBox109)
            mtick115 = tick115.wb2.get_sheet_by_name('Audit Checklist')
            put115 = mtick115['F21']
            box115 = self.checkBox109.GetValue()
    
            if box115 == True:
                put115.value = "/"
            else:
                put115.value = ""
    
    def check3( self, event ):
            #event.Skip()
            tick116 = LoadWorkbook2(parent=self.checkBox110)
            mtick116 = tick116.wb2.get_sheet_by_name('Audit Checklist')
            put116 = mtick116['G21']
            box116 = self.checkBox110.GetValue()
    
            if box116 == True:
                put116.value = "/"
            else:
                put116.value = ""
    
    def check4( self, event ):
            #event.Skip()
            tick117 = LoadWorkbook2(parent=self.checkBox111)
            mtick117 = tick117.wb2.get_sheet_by_name('Audit Checklist')
            put117 = mtick117['E24']
            box117 = self.checkBox111.GetValue()
    
            if box117 == True:
                put117.value = "/"
            else:
                put117.value = ""
    
    def check5( self, event ):
            #event.Skip()
            tick118 = LoadWorkbook2(parent=self.checkBox112)
            mtick118 = tick118.wb2.get_sheet_by_name('Audit Checklist')
            put118 = mtick118['F24']
            box118 = self.checkBox112.GetValue()
    
            if box118 == True:
                put118.value = "/"
            else:
                put118.value = ""
    
    def check6( self, event ):
            #event.Skip()
            tick119 = LoadWorkbook2(parent=self.checkBox113)
            mtick119 = tick119.wb2.get_sheet_by_name('Audit Checklist')
            put119 = mtick119['G24']
            box119 = self.checkBox113.GetValue()
    
            if box119 == True:
                put119.value = "/"
            else:
                put119.value = ""
    
    def check7( self, event ):
            #event.Skip()
            tick120 = LoadWorkbook2(parent=self.checkBox114)
            mtick120 = tick120.wb2.get_sheet_by_name('Audit Checklist')
            put120 = mtick120['E26']
            box120 = self.checkBox114.GetValue()
    
            if box120 == True:
                put120.value = "/"
            else:
                put120.value = ""
    
    def check8( self, event ):
            #event.Skip()
            tick121 = LoadWorkbook2(parent=self.checkBox115)
            mtick121 = tick121.wb2.get_sheet_by_name('Audit Checklist')
            put121 = mtick121['F26']
            box121 = self.checkBox115.GetValue()
    
            if box121 == True:
                put121.value = "/"
            else:
                put121.value = ""
    
    def check9( self, event ):
            #event.Skip()
            tick122 = LoadWorkbook2(parent=self.checkBox116)
            mtick122 = tick122.wb2.get_sheet_by_name('Audit Checklist')
            put122 = mtick122['G26']
            box122 = self.checkBox116.GetValue()
    
            if box122 == True:
                put122.value = "/"
            else:
                put122.value = ""
    
    def check10( self, event ):
            #event.Skip()
            tick123 = LoadWorkbook2(parent=self.checkBox117)
            mtick123 = tick123.wb2.get_sheet_by_name('Audit Checklist')
            put123 = mtick123['E28']
            box123 = self.checkBox117.GetValue()
    
            if box123 == True:
                put123.value = "/"
            else:
                put123.value = ""
    
    def check11( self, event ):
            #event.Skip()
            tick124 = LoadWorkbook2(parent=self.checkBox118)
            mtick124 = tick124.wb2.get_sheet_by_name('Audit Checklist')
            put124 = mtick124['F28']
            box124 = self.checkBox118.GetValue()
    
            if box124 == True:
                put124.value = "/"
            else:
                put124.value = ""
    
    def check12( self, event ):
            #event.Skip()
            tick125 = LoadWorkbook2(parent=self.checkBox119)
            mtick125 = tick125.wb2.get_sheet_by_name('Audit Checklist')
            put125 = mtick125['G28']
            box125 = self.checkBox119.GetValue()
    
            if box125 == True:
                put125.value = "/"
            else:
                put125.value = ""
    
    def check13( self, event ):
            #event.Skip()
            tick126 = LoadWorkbook2(parent=self.checkBox120)
            mtick126 = tick126.wb2.get_sheet_by_name('Audit Checklist')
            put126 = mtick126['E30']
            box126 = self.checkBox120.GetValue()
    
            if box126 == True:
                put126.value = "/"
            else:
                put126.value = ""
    
    def check14( self, event ):
            #event.Skip()
            tick127 = LoadWorkbook2(parent=self.checkBox121)
            mtick127 = tick127.wb2.get_sheet_by_name('Audit Checklist')
            put127 = mtick127['F30']
            box127 = self.checkBox121.GetValue()
    
            if box127 == True:
                put127.value = "/"
            else:
                put127.value = ""
    
    def check15( self, event ):
            #event.Skip()
            tick128 = LoadWorkbook2(parent=self.checkBox122)
            mtick128 = tick128.wb2.get_sheet_by_name('Audit Checklist')
            put128 = mtick128['G30']
            box128 = self.checkBox122.GetValue()
    
            if box128 == True:
                put128.value = "/"
            else:
                put128.value = ""
    
    def check16( self, event ):
            #event.Skip()
            tick129 = LoadWorkbook2(parent=self.checkBox123)
            mtick129 = tick129.wb2.get_sheet_by_name('Audit Checklist')
            put129 = mtick129['E31']
            box129 = self.checkBox123.GetValue()
    
            if box129 == True:
                put129.value = "/"
            else:
                put129.value = ""
    
    def check17( self, event ):
            #event.Skip()
            tick130 = LoadWorkbook2(parent=self.checkBox124)
            mtick130 = tick130.wb2.get_sheet_by_name('Audit Checklist')
            put130 = mtick130['F31']
            box130 = self.checkBox124.GetValue()
    
            if box130 == True:
                put130.value = "/"
            else:
                put130.value = ""
    
    def check18( self, event ):
            #event.Skip()
            tick131 = LoadWorkbook2(parent=self.checkBox125)
            mtick131 = tick131.wb2.get_sheet_by_name('Audit Checklist')
            put131 = mtick131['G31']
            box131 = self.checkBox125.GetValue()
    
            if box131 == True:
                put131.value = "/"
            else:
                put131.value = ""
    
    def check19( self, event ):
            #event.Skip()
            tick132 = LoadWorkbook2(parent=self.checkBox126)
            mtick132 = tick132.wb2.get_sheet_by_name('Audit Checklist')
            put132 = mtick132['E32']
            box132 = self.checkBox126.GetValue()
    
            if box132 == True:
                put132.value = "/"
            else:
                put132.value = ""
    
    def check20( self, event ):
            #event.Skip()
            tick133 = LoadWorkbook2(parent=self.checkBox127)
            mtick133 = tick133.wb2.get_sheet_by_name('Audit Checklist')
            put133 = mtick133['F32']
            box133 = self.checkBox133.GetValue()
    
            if box133 == True:
                put133.value = "/"
            else:
                put133.value = ""
    
    def check21( self, event ):
            #event.Skip()
            tick134 = LoadWorkbook2(parent=self.checkBox128)
            mtick134 = tick134.wb2.get_sheet_by_name('Audit Checklist')
            put134 = mtick134['G32']
            box134 = self.checkBox128.GetValue()
    
            if box134 == True:
                put134.value = "/"
            else:
                put134.value = ""
    
    def check22( self, event ):
            #event.Skip()
            tick135 = LoadWorkbook2(parent=self.checkBox129)
            mtick135 = tick135.wb2.get_sheet_by_name('Audit Checklist')
            put135 = mtick135['E35']
            box135 = self.checkBox129.GetValue()
    
            if box135 == True:
                put135.value = "/"
            else:
                put135.value = ""
    
    def check23( self, event ):
            #event.Skip()
            tick136 = LoadWorkbook2(parent=self.checkBox130)
            mtick136 = tick136.wb2.get_sheet_by_name('Audit Checklist')
            put136 = mtick136['F35']
            box136 = self.checkBox130.GetValue()
    
            if box136 == True:
                put136.value = "/"
            else:
                put136.value = ""
    
    def check24( self, event ):
            #event.Skip()
            tick137 = LoadWorkbook2(parent=self.checkBox131)
            mtick137 = tick137.wb2.get_sheet_by_name('Audit Checklist')
            put137 = mtick137['G35']
            box137 = self.checkBox131.GetValue()
    
            if box137 == True:
                put137.value = "/"
            else:
                put137.value = ""
    
    def check25( self, event ):
            #event.Skip()
            tick138 = LoadWorkbook2(parent=self.checkBox132)
            mtick138 = tick138.wb2.get_sheet_by_name('Audit Checklist')
            put138 = mtick138['E37']
            box138 = self.checkBox132.GetValue()
    
            if box138 == True:
                put138.value = "/"
            else:
                put138.value = ""
    
    def check26( self, event ):
            #event.Skip()
            tick139 = LoadWorkbook2(parent=self.checkBox133)
            mtick139 = tick139.wb2.get_sheet_by_name('Audit Checklist')
            put139 = mtick139['F37']
            box139 = self.checkBox133.GetValue()
    
            if box139 == True:
                put139.value = "/"
            else:
                put139.value = ""
    
    def check27( self, event ):
            #event.Skip()
            tick140 = LoadWorkbook2(parent=self.checkBox134)
            mtick140 = tick140.wb2.get_sheet_by_name('Audit Checklist')
            put140 = mtick140['G37']
            box140 = self.checkBox134.GetValue()
    
            if box140 == True:
                put140.value = "/"
            else:
                put140.value = ""
    
    def check28( self, event ):
            #event.Skip()
            tick141 = LoadWorkbook2(parent=self.checkBox135)
            mtick141 = tick141.wb2.get_sheet_by_name('Audit Checklist')
            put141 = mtick141['E39']
            box141 = self.checkBox135.GetValue()
    
            if box141 == True:
                put141.value = "/"
            else:
                put141.value = ""
    
    def check29( self, event ):
            #event.Skip()
            tick142 = LoadWorkbook2(parent=self.checkBox136)
            mtick142 = tick142.wb2.get_sheet_by_name('Audit Checklist')
            put142 = mtick142['F39']
            box142 = self.checkBox136.GetValue()
    
            if box142 == True:
                put142.value = "/"
            else:
                put142.value = ""
    
    def check30( self, event ):
            #event.Skip()
            tick143 = LoadWorkbook2(parent=self.checkBox137)
            mtick143 = tick143.wb2.get_sheet_by_name('Audit Checklist')
            put143 = mtick143['G39']
            box143 = self.checkBox137.GetValue()
    
            if box143 == True:
                put143.value = "/"
            else:
                put143.value = ""
    
    def Reset( self, event ):
    		#event.Skip()
            self.checkBox108.SetValue(False)
            self.checkBox109.SetValue(False)
            self.checkBox110.SetValue(False)
            self.checkBox111.SetValue(False)
            self.checkBox112.SetValue(False)
            self.checkBox113.SetValue(False)
            self.checkBox114.SetValue(False)
            self.checkBox115.SetValue(False)
            self.checkBox116.SetValue(False)
            self.checkBox117.SetValue(False)
            self.checkBox118.SetValue(False)
            self.checkBox119.SetValue(False)
            self.checkBox120.SetValue(False)
            self.checkBox121.SetValue(False)
            self.checkBox122.SetValue(False)
            self.checkBox123.SetValue(False)
            self.checkBox124.SetValue(False)
            self.checkBox125.SetValue(False)
            self.checkBox126.SetValue(False)
            self.checkBox127.SetValue(False)
            self.checkBox128.SetValue(False)
            self.checkBox129.SetValue(False)
            self.checkBox130.SetValue(False)
            self.checkBox131.SetValue(False)
            self.checkBox132.SetValue(False)
            self.checkBox133.SetValue(False)
            self.checkBox134.SetValue(False)
            self.checkBox135.SetValue(False)
            self.checkBox136.SetValue(False)
            self.checkBox137.SetValue(False)
    
    
    def Save( self, event ):
            #event.Skip()
            check114 = self.checkBox108.GetValue()
            check115 = self.checkBox109.GetValue()
            check116 = self.checkBox110.GetValue()
            check117 = self.checkBox111.GetValue()
            check118 = self.checkBox112.GetValue()
            check119 = self.checkBox113.GetValue()
            check120 = self.checkBox114.GetValue()
            check121 = self.checkBox115.GetValue()
            check122 = self.checkBox116.GetValue()
            check123 = self.checkBox117.GetValue()
            check124 = self.checkBox118.GetValue()
            check125 = self.checkBox119.GetValue()
            check126 = self.checkBox120.GetValue()
            check127 = self.checkBox121.GetValue()
            check128 = self.checkBox122.GetValue()
            check129 = self.checkBox123.GetValue()
            check130 = self.checkBox124.GetValue()
            check131 = self.checkBox125.GetValue()
            check132 = self.checkBox126.GetValue()
            check133 = self.checkBox127.GetValue()
            check134 = self.checkBox128.GetValue()
            check135 = self.checkBox129.GetValue()
            check136 = self.checkBox130.GetValue()
            check137 = self.checkBox131.GetValue()
            check138 = self.checkBox132.GetValue()
            check139 = self.checkBox133.GetValue()
            check140 = self.checkBox134.GetValue()
            check141 = self.checkBox135.GetValue()
            check142 = self.checkBox136.GetValue()
            check143 = self.checkBox137.GetValue()
    
            if check114 == True:
                b17 = 1
            else:
                b17 = 0
    
            if check115 == True:
                b18 = 1
            else:
                b18 = 0
    
            if check116 == True:
                b19 = 1
            else:
                b19 = 0
    
            if check117 == True:
                b20 = 1
            else:
                b20 = 0
    
            if check118 == True:
                b21 = 1
            else:
                b21 = 0
    
            if check119 == True:
                b22 = 1
            else:
                b22 = 0
    
            if check120 == True:
                b23 = 1
            else:
                b23 = 0
    
            if check121 == True:
                b24 = 1
            else:
                b24 = 0
    
            if check122 == True:
                b24 = 1
            else:
                b24 = 0
    
            if check123 == True:
                b25 = 1
            else:
                b25 = 0
    
            if check124 == True:
                b26 = 1
            else:
                b26 = 0
    
            if check125 == True:
                b27 = 1
            else:
                b27 = 0
    
            if check126 == True:
                b28 = 1
            else:
                b28 = 0
    
            if check127 == True:
                b29 = 1
            else:
                b29 = 0
    
            if check128 == True:
                b30 = 1
            else:
                b30 = 0
    
            if check129 == True:
                b31 = 1
            else:
                b31 = 0
    
            if check130 == True:
                b32 = 1
            else:
                b32 = 0
    
            if check131 == True:
                b33 = 1
            else:
                b33 = 0
    
            if check132 == True:
                b34 = 1
            else:
                b34 = 0
    
            if check133 == True:
                b35 = 1
            else:
                b35 = 0
    
            if check134 == True:
                b36 = 1
            else:
                b36 = 0
    
            if check135 == True:
                b37 = 1
            else:
                b37 = 0
    
            if check136 == True:
                b38 = 1
            else:
                b38 = 0
    
            if check137 == True:
                b39 = 1
            else:
                b39 = 0
    
            if check138 == True:
                b40 = 1
            else:
                b40 = 0
    
            if check139 == True:
                b41 = 1
            else:
                b41 = 0
    
            if check140 == True:
                b42 = 1
            else:
                b42 = 0
    
            if check141 == True:
                b43 = 1
            else:
                b43 = 0
    
            if check142 == True:
                b44 = 1
            else:
                b44 = 0
    
            if check143 == True:
                b45 = 1
            else:
                b45 = 0





###########################################################################
## Class Clause91a
###########################################################################

class Clause91a ( wx.Frame ):

    def __init__( self, parent ):
    	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Labelling And Packaging", pos = wx.DefaultPosition, size = wx.Size( 983,600 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
    
    	self.SetSizeHintsSz( wx.Size(983,600), wx.Size(983,600) )
    	self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
    
    	bSizer95 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer117 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.1", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText64.Wrap( -1 )
    	self.staticText64.SetFont( wx.Font( 11, 70, 90, 92, False, "Arial" ) )
    
    	bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )
    
    
    	bSizer117.Add( ( 330, 0), 0, 0, 5 )
    
    	self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText65.Wrap( -1 )
    	self.staticText65.SetFont( wx.Font( 11, 70, 90, 92, False, "Arial" ) )
    
    	bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )
    
    
    	bSizer95.Add( bSizer117, 0, wx.EXPAND, 5 )
    
    
    	bSizer95.Add( ( 0, 20), 0, 0, 5 )
    
    	bSizer116 = wx.BoxSizer( wx.HORIZONTAL )
    
    	bSizer104 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer96 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText49.Wrap( -1 )
    	self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )
    
    
    	bSizer96.Add( ( 75, 0), 0, 0, 5 )
    
    	self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"Labelling shall include the following\ninformation:\n", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText103.Wrap( -1 )
    	self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )
    
    
    	bSizer96.Add( ( 90, 0), 1, 0, 5 )
    
    	self.staticText50 = wx.StaticText( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText50.Wrap( -1 )
    	bSizer96.Add( self.staticText50, 1, wx.ALL|wx.EXPAND, 5 )
    
    
    	bSizer104.Add( bSizer96, 0, 0, 5 )
    
    	bSizer97 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(a)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText51.Wrap( -1 )
    	self.staticText51.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer97.Add( self.staticText51, 0, wx.ALL, 5 )
    
    
    	bSizer97.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText105 = wx.StaticText( self, wx.ID_ANY, u"Company name", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText105.Wrap( -1 )
    	self.staticText105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer97.Add( self.staticText105, 0, wx.ALL, 5 )
    
    
    	bSizer97.Add( ( 100, 0), 0, 0, 5 )
    
    
    	bSizer104.Add( bSizer97, 1, wx.EXPAND, 5 )
    
    	bSizer98 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText57 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(b)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText57.Wrap( -1 )
    	self.staticText57.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer98.Add( self.staticText57, 0, wx.ALL, 5 )
    
    
    	bSizer98.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText106 = wx.StaticText( self, wx.ID_ANY, u"Parts identification\nnumber, if applicable", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText106.Wrap( -1 )
    	self.staticText106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer98.Add( self.staticText106, 0, wx.ALL, 5 )
    
    
    	bSizer98.Add( ( 95, 0), 0, 0, 5 )
    
    
    	bSizer104.Add( bSizer98, 1, wx.EXPAND, 5 )
    
    	bSizer100 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText59 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(c)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText59.Wrap( -1 )
    	self.staticText59.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText59, 0, wx.ALL, 5 )
    
    
    	bSizer100.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText107 = wx.StaticText( self, wx.ID_ANY, u"Make and model", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText107.Wrap( -1 )
    	self.staticText107.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText107, 0, wx.ALL, 5 )
    
    
    	bSizer100.Add( ( 140, 0), 0, 0, 5 )
    
    	self.staticText60 = wx.StaticText( self, wx.ID_ANY, u"Products shall be properly\npackaged, labelled and\nidentified as reused,\nrepaired or remanufactured", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText60.Wrap( -1 )
    	self.staticText60.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer100.Add( self.staticText60, 0, wx.ALL, 5 )
    
    
    	bSizer104.Add( bSizer100, 1, wx.EXPAND, 5 )
    
    	bSizer101 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText61 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(d)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText61.Wrap( -1 )
    	self.staticText61.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer101.Add( self.staticText61, 0, wx.ALL, 5 )
    
    
    	bSizer101.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText108 = wx.StaticText( self, wx.ID_ANY, u"Engine capacity,\nif applicable", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText108.Wrap( -1 )
    	self.staticText108.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer101.Add( self.staticText108, 0, wx.ALL, 5 )
    
    
    	bSizer101.Add( ( 85, 0), 0, 0, 5 )
    
    
    	bSizer104.Add( bSizer101, 1, wx.EXPAND, 5 )
    
    	bSizer218 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText160 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(e)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText160.Wrap( -1 )
    	self.staticText160.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer218.Add( self.staticText160, 0, wx.ALL, 5 )
    
    
    	bSizer218.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText162 = wx.StaticText( self, wx.ID_ANY, u"Classifications according to\nindustry best practices e.g.\nreuse, remanufacturing or repair", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText162.Wrap( -1 )
    	self.staticText162.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer218.Add( self.staticText162, 0, wx.ALL, 5 )
    
    
    	bSizer218.Add( ( 110, 0), 0, 0, 5 )
    
    
    	bSizer104.Add( bSizer218, 1, wx.EXPAND, 5 )
    
    	bSizer1011 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.staticText611 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(f)", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText611.Wrap( -1 )
    	self.staticText611.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1011.Add( self.staticText611, 0, wx.ALL, 5 )
    
    
    	bSizer1011.Add( ( 60, 0), 0, 0, 5 )
    
    	self.staticText1081 = wx.StaticText( self, wx.ID_ANY, u"Recovered from local\nvehicles or importe", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.staticText1081.Wrap( -1 )
    	self.staticText1081.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1011.Add( self.staticText1081, 0, wx.ALL, 5 )
    
    
    	bSizer1011.Add( ( 115, 0), 0, 0, 5 )
    
    
    	bSizer104.Add( bSizer1011, 1, wx.EXPAND, 5 )
    
    
    	bSizer116.Add( bSizer104, 1, 0, 5 )
    
    	bSizer105 = wx.BoxSizer( wx.VERTICAL )
    
    	bSizer106 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox138 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox138.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox138, 0, wx.ALL, 5 )
    
    	self.checkBox139 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox139.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox139, 0, wx.ALL, 5 )
    
    	self.checkBox140 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox140.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer106.Add( self.checkBox140, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer106, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 30), 0, 0, 5 )
    
    	bSizer107 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox141 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox141.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox141, 0, wx.ALL, 5 )
    
    	self.checkBox142 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox142.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox142, 0, wx.ALL, 5 )
    
    	self.checkBox143 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox143.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer107.Add( self.checkBox143, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer107, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 60), 0, 0, 5 )
    
    	bSizer108 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox144 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox144.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox144, 0, wx.ALL, 5 )
    
    	self.checkBox145 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox145.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox145, 0, wx.ALL, 5 )
    
    	self.checkBox146 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox146.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer108.Add( self.checkBox146, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer108, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 40), 0, 0, 5 )
    
    	bSizer114 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox147 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox147.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox147, 0, wx.ALL, 5 )
    
    	self.checkBox148 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox148.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox148, 0, wx.ALL, 5 )
    
    	self.checkBox149 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox149.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer114.Add( self.checkBox149, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer114, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 50), 0, 0, 5 )
    
    	bSizer115 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox150 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox150.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox150, 0, wx.ALL, 5 )
    
    	self.checkBox151 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox151.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox151, 0, wx.ALL, 5 )
    
    	self.checkBox152 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox152.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer115.Add( self.checkBox152, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer115, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 40), 0, 0, 5 )
    
    	bSizer1061 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox153 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox153.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1061.Add( self.checkBox153, 0, wx.ALL, 5 )
    
    	self.checkBox154 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox154.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1061.Add( self.checkBox154, 0, wx.ALL, 5 )
    
    	self.checkBox155 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox155.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1061.Add( self.checkBox155, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer1061, 1, wx.EXPAND, 5 )
    
    
    	bSizer105.Add( ( 0, 40), 0, 0, 5 )
    
    	bSizer1062 = wx.BoxSizer( wx.HORIZONTAL )
    
    	self.checkBox156 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox156.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1062.Add( self.checkBox156, 0, wx.ALL, 5 )
    
    	self.checkBox157 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox157.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1062.Add( self.checkBox157, 0, wx.ALL, 5 )
    
    	self.checkBox158 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
    	self.checkBox158.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
    
    	bSizer1062.Add( self.checkBox158, 0, wx.ALL, 5 )
    
    
    	bSizer105.Add( bSizer1062, 1, wx.EXPAND, 5 )
    
    
    	bSizer116.Add( bSizer105, 0, 0, 5 )
    
    
    	bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )
    
    	bSizer103 = wx.BoxSizer( wx.HORIZONTAL )
    
    
    	bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )
    
    	self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer103.Add( self.button33, 0, wx.ALL, 5 )
    
    	self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
    	self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
    
    	bSizer103.Add( self.button34, 0, wx.ALL, 5 )
    
    
    	bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )
    
    
    	self.SetSizer( bSizer95 )
    	self.Layout()
    
    	self.Centre( wx.BOTH )
    
    	# Connect Events
    	self.checkBox138.Bind( wx.EVT_CHECKBOX, self.check1 )
    	self.checkBox139.Bind( wx.EVT_CHECKBOX, self.check2 )
    	self.checkBox140.Bind( wx.EVT_CHECKBOX, self.check3 )
    	self.checkBox141.Bind( wx.EVT_CHECKBOX, self.check4 )
    	self.checkBox142.Bind( wx.EVT_CHECKBOX, self.check5 )
    	self.checkBox143.Bind( wx.EVT_CHECKBOX, self.check6 )
    	self.checkBox144.Bind( wx.EVT_CHECKBOX, self.check7 )
    	self.checkBox145.Bind( wx.EVT_CHECKBOX, self.check8 )
    	self.checkBox146.Bind( wx.EVT_CHECKBOX, self.check9 )
    	self.checkBox147.Bind( wx.EVT_CHECKBOX, self.check10 )
    	self.checkBox148.Bind( wx.EVT_CHECKBOX, self.check11 )
    	self.checkBox149.Bind( wx.EVT_CHECKBOX, self.check12 )
    	self.checkBox150.Bind( wx.EVT_CHECKBOX, self.check13 )
    	self.checkBox151.Bind( wx.EVT_CHECKBOX, self.check14 )
    	self.checkBox152.Bind( wx.EVT_CHECKBOX, self.check15 )
    	self.checkBox153.Bind( wx.EVT_CHECKBOX, self.check16 )
    	self.checkBox154.Bind( wx.EVT_CHECKBOX, self.check17 )
    	self.checkBox155.Bind( wx.EVT_CHECKBOX, self.check18 )
    	self.checkBox156.Bind( wx.EVT_CHECKBOX, self.check19 )
    	self.checkBox157.Bind( wx.EVT_CHECKBOX, self.check20 )
    	self.checkBox158.Bind( wx.EVT_CHECKBOX, self.check21 )
    	self.button33.Bind( wx.EVT_BUTTON, self.Reset )
    	self.button34.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
    	pass
    
    
    # Virtual event handlers, overide them in your derived class
    def check1( self, event ):
            #event.Skip()
            tick144 = LoadWorkbook2(parent=self.checkBox138)
            mtick144 = tick144.wb2.get_sheet_by_name('Audit Checklist')
            put144 = mtick144['E42']
            box144 = self.checkBox138.GetValue()
    
            if box144 == True:
                put144.value = "/"
            else:
                put144.value = ""
    
    def check2( self, event ):
            #event.Skip()
            tick145 = LoadWorkbook2(parent=self.checkBox139)
            mtick145 = tick145.wb2.get_sheet_by_name('Audit Checklist')
            put145 = mtick145['F42']
            box145 = self.checkBox139.GetValue()
    
            if box145 == True:
                put145.value = "/"
            else:
                put145.value = ""
    
    def check3( self, event ):
            #event.Skip()
            tick146 = LoadWorkbook2(parent=self.checkBox140)
            mtick146 = tick146.wb2.get_sheet_by_name('Audit Checklist')
            put146 = mtick146['G42']
            box146 = self.checkBox140.GetValue()
    
            if box146 == True:
                put146.value = "/"
            else:
                put146.value = ""
    
    def check4( self, event ):
            #event.Skip()
            tick147 = LoadWorkbook2(parent=self.checkBox141)
            mtick147 = tick147.wb2.get_sheet_by_name('Audit Checklist')
            put147 = mtick147['E42']
            box147 = self.checkBox141.GetValue()
    
            if box147 == True:
                put147.value = "/"
            else:
                put147.value = ""
    
    def check5( self, event ):
            #event.Skip()
            tick148 = LoadWorkbook2(parent=self.checkBox142)
            mtick148 = tick148.wb2.get_sheet_by_name('Audit Checklist')
            put148 = mtick148['F42']
            box148 = self.checkBox142.GetValue()
    
            if box148 == True:
                put148.value = "/"
            else:
                put148.value = ""
    
    def check6( self, event ):
            #event.Skip()
            tick149 = LoadWorkbook2(parent=self.checkBox143)
            mtick149 = tick149.wb2.get_sheet_by_name('Audit Checklist')
            put149 = mtick149['G42']
            box149 = self.checkBox143.GetValue()
    
            if box149 == True:
                put149.value = "/"
            else:
                put149.value = ""
    
    def check7( self, event ):
            #event.Skip()
            tick150 = LoadWorkbook2(parent=self.checkBox144)
            mtick150 = tick150.wb2.get_sheet_by_name('Audit Checklist')
            put150 = mtick150['E45']
            box150 = self.checkBox144.GetValue()
    
            if box150 == True:
                put150.value = "/"
            else:
                put150.value = ""
    
    def check8( self, event ):
            #event.Skip()
            tick151 = LoadWorkbook2(parent=self.checkBox145)
            mtick151 = tick151.wb2.get_sheet_by_name('Audit Checklist')
            put151 = mtick151['F45']
            box151 = self.checkBox145.GetValue()
    
            if box151 == True:
                put151.value = "/"
            else:
                put151.value = ""
    
    def check9( self, event ):
            #event.Skip()
            tick152 = LoadWorkbook2(parent=self.checkBox146)
            mtick152 = tick152.wb2.get_sheet_by_name('Audit Checklist')
            put152 = mtick152['G45']
            box152 = self.checkBox146.GetValue()
    
            if box152 == True:
                put152.value = "/"
            else:
                put152.value = ""
    
    def check10( self, event ):
            #event.Skip()
            tick153 = LoadWorkbook2(parent=self.checkBox147)
            mtick153 = tick153.wb2.get_sheet_by_name('Audit Checklist')
            put153 = mtick153['E46']
            box153 = self.checkBox147.GetValue()
    
            if box153 == True:
                put153.value = "/"
            else:
                put153.value = ""
    
    def check11( self, event ):
            #event.Skip()
            tick154 = LoadWorkbook2(parent=self.checkBox148)
            mtick154 = tick154.wb2.get_sheet_by_name('Audit Checklist')
            put154 = mtick154['F46']
            box154 = self.checkBox148.GetValue()
    
            if box154 == True:
                put154.value = "/"
            else:
                put154.value = ""
    
    def check12( self, event ):
            #event.Skip()
            tick155 = LoadWorkbook2(parent=self.checkBox149)
            mtick155 = tick155.wb2.get_sheet_by_name('Audit Checklist')
            put155 = mtick155['G46']
            box155 = self.checkBox149.GetValue()
    
            if box155 == True:
                put155.value = "/"
            else:
                put155.value = ""
    
    def check13( self, event ):
            #event.Skip()
            tick156 = LoadWorkbook2(parent=self.checkBox150)
            mtick156 = tick156.wb2.get_sheet_by_name('Audit Checklist')
            put156 = mtick156['E47']
            box156 = self.checkBox150.GetValue()
    
            if box156 == True:
                put156.value = "/"
            else:
                put156.value = ""
    
    def check14( self, event ):
            #event.Skip()
            tick157 = LoadWorkbook2(parent=self.checkBox151)
            mtick157 = tick157.wb2.get_sheet_by_name('Audit Checklist')
            put157 = mtick157['F47']
            box157 = self.checkBox151.GetValue()
    
            if box157 == True:
                put157.value = "/"
            else:
                put157.value = ""
    
    def check15( self, event ):
            #event.Skip()
            tick158 = LoadWorkbook2(parent=self.checkBox152)
            mtick158 = tick158.wb2.get_sheet_by_name('Audit Checklist')
            put158 = mtick158['G47']
            box158 = self.checkBox152.GetValue()
    
            if box158 == True:
                put158.value = "/"
            else:
                put158.value = ""
    
    def check16( self, event ):
            #event.Skip()
            tick159 = LoadWorkbook2(parent=self.checkBox153)
            mtick159 = tick159.wb2.get_sheet_by_name('Audit Checklist')
            put159 = mtick159['E48']
            box159 = self.checkBox153.GetValue()
    
            if box159 == True:
                put159.value = "/"
            else:
                put159.value = ""
    
    def check17( self, event ):
            #event.Skip()
            tick160 = LoadWorkbook2(parent=self.checkBox154)
            mtick160 = tick160.wb2.get_sheet_by_name('Audit Checklist')
            put160 = mtick160['F48']
            box160 = self.checkBox154.GetValue()
    
            if box160 == True:
                put160.value = "/"
            else:
                put160.value = ""
    
    def check18( self, event ):
            #event.Skip()
            tick161 = LoadWorkbook2(parent=self.checkBox155)
            mtick161 = tick161.wb2.get_sheet_by_name('Audit Checklist')
            put161 = mtick161['G48']
            box161 = self.checkBox155.GetValue()
    
            if box161 == True:
                put161.value = "/"
            else:
                put161.value = ""
    
    def check19( self, event ):
            #event.Skip()
            tick162 = LoadWorkbook2(parent=self.checkBox156)
            mtick162 = tick162.wb2.get_sheet_by_name('Audit Checklist')
            put162 = mtick162['E49']
            box162 = self.checkBox156.GetValue()
    
            if box162 == True:
                put162.value = "/"
            else:
                put162.value = ""
    
    def check20( self, event ):
            #event.Skip()
            tick163 = LoadWorkbook2(parent=self.checkBox157)
            mtick163 = tick163.wb2.get_sheet_by_name('Audit Checklist')
            put163 = mtick163['F49']
            box163 = self.checkBox157.GetValue()
    
            if box163 == True:
                put163.value = "/"
            else:
                put163.value = ""
    
    def check21( self, event ):
            #event.Skip()
            tick164 = LoadWorkbook2(parent=self.checkBox158)
            mtick164 = tick164.wb2.get_sheet_by_name('Audit Checklist')
            put164 = mtick164['G49']
            box164 = self.checkBox158.GetValue()
    
            if box164 == True:
                put164.value = "/"
            else:
                put164.value = ""
    
    def Reset( self, event ):
    		#event.Skip()
            self.checkBox138.SetValue(False)
            self.checkBox139.SetValue(False)
            self.checkBox140.SetValue(False)
            self.checkBox141.SetValue(False)
            self.checkBox142.SetValue(False)
            self.checkBox143.SetValue(False)
            self.checkBox144.SetValue(False)
            self.checkBox145.SetValue(False)
            self.checkBox146.SetValue(False)
            self.checkBox147.SetValue(False)
            self.checkBox148.SetValue(False)
            self.checkBox149.SetValue(False)
            self.checkBox150.SetValue(False)
            self.checkBox151.SetValue(False)
            self.checkBox152.SetValue(False)
            self.checkBox153.SetValue(False)
            self.checkBox154.SetValue(False)
            self.checkBox155.SetValue(False)
            self.checkBox156.SetValue(False)
            self.checkBox157.SetValue(False)
            self.checkBox158.SetValue(False)


    def Save( self, event ):
            #event.Skip()
            check144 = self.checkBox138.GetValue()
            check145 = self.checkBox139.GetValue()
            check146 = self.checkBox140.GetValue()
            check147 = self.checkBox141.GetValue()
            check148 = self.checkBox142.GetValue()
            check149 = self.checkBox143.GetValue()
            check150 = self.checkBox144.GetValue()
            check151 = self.checkBox145.GetValue()
            check152 = self.checkBox146.GetValue()
            check153 = self.checkBox147.GetValue()
            check154 = self.checkBox148.GetValue()
            check155 = self.checkBox149.GetValue()
            check156 = self.checkBox150.GetValue()
            check157 = self.checkBox151.GetValue()
            check158 = self.checkBox152.GetValue()
            check159 = self.checkBox153.GetValue()
            check160 = self.checkBox154.GetValue()
            check161 = self.checkBox155.GetValue()
            check162 = self.checkBox156.GetValue()
            check163 = self.checkBox157.GetValue()
            check164 = self.checkBox158.GetValue()
    
            if check144 == True:
                b46 = 1
            else:
                b46 = 0
    
            if check145 == True:
                b47 = 1
            else:
                b47 = 0
    
            if check146 == True:
                b48 = 1
            else:
                b48 = 0
    
            if check147 == True:
                b49 = 1
            else:
                b49 = 0
    
            if check148 == True:
                b50 = 1
            else:
                b50 = 0
    
            if check149 == True:
                b51 = 1
            else:
                b51 = 0
    
            if check150 == True:
                b52 = 1
            else:
                b52 = 0
    
            if check151 == True:
                b53 = 1
            else:
                b53 = 0
    
            if check152 == True:
                b54 = 1
            else:
                b54 = 0
    
            if check153 == True:
                b55 = 1
            else:
                b55 = 0
    
            if check154 == True:
                b56 = 1
            else:
                b56 = 0
    
            if check155 == True:
                b57 = 1
            else:
                b57 = 0
    
            if check156 == True:
                b58 = 1
            else:
                b58 = 0
    
            if check157 == True:
                b59 = 1
            else:
                b59 = 0
    
            if check158 == True:
                b60 = 1
            else:
                b60 = 0
    
            if check159 == True:
                b61 = 1
            else:
                b61 = 0
    
            if check160 == True:
                b62 = 1
            else:
                b62 = 0
    
            if check161 == True:
                b63 = 1
            else:
                b63 = 0
    
            if check162 == True:
                b64 = 1
            else:
                b64 = 0
    
            if check163 == True:
                b65 = 1
            else:
                b65 = 0
    
            if check164 == True:
                b66 = 1
            else:
                b66 = 0


###########################################################################
## Class Clause10
###########################################################################

class Clause10 ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Warranty", pos = wx.DefaultPosition, size = wx.Size( 1103,426 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(1103,426), wx.Size(1103,426) )
	self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )

	bSizer95 = wx.BoxSizer( wx.VERTICAL )

	bSizer117 = wx.BoxSizer( wx.HORIZONTAL )

	self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 10", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText64.Wrap( -1 )
	self.staticText64.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )

	bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )


	bSizer117.Add( ( 375, 0), 0, 0, 5 )

	self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText65.Wrap( -1 )
	self.staticText65.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )

	bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )


	bSizer95.Add( bSizer117, 0, wx.EXPAND, 5 )

	bSizer116 = wx.BoxSizer( wx.HORIZONTAL )

	bSizer104 = wx.BoxSizer( wx.VERTICAL )

	bSizer96 = wx.BoxSizer( wx.HORIZONTAL )

	self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 10(a)", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText49.Wrap( -1 )
	self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )


	bSizer96.Add( ( 60, 0), 0, 0, 5 )

	self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"Handling remanufacturing of the\nused parts and components", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText103.Wrap( -1 )
	self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )


	bSizer96.Add( ( 90, 0), 1, 0, 5 )

	self.staticText50 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing\nof the used parts and components shall \nbe to  demonstrate that  it has the legal \nright to transfer their ownership to \nanother party", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText50.Wrap( -1 )
	self.staticText50.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer96.Add( self.staticText50, 0, wx.ALL, 5 )


	bSizer104.Add( bSizer96, 0, 0, 5 )

	bSizer97 = wx.BoxSizer( wx.HORIZONTAL )

	self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"Clause 10(b)", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText51.Wrap( -1 )
	self.staticText51.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer97.Add( self.staticText51, 0, wx.ALL, 5 )


	bSizer97.Add( ( 60, 0), 0, 0, 5 )

	self.staticText105 = wx.StaticText( self, wx.ID_ANY, u"Provide specified information \nof remanufactured parts\nand components", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText105.Wrap( -1 )
	self.staticText105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer97.Add( self.staticText105, 0, wx.ALL, 5 )


	bSizer97.Add( ( 105, 0), 0, 0, 5 )

	self.staticText52 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing\nof the used parts and components shall \nbe to  demonstrate that  it has the legal \nright to transfer their ownership to \nanother party", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText52.Wrap( -1 )
	self.staticText52.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer97.Add( self.staticText52, 0, wx.ALL, 5 )


	bSizer104.Add( bSizer97, 1, wx.EXPAND, 5 )

	bSizer98 = wx.BoxSizer( wx.HORIZONTAL )

	self.staticText57 = wx.StaticText( self, wx.ID_ANY, u"Clause 10(c)", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText57.Wrap( -1 )
	self.staticText57.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer98.Add( self.staticText57, 0, wx.ALL, 5 )


	bSizer98.Add( ( 60, 0), 0, 0, 5 )

	self.staticText106 = wx.StaticText( self, wx.ID_ANY, u"Manpower certification ", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText106.Wrap( -1 )
	self.staticText106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer98.Add( self.staticText106, 0, wx.ALL, 5 )


	bSizer98.Add( ( 147, 0), 0, 0, 5 )

	self.staticText58 = wx.StaticText( self, wx.ID_ANY, u"Person dismantling, cleaning, examining,\nremediating, re-assembling, testing and \nhandling of the core and remanufactured\nparts or components shall be of \ntechnical specialist", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText58.Wrap( -1 )
	self.staticText58.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer98.Add( self.staticText58, 0, wx.ALL, 5 )


	bSizer104.Add( bSizer98, 0, wx.EXPAND, 5 )


	bSizer116.Add( bSizer104, 1, wx.EXPAND, 5 )

	bSizer105 = wx.BoxSizer( wx.VERTICAL )


	bSizer105.Add( ( 0, 20), 0, 0, 5 )

	bSizer106 = wx.BoxSizer( wx.HORIZONTAL )

	self.checkBox159 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox159.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer106.Add( self.checkBox159, 0, wx.ALL, 5 )

	self.checkBox160 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox160.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer106.Add( self.checkBox160, 0, wx.ALL, 5 )

	self.checkBox161 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox161.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer106.Add( self.checkBox161, 0, wx.ALL, 5 )


	bSizer105.Add( bSizer106, 0, wx.ALIGN_CENTER_HORIZONTAL, 5 )


	bSizer105.Add( ( 0, 50), 0, 0, 5 )

	bSizer107 = wx.BoxSizer( wx.HORIZONTAL )

	self.checkBox162 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox162.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer107.Add( self.checkBox162, 0, wx.ALL, 5 )

	self.checkBox163 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox163.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer107.Add( self.checkBox163, 0, wx.ALL, 5 )

	self.checkBox164 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox164.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer107.Add( self.checkBox164, 0, wx.ALL, 5 )


	bSizer105.Add( bSizer107, 0, wx.ALIGN_CENTER_HORIZONTAL, 5 )


	bSizer105.Add( ( 0, 100), 0, 0, 5 )

	bSizer108 = wx.BoxSizer( wx.HORIZONTAL )

	self.checkBox165 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox165.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer108.Add( self.checkBox165, 0, wx.ALL, 5 )

	self.checkBox166 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox166.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer108.Add( self.checkBox166, 0, wx.ALL, 5 )

	self.checkBox167 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox167.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer108.Add( self.checkBox167, 0, wx.ALL, 5 )


	bSizer105.Add( bSizer108, 0, wx.ALIGN_CENTER_HORIZONTAL, 5 )


	bSizer116.Add( bSizer105, 1, 0, 5 )


	bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )

	bSizer103 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer103.Add( self.button33, 0, wx.ALL, 5 )

	self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer103.Add( self.button34, 0, wx.ALL, 5 )


	bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )


	self.SetSizer( bSizer95 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.checkBox159.Bind( wx.EVT_CHECKBOX, self.check1 )
	self.checkBox160.Bind( wx.EVT_CHECKBOX, self.check2 )
	self.checkBox161.Bind( wx.EVT_CHECKBOX, self.check3 )
	self.checkBox162.Bind( wx.EVT_CHECKBOX, self.check4 )
	self.checkBox163.Bind( wx.EVT_CHECKBOX, self.check5 )
	self.checkBox164.Bind( wx.EVT_CHECKBOX, self.check6 )
	self.checkBox165.Bind( wx.EVT_CHECKBOX, self.check7 )
	self.checkBox166.Bind( wx.EVT_CHECKBOX, self.check8 )
	self.checkBox167.Bind( wx.EVT_CHECKBOX, self.check9 )
	self.button33.Bind( wx.EVT_BUTTON, self.Reset )
	self.button34.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def check1( self, event ):
        #event.Skip()
        tick165 = LoadWorkbook2(parent=self.checkBox159)
        mtick165 = tick165.wb2.get_sheet_by_name('Audit Checklist')
        put165 = mtick165['E50']
        box165 = self.checkBox159.GetValue()

        if box165 == True:
            put165.value = "/"
        else:
            put165.value = ""

def check2( self, event ):
        #event.Skip()
        tick166 = LoadWorkbook2(parent=self.checkBox160)
        mtick166 = tick166.wb2.get_sheet_by_name('Audit Checklist')
        put166 = mtick166['F50']
        box166 = self.checkBox160.GetValue()

        if box166 == True:
            put166.value = "/"
        else:
            put166.value = ""

def check3( self, event ):
        #event.Skip()
        tick167 = LoadWorkbook2(parent=self.checkBox161)
        mtick167 = tick167.wb2.get_sheet_by_name('Audit Checklist')
        put167 = mtick167['G50']
        box167 = self.checkBox161.GetValue()

        if box167 == True:
            put167.value = "/"
        else:
            put167.value = ""

def check4( self, event ):
        #event.Skip()
        tick168 = LoadWorkbook2(parent=self.checkBox162)
        mtick168 = tick168.wb2.get_sheet_by_name('Audit Checklist')
        put168 = mtick168['E53']
        box168 = self.checkBox162.GetValue()

        if box168 == True:
            put168.value = "/"
        else:
            put168.value = ""

def check5( self, event ):
        #event.Skip()
        global team169
        tick169 = LoadWorkbook2(parent=self.checkBox163)
        mtick169 = tick169.wb2.get_sheet_by_name('Audit Checklist')
        put169 = mtick169['F53']
        box169 = self.checkBox163.GetValue()

        if box169 == True:
            put169.value = "/"
        else:
            put169.value = ""

def check6( self, event ):
        #event.Skip()
        global team170
        tick170 = LoadWorkbook2(parent=self.checkBox164)
        mtick170 = tick170.wb2.get_sheet_by_name('Audit Checklist')
        put170 = mtick170['G53']
        box170 = self.checkBox164.GetValue()

        if box170 == True:
            put170.value = "/"
        else:
            put170.value = ""

def check7( self, event ):
        #event.Skip()
        global team171
        tick171 = LoadWorkbook2(parent=self.checkBox165)
        mtick171 = tick171.wb2.get_sheet_by_name('Audit Checklist')
        put171 = mtick171['E55']
        box171 = self.checkBox165.GetValue()

        if box171 == True:
            put171.value = "/"
        else:
            put171.value = ""


def check8( self, event ):
        #event.Skip()
        global team172
        tick172 = LoadWorkbook2(parent=self.checkBox166)
        mtick172 = tick172.wb2.get_sheet_by_name('Audit Checklist')
        put172 = mtick172['F55']
        box172 = self.checkBox166.GetValue()

        if box172 == True:
            put172.value = "/"
        else:
            put172.value = ""

def check9( self, event ):
        #event.Skip()
        global team173
        tick173 = LoadWorkbook2(parent=self.checkBox167)
        mtick173 = tick173.wb2.get_sheet_by_name('Audit Checklist')
        put173 = mtick173['G55']
        box173 = self.checkBox167.GetValue()

        if box173 == True:
            put173.value = "/"
        else:
            put173.value = ""

def Reset( self, event ):
		#event.Skip()
        self.checkBox159.SetValue(False)
        self.checkBox160.SetValue(False)
        self.checkBox161.SetValue(False)
        self.checkBox162.SetValue(False)
        self.checkBox163.SetValue(False)
        self.checkBox164.SetValue(False)
        self.checkBox165.SetValue(False)
        self.checkBox166.SetValue(False)
        self.checkBox167.SetValue(False)

def Save( self, event ):
        #event.Skip()
        check165 = self.checkBox159.GetValue()
        check166 = self.checkBox160.GetValue()
        check167 = self.checkBox161.GetValue()
        check168 = self.checkBox162.GetValue()
        check169 = self.checkBox163.GetValue()
        check170 = self.checkBox164.GetValue()
        check171 = self.checkBox165.GetValue()
        check172 = self.checkBox166.GetValue()
        check173 = self.checkBox167.GetValue()

        if check165 == True:
            b67 = 1
        else:
            b67 = 0

        if check166 == True:
            b68 = 1
        else:
            b68 = 0

        if check167 == True:
            b69 = 1
        else:
            b69 = 0

        if check168 == True:
            b70 = 1
        else:
            b70 = 0

        if check169 == True:
            b71 = 1
        else:
            b71 = 0

        if check170 == True:
            b72 = 1
        else:
            b72 = 0

        if check171 == True:
            b73 = 1
        else:
            b73 = 0

        if check172 == True:
            b74 = 1
        else:
            b74 = 0

        if check173 == True:
            b75 = 1
        else:
            b75 = 0


###########################################################################
## Class Clause11
###########################################################################

class Clause11 ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Supplier Mark", pos = wx.DefaultPosition, size = wx.Size( 887,213 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(887,213), wx.Size(887,213) )
	self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )

	bSizer95 = wx.BoxSizer( wx.VERTICAL )

	bSizer117 = wx.BoxSizer( wx.HORIZONTAL )

	self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 11", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText64.Wrap( -1 )
	self.staticText64.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )

	bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )


	bSizer117.Add( ( 200, 0), 0, 0, 5 )

	self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText65.Wrap( -1 )
	self.staticText65.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )

	bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )


	bSizer95.Add( bSizer117, 0, wx.EXPAND, 5 )

	bSizer116 = wx.BoxSizer( wx.HORIZONTAL )

	bSizer104 = wx.BoxSizer( wx.VERTICAL )

	bSizer96 = wx.BoxSizer( wx.HORIZONTAL )

	self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 11", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText49.Wrap( -1 )
	self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )


	bSizer96.Add( ( 60, 0), 0, 0, 5 )

	self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"Supplier mark", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText103.Wrap( -1 )
	self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )


	bSizer96.Add( ( 50, 0), 1, 0, 5 )

	self.staticText50 = wx.StaticText( self, wx.ID_ANY, u"Each product shall be marked with\nits supplier mark and conforms to\nthe requirements of this standard\nupon of sales. Supplier mark shall\nbe affixed to the product.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText50.Wrap( -1 )
	self.staticText50.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer96.Add( self.staticText50, 0, wx.ALL, 5 )


	bSizer104.Add( bSizer96, 0, 0, 5 )


	bSizer116.Add( bSizer104, 1, wx.EXPAND, 5 )

	bSizer105 = wx.BoxSizer( wx.VERTICAL )

	bSizer106 = wx.BoxSizer( wx.HORIZONTAL )

	self.checkBox168 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox168.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer106.Add( self.checkBox168, 0, wx.ALL, 5 )

	self.checkBox169 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox169.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer106.Add( self.checkBox169, 0, wx.ALL, 5 )

	self.checkBox170 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.checkBox170.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer106.Add( self.checkBox170, 0, wx.ALL, 5 )


	bSizer105.Add( bSizer106, 0, wx.ALIGN_CENTER_HORIZONTAL, 5 )


	bSizer116.Add( bSizer105, 1, 0, 5 )


	bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )

	bSizer103 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer103.Add( self.button33, 0, wx.ALL, 5 )

	self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer103.Add( self.button34, 0, wx.ALL, 5 )


	bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )


	self.SetSizer( bSizer95 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.checkBox168.Bind( wx.EVT_CHECKBOX, self.check1 )
	self.checkBox169.Bind( wx.EVT_CHECKBOX, self.check2 )
	self.checkBox170.Bind( wx.EVT_CHECKBOX, self.check3 )
	self.button33.Bind( wx.EVT_BUTTON, self.Reset )
	self.button34.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def check1( self, event ):
        #event.Skip()
        tick174 = LoadWorkbook2(parent=self.checkBox168)
        mtick174 = tick174.wb2.get_sheet_by_name('Audit Checklist')
        put174 = mtick174['E57']
        box174 = self.checkBox168.GetValue()

        if box174 == True:
            put174.value = "/"
        else:
            put174.value = ""

def check2( self, event ):
        #event.Skip()
        global team175
        tick175 = LoadWorkbook2(parent=self.checkBox169)
        mtick175 = tick175.wb2.get_sheet_by_name('Audit Checklist')
        put175 = mtick175['F57']
        box175 = self.checkBox169.GetValue()

        if box175 == True:
            put175.value = "/"
        else:
            put175.value = ""

def check3( self, event ):
        #event.Skip()
        global team176
        tick176 = LoadWorkbook2(parent=self.checkBox170)
        mtick176 = tick176.wb2.get_sheet_by_name('Audit Checklist')
        put176 = mtick176['G57']
        box176 = self.checkBox170.GetValue()

        if box176 == True:
            put176.value = "/"
        else:
            put176.value = ""

def Reset( self, event ):
		#event.Skip()
        self.checkBox168.SetValue(False)
        self.checkBox169.SetValue(False)
        self.checkBox170.SetValue(False)

def Save( self, event ):
        #event.Skip()
        check174 = self.checkBox168.GetValue()
        check175 = self.checkBox169.GetValue()
        check176 = self.checkBox170.GetValue()

        if check174 == True:
            b76 = 1
        else:
            b76 = 0

        if check175 == True:
            b77 = 1
        else:
            b77 = 0

        if check176 == True:
            b78 = 1
        else:
            b78 = 0


###########################################################################
## Class AuditScore
###########################################################################

class AuditScore ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Audit Score", pos = wx.DefaultPosition, size = wx.Size( 491,831 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(491,831), wx.Size(491,831) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

	bSizer244 = wx.BoxSizer( wx.VERTICAL )

	bSizer255 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer255.Add( ( 120, 0), 0, 0, 5 )

	self.staticText224 = wx.StaticText( self, wx.ID_ANY, u"State Date", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText224.Wrap( -1 )
	self.staticText224.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer255.Add( self.staticText224, 0, wx.ALL, 5 )

	self.datePicker3 = wx.adv.DatePickerCtrl( self, wx.ID_ANY, wx.DefaultDateTime, wx.DefaultPosition, wx.DefaultSize, wx.adv.DP_DEFAULT )
	self.datePicker3.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer255.Add( self.datePicker3, 0, wx.ALL, 5 )


	bSizer244.Add( bSizer255, 1, wx.EXPAND, 5 )

	bSizer245 = wx.BoxSizer( wx.VERTICAL )

	self.m_button48 = wx.Button( self, wx.ID_ANY, u"Section 7.1 - GENERAL REQUIREMENT", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button48.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button48, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button49 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (a) - CORE MANAGEMENT\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\n", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button49.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button49, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button50 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (b) - CLEANING OF ALL INTERNAL AND EXTERNAL COMPONENTS\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\n", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button50.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button50, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button51 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (c) - REPLACEMENT AND/OR\nRESTORATION OF COMPONENTS", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button51.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button51, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button52 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (d) - MACHINING, INSPECTION AND TESTING", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button52.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button52, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button53 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (e) - COMPONENT ASSEMBLY", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button53.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button53, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button54 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (f) - FINAL INSPECTION", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button54.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button54, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button55 = wx.Button( self, wx.ID_ANY, u"Section 9.1 - LABELLING AND PACKAGING\t", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button55.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button55, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button56 = wx.Button( self, wx.ID_ANY, u"Section 10 - WARRANTY", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button56.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button56, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

	self.m_button57 = wx.Button( self, wx.ID_ANY, u"Section 11 - SUPPLIER MARK", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
	self.m_button57.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer245.Add( self.m_button57, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


	bSizer244.Add( bSizer245, 1, wx.EXPAND, 5 )

	bSizer246 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer246.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.m_button58 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.m_button58.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer246.Add( self.m_button58, 0, wx.ALL, 5 )


	bSizer244.Add( bSizer246, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer244 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.datePicker3.Bind( wx.adv.EVT_DATE_CHANGED, self.Date3 )
	self.m_button48.Bind( wx.EVT_BUTTON, self.Section1 )
	self.m_button49.Bind( wx.EVT_BUTTON, self.Section2 )
	self.m_button50.Bind( wx.EVT_BUTTON, self.Section3 )
	self.m_button51.Bind( wx.EVT_BUTTON, self.Section4 )
	self.m_button52.Bind( wx.EVT_BUTTON, self.Section5 )
	self.m_button53.Bind( wx.EVT_BUTTON, self.Section6 )
	self.m_button54.Bind( wx.EVT_BUTTON, self.Section7 )
	self.m_button55.Bind( wx.EVT_BUTTON, self.Section8 )
	self.m_button56.Bind( wx.EVT_BUTTON, self.Section9 )
	self.m_button57.Bind( wx.EVT_BUTTON, self.Section10 )
	self.m_button58.Bind( wx.EVT_BUTTON, self.Reset )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Date3( self, event ):
	event.Skip()

def Section1( self, event ):
        #event.Skip()
        Section1Form = Section71(parent=self.m_button48)
        Section1Form.Show()

def Section2( self, event ):
        #event.Skip()
        Section2Form = Section731a(parent=self.m_button49)
        Section2Form.Show()

def Section3( self, event ):
        #event.Skip()
        Section3Form = Section731b(parent=self.m_button50)
        Section3Form.Show()

def Section4( self, event ):
        #event.Skip()
        Section4Form = Seksyen731c(parent=self.m_button51)
        Section4Form.Show()

def Section5( self, event ):
        #event.Skip()
        Section5Form = Section731d(parent=self.m_button52)
        Section5Form.Show()

def Section6( self, event ):
        #event.Skip()
        Section6Form = Section731e(parent=self.m_button53)
        Section6Form.Show()

def Section7( self, event ):
        #event.Skip()
        Section7Form = Section731f(parent = self.m_button54)
        Section7Form.Show()

def Section8( self, event ):
        #event.Skip()
        Section8Form = Section91(parent = self.m_button55)
        Section8Form.Show()

def Section9( self, event ):
        #event.Skip()
        Section9Form = Section10(parent = self.m_button56)
        Section9Form.Show()

def Section10( self, event ):
        #event.Skip()
        Section10Form = Section11(parent = self.m_button57)
        Section10Form.Show()

def Reset( self, event ):
        event.Skip()


###########################################################################
## Class Section71
###########################################################################

class Section71 ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"General Requirement", pos = wx.DefaultPosition, size = wx.Size( 1151,750 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(1151,750), wx.Size(1151,750) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Internal Audit and Management Review", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 50), 0, 0, 5 )

	self.staticText213 = wx.StaticText( self, wx.ID_ANY, u"II. Safety, Health and Environmental Regulation Compliance", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText213.Wrap( -1 )
	self.staticText213.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText213, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 50), 0, 0, 5 )

	self.staticText214 = wx.StaticText( self, wx.ID_ANY, u"III. Competency Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText214.Wrap( -1 )
	self.staticText214.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText214, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 50), 0, 0, 5 )

	self.staticText215 = wx.StaticText( self, wx.ID_ANY, u"IV. Facilities Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText215.Wrap( -1 )
	self.staticText215.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText215, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 45), 0, 0, 5 )

	self.staticText216 = wx.StaticText( self, wx.ID_ANY, u"V. Environmental Quality Act 1974.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText216.Wrap( -1 )
	self.staticText216.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText216, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 35), 0, 0, 5 )

	self.staticText217 = wx.StaticText( self, wx.ID_ANY, u"VI. Standard Operating Procedure", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText217.Wrap( -1 )
	self.staticText217.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText217, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 45), 0, 0, 5 )

	self.staticText218 = wx.StaticText( self, wx.ID_ANY, u"VII. Calibration Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText218.Wrap( -1 )
	self.staticText218.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText218, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 40), 0, 0, 5 )

	self.staticText219 = wx.StaticText( self, wx.ID_ANY, u"VIII. Tools and Equipments For Assembly", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText219.Wrap( -1 )
	self.staticText219.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText219, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 40), 0, 0, 5 )

	self.staticText220 = wx.StaticText( self, wx.ID_ANY, u"IX. Documentation", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText220.Wrap( -1 )
	self.staticText220.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText220, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 1, wx.EXPAND, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Conduct for Management review for the 4R2S implementation performance. \n4R2S Management review record and conduct 4R2S internal self audit.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText222 = wx.StaticText( self, wx.ID_ANY, u"Machining of remanufacturing automotive parts and component \nshall be done in compliance with the relevant safety, health and \nenvironmental regulations. Parts that cannot be reused must be replaced.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText222.Wrap( -1 )
	self.staticText222.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText222, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText223 = wx.StaticText( self, wx.ID_ANY, u"Documented, establish the personnel competency training requirement.\nIdentify any personnel under training certification from SKM.\nRemanufacturing process shall be conducted in accordance to industry best practices.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText223.Wrap( -1 )
	self.staticText223.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText223, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText224 = wx.StaticText( self, wx.ID_ANY, u"The facilities shall have a dedicated area/section in their compound for each\nof the remanufacturing operation. The area should be in compliance with the\nrelevant safety, health and environmental regulations.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText224.Wrap( -1 )
	self.staticText224.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText224, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText225 = wx.StaticText( self, wx.ID_ANY, u"Storage, labelling and disposal of hazardous wastes shall be managed in\naccordance with the Environmental Quality Act 1974.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText225.Wrap( -1 )
	self.staticText225.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText225, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText226 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing processes required as good as new per Standard\nOperating Procedure/ OEM manual.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText226.Wrap( -1 )
	self.staticText226.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText226, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText227 = wx.StaticText( self, wx.ID_ANY, u"Tools and equipment used for inspection and testing and machines used for\nremanufacturing process must be calibrated and the calibration reports\nare able to be presented.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText227.Wrap( -1 )
	self.staticText227.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText227, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText228 = wx.StaticText( self, wx.ID_ANY, u"Personnel shall be equipped with necessary tools and equipments to\nperform the assembly.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText228.Wrap( -1 )
	self.staticText228.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText228, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText229 = wx.StaticText( self, wx.ID_ANY, u"All inspection and testing results shall be fully documented, recorded and\nretained for at least 7 years upon completion of sales.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText229.Wrap( -1 )
	self.staticText229.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText229, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer249, 1, wx.EXPAND, 5 )

	bSizer185 = wx.BoxSizer( wx.VERTICAL )

	bSizer258 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer258.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText228 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText228.Wrap( -1 )
	bSizer258.Add( self.staticText228, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer258, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 40), 0, 0, 5 )

	bSizer259 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl7 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer259.Add( self.textCtrl7, 0, wx.ALL, 5 )

	self.staticText229 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText229.Wrap( -1 )
	bSizer259.Add( self.staticText229, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer259, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 40), 0, 0, 5 )

	bSizer260 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl8 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer260.Add( self.textCtrl8, 0, wx.ALL, 5 )

	self.staticText230 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText230.Wrap( -1 )
	bSizer260.Add( self.staticText230, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer260, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 40), 0, 0, 5 )

	bSizer261 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl9 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer261.Add( self.textCtrl9, 0, wx.ALL, 5 )

	self.m_staticText231 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.m_staticText231.Wrap( -1 )
	bSizer261.Add( self.m_staticText231, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer261, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 45), 0, 0, 5 )

	bSizer262 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl10 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer262.Add( self.textCtrl10, 0, wx.ALL, 5 )

	self.staticText232 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText232.Wrap( -1 )
	bSizer262.Add( self.staticText232, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer262, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 25), 0, 0, 5 )

	bSizer263 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl11 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer263.Add( self.textCtrl11, 0, wx.ALL, 5 )

	self.staticText233 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText233.Wrap( -1 )
	bSizer263.Add( self.staticText233, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer263, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 30), 0, 0, 5 )

	bSizer264 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl12 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer264.Add( self.textCtrl12, 0, wx.ALL, 5 )

	self.staticText234 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText234.Wrap( -1 )
	bSizer264.Add( self.staticText234, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer264, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 45), 0, 0, 5 )

	bSizer265 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl13 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer265.Add( self.textCtrl13, 0, wx.ALL, 5 )

	self.staticText235 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText235.Wrap( -1 )
	bSizer265.Add( self.staticText235, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer265, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 25), 0, 0, 5 )

	bSizer266 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl14 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer266.Add( self.textCtrl14, 0, wx.ALL, 5 )

	self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText236.Wrap( -1 )
	bSizer266.Add( self.staticText236, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer266, 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer185, 1, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 1, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.textCtrl7.Bind( wx.EVT_TEXT, self.Score2 )
	self.textCtrl8.Bind( wx.EVT_TEXT, self.Score3 )
	self.textCtrl9.Bind( wx.EVT_TEXT, self.Score4 )
	self.textCtrl10.Bind( wx.EVT_TEXT, self.Score5 )
	self.textCtrl11.Bind( wx.EVT_TEXT, self.Score6 )
	self.textCtrl12.Bind( wx.EVT_TEXT, self.Score7 )
	self.textCtrl13.Bind( wx.EVT_TEXT, self.Score8 )
	self.textCtrl14.Bind( wx.EVT_TEXT, self.Score9 )
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )
def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get1
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 7.1')
        input1 = mscore1['C8']
        get1 = self.textCtrl6.GetValue()

def Score2( self, event ):
        #event.Skip()
        global input2
        global get2
        score2 = LoadWorkbook(parent=self.textCtrl7)
        mscore2 = score2.wb.get_sheet_by_name('Sec 7.1')
        input2 = mscore2['C12']
        get2 = self.textCtrl7.GetValue()

def Score3( self, event ):
        #event.Skip()
        global input3
        global get3
        score3 = LoadWorkbook(parent=self.textCtrl8)
        mscore3 = score3.wb.get_sheet_by_name('Sec 7.1')
        input3 = mscore3['C16']
        get3 = self.textCtrl8.GetValue()

def Score4( self, event ):
        #event.Skip()
        global input4
        global get4
        score4 = LoadWorkbook(parent=self.textCtrl9)
        mscore4 = score4.wb.get_sheet_by_name('Sec 7.1')
        input4 = mscore4['C21']
        get4 = self.textCtrl9.GetValue()

def Score5( self, event ):
        #event.Skip()
        global input5
        global get5
        score5 = LoadWorkbook(parent=self.textCtrl10)
        mscore5 = score5.wb.get_sheet_by_name('Sec 7.1')
        input5 = mscore5['C25']
        get5 = self.textCtrl10.GetValue()

def Score6( self, event ):
        #event.Skip()
        global input6
        global get6
        score6 = LoadWorkbook(parent=self.textCtrl11)
        mscore6 = score6.wb.get_sheet_by_name('Sec 7.1')
        input6 = mscore6['C29']
        get6 = self.textCtrl11.GetValue()

def Score7( self, event ):
        #event.Skip()
        global input7
        global get7
        score7 = LoadWorkbook(parent=self.textCtrl12)
        mscore7 = score7.wb.get_sheet_by_name('Sec 7.1')
        input7 = mscore7['C33']
        get7 = self.textCtrl12.GetValue()

def Score8( self, event ):
        #event.Skip()
        global input8
        global get8
        score8 = LoadWorkbook(parent=self.textCtrl13)
        mscore8 = score8.wb.get_sheet_by_name('Sec 7.1')
        input8 = mscore8['C37']
        get8 = self.textCtrl13.GetValue()

def Score9( self, event ):
        #event.Skip()
        global input9
        global get9
        score9 = LoadWorkbook(parent=self.textCtrl14)
        mscore9 = score9.wb.get_sheet_by_name('Sec 7.1')
        input9 = mscore9['C41']
        get9 = self.textCtrl14.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get1
        input2.value = get2
        input3.value = get3
        input4.value = get4
        input5.value = get5
        input6.value = get6
        input7.value = get7
        input8.value = get8
        input9.value = get9

        word1 = get1
        word2 = get2
        word3 = get3
        word4 = get4
        word5 = get5
        word6 = get6
        word7 = get7
        word8 = get8
        word9 = get9

###########################################################################
## Class Section731a
###########################################################################

class Section731a ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Core Management", pos = wx.DefaultPosition, size = wx.Size( 857,980 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(857,980), wx.Size(857,980) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )

	bSizer196 = wx.BoxSizer( wx.HORIZONTAL )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Data Retention", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 30), 0, 0, 5 )

	self.staticText213 = wx.StaticText( self, wx.ID_ANY, u"II. Documented Information", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText213.Wrap( -1 )
	self.staticText213.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText213, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 25), 0, 0, 5 )

	self.staticText214 = wx.StaticText( self, wx.ID_ANY, u"III. Parts and Components Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText214.Wrap( -1 )
	self.staticText214.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText214, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 20), 0, 0, 5 )

	self.staticText215 = wx.StaticText( self, wx.ID_ANY, u"IV. Core Sorting", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText215.Wrap( -1 )
	self.staticText215.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText215, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 25), 0, 0, 5 )

	self.staticText216 = wx.StaticText( self, wx.ID_ANY, u"V. Disassembly Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText216.Wrap( -1 )
	self.staticText216.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText216, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 30), 0, 0, 5 )

	self.staticText217 = wx.StaticText( self, wx.ID_ANY, u"VI. Disassembly Selection", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText217.Wrap( -1 )
	self.staticText217.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText217, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 35), 0, 0, 5 )

	self.staticText218 = wx.StaticText( self, wx.ID_ANY, u"VII. Physical Examination", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText218.Wrap( -1 )
	self.staticText218.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText218, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 260), 0, 0, 5 )

	self.staticText219 = wx.StaticText( self, wx.ID_ANY, u"VIII. Test Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText219.Wrap( -1 )
	self.staticText219.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText219, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 125), 0, 0, 5 )

	self.staticText220 = wx.StaticText( self, wx.ID_ANY, u"IX.  Component for Recycle", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText220.Wrap( -1 )
	self.staticText220.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText220, 0, wx.ALL, 5 )


	bSizer196.Add( bSizer248, 1, wx.EXPAND, 5 )


	bSizer196.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Essential information such as and not limited to serial number\nand part name are recorded to ensure traceability of components.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 10), 0, 0, 5 )

	self.staticText222 = wx.StaticText( self, wx.ID_ANY, u"Document such as quality manual, procedure, SOP or form and\netc are well maintained.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText222.Wrap( -1 )
	self.staticText222.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText222, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 10), 0, 0, 5 )

	self.staticText223 = wx.StaticText( self, wx.ID_ANY, u"Parts and components which are not allowed as used parts and\ncomponents under the relevant regulation shall not be repaired.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText223.Wrap( -1 )
	self.staticText223.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText223, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 10), 0, 0, 5 )

	self.staticText224 = wx.StaticText( self, wx.ID_ANY, u"Cores are sorted to relevant categories and are tagged\naccordingly to reuse the core or recycle the core.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText224.Wrap( -1 )
	self.staticText224.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText224, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 10), 0, 0, 5 )

	self.staticText225 = wx.StaticText( self, wx.ID_ANY, u"Disassembly is required for used parts and components to\ndetermine its usability.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText225.Wrap( -1 )
	self.staticText225.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText225, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 10), 0, 0, 5 )

	self.staticText226 = wx.StaticText( self, wx.ID_ANY, u"Consideration shall be taken with regards to the type of\ndisassembly method, tools and machine selected for\ndisassembly processes.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText226.Wrap( -1 )
	self.staticText226.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText226, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 10), 0, 0, 5 )

	self.staticText190 = wx.StaticText( self, wx.ID_ANY, u"a. Regulated by relevant authorities with an identification mark", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText190.Wrap( -1 )
	self.staticText190.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

	bSizer249.Add( self.staticText190, 0, wx.TOP|wx.RIGHT|wx.LEFT, 5 )

	self.staticText227 = wx.StaticText( self, wx.ID_ANY, u"Any reused parts and components that are regulated by relevant\nauthorities shall be affixed with an identification mark. This\nidentification mark shall identify the source of the used parts and\ncomponents. A detailed record of the identification mark shall be\nmaintained.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText227.Wrap( -1 )
	self.staticText227.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText227, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 5), 0, 0, 5 )

	self.staticText191 = wx.StaticText( self, wx.ID_ANY, u"b. Part and component", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText191.Wrap( -1 )
	self.staticText191.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

	bSizer249.Add( self.staticText191, 0, wx.TOP|wx.RIGHT|wx.LEFT, 5 )

	self.staticText228 = wx.StaticText( self, wx.ID_ANY, u"The serial number for the parts shall be recorded. If the partsor \ncomponents serial numbers cannot be read or do not match\nexisting available records for that vehicle. The safety parts and\ncomponents shall not be sold as used, shall be recycled.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText228.Wrap( -1 )
	self.staticText228.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText228, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 5), 0, 0, 5 )

	self.staticText192 = wx.StaticText( self, wx.ID_ANY, u"c. Part visual inspection", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText192.Wrap( -1 )
	self.staticText192.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

	bSizer249.Add( self.staticText192, 0, wx.TOP|wx.RIGHT|wx.LEFT, 5 )

	self.staticText229 = wx.StaticText( self, wx.ID_ANY, u"Visual inspection of parts and components shall be undertaken to \nlook for any critical or physical damages.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText229.Wrap( -1 )
	self.staticText229.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText229, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 10), 0, 0, 5 )

	self.staticText193 = wx.StaticText( self, wx.ID_ANY, u"a. Test requirement -Functional test", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText193.Wrap( -1 )
	self.staticText193.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

	bSizer249.Add( self.staticText193, 0, 0, 5 )

	self.staticText188 = wx.StaticText( self, wx.ID_ANY, u"General condition of the used parts and components shall be\nassessed and recorded in order to determine its usability\ncondition. The used parts and components shall be subjected to\na functional/bench test in order to ensure that the parts are fit to\nbe reused. Proof of functionality shall be recorded, where\nnecessary. (Include visual and dimensional check, Fluids\ninspection)", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText188.Wrap( -1 )
	self.staticText188.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText188, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 10), 0, 0, 5 )

	self.staticText194 = wx.StaticText( self, wx.ID_ANY, u"a. Depolluted consideration", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText194.Wrap( -1 )
	self.staticText194.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

	bSizer249.Add( self.staticText194, 0, wx.TOP|wx.RIGHT|wx.LEFT, 5 )

	self.staticText189 = wx.StaticText( self, wx.ID_ANY, u"Consideration shall be taken to ensure that the products of the\nrecyclable materials are depolluted prior to dismantling, handling\nand segregation.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText189.Wrap( -1 )
	self.staticText189.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText189, 0, wx.ALL, 5 )


	bSizer196.Add( bSizer249, 1, wx.EXPAND, 5 )


	bSizer196.Add( ( 20, 0), 0, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.VERTICAL )

	bSizer267 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer267.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText237.Wrap( -1 )
	bSizer267.Add( self.staticText237, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer267, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 15), 0, 0, 5 )

	bSizer268 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl7 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer268.Add( self.textCtrl7, 0, wx.ALL, 5 )

	self.staticText238 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText238.Wrap( -1 )
	bSizer268.Add( self.staticText238, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer268, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 20), 0, 0, 5 )

	bSizer269 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl8 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer269.Add( self.textCtrl8, 0, wx.ALL, 5 )

	self.staticText239 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText239.Wrap( -1 )
	bSizer269.Add( self.staticText239, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer269, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 15), 0, 0, 5 )

	bSizer270 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl9 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer270.Add( self.textCtrl9, 0, wx.ALL, 5 )

	self.staticText240 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText240.Wrap( -1 )
	bSizer270.Add( self.staticText240, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer270, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 15), 0, 0, 5 )

	bSizer271 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl10 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer271.Add( self.textCtrl10, 0, wx.ALL, 5 )

	self.staticText241 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText241.Wrap( -1 )
	bSizer271.Add( self.staticText241, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer271, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 15), 0, 0, 5 )

	bSizer272 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl11 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer272.Add( self.textCtrl11, 0, wx.ALL, 5 )

	self.staticText242 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText242.Wrap( -1 )
	bSizer272.Add( self.staticText242, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer272, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 30), 0, 0, 5 )

	bSizer273 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl12 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer273.Add( self.textCtrl12, 0, wx.ALL, 5 )

	self.staticText243 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText243.Wrap( -1 )
	bSizer273.Add( self.staticText243, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer273, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 75), 0, 0, 5 )

	bSizer274 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl13 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer274.Add( self.textCtrl13, 0, wx.ALL, 5 )

	self.staticText244 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText244.Wrap( -1 )
	bSizer274.Add( self.staticText244, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer274, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 60), 0, 0, 5 )

	bSizer275 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl14 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer275.Add( self.textCtrl14, 0, wx.ALL, 5 )

	self.staticText245 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText245.Wrap( -1 )
	bSizer275.Add( self.staticText245, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer275, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 30), 0, 0, 5 )

	bSizer276 = wx.BoxSizer( wx.HORIZONTAL )

	self.m_textCtrl33 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer276.Add( self.m_textCtrl33, 0, wx.ALL, 5 )

	self.staticText246 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText246.Wrap( -1 )
	bSizer276.Add( self.staticText246, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer276, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 100), 0, 0, 5 )

	bSizer277 = wx.BoxSizer( wx.HORIZONTAL )

	self.m_textCtrl34 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer277.Add( self.m_textCtrl34, 0, wx.ALL, 5 )

	self.staticText247 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText247.Wrap( -1 )
	bSizer277.Add( self.staticText247, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer277, 1, wx.EXPAND, 5 )


	bSizer196.Add( bSizer185, 1, wx.EXPAND, 5 )


	bSizer247.Add( bSizer196, 1, wx.EXPAND, 5 )

	bSizer197 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer197.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button53 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button53.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer197.Add( self.button53, 0, wx.ALL, 5 )

	self.button54 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button54.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer197.Add( self.button54, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer197, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.textCtrl7.Bind( wx.EVT_TEXT, self.Score2 )
	self.textCtrl8.Bind( wx.EVT_TEXT, self.Score3 )
	self.textCtrl9.Bind( wx.EVT_TEXT, self.Score4 )
	self.textCtrl10.Bind( wx.EVT_TEXT, self.Score5 )
	self.textCtrl11.Bind( wx.EVT_TEXT, self.Score6 )
	self.textCtrl12.Bind( wx.EVT_TEXT, self.Score7 )
	self.textCtrl13.Bind( wx.EVT_TEXT, self.Score8 )
	self.textCtrl14.Bind( wx.EVT_TEXT, self.Score9 )
	self.m_textCtrl33.Bind( wx.EVT_TEXT, self.Score10 )
	self.m_textCtrl34.Bind( wx.EVT_TEXT, self.Score11 )
	self.button53.Bind( wx.EVT_BUTTON, self.Reset )
	self.button54.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get10
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input1 = mscore1['C9']
        get10 = self.textCtrl6.GetValue()

def Score2( self, event ):
        #event.Skip()
        global input2
        global get11
        score2 = LoadWorkbook(parent=self.textCtrl7)
        mscore2 = score2.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input2 = mscore2['C13']
        get11 = self.textCtrl7.GetValue()

def Score3( self, event ):
        #event.Skip()
        global input3
        global get12
        score3 = LoadWorkbook(parent=self.textCtrl8)
        mscore3 = score3.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input3 = mscore3['C17']
        get12 = self.textCtrl8.GetValue()

def Score4( self, event ):
        #event.Skip()
        global input4
        global get13
        score4 = LoadWorkbook(parent=self.textCtrl9)
        mscore4 = score4.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input4 = mscore4['C21']
        get13 = self.textCtrl9.GetValue()

def Score5( self, event ):
        #event.Skip()
        global input5
        global get14
        score5 = LoadWorkbook(parent=self.textCtrl10)
        mscore5 = score5.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input5 = mscore5['C25']
        get14 = self.textCtrl10.GetValue()

def Score6( self, event ):
        #event.Skip()
        global input6
        global get15
        score6 = LoadWorkbook(parent=self.textCtrl11)
        mscore6 = score6.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input6 = mscore6['C29']
        get15 = self.textCtrl11.GetValue()

def Score7( self, event ):
        #event.Skip()
        global input7
        global get16
        score7 = LoadWorkbook(parent=self.textCtrl12)
        mscore7 = score7.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input7 = mscore7['C34']
        get16 = self.textCtrl12.GetValue()

def Score8( self, event ):
        #event.Skip()
        global input8
        global get17
        score8 = LoadWorkbook(parent=self.textCtrl13)
        mscore8 = score8.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input8 = mscore8['C38']
        get17 = self.textCtrl13.GetValue()

def Score9( self, event ):
        #event.Skip()
        global input9
        global get18
        score9 = LoadWorkbook(parent=self.textCtrl14)
        mscore9 = score9.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input9 = mscore9['C42']
        get18 = self.textCtrl14.GetValue()

def Score10( self, event ):
        #event.Skip()
        global input10
        global get19
        score10 = LoadWorkbook(parent=self.m_textCtrl33)
        mscore10 = score10.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input10 = mscore10['C47']
        get19 = self.m_textCtrl33.GetValue()

def Score11( self, event ):
        #event.Skip()
        global input11
        global get20
        score11 = LoadWorkbook(parent=self.m_textCtrl34)
        mscore11 = score11.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input11 = mscore11['C52']
        get20 = self.m_textCtrl34.GetValue()

def Reset( self, event ):
        event.Skip()


def Save( self, event ):
        #event.Skip()
        input1.value = get10
        input2.value = get11
        input3.value = get12
        input4.value = get13
        input5.value = get14
        input6.value = get15
        input7.value = get16
        input8.value = get17
        input9.value = get18
        input10.value = get19
        input11.value = get20

        word10 = get10
        word11 = get11
        word12 = get12
        word13 = get13
        word14 = get14
        word15 = get15
        word16 = get16
        word17 = get17
        word18 = get18
        word19 = get19
        word20 = get20

###########################################################################
## Class Section731b
###########################################################################

class Section731b ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Cleaning Of All Internal And External Components", pos = wx.DefaultPosition, size = wx.Size( 779,172 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(779,172), wx.Size(779,172) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 10, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Cleaning Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Internal and external components that will undergo\nremanufacturing process shall be cleaned.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 0), 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer249, 1, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer185.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText248 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText248.Wrap( -1 )
	bSizer185.Add( self.staticText248, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 10), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )


	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get21
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 7.3.1 (b)')
        input1 = mscore1['C8']
        get21 = self.textCtrl6.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get21

        word21 = get21



###########################################################################
## Class Seksyen731c
###########################################################################

class Seksyen731c ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Replacement And/Or Restoration of Components", pos = wx.DefaultPosition, size = wx.Size( 758,310 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(758,310), wx.Size(758,310) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 10, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Replacement of Components", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 65), 0, 0, 5 )

	self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Restoration of Components", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText236.Wrap( -1 )
	self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Parts and components which are not allowed\nas used parts and components under the \nrelevant regulation(brake lining, battery,\ntires) and/or missing shall be replaced.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"Malfunction parts and components that is fit\nfor restoring within approved tolerance\naccording to established industrial\nspecifications shall be repaired.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText237.Wrap( -1 )
	self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText237, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer249, 1, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.VERTICAL )

	bSizer278 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer278.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText249 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText249.Wrap( -1 )
	bSizer278.Add( self.staticText249, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer278, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 60), 0, 0, 5 )

	bSizer279 = wx.BoxSizer( wx.HORIZONTAL )

	self.m_textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer279.Add( self.m_textCtrl57, 0, wx.ALL, 5 )

	self.staticText250 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText250.Wrap( -1 )
	bSizer279.Add( self.staticText250, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer279, 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 10), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )


	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.m_textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get22
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 7.3.1 (c)')
        input1 = mscore1['C8']
        get22 = self.textCtrl6.GetValue()

def Score2( self, event ):
        #event.Skip()
        global input2
        global get23
        score2 = LoadWorkbook(parent=self.m_textCtrl57)
        mscore2 = score2.wb.get_sheet_by_name('Sec 7.3.1 (c)')
        input2 = mscore2['C12']
        get23 = self.m_textCtrl57.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get22
        input2.value = get23

        word22 = get22
        word23 = get23



###########################################################################
## Class Section731d
###########################################################################

class Section731d ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Machining, Inspection and Testing", pos = wx.DefaultPosition, size = wx.Size( 770,360 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(770,360), wx.Size(770,360) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 10, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Replacement of Components", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 50), 0, 0, 5 )

	self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Restoration of Components", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText236.Wrap( -1 )
	self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 50), 0, 0, 5 )

	self.staticText246 = wx.StaticText( self, wx.ID_ANY, u"III. Tools and Equipments", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText246.Wrap( -1 )
	self.staticText246.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText246, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"All remanufacturing components are required to undergo\ninspection and testing before assembly to ensure\nall the internals are in working condition.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"Reworking, machining or performing such other operations\nas are necessary to put the part in original working condition\nor better.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText237.Wrap( -1 )
	self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText237, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

	self.staticText247 = wx.StaticText( self, wx.ID_ANY, u"Personnel shall be equipped with necessary PPE to perform\npainting process as per relevant Standard Operating Procedure (SOP).", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText247.Wrap( -1 )
	self.staticText247.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText247, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer249, 1, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.VERTICAL )

	bSizer280 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer280.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText251 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText251.Wrap( -1 )
	bSizer280.Add( self.staticText251, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer280, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 40), 0, 0, 5 )

	bSizer281 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer281.Add( self.textCtrl57, 0, wx.ALL, 5 )

	self.staticText252 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText252.Wrap( -1 )
	bSizer281.Add( self.staticText252, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer281, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 40), 0, 0, 5 )

	bSizer282 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl62 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer282.Add( self.textCtrl62, 0, wx.ALL, 5 )

	self.staticText253 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText253.Wrap( -1 )
	bSizer282.Add( self.staticText253, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer282, 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 10), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )


	bSizer199.Add( ( 0, 30), 0, 0, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )


	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
	self.textCtrl62.Bind( wx.EVT_TEXT, self.Score3 )
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get24
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 7.3.1 (d)')
        input1 = mscore1['C7']
        get24 = self.textCtrl6.GetValue()

def Score2( self, event ):
        #event.Skip()
        global input2
        global get25
        score2 = LoadWorkbook(parent=self.textCtrl57)
        mscore2 = score2.wb.get_sheet_by_name('Sec 7.3.1 (d)')
        input2 = mscore2['C11']
        get25 = self.textCtrl57.GetValue()

def Score3( self, event ):
        #event.Skip()
        global input3
        global get26
        score3 = LoadWorkbook(parent=self.textCtrl62)
        mscore3 = score3.wb.get_sheet_by_name('Sec 7.3.1 (d)')
        input3 = mscore3['C15']
        get26 = self.textCtrl62.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get24
        input2.value = get25
        input3.value = get26

        word24 = get24
        word25 = get25
        word26 = get26


###########################################################################
## Class Section731e
###########################################################################

class Section731e ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Component Assembly", pos = wx.DefaultPosition, size = wx.Size( 785,180 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(785,180), wx.Size(785,180) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 10, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Assembly Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Assembly process was conducted in accordance to standard\noperating procedure.\t", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 0), 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer249, 1, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer185.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText254 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText254.Wrap( -1 )
	bSizer185.Add( self.staticText254, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 10), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get27
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 7.3.1 (e)')
        input1 = mscore1['C8']
        get27 = self.textCtrl6.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get27

        word27 = get27



###########################################################################
## Class Section731f
###########################################################################

class Section731f ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Final Inspection", pos = wx.DefaultPosition, size = wx.Size( 785,322 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(785,322), wx.Size(785,322) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 10, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Inspection Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 110), 0, 0, 5 )

	self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Inspection and Testing Information", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText236.Wrap( -1 )
	self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Components shall undergo final inspection that\ninclude testing. Components that comply\nrelevant testing standards shall be recorded with\nQC inspection, provided the product conforms\nto the requirements of the standard with inspection label.\nInspection label may be affixed to the product or\npackaging, where appropriate.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"Components were tested to its nature of working condition\naccording to relevant standard information.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText237.Wrap( -1 )
	self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText237, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer249, 1, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.VERTICAL )

	bSizer283 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer283.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText255 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText255.Wrap( -1 )
	bSizer283.Add( self.staticText255, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer283, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	bSizer284 = wx.BoxSizer( wx.HORIZONTAL )

	self.m_textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer284.Add( self.m_textCtrl57, 0, wx.ALL, 5 )

	self.staticText256 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText256.Wrap( -1 )
	bSizer284.Add( self.staticText256, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer284, 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 10), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class

def Score1( self, event ):
        #event.Skip()
        global input1
        global get28
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 7.3.1 (f)')
        input1 = mscore1['C8']
        get28 = self.textCtrl6.GetValue()

def Score2( self, event ):
        #event.Skip()
        global input2
        global get29
        score2 = LoadWorkbook(parent=self.m_textCtrl57)
        mscore2 = score2.wb.get_sheet_by_name('Sec 7.3.1 (f)')
        input2 = mscore2['C12']
        get29 = self.textCtrl57.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get28
        input2.value = get29

        word28 = get28
        word29 = get29



###########################################################################
## Class Section91
###########################################################################

class Section91 ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"labelling And Packaging", pos = wx.DefaultPosition, size = wx.Size( 839,547 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(839,547), wx.Size(839,547) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 10, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Labelling and Identification", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 35), 0, 0, 5 )

	self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Labelling Information", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText236.Wrap( -1 )
	self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 130), 0, 0, 5 )

	self.staticText246 = wx.StaticText( self, wx.ID_ANY, u"III. Packaging Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText246.Wrap( -1 )
	self.staticText246.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText246, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 53), 0, 0, 5 )

	self.staticText272 = wx.StaticText( self, wx.ID_ANY, u"IV. Packaging information", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText272.Wrap( -1 )
	self.staticText272.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText272, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Products shall have a permanent label to be identified\nas reused, repaired or remanufactured parts or components.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"\"Labelling shall include, but not limited to, the following information:\n\na) company name\nb) parts identification, if applicable\nc) make and model\nd) engine capacity, if applicable\ne) classifications according to industry best practices (reuse, remanufacturing or repair)\nf) recovered from local vehicles or imported.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText237.Wrap( -1 )
	self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText237, 0, wx.RIGHT|wx.LEFT, 5 )


	bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

	self.staticText247 = wx.StaticText( self, wx.ID_ANY, u"Types of packaging used shall be able to protect the reused,\nrepaired and remanufactured parts from damage during\nlogistics and transportations.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText247.Wrap( -1 )
	self.staticText247.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText247, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

	self.staticText273 = wx.StaticText( self, wx.ID_ANY, u"Packaging shall include, but not limited to, the following information:\n\na) name of product\nb) manufacturer's details\nc) type of warranty covered\t\t\n", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText273.Wrap( -1 )
	self.staticText273.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText273, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer249, 1, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.VERTICAL )

	bSizer285 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer285.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.m_staticText257 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.m_staticText257.Wrap( -1 )
	bSizer285.Add( self.m_staticText257, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer285, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 25), 0, 0, 5 )

	bSizer286 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer286.Add( self.textCtrl57, 0, wx.ALL, 5 )

	self.staticText258 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText258.Wrap( -1 )
	bSizer286.Add( self.staticText258, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer286, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 130), 0, 0, 5 )

	bSizer287 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl62 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer287.Add( self.textCtrl62, 0, wx.ALL, 5 )

	self.staticText259 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText259.Wrap( -1 )
	bSizer287.Add( self.staticText259, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer287, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 50), 0, 0, 5 )

	bSizer288 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl75 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer288.Add( self.textCtrl75, 0, wx.ALL, 5 )

	self.staticText260 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText260.Wrap( -1 )
	bSizer288.Add( self.staticText260, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer288, 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 10), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )


	bSizer199.Add( ( 0, 30), 0, 0, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
	self.textCtrl62.Bind( wx.EVT_TEXT, self.Score3 )
	self.textCtrl75.Bind( wx.EVT_TEXT, self.Score4 )
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get30
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 9.1')
        input1 = mscore1['C8']
        get30 = self.textCtrl6.GetValue()


def Score2( self, event ):
        #event.Skip()
        global input2
        global get31
        score2 = LoadWorkbook(parent=self.textCtrl57)
        mscore2 = score2.wb.get_sheet_by_name('Sec 9.1')
        input2 = mscore2['C12']
        get31 = self.textCtrl57.GetValue()

def Score3( self, event ):
        #event.Skip()
        global input3
        global get32
        score3 = LoadWorkbook(parent=self.textCtrl62)
        mscore3 = score3.wb.get_sheet_by_name('Sec 9.1')
        input3 = mscore3['C16']
        get32 = self.textCtrl62.GetValue()

def Score4( self, event ):
        #event.Skip()
        global input4
        global get33
        score4 = LoadWorkbook(parent=self.textCtrl75)
        mscore4 = score4.wb.get_sheet_by_name('Sec 9.1')
        input4 = mscore4['C20']
        get33 = self.textCtrl75.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get30
        input2.value = get31
        input3.value = get32
        input4.value = get33

        word30 = get30
        word31 = get31
        word32 = get32
        word33 = get33



###########################################################################
## Class Section10
###########################################################################

class Section10 ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Warranty", pos = wx.DefaultPosition, size = wx.Size( 935,478 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(935,478), wx.Size(935,478) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 10, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Warranty Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 35), 0, 0, 5 )

	self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Warranty Information", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText236.Wrap( -1 )
	self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 95), 0, 0, 5 )

	self.staticText246 = wx.StaticText( self, wx.ID_ANY, u"III. Record of Warranty Claim", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText246.Wrap( -1 )
	self.staticText246.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText246, 0, wx.ALL, 5 )


	bSizer248.Add( ( 0, 36), 0, 0, 5 )

	self.staticText272 = wx.StaticText( self, wx.ID_ANY, u"IV. Record Information", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText272.Wrap( -1 )
	self.staticText272.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText272, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Warranty shall be provided for reused parts, repaired parts and\nremanufacturing parts.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 0, 0, 5 )

	self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"Remanufactured parts and components - not less than 90 days\nfrom the date of purchase\nRepaired parts and components - not less than 45 days from the date of purchase\nReused parts and components - not less than 30 days from the date of purchase\nWarranty shall be in writing and assured by the warranty\nprovider.\" cal vehicles or imported.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText237.Wrap( -1 )
	self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText237, 0, wx.RIGHT|wx.LEFT, 5 )


	bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

	self.staticText247 = wx.StaticText( self, wx.ID_ANY, u"Record of warranty claim is kept and retained for 7 years for\nreferences.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText247.Wrap( -1 )
	self.staticText247.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText247, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

	self.staticText273 = wx.StaticText( self, wx.ID_ANY, u"Records shall include, but not limited to, the following information:\n\na) customer's particular\nb) component type\nc) component serial number", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText273.Wrap( -1 )
	self.staticText273.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText273, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer249, 1, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.VERTICAL )

	bSizer289 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer289.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText261 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText261.Wrap( -1 )
	bSizer289.Add( self.staticText261, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer289, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 25), 0, 0, 5 )

	bSizer290 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer290.Add( self.textCtrl57, 0, wx.ALL, 5 )

	self.staticText262 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText262.Wrap( -1 )
	bSizer290.Add( self.staticText262, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer290, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 95), 0, 0, 5 )

	bSizer291 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl62 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer291.Add( self.textCtrl62, 0, wx.ALL, 5 )

	self.staticText263 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText263.Wrap( -1 )
	bSizer291.Add( self.staticText263, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer291, 1, wx.EXPAND, 5 )


	bSizer185.Add( ( 0, 30), 0, 0, 5 )

	bSizer292 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl75 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer292.Add( self.textCtrl75, 0, wx.ALL, 5 )

	self.staticText264 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText264.Wrap( -1 )
	bSizer292.Add( self.staticText264, 0, wx.ALL, 5 )


	bSizer185.Add( bSizer292, 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 10), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )


	bSizer199.Add( ( 0, 30), 0, 0, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )

	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
	self.textCtrl62.Bind( wx.EVT_TEXT, self.Score3 )
	self.textCtrl75.Bind( wx.EVT_TEXT, self.Score4 )
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get34
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 10')
        input1 = mscore1['C8']
        get34 = self.textCtrl6.GetValue()

def Score2( self, event ):
        #event.Skip()
        global input2
        global get35
        score2 = LoadWorkbook(parent=self.textCtrl57)
        mscore2 = score2.wb.get_sheet_by_name('Sec 10')
        input2 = mscore2['C12']
        get35 = self.textCtrl57.GetValue()

def Score3( self, event ):
        #event.Skip()
        global input3
        global get36
        score3 = LoadWorkbook(parent=self.textCtrl62)
        mscore3 = score3.wb.get_sheet_by_name('Sec 10')
        input3 = mscore3['C16']
        get36 = self.textCtrl62.GetValue()

def Score4( self, event ):
        #event.Skip()
        global input4
        global get37
        score4 = LoadWorkbook(parent=self.textCtrl75)
        mscore4 = score4.wb.get_sheet_by_name('Sec 10')
        input4 = mscore4['C20']
        get37 = self.textCtrl75.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get34
        input2.value = get35
        input3.value = get36
        input4.value = get37

        word34 = get34
        word35 = get35
        word36 = get36
        word37 = get37



###########################################################################
## Class Section11
###########################################################################

class Section11 ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Supplier Mark", pos = wx.DefaultPosition, size = wx.Size( 782,172 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(782,172), wx.Size(782,172) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

	bSizer247 = wx.BoxSizer( wx.VERTICAL )


	bSizer247.Add( ( 0, 20), 0, 0, 5 )

	bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer198.Add( ( 10, 0), 0, 0, 5 )

	bSizer248 = wx.BoxSizer( wx.VERTICAL )

	self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Supplier Mark", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText212.Wrap( -1 )
	self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

	bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


	bSizer198.Add( ( 20, 0), 0, 0, 5 )

	bSizer249 = wx.BoxSizer( wx.VERTICAL )

	self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Each product be marked with its certification mark based on company standard. \nSupplier mark be affixed to the product or packaging, where appropriate.", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText221.Wrap( -1 )
	self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


	bSizer249.Add( ( 0, 0), 1, wx.EXPAND, 5 )


	bSizer198.Add( bSizer249, 1, 0, 5 )

	bSizer185 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
	bSizer185.Add( self.textCtrl6, 0, wx.ALL, 5 )

	self.staticText265 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText265.Wrap( -1 )
	bSizer185.Add( self.staticText265, 0, wx.ALL, 5 )


	bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


	bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


	bSizer247.Add( ( 0, 10), 0, 0, 5 )

	bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button55, 0, wx.ALL, 5 )

	self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer199.Add( self.button56, 0, wx.ALL, 5 )


	bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer247 )
	self.Layout()

	self.Centre( wx.BOTH )


	# Connect Events
	self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
	self.button55.Bind( wx.EVT_BUTTON, self.Reset )
	self.button56.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Score1( self, event ):
        #event.Skip()
        global input1
        global get38
        score1 = LoadWorkbook(parent=self.textCtrl6)
        mscore1 = score1.wb.get_sheet_by_name('Sec 11')
        input1 = mscore1['C8']
        get38 = self.textCtrl6.GetValue()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
        #event.Skip()
        input1.value = get38

        word38 = get38



###########################################################################
## Class ExecutiveSummary
###########################################################################

class ExecutiveSummary ( wx.Frame ):

def __init__( self, parent ):
	wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Executive Summary", pos = wx.DefaultPosition, size = wx.Size( 611,429 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

	self.SetSizeHintsSz( wx.Size(611,429), wx.Size(611,429) )
	self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

	bSizer304 = wx.BoxSizer( wx.VERTICAL )

	bSizer256 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer256.Add( ( 240, 0), 0, 0, 5 )

	self.staticText225 = wx.StaticText( self, wx.ID_ANY, u"State Date", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText225.Wrap( -1 )
	self.staticText225.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer256.Add( self.staticText225, 0, wx.ALL, 5 )

	self.datePicker4 = wx.adv.DatePickerCtrl( self, wx.ID_ANY, wx.DefaultDateTime, wx.DefaultPosition, wx.DefaultSize, wx.adv.DP_DEFAULT )
	self.datePicker4.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

	bSizer256.Add( self.datePicker4, 0, wx.ALL, 5 )


	bSizer304.Add( bSizer256, 1, wx.EXPAND, 5 )

	bSizer314 = wx.BoxSizer( wx.HORIZONTAL )

	self.staticText300 = wx.StaticText( self, wx.ID_ANY, u"Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText300.Wrap( -1 )
	self.staticText300.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )

	bSizer314.Add( self.staticText300, 0, wx.ALL, 5 )


	bSizer314.Add( ( 190, 0), 0, 0, 5 )

	self.staticText301 = wx.StaticText( self, wx.ID_ANY, u"Available", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText301.Wrap( -1 )
	self.staticText301.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )

	bSizer314.Add( self.staticText301, 0, wx.ALL, 5 )


	bSizer314.Add( ( 10, 0), 0, wx.EXPAND, 5 )

	self.staticText302 = wx.StaticText( self, wx.ID_ANY, u"In Process", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText302.Wrap( -1 )
	self.staticText302.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )

	bSizer314.Add( self.staticText302, 0, wx.ALL, 5 )


	bSizer314.Add( ( 5, 0), 0, wx.EXPAND, 5 )

	self.staticText303 = wx.StaticText( self, wx.ID_ANY, u"Not Available", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText303.Wrap( -1 )
	self.staticText303.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )

	bSizer314.Add( self.staticText303, 0, wx.ALL, 5 )


	bSizer314.Add( ( 40, 0), 0, wx.EXPAND, 5 )


	bSizer314.Add( ( 0, 0), 0, wx.EXPAND, 5 )


	bSizer304.Add( bSizer314, 0, wx.EXPAND, 5 )

	bSizer305 = wx.BoxSizer( wx.HORIZONTAL )

	bSizer306 = wx.BoxSizer( wx.VERTICAL )

	self.staticText294 = wx.StaticText( self, wx.ID_ANY, u"ISO 9001: 2015", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText294.Wrap( -1 )
	self.staticText294.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer306.Add( self.staticText294, 1, wx.ALL|wx.EXPAND, 5 )


	bSizer306.Add( ( 0, 20), 1, wx.EXPAND, 5 )

	self.staticText295 = wx.StaticText( self, wx.ID_ANY, u"Improvement in Remanufacturing Activities", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText295.Wrap( -1 )
	self.staticText295.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer306.Add( self.staticText295, 1, wx.ALL|wx.EXPAND, 5 )


	bSizer306.Add( ( 0, 20), 1, wx.EXPAND, 5 )

	self.staticText296 = wx.StaticText( self, wx.ID_ANY, u"VTA Compliance", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText296.Wrap( -1 )
	self.staticText296.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer306.Add( self.staticText296, 1, wx.ALL|wx.EXPAND, 5 )


	bSizer306.Add( ( 0, 23), 1, wx.EXPAND, 5 )

	self.staticText297 = wx.StaticText( self, wx.ID_ANY, u"4R2S Training", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText297.Wrap( -1 )
	self.staticText297.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer306.Add( self.staticText297, 1, wx.ALL|wx.EXPAND, 5 )


	bSizer306.Add( ( 0, 23), 1, wx.EXPAND, 5 )

	self.staticText298 = wx.StaticText( self, wx.ID_ANY, u"Score", wx.DefaultPosition, wx.DefaultSize, 0 )
	self.staticText298.Wrap( -1 )
	self.staticText298.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

	bSizer306.Add( self.staticText298, 1, wx.ALL|wx.EXPAND, 5 )


	bSizer305.Add( bSizer306, 1, wx.EXPAND, 5 )

	bSizer307 = wx.BoxSizer( wx.VERTICAL )

	bSizer309 = wx.BoxSizer( wx.HORIZONTAL )

	self.checkBox171 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer309.Add( self.checkBox171, 0, wx.ALL, 5 )


	bSizer309.Add( ( 60, 0), 0, 0, 5 )

	self.checkBox172 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer309.Add( self.checkBox172, 0, wx.ALL, 5 )


	bSizer309.Add( ( 60, 0), 0, 0, 5 )

	self.checkBox173 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer309.Add( self.checkBox173, 0, wx.ALL, 5 )


	bSizer307.Add( bSizer309, 0, wx.EXPAND, 5 )


	bSizer307.Add( ( 0, 33), 0, 0, 5 )

	bSizer310 = wx.BoxSizer( wx.HORIZONTAL )

	self.checkBox174 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer310.Add( self.checkBox174, 0, wx.ALL, 5 )


	bSizer310.Add( ( 60, 0), 0, 0, 5 )

	self.checkBox175 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer310.Add( self.checkBox175, 0, wx.ALL, 5 )


	bSizer310.Add( ( 60, 0), 0, 0, 5 )

	self.checkBox176 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer310.Add( self.checkBox176, 0, wx.ALL, 5 )


	bSizer307.Add( bSizer310, 0, wx.EXPAND, 5 )


	bSizer307.Add( ( 0, 30), 0, 0, 5 )

	bSizer311 = wx.BoxSizer( wx.HORIZONTAL )

	self.checkBox177 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer311.Add( self.checkBox177, 0, wx.ALL, 5 )


	bSizer311.Add( ( 60, 0), 0, 0, 5 )

	self.checkBox178 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer311.Add( self.checkBox178, 0, wx.ALL, 5 )


	bSizer311.Add( ( 60, 0), 0, 0, 5 )

	self.checkBox179 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer311.Add( self.checkBox179, 0, wx.ALL, 5 )


	bSizer307.Add( bSizer311, 0, wx.EXPAND, 5 )


	bSizer307.Add( ( 0, 30), 0, 0, 5 )

	bSizer312 = wx.BoxSizer( wx.HORIZONTAL )

	self.checkBox180 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer312.Add( self.checkBox180, 0, wx.ALL, 5 )


	bSizer312.Add( ( 60, 0), 0, 0, 5 )

	self.checkBox181 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer312.Add( self.checkBox181, 0, wx.ALL, 5 )


	bSizer312.Add( ( 60, 0), 0, 0, 5 )

	self.checkBox182 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer312.Add( self.checkBox182, 0, wx.ALL, 5 )


	bSizer307.Add( bSizer312, 0, wx.EXPAND, 5 )


	bSizer307.Add( ( 0, 25), 0, 0, 5 )

	bSizer313 = wx.BoxSizer( wx.HORIZONTAL )

	self.textCtrl91 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
	bSizer313.Add( self.textCtrl91, 0, wx.ALL, 5 )


	bSizer307.Add( bSizer313, 0, 0, 5 )


	bSizer305.Add( bSizer307, 1, wx.EXPAND, 5 )


	bSizer304.Add( bSizer305, 0, wx.EXPAND, 5 )

	bSizer315 = wx.BoxSizer( wx.HORIZONTAL )


	bSizer315.Add( ( 0, 0), 1, wx.EXPAND, 5 )

	self.button89 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button89.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer315.Add( self.button89, 0, wx.ALL, 5 )

	self.button90 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
	self.button90.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

	bSizer315.Add( self.button90, 0, wx.ALL, 5 )


	bSizer304.Add( bSizer315, 1, wx.EXPAND, 5 )


	self.SetSizer( bSizer304 )
	self.Layout()

	self.Centre( wx.BOTH )


	# Connect Events
	self.datePicker4.Bind( wx.adv.EVT_DATE_CHANGED, self.Date4 )
	self.checkBox171.Bind( wx.EVT_CHECKBOX, self.check1 )
	self.checkBox172.Bind( wx.EVT_CHECKBOX, self.check2 )
	self.checkBox173.Bind( wx.EVT_CHECKBOX, self.check3 )
	self.checkBox174.Bind( wx.EVT_CHECKBOX, self.check4 )
	self.checkBox175.Bind( wx.EVT_CHECKBOX, self.check5 )
	self.checkBox176.Bind( wx.EVT_CHECKBOX, self.check6 )
	self.checkBox177.Bind( wx.EVT_CHECKBOX, self.check7 )
	self.checkBox178.Bind( wx.EVT_CHECKBOX, self.check8 )
	self.checkBox179.Bind( wx.EVT_CHECKBOX, self.check9 )
	self.checkBox180.Bind( wx.EVT_CHECKBOX, self.check10 )
	self.checkBox181.Bind( wx.EVT_CHECKBOX, self.check11 )
	self.checkBox182.Bind( wx.EVT_CHECKBOX, self.check12 )
	self.textCtrl91.Bind( wx.EVT_TEXT, self.score1 )
	self.button89.Bind( wx.EVT_BUTTON, self.Reset )
	self.button90.Bind( wx.EVT_BUTTON, self.Save )

def __del__( self ):
	pass


# Virtual event handlers, overide them in your derived class
def Date4( self, event ):
	event.Skip()

def check1( self, event ):
	event.Skip()

def check2( self, event ):
	event.Skip()

def check3( self, event ):
	event.Skip()

def check4( self, event ):
	event.Skip()

def check5( self, event ):
	event.Skip()

def check6( self, event ):
	event.Skip()

def check7( self, event ):
	event.Skip()

def check8( self, event ):
	event.Skip()

def check9( self, event ):
	event.Skip()

def check10( self, event ):
	event.Skip()

def check11( self, event ):
	event.Skip()

def check12( self, event ):
	event.Skip()

def score1( self, event ):
	event.Skip()

def Reset( self, event ):
	event.Skip()

def Save( self, event ):
	event.Skip()




class MyApp(wx.App):
def OnInit(self):
    myframe = MainSystem(None)
    myframe.Show(True)
    return True


if __name__ == '__main__':
app = MyApp()
app.MainLoop()