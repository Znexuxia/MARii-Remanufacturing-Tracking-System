###########################################################################
## Python code generated with wxFormBuilder (version Jun 17 2015)
## http://www.wxformbuilder.org/
##
## PLEASE DO "NOT" EDIT THIS FILE!
###########################################################################

import wx
import wx.xrc
import wx.adv
import openpyxl
from win32com.client import DispatchEx
import pymongo

###########################################################################
## Load The Workbook(Experimental and Will be updated in GUI)
###########################################################################

class LoadWorkbook(object):
    def __init__ (self, parent):
        self.wb = openpyxl.load_workbook('Audit Report Score (fix).xlsx')
        self.wb.get_sheet_names()

class LoadWorkbook1(object):
    def __init__(self, parent):
        self.wb1 = openpyxl.load_workbook('Checklist Old.xlsx')
        self.wb1.get_sheet_names()

class LoadWorkbook2(object):
    def __init__ (self, parent):
        self.wb2 = openpyxl.load_workbook('Checklist New.xlsx')
        self.wb2.get_sheet_names()

class LoadWorkbook3(object):
    def __init__(self,parent):
        self.wb3 = openpyxl.load_workbook('Executive Summary.xlsx')
        self.wb3.get_sheet_names()
        
###########################################################################
## Class MainSystem
###########################################################################

class MainSystem ( wx.Frame ):
    # For Checklist Old
    global a1
    global b1
    global c1
    global d1
    
    global e1
    global f1
    global g1
    global h1
    global i1
    global j1
    global k1
    global l1
    global m1
    global n1
    global o1
    global p1
    global q1
    global r1
    global s1
    global t1
    global u1
    global v1
    global w1
    global x1
    global y1
    global z1
    global a2
    global b2
    
    global c2
    global d2
    global e2
    global f2
    global g2
    global h2
    global i2
    global j2
    global k2
    global l2
    global m2
    global n2
    global o2
    global p2
    global q2
    global r2
    global s2
    global t2
    global u2
    global v2
    global w2
    global x2
    global y2
    global z2
    global a3
    global b3
    global c3
    global d3
    global e3
    global f3
    global g3
    global h3
    global i3
    global j3
    global k3
    global l3
    global m3
    global n3
    global o3
    global p3
    global q3
    global r3
    global s3
    global t3
    global u3
    global v3
    global w3
    global x3
    global y3
    global z3
    global a4
    global b4
    global c4
    global d4
    global e4
    global f4
    global g4
    global h4
    global i4
    global j4
    global k4
    global l4
    global m4
    
    global n4
    global o4
    global p4
    global q4
    global r4
    global s4
    global t4
    
    
    global a1x
    global a2x
    global a3x
    global a4x
    
    global a5x
    global a6x
    global a7x
    global a8x
    global a9x
    global a10x
    global a11x
    global a12x
    global a13x
    global a14x
    global a15x
    global a16x
    global a17x
    global a18x
    global a19x
    global a20x
    global a21x
    global a22x
    global a23x
    global a24x
    global a25x
    global a26x
    global a27x
    global a28x
    
    global a29x
    global a30x
    global a31x
    global a32x
    global a33x
    global a34x
    global a35x
    global a36x
    global a37x
    global a38x
    global a39x
    global a40x
    global a41x
    global a42x
    global a43x
    global a44x
    global a45x
    global a46x
    global a47x
    global a48x
    global a49x
    global a50x
    global a51x
    global a52x
    global a53x
    global a54x
    global a55x
    global a56x
    global a57x
    global a58x
    global a59x
    global a60x
    global a61x
    global a62x
    global a63x
    global a64x
    global a65x
    global a66x
    global a67x
    global a68x
    global a69x
    global a70x
    global a71x
    global a72x
    global a73x
    global a74x
    global a75x
    global a76x
    global a77x
    global a78x
    global a79x
    global a80x
    global a81x
    global a82x
    global a83x
    global a84x
    global a85x
    global a86x
    global a87x
    global a88x
    global a89x
    global a90x
    global a91x
    
    
    global a92x
    global a93x
    global a94x
    global a95x
    global a96x
    global a97x
    global a98x
    
    # For Checklist New
    global aa
    global ba
    global ca
    global da
    global ea
    
    global ab
    global bb
    global cb
    global db
    global eb
    global fb
    global gb
    global hb
    global ib
    global jb
    
    global ac
    global bc
    global cc
    global dc
    global ec
    global fc
    global gc
    
    global ad
    global bd
    global cd
    
    global ae
    
    
    global a12
    global b12
    global c12
    global d12
    global e12
    global f12
    global g12
    global h12
    global i12
    global j12
    global k12
    global l12
    global m12
    global n12
    global o12
    
    global p12
    global q12
    global r12
    global s12
    global t12
    global u12
    global v12
    global w12
    global x12
    global y12
    global z12
    global a13
    global b13
    global c13
    global d13
    global e13
    global f13
    global g13
    global h13
    global i13
    global j13
    global k13
    global l13
    global m13
    global n13
    global o13
    global p13
    global q13
    global r13
    global s13
    
    global t13
    global u13
    global v13
    global w13
    global x13
    global y13
    global z13
    global a14
    global b14
    global c14
    global d14
    global e14
    global f14
    global g14
    global h14
    global i14
    global j14
    global k14
    global l14
    global m14
    global n14
    
    global o14
    global p14
    global q14
    global r14
    global s14
    global t14
    global u14
    global v14
    global w14
    
    global x14
    global y14
    global z14
    
    global mop1
    global mop2
    global mop3
    global mop4
    global mop5
    
    global mop6
    global mop7
    global mop8
    global mop9
    global mop10
    global mop11
    global mop12
    global mop13
    global mop14
    global mop15
    
    global mop16
    global mop17
    global mop18
    global mop19
    global mop20
    global mop21
    global mop22
    
    global mop23
    global mop24
    global mop25
    
    global mop26
    
    
    global b1x
    global b2x
    global b3x
    global b4x
    global b5x
    global b6x
    global b7x
    global b8x
    global b9x
    global b10x
    global b11x
    global b12x
    global b13x
    global b14x
    global b15x
    
    global b16x
    global b17x
    global b18x
    global b19x
    global b20x
    global b21x
    global b22x
    global b23x
    global b24x
    global b25x
    global b26x
    global b27x

    
    #For Audit Score
    global word1
    global word2
    global word3
    global word4
    global word5
    global word6
    global word7
    global word8
    global word9
    global total1
    
    global word10
    global word11
    global word12
    global word13
    global word14
    global word15
    global word16
    global word17
    global word18
    global word19
    global word20
    global total2
    
    global word21
    global total3
    
    global word22
    global word23
    global total4
    
    global word24
    global word25
    global word26
    global total5
    
    global word27
    global total6
    
    global word28
    global word29
    global total7
    
    global word30
    global word31
    global word32
    global word33
    global total8
    
    global word34
    global word35
    global word36
    global word37
    global total9
    
    global word38
    global total10
    
    global a15
    global b15
    global c15
    global d15
    global e15
    global f15
    global g15
    global h15
    global i15
    
    global j15
    global k15
    global l15
    global m15
    global n15
    global o15
    global p15
    global q15
    global r15
    global s15
    global t15
    
    global u15

    global v15
    global w15

    global x15
    global y15
    global z15

    global a16
   
    global b16
    global c16
    
    global d16
    global e16
    global f16
    global g16

    global h16
    global i16
    global j16
    global k16

    global l16
    
    # For Executive Summary
    global word39
    
    global b79x
    global b80x
    global b81x
    global b82x
    global b83x
    global b84x
    global b85x
    global b86x
    global b87x
    global b88x
    global b89x
    global b90x
    global b91x
    
    
    

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Main System", pos = wx.DefaultPosition, size = wx.Size( 722,456 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(722,456),wx.Size(722,456))
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_BTNHIGHLIGHT ) )

        bSizer1 = wx.BoxSizer( wx.VERTICAL )

        bSizer258 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer258.Add( ( 5, 0), 0, 0, 5 )

        self.bitmap1 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"MARii logo(Small).jpg", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.Size( 300,80 ), 0 )
        bSizer258.Add( self.bitmap1, 0, 0, 5 )

        self.staticText226 = wx.StaticText( self, wx.ID_ANY, u"REMANUFACTURING \nTRACKING SYSTEM", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText226.Wrap( -1 )
        self.staticText226.SetFont( wx.Font( 18, 74, 93, 92, False, "Arial" ) )

        bSizer258.Add( self.staticText226, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )


        bSizer1.Add( bSizer258, 1, wx.EXPAND, 5 )

        bSizer257 = wx.BoxSizer( wx.VERTICAL )

        bSizer2 = wx.BoxSizer( wx.HORIZONTAL )

        bSizer4 = wx.BoxSizer( wx.VERTICAL )


        bSizer4.Add( ( 0, 20), 0, 0, 5 )

        self.button1 = wx.Button( self, wx.ID_ANY, u"Details of\nCompany", wx.Point( -1,-1 ), wx.Size( 180,100 ), 0 )
        self.button1.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

        bSizer4.Add( self.button1, 0, wx.ALIGN_BOTTOM|wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


        bSizer2.Add( bSizer4, 1, wx.EXPAND, 5 )

        bSizer6 = wx.BoxSizer( wx.VERTICAL )


        bSizer6.Add( ( 0, 20), 0, 0, 5 )

        self.button2 = wx.Button( self, wx.ID_ANY, u"Checklist", wx.Point( 0,0 ), wx.Size( 180,100 ), 0 )
        self.button2.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

        bSizer6.Add( self.button2, 0, wx.ALIGN_CENTER_HORIZONTAL|wx.ALL, 5 )


        bSizer2.Add( bSizer6, 1, wx.EXPAND, 5 )


        bSizer257.Add( bSizer2, 1, wx.EXPAND, 5 )

        bSizer5 = wx.BoxSizer( wx.HORIZONTAL )

        bSizer7 = wx.BoxSizer( wx.VERTICAL )


        bSizer7.Add( ( 0, 20), 0, 0, 5 )

        self.button3 = wx.Button( self, wx.ID_ANY, u"Audit \nScore", wx.DefaultPosition, wx.Size( 180,100 ), 0 )
        self.button3.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

        bSizer7.Add( self.button3, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer5.Add( bSizer7, 1, wx.EXPAND, 5 )

        bSizer10 = wx.BoxSizer( wx.VERTICAL )


        bSizer10.Add( ( 0, 20), 0, 0, 5 )

        self.button4 = wx.Button( self, wx.ID_ANY, u"Executive\nSummary", wx.DefaultPosition, wx.Size( 180,100 ), 0 )
        self.button4.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

        bSizer10.Add( self.button4, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer5.Add( bSizer10, 1, wx.EXPAND, 5 )


        bSizer257.Add( bSizer5, 1, wx.EXPAND, 5 )


        bSizer257.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        bSizer91 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer91.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button23 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button23.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer91.Add( self.button23, 0, wx.ALL, 5 )

        self.button26 = wx.Button( self, wx.ID_ANY, u"Generate", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button26.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer91.Add( self.button26, 0, wx.ALL, 5 )

        self.button24 = wx.Button( self, wx.ID_ANY, u"Check Report", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button24.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer91.Add( self.button24, 0, wx.ALL, 5 )

        self.button25 = wx.Button( self, wx.ID_ANY, u"Upload", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button25.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer91.Add( self.button25, 0, wx.ALL, 5 )


        bSizer257.Add( bSizer91, 0, wx.EXPAND, 5 )


        bSizer1.Add( bSizer257, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer1 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button1.Bind( wx.EVT_BUTTON, self.DetailnReportWindows )
        self.button2.Bind( wx.EVT_BUTTON, self.ChecklistWindows )
        self.button3.Bind( wx.EVT_BUTTON, self.AuditWindows )
        self.button4.Bind( wx.EVT_BUTTON, self.ExecutiveWindows )
        self.button23.Bind( wx.EVT_BUTTON, self.Reset )
        self.button26.Bind( wx.EVT_BUTTON, self.Generate )
        self.button24.Bind( wx.EVT_BUTTON, self.Check )
        self.button25.Bind( wx.EVT_BUTTON, self.Upload )



    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class                   
    def DetailnReportWindows( self, event ):
            Detail = DetailsnReport(parent=self.button1)
            Detail.Show()
            
    def ChecklistWindows( self, event ):
            Checklist = ChecklistMain(parent=self.button2)
            Checklist.Show()

    def AuditWindows( self, event ):
            Audit = AuditScore(parent=self.button3)
            Audit.Show()

    def ExecutiveWindows( self, event ):
            Executive = ExecutiveSummary(parent=self.button4)
            Executive.Show()

    def Reset( self, event ):
        
            # For Checklist Old
            global a1
            global b1
            global c1
            global d1
            
            global e1
            global f1
            global g1
            global h1
            global i1
            global j1
            global k1
            global l1
            global m1
            global n1
            global o1
            global p1
            global q1
            global r1
            global s1
            global t1
            global u1
            global v1
            global w1
            global x1
            global y1
            global z1
            global a2
            global b2
            
            global c2
            global d2
            global e2
            global f2
            global g2
            global h2
            global i2
            global j2
            global k2
            global l2
            global m2
            global n2
            global o2
            global p2
            global q2
            global r2
            global s2
            global t2
            global u2
            global v2
            global w2
            global x2
            global y2
            global z2
            global a3
            global b3
            global c3
            global d3
            global e3
            global f3
            global g3
            global h3
            global i3
            global j3
            global k3
            global l3
            global m3
            global n3
            global o3
            global p3
            global q3
            global r3
            global s3
            global t3
            global u3
            global v3
            global w3
            global x3
            global y3
            global z3
            global a4
            global b4
            global c4
            global d4
            global e4
            global f4
            global g4
            global h4
            global i4
            global j4
            global k4
            global l4
            global m4
            
            global n4
            global o4
            global p4
            global q4
            global r4
            global s4
            global t4
            
            
            global a1x
            global a2x
            global a3x
            global a4x
            
            global a5x
            global a6x
            global a7x
            global a8x
            global a9x
            global a10x
            global a11x
            global a12x
            global a13x
            global a14x
            global a15x
            global a16x
            global a17x
            global a18x
            global a19x
            global a20x
            global a21x
            global a22x
            global a23x
            global a24x
            global a25x
            global a26x
            global a27x
            global a28x
            
            global a29x
            global a30x
            global a31x
            global a32x
            global a33x
            global a34x
            global a35x
            global a36x
            global a37x
            global a38x
            global a39x
            global a40x
            global a41x
            global a42x
            global a43x
            global a44x
            global a45x
            global a46x
            global a47x
            global a48x
            global a49x
            global a50x
            global a51x
            global a52x
            global a53x
            global a54x
            global a55x
            global a56x
            global a57x
            global a58x
            global a59x
            global a60x
            global a61x
            global a62x
            global a63x
            global a64x
            global a65x
            global a66x
            global a67x
            global a68x
            global a69x
            global a70x
            global a71x
            global a72x
            global a73x
            global a74x
            global a75x
            global a76x
            global a77x
            global a78x
            global a79x
            global a80x
            global a81x
            global a82x
            global a83x
            global a84x
            global a85x
            global a86x
            global a87x
            global a88x
            global a89x
            global a90x
            global a91x
            
            
            global a92x
            global a93x
            global a94x
            global a95x
            global a96x
            global a97x
            global a98x
            
            
            
            
            a1 = 0
            b1 = 0
            c1 = 0
            d1 = 0
            
            e1 = 0
            f1 = 0
            g1 = 0
            h1 = 0
            i1 = 0
            j1 = 0
            k1 = 0
            l1 = 0
            m1 = 0
            n1 = 0
            o1 = 0
            p1 = 0
            q1 = 0
            r1 = 0
            s1 = 0
            t1 = 0
            u1 = 0
            v1 = 0
            w1 = 0
            x1 = 0
            y1 = 0
            z1 = 0
            a2 = 0
            b2 = 0
            
            c2 = 0
            d2 = 0
            e2 = 0
            f2 = 0
            g2 = 0
            h2 = 0
            i2 = 0
            j2 = 0
            k2 = 0
            l2 = 0
            m2 = 0
            n2 = 0
            o2 = 0
            p2 = 0
            q2 = 0
            r2 = 0
            s2 = 0
            t2 = 0
            u2 = 0
            v2 = 0
            w2 = 0
            x2 = 0
            y2 = 0
            z2 = 0
            a3 = 0
            b3 = 0
            c3 = 0
            d3 = 0
            e3 = 0
            f3 = 0
            g3 = 0
            h3 = 0
            i3 = 0
            j3 = 0
            k3 = 0
            l3 = 0
            m3 = 0
            n3 = 0
            o3 = 0
            p3 = 0
            q3 = 0
            r3 = 0
            s3 = 0
            t3 = 0
            u3 = 0
            v3 = 0
            w3 = 0
            x3 = 0
            y3 = 0
            z3 = 0
            a4 = 0
            b4 = 0
            c4 = 0
            d4 = 0
            e4 = 0
            f4 = 0
            g4 = 0
            h4 = 0
            i4 = 0
            j4 = 0
            k4 = 0
            l4 = 0
            m4 = 0
            
            n4 = 0
            o4 = 0
            p4 = 0
            q4 = 0
            r4 = 0
            s4 = 0
            t4 = 0
            
            
            a1x = 0
            a2x = 0
            a3x = 0
            a4x = 0
            
            a5x = 0
            a6x = 0
            a7x = 0
            a8x = 0
            a9x = 0
            a10x = 0
            a11x = 0
            a12x = 0
            a13x = 0
            a14x = 0
            a15x = 0
            a16x = 0
            a17x = 0
            a18x = 0
            a19x = 0
            a20x = 0
            a21x = 0
            a22x = 0
            a23x = 0
            a24x = 0
            a25x = 0
            a26x = 0
            a27x = 0
            a28x = 0
            
            a29x = 0
            a30x = 0
            a31x = 0
            a32x = 0
            a33x = 0
            a34x = 0
            a35x = 0
            a36x = 0
            a37x = 0
            a38x = 0
            a39x = 0
            a40x = 0
            a41x = 0
            a42x = 0
            a43x = 0
            a44x = 0
            a45x = 0
            a46x = 0
            a47x = 0
            a48x = 0
            a49x = 0
            a50x = 0
            a51x = 0
            a52x = 0
            a53x = 0
            a54x = 0
            a55x = 0
            a56x = 0
            a57x = 0
            a58x = 0
            a59x = 0
            a60x = 0
            a61x = 0
            a62x = 0
            a63x = 0
            a64x = 0
            a65x = 0
            a66x = 0
            a67x = 0
            a68x = 0
            a69x = 0
            a70x = 0
            a71x = 0
            a72x = 0
            a73x = 0
            a74x = 0
            a75x = 0
            a76x = 0
            a77x = 0
            a78x = 0
            a79x = 0
            a80x = 0
            a81x = 0
            a82x = 0
            a83x = 0
            a84x = 0
            a85x = 0
            a86x = 0
            a87x = 0
            a88x = 0
            a89x = 0
            a90x = 0
            
            a92x = 0
            a93x = 0
            a94x = 0
            a95x = 0
            a96x = 0
            a97x = 0
            a98x = 0
            
            
            # For Checklist New
            global aa
            global ba
            global ca
            global da
            global ea
            
            global ab
            global bb
            global cb
            global db
            global eb
            global fb
            global gb
            global hb
            global ib
            global jb
            
            global ac
            global bc
            global cc
            global dc
            global ec
            global fc
            global gc
            
            global ad
            global bd
            global cd
            
            global ae
            
            
            global a12
            global b12
            global c12
            global d12
            global e12
            global f12
            global g12
            global h12
            global i12
            global j12
            global k12
            global l12
            global m12
            global n12
            global o12
            
            global p12
            global q12
            global r12
            global s12
            global t12
            global u12
            global v12
            global w12
            global x12
            global y12
            global z12
            global a13
            global b13
            global c13
            global d13
            global e13
            global f13
            global g13
            global h13
            global i13
            global j13
            global k13
            global l13
            global m13
            global n13
            global o13
            global p13
            global q13
            global r13
            global s13
            
            global t13
            global u13
            global v13
            global w13
            global x13
            global y13
            global z13
            global a14
            global b14
            global c14
            global d14
            global e14
            global f14
            global g14
            global h14
            global i14
            global j14
            global k14
            global l14
            global m14
            global n14
            
            global o14
            global p14
            global q14
            global r14
            global s14
            global t14
            global u14
            global v14
            global w14
            
            global x14
            global y14
            global z14
            
            global mop1
            global mop2
            global mop3
            global mop4
            global mop5
            
            global mop6
            global mop7
            global mop8
            global mop9
            global mop10
            global mop11
            global mop12
            global mop13
            global mop14
            global mop15
            
            global mop16
            global mop17
            global mop18
            global mop19
            global mop20
            global mop21
            global mop22
            
            global mop23
            global mop24
            global mop25
            
            global mop26
            
            global mup1
            global mup2
            global mup3
            global mup4
            global mup5
            global mup6
            global mup7
            global mup8
            global mup9
            
            global b1x
            global b2x
            global b3x
            global b4x
            global b5x
            global b6x
            global b7x
            global b8x
            global b9x
            global b10x
            global b11x
            global b12x
            global b13x
            global b14x
            global b15x
            
            global b16x
            global b17x
            global b18x
            global b19x
            global b20x
            global b21x
            global b22x
            global b23x
            global b24x
            global b25x
            global b26x
            global b27x



            aa = 0
            ba = 0
            ca = 0
            da = 0
            ea = 0
            
            ab = 0
            bb = 0
            cb = 0
            db = 0
            eb = 0
            fb = 0
            gb = 0
            hb = 0
            ib = 0
            jb = 0
            
            ac = 0
            bc = 0
            cc = 0
            dc = 0
            ec = 0
            fc = 0
            gc = 0
            
            ad = 0
            bd = 0
            cd = 0
            
            ae = 0
            
            
            a12 = 0
            b12 = 0
            c12 = 0
            d12 = 0
            e12 = 0
            f12 = 0
            g12 = 0
            h12 = 0
            i12 = 0
            j12 = 0
            k12 = 0
            l12 = 0
            m12 = 0
            n12 = 0
            o12 = 0
            
            p12 = 0
            q12 = 0
            r12 = 0
            s12 = 0
            t12 = 0
            u12 = 0
            v12 = 0
            w12 = 0
            x12 = 0
            y12 = 0
            z12 = 0
            a13 = 0
            b13 = 0
            c13 = 0
            d13 = 0
            e13 = 0
            f13 = 0
            g13 = 0
            h13 = 0
            i13 = 0
            j13 = 0
            k13 = 0
            l13 = 0
            m13 = 0
            n13 = 0
            o13 = 0
            p13 = 0
            q13 = 0
            r13 = 0
            s13 = 0
            
            t13 = 0
            u13 = 0
            v13 = 0
            w13 = 0
            x13 = 0
            y13 = 0
            z13 = 0
            a14 = 0
            b14 = 0
            c14 = 0
            d14 = 0
            e14 = 0
            f14 = 0
            g14 = 0
            h14 = 0
            i14 = 0
            j14 = 0
            k14 = 0
            l14 = 0
            m14 = 0
            n14 = 0
            
            o14 = 0
            p14 = 0
            q14 = 0
            r14 = 0
            s14 = 0
            t14 = 0
            u14 = 0
            v14 = 0
            w14 = 0
            
            x14 = 0
            y14 = 0
            z14 = 0
            
            
            b1x = ""
            b2x = ""
            b3x = ""
            b4x = ""
            b5x = ""
            b6x = ""
            b7x = ""
            b8x = ""
            b9x = ""
            b10x = ""
            b11x = ""
            b12x = ""
            b13x = ""
            b14x = ""
            b15x = ""
            
            b16x = ""
            b17x = ""
            b18x = ""
            b19x = ""
            b20x = ""
            b21x = ""
            b22x = ""
            b23x = ""
            b24x = ""
            b25x = ""
            b26x = ""
            b27x = ""

            
            mop1 = ""
            mop2 = ""
            mop3 = ""
            mop4 = ""
            mop5 = ""
            
            mop6 = ""
            mop7 = ""
            mop8 = ""
            mop9 = ""
            mop10 = ""
            mop11 = ""
            mop12 = ""
            mop13 = ""
            mop14 = ""
            mop15 = ""
        
            mop16 = ""
            mop17 = ""
            mop18 = ""
            mop19 = ""
            mop20 = ""
            mop21 = ""
            mop22 = ""
            
            mop23 = ""
            mop24 = ""
            mop25 = ""
            
            mop26 = ""
            
            mup1 = ""
            mup2 = ""
            mup3 = ""
            mup4 = ""
            mup5 = ""
            mup6 = ""
            mup7 = ""
            mup8 = ""
            mup9 = ""
            
            #For Audit Score
            global word1
            global word2
            global word3
            global word4
            global word5
            global word6
            global word7
            global word8
            global word9
            global total1
            
            global word10
            global word11
            global word12
            global word13
            global word14
            global word15
            global word16
            global word17
            global word18
            global word19
            global word20
            global total2
            
            global word21
            global total3
            
            global word22
            global word23
            global total4
            
            global word24
            global word25
            global word26
            global total5
            
            global word27
            global total6
            
            global word28
            global word29
            global total7
            
            global word30
            global word31
            global word32
            global word33
            global total8
            
            global word34
            global word35
            global word36
            global word37
            global total9
            
            global word38
            global total10
            
            global a15
            global b15
            global c15
            global d15
            global e15
            global f15
            global g15
            global h15
            global i15
            
            global j15
            global k15
            global l15
            global m15
            global n15
            global o15
            global p15
            global q15
            global r15
            global s15
            global t15
            
            global u15

            global v15
            global w15

            global x15
            global y15
            global z15

            global a16
   
            global b16
            global c16
            
            global d16
            global e16
            global f16
            global g16

            global h16
            global i16
            global j16
            global k16
    
            global l16

            
            a15 = 0
            b15 = 0
            c15 = 0
            d15 = 0
            e15 = 0
            f15 = 0
            g15 = 0
            h15 = 0
            i15 = 0
            
            j15 = 0
            k15 = 0
            l15 = 0
            m15 = 0
            n15 = 0
            o15 = 0
            p15 = 0
            q15 = 0
            r15 = 0
            s15 = 0
            t15 = 0
            
            u15 = 0
            
            v15 = 0
            w15 = 0
            
            x15 = 0
            y15 = 0
            z15 = 0
            
            a16 = 0
            
            b16 = 0
            c16 = 0
            
            d16 = 0
            e16 = 0
            f16 = 0
            g16 = 0
            
            h16 = 0
            i16 = 0
            j16 = 0
            k16 = 0
            
            l16 = 0
            
            
            word1 = 0
            word2 = 0
            word3 = 0
            word4 = 0
            word5 = 0
            word6 = 0
            word7 = 0
            word8 = 0
            word9 = 0
            total1 = 0
            
            word10 = 0
            word11 = 0
            word12 = 0
            word13 = 0
            word14 = 0
            word15 = 0
            word16 = 0
            word17 = 0
            word18 = 0
            word19 = 0
            word20 = 0
            total2 = 0
            
            word21 = 0
            total3 = 0
            
            word22 = 0
            word23 = 0
            total4 = 0
            
            word24 = 0
            word25 = 0
            word26 = 0
            total5 = 0
            
            word27 = 0
            total6 = 0
            
            word28 = 0
            word29 = 0
            total7 = 0
            
            word30 = 0
            word31 = 0
            word32 = 0
            word33 = 0
            total8 = 0
            
            word34 = 0
            word35 = 0
            word36 = 0
            word37 = 0
            total9 = 0
            
            word38 = 0
            total10 = 0
            
            # For Executive Summary
            global word39
            
            global b79x
            global b80x
            global b81x
            global b82x
            global b83x
            global b84x
            global b85x
            global b86x
            global b87x
            global b88x
            global b89x
            global b90x
            global b91x
            
            b79x = 0
            b80x = 0
            b81x = 0
            b82x = 0
            b83x = 0
            b84x = 0
            b85x = 0
            b86x = 0
            b87x = 0
            b88x = 0
            b89x = 0
            b90x = 0
            b91x = 0
            
            word39 = 0
            
            DataReset = Reset(parent = self.button23)
            SoundReset = soundReset(parent = self.button23)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
            

    def Generate( self, event ):
            GenerateFile = Generate(parent = self.button26)
            GenerateFile.Show()

    def Check( self, event ):
            CheckFile1 = FileName1
            CheckFile2 = FileName2
            
            excel1 = DispatchEx("Excel.Application")
            excel2 = DispatchEx("Excel.Application")
            excel1.Visible = 1
            excel2.Visible = 1
            excel1.Workbooks.Open(CheckFile1)
            excel2.Workbooks.Open(CheckFile2)
            
            
            

    def Upload( self, event ):
            Serverconnect = Server(parent = self.button25)
            Serverconnect.Show()


###########################################################################
## Class DetailsnReport
###########################################################################
class DetailsnReport ( wx.Frame ):
    
    def __init__( self, parent ):
        global a
        global b
        global c
        global d
        global e
        global f
        global g
        global h
        global i
        global j
        global k
        global l
        global m
        global n
        
        global map1
        global map2
        global map3
        global map4
        global map5
        global map6
        global map7
        global map8
        global map9
        global map10
        global map11
        global map12
        global map13
        global map14
        
        global mailelement
        
        a = 0
        b = 0
        c = 0
        d = 0
        e = 0
        f = 0
        g = 0
        h = 0
        i = 0
        j = 0
        k = 0
        l = 0
        m = 0
        n = 0
        
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Details of Company", pos = wx.DefaultPosition, size = wx.Size( 608,642 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
        
        self.SetSizeHintsSz( wx.Size(608,642), wx.Size(608,642) )
        self.SetForegroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_WINDOW ) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )
        
        bSizer8 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer8.SetMinSize( wx.Size( -1,500 ) ) 
        
        bSizer8.Add( ( 0, 50), 0, 0, 5 )
        
        bSizer9 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.text1 = wx.StaticText( self, wx.ID_ANY, u"Company Name ", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.text1.Wrap( -1 )
        self.text1.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer9.Add( self.text1, 0, wx.ALL, 5 )
        
        bSizer14 = wx.BoxSizer( wx.VERTICAL )
        
        self.input1 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 200,-1 ), 0 )
        bSizer14.Add( self.input1, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer9.Add( bSizer14, 1, 0, 5 )
        
        
        bSizer8.Add( bSizer9, 0, wx.EXPAND, 5 )
        
        
        bSizer8.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        bSizer11 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.text3 = wx.StaticText( self, wx.ID_ANY, u"Address \t\t\t\t ", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.text3.Wrap( -1 )
        self.text3.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer11.Add( self.text3, 0, wx.ALL, 5 )
        
        bSizer15 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer15.SetMinSize( wx.Size( 100,-1 ) ) 
        self.input2 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 200,-1 ), 0 )
        bSizer15.Add( self.input2, 0, wx.ALL|wx.EXPAND, 5 )
        
        self.textCtrl49 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer15.Add( self.textCtrl49, 0, wx.ALL|wx.EXPAND, 5 )
        
        self.textCtrl50 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer15.Add( self.textCtrl50, 0, wx.ALL|wx.EXPAND, 5 )
        
        self.textCtrl51 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer15.Add( self.textCtrl51, 0, wx.ALL|wx.EXPAND, 5 )
        
        self.textCtrl52 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer15.Add( self.textCtrl52, 0, wx.ALL|wx.EXPAND, 5 )
        
        self.textCtrl53 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer15.Add( self.textCtrl53, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer11.Add( bSizer15, 1, wx.EXPAND, 5 )
        
        
        bSizer8.Add( bSizer11, 0, wx.EXPAND, 5 )
        
        
        bSizer8.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        bSizer300 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText268 = wx.StaticText( self, wx.ID_ANY, u"Auditee Name", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText268.Wrap( -1 )
        self.staticText268.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer300.Add( self.staticText268, 0, wx.ALL, 5 )
        
        
        bSizer300.Add( ( 18, 0), 0, 0, 5 )
        
        self.textCtrl101 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 450,-1 ), 0 )
        bSizer300.Add( self.textCtrl101, 0, wx.ALL, 5 )
        
        
        bSizer8.Add( bSizer300, 1, wx.EXPAND, 5 )
        
        
        bSizer8.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        bSizer16 = wx.BoxSizer( wx.HORIZONTAL )
        
        
        bSizer16.Add( ( 100, 0), 0, 0, 5 )
        
        bSizer17 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer20 = wx.BoxSizer( wx.VERTICAL )
        
        self.text5 = wx.StaticText( self, wx.ID_ANY, u"Contact No", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.text5.Wrap( -1 )
        self.text5.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer20.Add( self.text5, 0, wx.ALL, 5 )
        
        self.input3 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 200,-1 ), 0 )
        bSizer20.Add( self.input3, 0, wx.ALL, 5 )
        
        
        bSizer17.Add( bSizer20, 1, wx.EXPAND, 5 )
        
        bSizer3001 = wx.BoxSizer( wx.VERTICAL )
        
        self.staticText269 = wx.StaticText( self, wx.ID_ANY, u"Email", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText269.Wrap( -1 )
        self.staticText269.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer3001.Add( self.staticText269, 0, wx.ALL, 5 )
        
        self.textCtrl46 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        self.textCtrl46.SetMinSize( wx.Size( 200,-1 ) )
        
        bSizer3001.Add( self.textCtrl46, 0, wx.ALL, 5 )
        
        self.textCtrl47 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        self.textCtrl47.SetMinSize( wx.Size( 200,-1 ) )
        
        bSizer3001.Add( self.textCtrl47, 0, wx.ALL, 5 )
        
        self.textCtrl48 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        self.textCtrl48.SetMinSize( wx.Size( 200,-1 ) )
        
        bSizer3001.Add( self.textCtrl48, 0, wx.ALL, 5 )
        
        
        bSizer17.Add( bSizer3001, 0, wx.EXPAND, 5 )
        
        
        bSizer16.Add( bSizer17, 1, wx.EXPAND, 5 )
        
        bSizer296 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer297 = wx.BoxSizer( wx.VERTICAL )
        
        self.staticText266 = wx.StaticText( self, wx.ID_ANY, u"Audit Standard", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText266.Wrap( -1 )
        self.staticText266.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer297.Add( self.staticText266, 0, wx.ALL, 5 )
        
        self.textCtrl45 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 250,-1 ), 0 )
        bSizer297.Add( self.textCtrl45, 0, wx.ALL, 5 )
        
        
        bSizer297.Add( ( 0, 25), 0, wx.EXPAND, 5 )
        
        self.staticText267 = wx.StaticText( self, wx.ID_ANY, u"Plant/Location", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText267.Wrap( -1 )
        self.staticText267.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer297.Add( self.staticText267, 0, wx.ALL, 5 )
        
        self.textCtrl100 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 250,-1 ), 0 )
        bSizer297.Add( self.textCtrl100, 0, wx.ALL, 5 )
        
        
        bSizer296.Add( bSizer297, 1, wx.EXPAND, 5 )
        
        bSizer299 = wx.BoxSizer( wx.VERTICAL )
        
        
        bSizer296.Add( bSizer299, 1, wx.EXPAND, 5 )
        
        
        bSizer16.Add( bSizer296, 1, wx.EXPAND, 5 )
        
        
        bSizer8.Add( bSizer16, 0, wx.EXPAND, 5 )
        
        bSizer13 = wx.BoxSizer( wx.VERTICAL )
        
        
        bSizer13.Add( ( 0, 10), 0, 0, 5 )
        
        
        bSizer8.Add( bSizer13, 0, 0, 5 )
        
        bSizer12 = wx.BoxSizer( wx.HORIZONTAL )
        
        
        bSizer12.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        self.button6 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button6.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer12.Add( self.button6, 0, wx.ALL, 5 )
        
        self.button5 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button5.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer12.Add( self.button5, 0, wx.ALL|wx.ALIGN_RIGHT, 5 )
        
        
        bSizer8.Add( bSizer12, 1, wx.EXPAND|wx.ALIGN_RIGHT, 5 )
        
        
        self.SetSizer( bSizer8 )
        self.Layout()
        
        self.Centre( wx.BOTH )
        
        # Connect Events
        self.input1.Bind( wx.EVT_TEXT, self.CompanyName )
        self.input2.Bind( wx.EVT_TEXT, self.Address )
        self.textCtrl49.Bind( wx.EVT_TEXT, self.Address1 )
        self.textCtrl50.Bind( wx.EVT_TEXT, self.Address2 )
        self.textCtrl51.Bind( wx.EVT_TEXT, self.Address3 )
        self.textCtrl52.Bind( wx.EVT_TEXT, self.Address4 )
        self.textCtrl53.Bind( wx.EVT_TEXT, self.Address5 )
        self.textCtrl101.Bind( wx.EVT_TEXT, self.Auditee )
        self.input3.Bind( wx.EVT_TEXT, self.Contact1 )
        self.textCtrl46.Bind( wx.EVT_TEXT, self.email1 )
        self.textCtrl47.Bind( wx.EVT_TEXT, self.email2 )
        self.textCtrl48.Bind( wx.EVT_TEXT, self.email3 )
        self.textCtrl45.Bind( wx.EVT_TEXT, self.Audit )
        self.textCtrl100.Bind( wx.EVT_TEXT, self.plant )
        self.button6.Bind( wx.EVT_BUTTON, self.Reset )
        self.button5.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
        pass
    
    
    # Virtual event handlers, overide them in your derived class
    def CompanyName( self, event ):
            global peach1
            global a
            peach1 = self.input1.GetValue()
            a = 1
    
    def Address( self, event ):
            global peach2
            global b
            peach2 = self.input2.GetValue()
            b = 1
    
    def Address1( self, event ):
            global peach3
            global c
            peach3 = self.textCtrl49.GetValue()
            c = 1
            
    
    def Address2( self, event ):
            global peach4
            global d
            peach4 = self.textCtrl50.GetValue()
            d = 1
    
    def Address3( self, event ):
            global peach5
            global e
            peach5 = self.textCtrl51.GetValue()
            e = 1
            
    
    def Address4( self, event ):
            global peach6
            global f
            peach6 = self.textCtrl52.GetValue()
            f = 1
            
    
    def Address5( self, event ):
            global peach7
            global g
            peach7 = self.textCtrl53.GetValue()
            g = 1
            
    
    def Auditee( self, event ):
            global peach8
            global h
            peach8 = self.textCtrl101.GetValue()
            h = 1
    
    def Contact1( self, event ):
            global peach9
            global i
            peach9 = self.input3.GetValue()
            i = 1
    
    def email1( self, event ):
            global peach10
            global j
            peach10 = self.textCtrl46.GetValue()
            j = 1
    
    def email2( self, event ):
            global peach11
            global k
            peach11 = self.textCtrl47.GetValue()
            k = 1
    
    def email3( self, event ):
            global peach12
            global l
            peach12 = self.textCtrl48.GetValue()
            l = 1
    
    def Audit( self, event ):
            global peach13
            global m
            peach13 = self.textCtrl45.GetValue()
            m = 1
    
    def plant( self, event ):
            global peach14
            global n
            peach14 = self.textCtrl100.GetValue()
            n = 1
    
    def Reset( self, event ):
            self.input1.SetValue("")
            self.input2.SetValue("")
            self.textCtrl49.SetValue("")
            self.textCtrl50.SetValue("")
            self.textCtrl51.SetValue("")
            self.textCtrl52.SetValue("")
            self.textCtrl53.SetValue("")
            self.textCtrl101.SetValue("")
            self.input3.SetValue("")
            self.textCtrl46.SetValue("")
            self.textCtrl47.SetValue("")
            self.textCtrl48.SetValue("")
            self.textCtrl45.SetValue("")
            self.textCtrl100.SetValue("")
            
            DataReset = Reset(parent = self.button6)
            SoundReset = soundReset(parent = self.button6)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
            
    
    def Save( self, event ):
            #event.Skip()
            global map1
            global map2
            global map3
            global map4
            global map5
            global map6
            global map7
            global map8
            global map9
            global map10
            global map11
            global map12
            global map13
            global map14
            
            global mailelement
            
            
            if a == 1:
                map1 = peach1
            else:
                map1 = ""
                
            if b == 1:  
                map2 = peach2
            else:
                map2 = ""
                
            if c == 1:
                map3 = peach3
            else:
                map3 = ""
                
            if d == 1:
                map4 = peach4
            else:
                map4 = ""
                
            if e == 1:
                map5 = peach5
            else:
                map5 = ""
                
            if f == 1:   
                 map6 = peach6
            else:
                 map6 = ""
            
            if g == 1:
                map7 = peach7
            else:
                map7 = ""
                
            if h == 1:    
                map8 = peach8
            else:
                map8 = ""
                
            if i == 1:
                map9 = peach9
            else:
                map9 = ""
                
            if j == 1:
                map10 = peach10
            else:
                map10 = ""
                
            if k == 1:    
                map11 = peach11
            else:
                map11 = ""
            
            if l == 1:
                map12 = peach12
            else:
                map12 = ""
                
            if m == 1:
                map13 = peach13
            else:
                map13 = ""
             
            if n == 1:    
                map14 = peach14
            else:
                map14 = ""
              
            # Element multiple add
            if j == 1:
                mailelement = map10
            
            elif (j == 1 and k == 1):
                element_array = list()
                number = 2
                
                for lot in range(int(number)):
                    multi1 = map[lot]
                    element_array.append(str(multi1))
                mailelement = element_array
                
            elif (j == 1 and k == 1 and l == 1):
                element_array = list()
                number = 3
                
                for lot in range(int(number)):
                    multi2 = map[lot]
                element_array.append(str(multi2))
                mailelement = element_array
                
            else:
                mailelement = ""
                
            SavingData = DataSaved(parent = self.button5)
            AcceptSound = soundAccept(parent = self.button5)
            AcceptSound.Accept.Play()
            SavingData.Show()


###########################################################################
## Class ChecklistMain
###########################################################################

class ChecklistMain ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Type Of Checklist", pos = wx.DefaultPosition, size = wx.Size( 550,219 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(550,219), wx.Size(550,219) )
        self.SetForegroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_WINDOW ) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

        bSizer24 = wx.BoxSizer( wx.VERTICAL )


        bSizer24.Add( ( 0, 30), 0, 0, 5 )

        self.text5 = wx.StaticText( self, wx.ID_ANY, u"Please Select The Format", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.text5.Wrap( -1 )
        self.text5.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

        bSizer24.Add( self.text5, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer24.Add( ( 0, 0), 0, 0, 5 )

        bSizer25 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer25.Add( ( 70, 0), 0, 0, 5 )

        self.button13 = wx.Button( self, wx.ID_ANY, u"Old", wx.DefaultPosition, wx.Size( 180,100 ), 0 )
        self.button13.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

        bSizer25.Add( self.button13, 0, wx.ALL, 5 )


        bSizer25.Add( ( 20, 0), 0, 0, 5 )

        self.button14 = wx.Button( self, wx.ID_ANY, u"New", wx.DefaultPosition, wx.Size( 180,100 ), 0 )
        self.button14.SetFont( wx.Font( 16, 74, 90, 90, False, "Arial" ) )

        bSizer25.Add( self.button14, 0, wx.ALL, 5 )


        bSizer24.Add( bSizer25, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer24 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button13.Bind( wx.EVT_BUTTON, self.Old )
        self.button14.Bind( wx.EVT_BUTTON, self.New )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Old( self, event ):
            #event.Skip()
            OldFormat = ChecklistOld(parent=self.button13)
            OldFormat.Show()

    def New( self, event ):
            #event.Skip()
            NewFormat = ChecklistNew(parent=self.button14)
            NewFormat.Show()

###########################################################################
## Class ChecklistOld
###########################################################################
class ChecklistOld ( wx.Frame ):
    global a1
    global b1
    global c1
    global d1
    
    global e1
    global f1
    global g1
    global h1
    global i1
    global j1
    global k1
    global l1
    global m1
    global n1
    global o1
    global p1
    global q1
    global r1
    global s1
    global t1
    global u1
    global v1
    global w1
    global x1
    global y1
    global z1
    global a2
    global b2
    
    global c2
    global d2
    global e2
    global f2
    global g2
    global h2
    global i2
    global j2
    global k2
    global l2
    global m2
    global n2
    global o2
    global p2
    global q2
    global r2
    global s2
    global t2
    global u2
    global v2
    global w2
    global x2
    global y2
    global z2
    global a3
    global b3
    global c3
    global d3
    global e3
    global f3
    global g3
    global h3
    global i3
    global j3
    global k3
    global l3
    global m3
    global n3
    global o3
    global p3
    global q3
    global r3
    global s3
    global t3
    global u3
    global v3
    global w3
    global x3
    global y3
    global z3
    global a4
    global b4
    global c4
    global d4
    global e4
    global f4
    global g4
    global h4
    global i4
    global j4
    global k4
    global l4
    global m4
    
    global n4
    global o4
    global p4
    global q4
    global r4
    global s4
    global t4
    
    
    global a1x
    global a2x
    global a3x
    global a4x
    
    global a5x
    global a6x
    global a7x
    global a8x
    global a9x
    global a10x
    global a11x
    global a12x
    global a13x
    global a14x
    global a15x
    global a16x
    global a17x
    global a18x
    global a19x
    global a20x
    global a21x
    global a22x
    global a23x
    global a24x
    global a25x
    global a26x
    global a27x
    global a28x
    
    global a29x
    global a30x
    global a31x
    global a32x
    global a33x
    global a34x
    global a35x
    global a36x
    global a37x
    global a38x
    global a39x
    global a40x
    global a41x
    global a42x
    global a43x
    global a44x
    global a45x
    global a46x
    global a47x
    global a48x
    global a49x
    global a50x
    global a51x
    global a52x
    global a53x
    global a54x
    global a55x
    global a56x
    global a57x
    global a58x
    global a59x
    global a60x
    global a61x
    global a62x
    global a63x
    global a64x
    global a65x
    global a66x
    global a67x
    global a68x
    global a69x
    global a70x
    global a71x
    global a72x
    global a73x
    global a74x
    global a75x
    global a76x
    global a77x
    global a78x
    global a79x
    global a80x
    global a81x
    global a82x
    global a83x
    global a84x
    global a85x
    global a86x
    global a87x
    global a88x
    global a89x
    global a90x
    global a91x
    
    
    global a92x
    global a93x
    global a94x
    global a95x
    global a96x
    global a97x
    global a98x
    
    global moo1

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Checklist Old", pos = wx.DefaultPosition, size = wx.Size( 371,480 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(371,480), wx.Size(371,480) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer20 = wx.BoxSizer( wx.VERTICAL )


        bSizer20.Add( ( 0, 20), 0, 0, 5 )

        bSizer253 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer253.Add( ( 30, 0), 0, 0, 5 )

        self.staticText222 = wx.StaticText( self, wx.ID_ANY, u"State Date", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText222.Wrap( -1 )

        self.staticText222.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer253.Add( self.staticText222, 0, wx.ALL, 5 )

        self.datePicker1 = wx.adv.DatePickerCtrl( self, wx.ID_ANY, wx.DefaultDateTime, wx.DefaultPosition, wx.DefaultSize, wx.adv.DP_DEFAULT )
        self.datePicker1.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer253.Add( self.datePicker1, 0, wx.ALL, 5 )

        self.button76 = wx.Button( self, wx.ID_ANY, u"Set", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.button76.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer253.Add( self.button76, 0, wx.ALL, 5 )


        bSizer20.Add( bSizer253, 1, wx.EXPAND, 5 )

        bSizer21 = wx.BoxSizer( wx.VERTICAL )

        self.button9 = wx.Button( self, wx.ID_ANY, u"Licences and Permit", wx.DefaultPosition, wx.Size( 200,60 ), 0 )
        self.button9.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer21.Add( self.button9, 1, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


        bSizer21.Add( ( 0, 5), 0, 0, 5 )

        self.button10 = wx.Button( self, wx.ID_ANY, u"Best industries Practices", wx.DefaultPosition, wx.Size( 200,60 ), 0 )
        self.button10.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer21.Add( self.button10, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


        bSizer21.Add( ( 0, 5), 0, 0, 5 )

        self.button11 = wx.Button( self, wx.ID_ANY, u"Remanufacturing Work Flow\nSystem and S.O.P", wx.DefaultPosition, wx.Size( 200,60 ), 0 )
        self.button11.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer21.Add( self.button11, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


        bSizer21.Add( ( 0, 5), 0, 0, 5 )

        self.button12 = wx.Button( self, wx.ID_ANY, u"Standard Compliance To\nVehicle Type Approval (VTA)\n-Safety Parts-", wx.DefaultPosition, wx.Size( 200,60 ), 0 )
        self.button12.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer21.Add( self.button12, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


        bSizer20.Add( bSizer21, 1, wx.EXPAND, 5 )

        bSizer90 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer90.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button21 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button21.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer90.Add( self.button21, 0, wx.ALL, 5 )


        bSizer20.Add( bSizer90, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer20 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.datePicker1.Bind( wx.adv.EVT_DATE_CHANGED, self.Date1 )
        self.button76.Bind( wx.EVT_BUTTON, self.Set1 )
        self.button9.Bind( wx.EVT_BUTTON, self.License )
        self.button10.Bind( wx.EVT_BUTTON, self.Best )
        self.button11.Bind( wx.EVT_BUTTON, self.Remanufacturing )
        self.button12.Bind( wx.EVT_BUTTON, self.Standard )
        self.button21.Bind( wx.EVT_BUTTON, self.Reset )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Date1( self, event ):
            #event.Skip()
            global mine1
            mine1 =  event.GetDate()
            

    def Set1( self, event ):
            global moo1
            moo1 = mine1.Format(("%d/%m/%Y"))

    def License( self, event ):
            global LicenseForm
            LicenseForm = PartA(parent = self.button9)
            LicenseForm.Show()

    def Best( self, event ):
            global BestForm
            BestForm = PartB(parent = self.button10)
            BestForm.Show()

    def Remanufacturing( self, event ):
            global RemanufacturingForm
            RemanufacturingForm = PartC(parent = self.button11)
            RemanufacturingForm.Show()
            

    def Standard( self, event ):
            global StandardForm
            StandardForm = PartD(parent = self.button12)
            StandardForm.Show()
            
    def Reset( self, event ):
            global a1
            global b1
            global c1
            global d1
            
            global e1
            global f1
            global g1
            global h1
            global i1
            global j1
            global k1
            global l1
            global m1
            global n1
            global o1
            global p1
            global q1
            global r1
            global s1
            global t1
            global u1
            global v1
            global w1
            global x1
            global y1
            global z1
            global a2
            global b2
            
            global c2
            global d2
            global e2
            global f2
            global g2
            global h2
            global i2
            global j2
            global k2
            global l2
            global m2
            global n2
            global o2
            global p2
            global q2
            global r2
            global s2
            global t2
            global u2
            global v2
            global w2
            global x2
            global y2
            global z2
            global a3
            global b3
            global c3
            global d3
            global e3
            global f3
            global g3
            global h3
            global i3
            global j3
            global k3
            global l3
            global m3
            global n3
            global o3
            global p3
            global q3
            global r3
            global s3
            global t3
            global u3
            global v3
            global w3
            global x3
            global y3
            global z3
            global a4
            global b4
            global c4
            global d4
            global e4
            global f4
            global g4
            global h4
            global i4
            global j4
            global k4
            global l4
            global m4
            
            global n4
            global o4
            global p4
            global q4
            global r4
            global s4
            global t4
            
            
            global a1x
            global a2x
            global a3x
            global a4x
            
            global a5x
            global a6x
            global a7x
            global a8x
            global a9x
            global a10x
            global a11x
            global a12x
            global a13x
            global a14x
            global a15x
            global a16x
            global a17x
            global a18x
            global a19x
            global a20x
            global a21x
            global a22x
            global a23x
            global a24x
            global a25x
            global a26x
            global a27x
            global a28x
            
            global a29x
            global a30x
            global a31x
            global a32x
            global a33x
            global a34x
            global a35x
            global a36x
            global a37x
            global a38x
            global a39x
            global a40x
            global a41x
            global a42x
            global a43x
            global a44x
            global a45x
            global a46x
            global a47x
            global a48x
            global a49x
            global a50x
            global a51x
            global a52x
            global a53x
            global a54x
            global a55x
            global a56x
            global a57x
            global a58x
            global a59x
            global a60x
            global a61x
            global a62x
            global a63x
            global a64x
            global a65x
            global a66x
            global a67x
            global a68x
            global a69x
            global a70x
            global a71x
            global a72x
            global a73x
            global a74x
            global a75x
            global a76x
            global a77x
            global a78x
            global a79x
            global a80x
            global a81x
            global a82x
            global a83x
            global a84x
            global a85x
            global a86x
            global a87x
            global a88x
            global a89x
            global a90x
            global a91x
            
            
            global a92x
            global a93x
            global a94x
            global a95x
            global a96x
            global a97x
            global a98x
            
            
            a1 = 0
            b1 = 0
            c1 = 0
            d1 = 0
            
            e1 = 0
            f1 = 0
            g1 = 0
            h1 = 0
            i1 = 0
            j1 = 0
            k1 = 0
            l1 = 0
            m1 = 0
            n1 = 0
            o1 = 0
            p1 = 0
            q1 = 0
            r1 = 0
            s1 = 0
            t1 = 0
            u1 = 0
            v1 = 0
            w1 = 0
            x1 = 0
            y1 = 0
            z1 = 0
            a2 = 0
            b2 = 0
            
            c2 = 0
            d2 = 0
            e2 = 0
            f2 = 0
            g2 = 0
            h2 = 0
            i2 = 0
            j2 = 0
            k2 = 0
            l2 = 0
            m2 = 0
            n2 = 0
            o2 = 0
            p2 = 0
            q2 = 0
            r2 = 0
            s2 = 0
            t2 = 0
            u2 = 0
            v2 = 0
            w2 = 0
            x2 = 0
            y2 = 0
            z2 = 0
            a3 = 0
            b3 = 0
            c3 = 0
            d3 = 0
            e3 = 0
            f3 = 0
            g3 = 0
            h3 = 0
            i3 = 0
            j3 = 0
            k3 = 0
            l3 = 0
            m3 = 0
            n3 = 0
            o3 = 0
            p3 = 0
            q3 = 0
            r3 = 0
            s3 = 0
            t3 = 0
            u3 = 0
            v3 = 0
            w3 = 0
            x3 = 0
            y3 = 0
            z3 = 0
            a4 = 0
            b4 = 0
            c4 = 0
            d4 = 0
            e4 = 0
            f4 = 0
            g4 = 0
            h4 = 0
            i4 = 0
            j4 = 0
            k4 = 0
            l4 = 0
            m4 = 0
            
            n4 = 0
            o4 = 0
            p4 = 0
            q4 = 0
            r4 = 0
            s4 = 0
            t4 = 0
            
            
            a1x = 0
            a2x = 0
            a3x = 0
            a4x = 0
            
            a5x = 0
            a6x = 0
            a7x = 0
            a8x = 0
            a9x = 0
            a10x = 0
            a11x = 0
            a12x = 0
            a13x = 0
            a14x = 0
            a15x = 0
            a16x = 0
            a17x = 0
            a18x = 0
            a19x = 0
            a20x = 0
            a21x = 0
            a22x = 0
            a23x = 0
            a24x = 0
            a25x = 0
            a26x = 0
            a27x = 0
            a28x = 0
            
            a29x = 0
            a30x = 0
            a31x = 0
            a32x = 0
            a33x = 0
            a34x = 0
            a35x = 0
            a36x = 0
            a37x = 0
            a38x = 0
            a39x = 0
            a40x = 0
            a41x = 0
            a42x = 0
            a43x = 0
            a44x = 0
            a45x = 0
            a46x = 0
            a47x = 0
            a48x = 0
            a49x = 0
            a50x = 0
            a51x = 0
            a52x = 0
            a53x = 0
            a54x = 0
            a55x = 0
            a56x = 0
            a57x = 0
            a58x = 0
            a59x = 0
            a60x = 0
            a61x = 0
            a62x = 0
            a63x = 0
            a64x = 0
            a65x = 0
            a66x = 0
            a67x = 0
            a68x = 0
            a69x = 0
            a70x = 0
            a71x = 0
            a72x = 0
            a73x = 0
            a74x = 0
            a75x = 0
            a76x = 0
            a77x = 0
            a78x = 0
            a79x = 0
            a80x = 0
            a81x = 0
            a82x = 0
            a83x = 0
            a84x = 0
            a85x = 0
            a86x = 0
            a87x = 0
            a88x = 0
            a89x = 0
            a90x = 0
            
            a92x = 0
            a93x = 0
            a94x = 0
            a95x = 0
            a96x = 0
            a97x = 0
            a98x = 0
            
            DataReset = Reset(parent = self.button21)
            SoundReset = soundReset(parent = self.button21)
            
            DataReset.Show()
            SoundReset.Resetto.Play()


###########################################################################
## Class PartA
###########################################################################

class PartA ( wx.Frame ):
    
    global a1x
    global a2x
    global a3x
    global a4x
    
    global a1
    global b1
    global c1
    global d1
    
    a1 = 0
    b1 = 0
    c1 = 0
    d1 = 0
    
    a1x = 0
    a2x = 0
    a3x = 0
    a4x = 0

    def __init__( self, parent ):
            wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Licences And Permit", pos = wx.DefaultPosition, size = wx.Size( 479,355 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

            self.SetSizeHintsSz( wx.Size(479,355), wx.Size(479,355) )
            self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )

            bSizer23 = wx.BoxSizer( wx.VERTICAL )

            self.staticText52 = wx.StaticText( self, wx.ID_ANY, u"AVAILABILITY OF LICENCES AND PERMIT", wx.DefaultPosition, wx.DefaultSize, 0 )
            self.staticText52.Wrap( -1 )
            self.staticText52.SetFont( wx.Font( 16, 74, 90, 92, False, "Arial" ) )

            bSizer23.Add( self.staticText52, 0, wx.ALL, 5 )


            bSizer23.Add( ( 0, 10), 0, 0, 5 )

            bSizer30 = wx.BoxSizer( wx.VERTICAL )

            bSizer31 = wx.BoxSizer( wx.HORIZONTAL )


            bSizer31.Add( ( 50, 0), 0, wx.ALIGN_CENTER_VERTICAL, 5 )

            bSizer24 = wx.BoxSizer( wx.VERTICAL )

            self.text5 = wx.StaticText( self, wx.ID_ANY, u"Number of AP allocated by MITI", wx.DefaultPosition, wx.DefaultSize, 0 )
            self.text5.Wrap( -1 )
            self.text5.SetFont( wx.Font( 14, 74, 90, 92, False, "Arial" ) )

            bSizer24.Add( self.text5, 0, wx.ALL, 5 )

            bSizer25 = wx.BoxSizer( wx.HORIZONTAL )


            bSizer25.Add( ( 20, 0), 0, 0, 5 )

            self.checkbox1 = wx.CheckBox( self, wx.ID_ANY, u"The number of AP used to import vehicle", wx.DefaultPosition, wx.DefaultSize, 0 )
            self.checkbox1.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

            bSizer25.Add( self.checkbox1, 0, wx.ALL, 5 )


            bSizer24.Add( bSizer25, 0, 0, 5 )

            self.Text6 = wx.StaticText( self, wx.ID_ANY, u"Manufacturing Licence (ML) from MITI", wx.DefaultPosition, wx.DefaultSize, 0 )
            self.Text6.Wrap( -1 )
            self.Text6.SetFont( wx.Font( 14, 74, 90, 92, False, "Arial" ) )

            bSizer24.Add( self.Text6, 0, wx.ALL, 5 )

            bSizer27 = wx.BoxSizer( wx.HORIZONTAL )


            bSizer27.Add( ( 20, 0), 0, 0, 5 )

            self.checkBox4 = wx.CheckBox( self, wx.ID_ANY, u"The business activities match the ML", wx.DefaultPosition, wx.DefaultSize, 0 )
            self.checkBox4.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

            bSizer27.Add( self.checkBox4, 0, wx.ALL, 5 )


            bSizer24.Add( bSizer27, 0, 0, 5 )

            self.m_staticText7 = wx.StaticText( self, wx.ID_ANY, u"Licenses from related local authorities (1,3,4,5,6)", wx.DefaultPosition, wx.DefaultSize, 0 )
            self.m_staticText7.Wrap( -1 )
            self.m_staticText7.SetFont( wx.Font( 14, 74, 90, 92, False, "Arial" ) )

            bSizer24.Add( self.m_staticText7, 0, wx.ALL, 5 )

            bSizer28 = wx.BoxSizer( wx.HORIZONTAL )


            bSizer28.Add( ( 20, 0), 0, 0, 5 )

            bSizer29 = wx.BoxSizer( wx.VERTICAL )

            self.checkBox6 = wx.CheckBox( self, wx.ID_ANY, u"Business License from PBT & SSM registration", wx.DefaultPosition, wx.DefaultSize, 0 )
            self.checkBox6.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

            bSizer29.Add( self.checkBox6, 0, wx.ALL, 5 )

            self.checkBox7 = wx.CheckBox( self, wx.ID_ANY, u"License/Certificate from Customs & GST", wx.DefaultPosition, wx.DefaultSize, 0 )
            self.checkBox7.SetFont( wx.Font( 12, 74, 90, 90, False, "Arial" ) )

            bSizer29.Add( self.checkBox7, 0, wx.ALL, 5 )


            bSizer28.Add( bSizer29, 1, wx.EXPAND, 5 )


            bSizer24.Add( bSizer28, 0, 0, 5 )


            bSizer31.Add( bSizer24, 0, 0, 5 )


            bSizer30.Add( bSizer31, 1, wx.EXPAND, 5 )


            bSizer23.Add( bSizer30, 1, wx.EXPAND, 5 )

            bSizer84 = wx.BoxSizer( wx.HORIZONTAL )


            bSizer84.Add( ( 0, 0), 1, 0, 5 )

            self.button13 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
            self.button13.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

            bSizer84.Add( self.button13, 0, wx.ALL, 5 )

            self.button14 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
            self.button14.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

            bSizer84.Add( self.button14, 0, wx.ALL, 5 )


            bSizer23.Add( bSizer84, 0, wx.EXPAND, 5 )


            self.SetSizer( bSizer23 )
            self.Layout()

            self.Centre( wx.BOTH )

            # Connect Events
            self.checkbox1.Bind( wx.EVT_CHECKBOX, self.Check1 )
            self.checkBox4.Bind( wx.EVT_CHECKBOX, self.Check2 )
            self.checkBox6.Bind( wx.EVT_CHECKBOX, self.Check3 )
            self.checkBox7.Bind( wx.EVT_CHECKBOX, self.Check4 )
            self.button13.Bind( wx.EVT_BUTTON, self.Reset )
            self.button14.Bind( wx.EVT_BUTTON, self.Save )
    def __del__( self ):
            pass



    ##    def Check1( self, event ):
    def Check1( self, event):
            global a1
            box1 = self.checkbox1.GetValue()
            
            if box1 == True:
                a1 = 1
            else:
                a1 = 0

    def Check2( self, event ):
            global b1
            box2 = self.checkBox4.GetValue()
            
            if box2 == True:
                b1 = 1
            else:
                b1 = 0

    def Check3( self, event ):
            global c1
            box3 = self.checkBox6.GetValue()
            
            if box3 == True:
                c1 = 1
            else:
                c1 = 0


    def Check4( self, event ):
            global d1
            box4 = self.checkBox7.GetValue()
            
            if box4 == True:
                d1 = 1
            else:
                d1 = 0

    def Reset( self, event ):
            self.checkbox1.SetValue(False)
            self.checkBox4.SetValue(False)
            self.checkBox6.SetValue(False)
            self.checkBox7.SetValue(False)
            
            DataReset = Reset(parent = self.button13)
            SoundReset = soundReset(parent = self.button13)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
            

    def Save( self, event):
            global a1x
            global a2x
            global a3x
            global a4x

            if a1 == 1:
                a1x = 1
            else:
                a1x = 0

            if b1 == 1:
                a2x = 1
            else:
                a2x = 0

            if c1 == 1:
                a3x = 1
            else:
                a3x = 0

            if d1 == True:
                a4x = 1
            else:
                a4x = 0
                
            SavingData = DataSaved(parent = self.button14)
            SavingData.Show()
            AcceptSound = soundAccept(parent = self.button14)
            AcceptSound.Accept.Play()
                
            


###########################################################################
## Class PartB
###########################################################################

class PartB ( wx.Frame ):
    global e1
    global f1
    global g1
    global h1
    global i1
    global j1
    global k1
    global l1
    global m1
    global n1
    global o1
    global p1
    global q1
    global r1
    global s1
    global t1
    global u1
    global v1
    global w1
    global x1
    global y1
    global z1
    global a2
    global b2
    
    global a5x
    global a6x
    global a7x
    global a8x
    global a9x
    global a10x
    global a11x
    global a12x
    global a13x
    global a14x
    global a15x
    global a16x
    global a17x
    global a18x
    global a19x
    global a20x
    global a21x
    global a22x
    global a23x
    global a24x
    global a25x
    global a26x
    global a27x
    global a28x
    
    e1 = 0
    f1 = 0
    g1 = 0
    h1 = 0
    i1 = 0
    j1 = 0
    k1 = 0
    l1 = 0
    m1 = 0
    n1 = 0
    o1 = 0
    p1 = 0
    q1 = 0
    r1 = 0
    s1 = 0
    t1 = 0
    u1 = 0
    v1 = 0
    w1 = 0
    x1 = 0
    y1 = 0
    z1 = 0
    a2 = 0
    b2 = 0
    
    a5x = 0
    a6x = 0
    a7x = 0
    a8x = 0
    a9x = 0
    a10x = 0
    a11x = 0
    a12x = 0
    a13x = 0
    a14x = 0
    a15x = 0
    a16x = 0
    a17x = 0
    a18x = 0
    a19x = 0
    a20x = 0
    a21x = 0
    a22x = 0
    a23x = 0
    a24x = 0
    a25x = 0
    a26x = 0
    a27x = 0
    a28x = 0
    
    
    def __init__( self, parent):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Best Industries Practices", pos = wx.DefaultPosition, size = wx.Size( 1180,772 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(1180,772), wx.Size(1180,772) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )

        bSizer32 = wx.BoxSizer( wx.VERTICAL )

        bSizer89 = wx.BoxSizer( wx.VERTICAL )

        self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"COMPLIANCES OF BEST INDUSTRY PRACTICES", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText51.Wrap( -1 )
        self.staticText51.SetFont( wx.Font( 18, 70, 90, 92, False, "Arial" ) )

        bSizer89.Add( self.staticText51, 0, wx.ALL, 5 )


        bSizer32.Add( bSizer89, 1, wx.EXPAND, 5 )

        bSizer85 = wx.BoxSizer( wx.HORIZONTAL )

        Left = wx.BoxSizer( wx.VERTICAL )

        Man = wx.BoxSizer( wx.VERTICAL )

        self.text8 = wx.StaticText( self, wx.ID_ANY, u"Man", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.text8.Wrap( -1 )
        self.text8.SetFont( wx.Font( 16, 70, 90, 92, False, "Arial" ) )

        Man.Add( self.text8, 0, wx.ALL, 5 )

        bSizer48 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer48.Add( ( 20, 0), 0, 0, 5 )

        bSizer47 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox9 = wx.CheckBox( self, wx.ID_ANY, u"Manpower Planning, *list of employee & Organisation Chart\n*to refer EPF forms (A KWSP) & SOCSO(8A -PERKESO) and document for foreign worker", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox9.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer47.Add( self.checkBox9, 0, wx.ALL, 5 )

        self.checkBox10 = wx.CheckBox( self, wx.ID_ANY, u"Manpower qualifications", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox10.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer47.Add( self.checkBox10, 0, wx.ALL, 5 )

        self.checkBox11 = wx.CheckBox( self, wx.ID_ANY, u"Manpower training and record", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox11.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer47.Add( self.checkBox11, 0, wx.ALL, 5 )

        self.checkBox12 = wx.CheckBox( self, wx.ID_ANY, u"Automotive Engineer (In-house)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox12.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer47.Add( self.checkBox12, 0, wx.ALL, 5 )

        self.checkBox13 = wx.CheckBox( self, wx.ID_ANY, u"Professional Engineer (PE) status", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox13.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer47.Add( self.checkBox13, 0, wx.ALL, 5 )


        bSizer48.Add( bSizer47, 1, wx.EXPAND, 5 )


        Man.Add( bSizer48, 1, wx.EXPAND, 5 )


        Left.Add( Man, 1, wx.EXPAND, 5 )

        Machine = wx.BoxSizer( wx.VERTICAL )

        self.staticText9 = wx.StaticText( self, wx.ID_ANY, u"Machine", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText9.Wrap( -1 )
        self.staticText9.SetFont( wx.Font( 16, 70, 90, 92, False, "Arial" ) )

        Machine.Add( self.staticText9, 0, wx.ALL, 5 )

        bSizer49 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer49.Add( ( 20, 0), 0, 0, 5 )

        self.checkBox14 = wx.CheckBox( self, wx.ID_ANY, u"Equipment and Testing equiptment based on JPJ & SIRIM Audit result", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox14.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer49.Add( self.checkBox14, 0, wx.ALL, 5 )


        Machine.Add( bSizer49, 1, wx.EXPAND, 5 )


        Left.Add( Machine, 1, wx.EXPAND, 5 )


        bSizer85.Add( Left, 1, wx.EXPAND, 5 )

        Right = wx.BoxSizer( wx.VERTICAL )

        Material = wx.BoxSizer( wx.VERTICAL )

        self.staticText10 = wx.StaticText( self, wx.ID_ANY, u"Material", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText10.Wrap( -1 )
        self.staticText10.SetFont( wx.Font( 16, 70, 90, 92, False, "Arial" ) )

        Material.Add( self.staticText10, 0, wx.ALL, 5 )

        bSizer51 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer51.Add( ( 20, 0), 0, 0, 5 )

        bSizer50 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox15 = wx.CheckBox( self, wx.ID_ANY, u"Incoming parts control", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox15.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer50.Add( self.checkBox15, 0, wx.ALL, 5 )

        self.checkBox16 = wx.CheckBox( self, wx.ID_ANY, u"Reject parts S.O.P", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox16.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer50.Add( self.checkBox16, 0, wx.ALL, 5 )

        self.checkBox17 = wx.CheckBox( self, wx.ID_ANY, u"Storage Management", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox17.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer50.Add( self.checkBox17, 0, wx.ALL, 5 )

        self.checkBox18 = wx.CheckBox( self, wx.ID_ANY, u"Regulatory & Statutory requirements", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox18.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer50.Add( self.checkBox18, 0, wx.ALL, 5 )


        bSizer51.Add( bSizer50, 1, wx.EXPAND, 5 )


        Material.Add( bSizer51, 1, wx.EXPAND, 5 )


        Right.Add( Material, 1, wx.EXPAND, 5 )

        Method = wx.BoxSizer( wx.VERTICAL )

        self.m_staticText11 = wx.StaticText( self, wx.ID_ANY, u"Method", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText11.Wrap( -1 )
        self.m_staticText11.SetFont( wx.Font( 16, 70, 90, 92, False, "Arial" ) )

        Method.Add( self.m_staticText11, 0, wx.ALL, 5 )

        bSizer52 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer52.Add( ( 20, 0), 0, 0, 5 )

        bSizer39 = wx.BoxSizer( wx.VERTICAL )

        bSizer40 = wx.BoxSizer( wx.VERTICAL )

        self.staticText12 = wx.StaticText( self, wx.ID_ANY, u"Process Flow & Control Plan", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText12.Wrap( -1 )
        self.staticText12.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer40.Add( self.staticText12, 0, wx.ALL, 5 )

        bSizer54 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer54.Add( ( 20, 0), 0, 0, 5 )

        bSizer53 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox31 = wx.CheckBox( self, wx.ID_ANY, u"SOP/Work Instruction/Checksheet /Record", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox31.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer53.Add( self.checkBox31, 0, wx.ALL, 5 )


        bSizer54.Add( bSizer53, 1, wx.EXPAND, 5 )


        bSizer40.Add( bSizer54, 1, wx.EXPAND, 5 )


        bSizer39.Add( bSizer40, 0, 0, 5 )

        bSizer41 = wx.BoxSizer( wx.VERTICAL )

        self.staticText13 = wx.StaticText( self, wx.ID_ANY, u"Safety, Health & Environment", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText13.Wrap( -1 )
        self.staticText13.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer41.Add( self.staticText13, 0, wx.ALL, 5 )

        bSizer55 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer55.Add( ( 20, 0), 0, 0, 5 )

        bSizer56 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox32 = wx.CheckBox( self, wx.ID_ANY, u"Manpower", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox32.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer56.Add( self.checkBox32, 0, wx.ALL, 5 )

        self.checkBox33 = wx.CheckBox( self, wx.ID_ANY, u"Workplace", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox33.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer56.Add( self.checkBox33, 0, wx.ALL, 5 )

        self.checkBox34 = wx.CheckBox( self, wx.ID_ANY, u"Environment", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox34.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer56.Add( self.checkBox34, 0, wx.ALL, 5 )


        bSizer55.Add( bSizer56, 1, wx.EXPAND, 5 )


        bSizer41.Add( bSizer55, 1, wx.EXPAND, 5 )


        bSizer39.Add( bSizer41, 0, 0, 5 )

        bSizer46 = wx.BoxSizer( wx.VERTICAL )

        self.staticText20 = wx.StaticText( self, wx.ID_ANY, u"Quality Management System - ISO 9001:2015", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText20.Wrap( -1 )
        self.staticText20.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer46.Add( self.staticText20, 0, wx.ALL, 5 )

        bSizer561 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer561.Add( ( 20, 0), 0, 0, 5 )

        bSizer58 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox36 = wx.CheckBox( self, wx.ID_ANY, u"Registration", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox36.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer58.Add( self.checkBox36, 0, wx.ALL, 5 )

        self.checkBox37 = wx.CheckBox( self, wx.ID_ANY, u"Training", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox37.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer58.Add( self.checkBox37, 0, wx.ALL, 5 )

        self.checkBox38 = wx.CheckBox( self, wx.ID_ANY, u"Certification", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox38.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer58.Add( self.checkBox38, 0, wx.ALL, 5 )


        bSizer561.Add( bSizer58, 1, wx.EXPAND, 5 )


        bSizer46.Add( bSizer561, 1, wx.EXPAND, 5 )


        bSizer39.Add( bSizer46, 1, wx.EXPAND, 5 )

        bSizer42 = wx.BoxSizer( wx.VERTICAL )

        self.staticText15 = wx.StaticText( self, wx.ID_ANY, u"MAI 4R2S Industry Standards", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText15.Wrap( -1 )
        self.staticText15.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer42.Add( self.staticText15, 0, wx.ALL, 5 )

        bSizer57 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer57.Add( ( 20, 0), 0, 0, 5 )

        self.checkBox35 = wx.CheckBox( self, wx.ID_ANY, u"Awareness", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox35.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer57.Add( self.checkBox35, 0, wx.ALL, 5 )


        bSizer42.Add( bSizer57, 1, wx.EXPAND, 5 )


        bSizer39.Add( bSizer42, 0, 0, 5 )

        bSizer43 = wx.BoxSizer( wx.VERTICAL )

        self.staticText16 = wx.StaticText( self, wx.ID_ANY, u"Vehicle Testing Report on Road Worthiness Test (In House & PUSPAKOM)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText16.Wrap( -1 )
        self.staticText16.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer43.Add( self.staticText16, 0, wx.ALL, 5 )

        bSizer45 = wx.BoxSizer( wx.VERTICAL )

        bSizer59 = wx.BoxSizer( wx.HORIZONTAL )

        self.staticText17 = wx.StaticText( self, wx.ID_ANY, u"Brake Test", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText17.Wrap( -1 )
        self.staticText17.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer59.Add( self.staticText17, 0, wx.ALL, 5 )


        bSizer59.Add( ( 100, 0), 1, 0, 5 )

        bSizer62 = wx.BoxSizer( wx.HORIZONTAL )

        self.checkBox183 = wx.CheckBox( self, wx.ID_ANY, u"InHouse", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox183.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer62.Add( self.checkBox183, 0, wx.ALL, 5 )

        self.checkBox184 = wx.CheckBox( self, wx.ID_ANY, u"External (PUSPAKOM)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox184.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer62.Add( self.checkBox184, 0, wx.ALL, 5 )


        bSizer59.Add( bSizer62, 1, wx.EXPAND, 5 )


        bSizer45.Add( bSizer59, 1, wx.EXPAND, 5 )

        bSizer60 = wx.BoxSizer( wx.HORIZONTAL )

        self.staticText18 = wx.StaticText( self, wx.ID_ANY, u"Side Slip Test", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText18.Wrap( -1 )
        self.staticText18.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer60.Add( self.staticText18, 0, wx.ALL, 5 )


        bSizer60.Add( ( 50, 0), 1, 0, 5 )

        bSizer61 = wx.BoxSizer( wx.HORIZONTAL )

        self.checkBox185 = wx.CheckBox( self, wx.ID_ANY, u"InHouse", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox185.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer61.Add( self.checkBox185, 0, wx.ALL, 5 )

        self.checkBox186 = wx.CheckBox( self, wx.ID_ANY, u"External (PUSPAKOM)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox186.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer61.Add( self.checkBox186, 0, wx.ALL, 5 )


        bSizer60.Add( bSizer61, 1, wx.EXPAND, 5 )


        bSizer45.Add( bSizer60, 1, wx.EXPAND, 5 )

        bSizer63 = wx.BoxSizer( wx.HORIZONTAL )

        self.staticText19 = wx.StaticText( self, wx.ID_ANY, u"Smoke Emission Test", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText19.Wrap( -1 )
        self.staticText19.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer63.Add( self.staticText19, 0, wx.ALL, 5 )

        bSizer64 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer64.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.checkBox187 = wx.CheckBox( self, wx.ID_ANY, u"InHouse", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox187.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer64.Add( self.checkBox187, 0, wx.ALL, 5 )

        self.checkBox188 = wx.CheckBox( self, wx.ID_ANY, u"External (PUSPAKOM)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox188.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer64.Add( self.checkBox188, 0, wx.ALL, 5 )


        bSizer63.Add( bSizer64, 1, wx.EXPAND, 5 )


        bSizer45.Add( bSizer63, 1, wx.EXPAND, 5 )


        bSizer43.Add( bSizer45, 1, wx.EXPAND, 5 )


        bSizer39.Add( bSizer43, 1, wx.EXPAND, 5 )


        bSizer52.Add( bSizer39, 1, wx.EXPAND, 5 )


        Method.Add( bSizer52, 1, wx.EXPAND, 5 )


        Right.Add( Method, 1, wx.EXPAND, 5 )


        bSizer85.Add( Right, 1, wx.EXPAND, 5 )


        bSizer32.Add( bSizer85, 1, wx.EXPAND, 5 )

        bSizer86 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer86.Add( ( 0, 0), 1, 0, 5 )

        self.button15 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button15.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.button15, 0, wx.ALL, 5 )

        self.button16 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button16.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.button16, 0, wx.ALL, 5 )


        bSizer32.Add( bSizer86, 0, wx.EXPAND, 5 )


        self.SetSizer( bSizer32 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.checkBox9.Bind( wx.EVT_CHECKBOX, self.Check1 )
        self.checkBox10.Bind( wx.EVT_CHECKBOX, self.Check2 )
        self.checkBox11.Bind( wx.EVT_CHECKBOX, self.Check3 )
        self.checkBox12.Bind( wx.EVT_CHECKBOX, self.Check4 )
        self.checkBox13.Bind( wx.EVT_CHECKBOX, self.Check5 )
        self.checkBox14.Bind( wx.EVT_CHECKBOX, self.Check6 )
        self.checkBox15.Bind( wx.EVT_CHECKBOX, self.Check7 )
        self.checkBox16.Bind( wx.EVT_CHECKBOX, self.Check8 )
        self.checkBox17.Bind( wx.EVT_CHECKBOX, self.Check9 )
        self.checkBox18.Bind( wx.EVT_CHECKBOX, self.Check10 )
        self.checkBox31.Bind( wx.EVT_CHECKBOX, self.Check11 )
        self.checkBox32.Bind( wx.EVT_CHECKBOX, self.Check12 )
        self.checkBox33.Bind( wx.EVT_CHECKBOX, self.Check13 )
        self.checkBox34.Bind( wx.EVT_CHECKBOX, self.Check14 )
        self.checkBox36.Bind( wx.EVT_CHECKBOX, self.Check15 )
        self.checkBox37.Bind( wx.EVT_CHECKBOX, self.Check16 )
        self.checkBox38.Bind( wx.EVT_CHECKBOX, self.Check17 )
        self.checkBox35.Bind( wx.EVT_CHECKBOX, self.Check18 )
        self.checkBox183.Bind( wx.EVT_CHECKBOX, self.Check19 )
        self.checkBox184.Bind( wx.EVT_CHECKBOX, self.Check20 )
        self.checkBox185.Bind( wx.EVT_CHECKBOX, self.Check21 )
        self.checkBox186.Bind( wx.EVT_CHECKBOX, self.Check22 )
        self.checkBox187.Bind( wx.EVT_CHECKBOX, self.Check23 )
        self.checkBox188.Bind( wx.EVT_CHECKBOX, self.Check24 )
        self.button15.Bind( wx.EVT_BUTTON, self.Reset )
        self.button16.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Check1( self, event ):
            global e1
            box5 = self.checkBox9.GetValue()
            
            if box5 == True:
                e1 = 1
            else:
                e1 = 0

    def Check2( self, event ):
            global f1
            box6 = self.checkBox10.GetValue()
            
            if box6 == True:
                f1 = 1
            else:
                f1 = 0

    def Check3( self, event ):
            global g1
            box7 = self.checkBox11.GetValue()
            
            if box7 == True:
                g1 = 1
            else:
                g1 = 0


    def Check4( self, event ):
            global h1
            box8 = self.checkBox12.GetValue()
            
            if box8 == True:
                h1 = 1
            else:
                h1 = 0

    def Check5( self, event ):
            global i1
            box9 = self.checkBox13.GetValue()
            
            if box9 == True:
                i1 = 1
            else:
                i1 = 0

    def Check6( self, event ):
            global j1
            box10 = self.checkBox14.GetValue()

            if box10 == True:
                j1 = 1
            else:
                j1 = 0

    def Check7( self, event ):
            global k1
            box11 = self.checkBox15.GetValue()

            if box11 == True:
                k1 = 1
            else:
                k1 = 0

    def Check8( self, event ):
            global l1
            box12 = self.checkBox16.GetValue()

            if box12 == True:
                l1 = 1
            else:
                l1 

    def Check9( self, event ):
            global m1
            box13 = self.checkBox17.GetValue()

            if box13 == True:
                m1 = 1
            else:
                m1 = 0

    def Check10( self, event ):
            global n1
            box14 = self.checkBox18.GetValue()

            if box14 == True:
                n1 = 1
            else:
                n1 = 0

    def Check11( self, event ):
            global o1
            box15 = self.checkBox31.GetValue()

            if box15 == True:
                o1 = 1
            else:
                o1 = 0

    def Check12( self, event ):
            global p1
            box16 = self.checkBox32

            if box16 == True:
                p1 = 1
            else:
                p1 = 0

    def Check13( self, event ):
            global q1
            box17 = self.checkBox33.GetValue()

            if box17 == True:
                q1 = 1
            else:
                q1 = 0


    def Check14( self, event ):
            global r1
            box18 = self.checkBox34.GetValue()

            if box18 == True:
                r1 = 1
            else:
                r1 = 0

    def Check15( self, event ):
            global s1
            box19 = self.checkBox36.GetValue()

            if box19 == True:
                s1 = 1
            else:
                s1 = 0

    def Check16( self, event ):
            global t1
            box20 = self.checkBox37.GetValue()

            if box20 == True:
                t1 = 1
            else:
                t1 = 0

    def Check17( self, event ):
            global u1
            box21 = self.checkBox38.GetValue()

            if box21 == True:
                u1 = 1
            else:
                u1 = 0

    def Check18( self, event ):
            global v1
            box22 = self.checkBox35.GetValue()

            if box22 == True:
                v1 = 1
            else:
                v1 = 0

    def Check19( self, event ):
            global w1
            box23 = self.checkBox183.GetValue()

            if box23 == True:
                w1 = 1
            else:
                w1 = 0


    def Check20( self, event ):
            global x1
            box24 = self.checkBox184.GetValue()

            if box24 == True:
                x1 = 1
            else:
                x1 = 0

    def Check21( self, event ):
            global y1
            box25 = self.checkBox185.GetValue()

            if box25 == True:
                y1 = 1
            else:
                y1 = 0

    def Check22( self, event ):
            global z1
            box26 = self.checkBox186.GetValue()

            if box26 == True:
                z1 = 1
            else:
                z1 = 0

    def Check23( self, event ):
            global a2
            box27 = self.checkBox187.GetValue()

            if box27 == True:
                a2 = 1
            else:
                a2 = 0

    def Check24( self, event ):
            global b2
            box28 = self.checkBox188.GetValue()

            if box28 == True:
                b2 = 1
            else:
                b2 = 0


    def Reset( self, event ):
            self.checkBox9.SetValue(False)
            self.checkBox10.SetValue(False)
            self.checkBox11.SetValue(False)
            self.checkBox12.SetValue(False)
            self.checkBox13.SetValue(False)
            self.checkBox14.SetValue(False)
            self.checkBox15.SetValue(False)
            self.checkBox16.SetValue(False)
            self.checkBox17.SetValue(False)
            self.checkBox18.SetValue(False)
            self.checkBox31.SetValue(False)
            self.checkBox32.SetValue(False)
            self.checkBox33.SetValue(False)
            self.checkBox34.SetValue(False)
            self.checkBox36.SetValue(False)
            self.checkBox37.SetValue(False)
            self.checkBox38.SetValue(False)
            self.checkBox35.SetValue(False)
            self.checkBox183.SetValue(False)
            self.checkBox184.SetValue(False)
            self.checkBox185.SetValue(False)
            self.checkBox186.SetValue(False)
            self.checkBox187.SetValue(False)
            self.checkBox188.SetValue(False)
            
            DataReset = Reset(parent = self.button15)
            SoundReset = soundReset(parent = self.button15)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global a5x
            global a6x
            global a7x
            global a8x
            global a9x
            global a10x
            global a11x
            global a12x
            global a13x
            global a14x
            global a15x
            global a16x
            global a17x
            global a18x
            global a19x
            global a20x
            global a21x
            global a22x
            global a23x
            global a24x
            global a25x
            global a26x
            global a27x
            global a28x

            if e1 == 1:
                a5x = 1
            else:
                a5x = 0

            if f1 == 1:
                a6x = 1
            else:
                a6x = 0

            if g1 == 1:
                a7x = 1
            else:
                a7x = 0

            if h1 == 1:
                a8x = 1
            else:
                a8x = 0

            if i1 == 1:
                a9x = 1
            else:
                a9x = 0

            if j1 == 1:
                a10x = 1
            else:
                a10x = 0

            if k1 == 1:
                a11x = 1
            else:
                a11x = 0

            if l1 == 1:
                a12x = 1
            else:
                a12x = 0

            if m1 == 1:
                a13x = 1
            else:
                a13x = 0

            if n1 == 1:
                a14x = 1
            else:
                a14x = 0

            if o1 == 1:
                a15x = 1
            else:
                a15x = 0

            if p1 == 1:
                a16x = 1
            else:
                a16x = 0

            if q1 == 1:
                a17x = 1
            else:
                a17x = 0

            if r1 == 1:
                a18x = 1
            else:
                a18x = 0

            if s1 == 1:
                a19x = 1
            else:
                a19x = 0

            if t1 == 1:
                a20x = 1
            else:
                a20x = 0

            if u1 == 1:
                a21x = 1
            else:
                a21x = 0

            if v1 == 1:
                a22x = 1
            else:
                a22x = 0

            if w1 == 1:
                a23x = 1
            else:
                a23x = 0

            if x1 == 1:
                a24x = 1
            else:
                a24x = 0
                
            if y1 == 1:
                a25x = 1
            else:
                a25x = 0
            
            if z1 == 1:
                a26x = 1
            else:
                a26x = 0
                
            if a2 == 1:
                a27x = 1
            else:
                a27x = 0
                
            if b2 == 1:
                a28x = 1
            else:
                a28x = 0
                
            SavingData = DataSaved(parent = self.button16)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button16)
            AcceptSound.Accept.Play()
            
###########################################################################
## Class PartC
###########################################################################

class PartC ( wx.Frame ):
    
    global a29x
    global a30x
    global a31x
    global a32x
    global a33x
    global a34x
    global a35x
    global a36x
    global a37x
    global a38x
    global a39x
    global a40x
    global a41x
    global a42x
    global a43x
    global a44x
    global a45x
    global a46x
    global a47x
    global a48x
    global a49x
    global a50x
    global a51x
    global a52x
    global a53x
    global a54x
    global a55x
    global a56x
    global a57x
    global a58x
    global a59x
    global a60x
    global a61x
    global a62x
    global a63x
    global a64x
    global a65x
    global a66x
    global a67x
    global a68x
    global a69x
    global a70x
    global a71x
    global a72x
    global a73x
    global a74x
    global a75x
    global a76x
    global a77x
    global a78x
    global a79x
    global a80x
    global a81x
    global a82x
    global a83x
    global a84x
    global a85x
    global a86x
    global a87x
    global a88x
    global a89x
    global a90x
    global a91x
    
    global c2
    global d2
    global e2
    global f2
    global g2
    global h2
    global i2
    global j2
    global k2
    global l2
    global m2
    global n2
    global o2
    global p2
    global q2
    global r2
    global s2
    global t2
    global u2
    global v2
    global w2
    global x2
    global y2
    global z2
    global a3
    global b3
    global c3
    global d3
    global e3
    global f3
    global g3
    global h3
    global i3
    global j3
    global k3
    global l3
    global m3
    global n3
    global o3
    global p3
    global q3
    global r3
    global s3
    global t3
    global u3
    global v3
    global w3
    global x3
    global y3
    global z3
    global a4
    global b4
    global c4
    global d4
    global e4
    global f4
    global g4
    global h4
    global i4
    global j4
    global k4
    global l4
    global m4
    
    c2 = 0
    d2 = 0
    e2 = 0
    f2 = 0
    g2 = 0
    h2 = 0
    i2 = 0
    j2 = 0
    k2 = 0
    l2 = 0
    m2 = 0
    n2 = 0
    o2 = 0
    p2 = 0
    q2 = 0
    r2 = 0
    s2 = 0
    t2 = 0
    u2 = 0
    v2 = 0
    w2 = 0
    x2 = 0
    y2 = 0
    z2 = 0
    a3 = 0
    b3 = 0
    c3 = 0
    d3 = 0
    e3 = 0
    f3 = 0
    g3 = 0
    h3 = 0
    i3 = 0
    j3 = 0
    k3 = 0
    l3 = 0
    m3 = 0
    n3 = 0
    o3 = 0
    p3 = 0
    q3 = 0
    r3 = 0
    s3 = 0
    t3 = 0
    u3 = 0
    v3 = 0
    w3 = 0
    x3 = 0
    y3 = 0
    z3 = 0
    a4 = 0
    b4 = 0
    c4 = 0
    d4 = 0
    e4 = 0
    f4 = 0
    g4 = 0
    h4 = 0
    i4 = 0
    j4 = 0
    k4 = 0
    l4 = 0
    m4 = 0
    
    a29x = 0
    a30x = 0
    a31x = 0
    a32x = 0
    a33x = 0
    a34x = 0
    a35x = 0
    a36x = 0
    a37x = 0
    a38x = 0
    a39x = 0
    a40x = 0
    a41x = 0
    a42x = 0
    a43x = 0
    a44x = 0
    a45x = 0
    a46x = 0
    a47x = 0
    a48x = 0
    a49x = 0
    a50x = 0
    a51x = 0
    a52x = 0
    a53x = 0
    a54x = 0
    a55x = 0
    a56x = 0
    a57x = 0
    a58x = 0
    a59x = 0
    a60x = 0
    a61x = 0
    a62x = 0
    a63x = 0
    a64x = 0
    a65x = 0
    a66x = 0
    a67x = 0
    a68x = 0
    a69x = 0
    a70x = 0
    a71x = 0
    a72x = 0
    a73x = 0
    a74x = 0
    a75x = 0
    a76x = 0
    a77x = 0
    a78x = 0
    a79x = 0
    a80x = 0
    a81x = 0
    a82x = 0
    a83x = 0
    a84x = 0
    a85x = 0
    a86x = 0
    a87x = 0
    a88x = 0
    a89x = 0
    a90x = 0

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Reamanufacturing Work Flow System And SOP", pos = wx.DefaultPosition, size = wx.Size( 764,591 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(764,561), wx.Size(764,561) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )

        bSizer65 = wx.BoxSizer( wx.VERTICAL )

        bSizer88 = wx.BoxSizer( wx.VERTICAL )

        self.staticText50 = wx.StaticText( self, wx.ID_ANY, u"COMPLIANCES OF REMANUFACTURING WORK FLOW SYSTEM & S.O.P", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText50.Wrap( -1 )
        self.staticText50.SetFont( wx.Font( 14, 70, 90, 92, False, "Arial" ) )

        bSizer88.Add( self.staticText50, 0, wx.ALL, 5 )


        bSizer65.Add( bSizer88, 1, wx.EXPAND, 5 )

        bSizer68 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer68.Add( ( 270, 0), 1, 0, 5 )

        self.staticText32 = wx.StaticText( self, wx.ID_ANY, u"Brake System", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText32.Wrap( -1 )
        self.staticText32.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer68.Add( self.staticText32, 0, wx.ALL, 5 )


        bSizer68.Add( ( 0, 0), 1, 0, 5 )

        self.staticText33 = wx.StaticText( self, wx.ID_ANY, u"Intercooler", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText33.Wrap( -1 )
        self.staticText33.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer68.Add( self.staticText33, 0, wx.ALL, 5 )


        bSizer68.Add( ( 0, 0), 1, 0, 5 )

        self.staticText34 = wx.StaticText( self, wx.ID_ANY, u"Turbocharger", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText34.Wrap( -1 )
        self.staticText34.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer68.Add( self.staticText34, 0, wx.ALL, 5 )


        bSizer68.Add( ( 0, 0), 1, 0, 5 )

        self.staticText35 = wx.StaticText( self, wx.ID_ANY, u"Stater Motor", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText35.Wrap( -1 )
        self.staticText35.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer68.Add( self.staticText35, 0, wx.ALL, 5 )


        bSizer68.Add( ( 0, 0), 1, 0, 5 )

        self.staticText36 = wx.StaticText( self, wx.ID_ANY, u"Alternator", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText36.Wrap( -1 )
        self.staticText36.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer68.Add( self.staticText36, 0, wx.ALL, 5 )


        bSizer68.Add( ( 0, 0), 1, 0, 5 )

        self.staticText37 = wx.StaticText( self, wx.ID_ANY, u"Radiator", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText37.Wrap( -1 )
        self.staticText37.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer68.Add( self.staticText37, 0, wx.ALL, 5 )


        bSizer68.Add( ( 0, 0), 1, 0, 5 )


        bSizer65.Add( bSizer68, 0, wx.EXPAND, 5 )

        bSizer66 = wx.BoxSizer( wx.VERTICAL )

        bSizer74 = wx.BoxSizer( wx.HORIZONTAL )

        bSizer73 = wx.BoxSizer( wx.VERTICAL )

        self.staticText21 = wx.StaticText( self, wx.ID_ANY, u"Core Management", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText21.Wrap( -1 )
        self.staticText21.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText21, 0, wx.ALL, 5 )

        self.staticText22 = wx.StaticText( self, wx.ID_ANY, u"Sorting", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText22.Wrap( -1 )
        self.staticText22.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText22, 0, wx.ALL, 5 )

        self.staticText23 = wx.StaticText( self, wx.ID_ANY, u"Disassembly of parts and components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText23.Wrap( -1 )
        self.staticText23.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText23, 0, wx.ALL, 5 )

        self.staticText24 = wx.StaticText( self, wx.ID_ANY, u"Cleaning of parts and components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText24.Wrap( -1 )
        self.staticText24.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText24, 0, wx.ALL, 5 )

        self.staticText25 = wx.StaticText( self, wx.ID_ANY, u"Machining and/or polishing", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText25.Wrap( -1 )
        self.staticText25.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText25, 0, wx.ALL, 5 )

        self.staticText26 = wx.StaticText( self, wx.ID_ANY, u"Inspection and Testing", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText26.Wrap( -1 )
        self.staticText26.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText26, 0, wx.ALL, 5 )

        self.staticText27 = wx.StaticText( self, wx.ID_ANY, u"Assembly", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText27.Wrap( -1 )
        self.staticText27.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText27, 0, wx.ALL, 5 )

        self.staticText28 = wx.StaticText( self, wx.ID_ANY, u"Painting/Polishing", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText28.Wrap( -1 )
        self.staticText28.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText28, 0, wx.ALL, 5 )

        self.staticText29 = wx.StaticText( self, wx.ID_ANY, u"Labelling", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText29.Wrap( -1 )
        self.staticText29.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText29, 0, wx.ALL, 5 )

        self.staticText31 = wx.StaticText( self, wx.ID_ANY, u"Warranty of parts (1 year waranty)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText31.Wrap( -1 )
        self.staticText31.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer73.Add( self.staticText31, 0, wx.ALL, 5 )


        bSizer74.Add( bSizer73, 0, wx.EXPAND, 5 )


        bSizer74.Add( ( 20, 0), 0, 0, 5 )

        bSizer72 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox34 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox34, 1, wx.ALL, 5 )

        self.checkBox35 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox35, 1, wx.ALL, 5 )

        self.checkBox36 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox36, 1, wx.ALL, 5 )

        self.checkBox37 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox37, 1, wx.ALL, 5 )

        self.checkBox38 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox38, 1, wx.ALL, 5 )

        self.checkBox39 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox39, 1, wx.ALL, 5 )

        self.checkBox40 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox40, 1, wx.ALL, 5 )

        self.checkBox41 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox41, 1, wx.ALL, 5 )

        self.checkBox42 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox42, 1, wx.ALL, 5 )

        self.checkBox43 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer72.Add( self.checkBox43, 1, wx.ALL, 5 )


        bSizer74.Add( bSizer72, 0, wx.EXPAND, 5 )


        bSizer74.Add( ( 55, 0), 0, 0, 5 )

        bSizer75 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox44 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox44, 1, wx.ALL, 5 )

        self.checkBox45 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox45, 1, wx.ALL, 5 )

        self.checkBox46 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox46, 1, wx.ALL, 5 )

        self.checkBox47 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox47, 1, wx.ALL, 5 )

        self.checkBox48 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox48, 1, wx.ALL, 5 )

        self.checkBox49 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox49, 1, wx.ALL, 5 )

        self.checkBox50 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox50, 1, wx.ALL, 5 )

        self.checkBox51 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox51, 1, wx.ALL, 5 )

        self.checkBox52 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox52, 1, wx.ALL, 5 )

        self.checkBox53 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer75.Add( self.checkBox53, 1, wx.ALL, 5 )


        bSizer74.Add( bSizer75, 0, wx.EXPAND, 5 )


        bSizer74.Add( ( 55, 0), 0, 0, 5 )

        bSizer77 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox54 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox54, 1, wx.ALL, 5 )

        self.checkBox55 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox55, 1, wx.ALL, 5 )

        self.checkBox56 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox56, 1, wx.ALL, 5 )

        self.checkBox57 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox57, 1, wx.ALL, 5 )

        self.checkBox58 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox58, 1, wx.ALL, 5 )

        self.checkBox59 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox59, 1, wx.ALL, 5 )

        self.checkBox60 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox60, 1, wx.ALL, 5 )

        self.checkBox61 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox61, 1, wx.ALL, 5 )

        self.checkBox62 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox62, 1, wx.ALL, 5 )

        self.checkBox63 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer77.Add( self.checkBox63, 1, wx.ALL, 5 )


        bSizer74.Add( bSizer77, 0, wx.EXPAND, 5 )


        bSizer74.Add( ( 70, 0), 0, 0, 5 )

        bSizer78 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox64 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox64, 1, wx.ALL, 5 )

        self.checkBox65 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox65, 1, wx.ALL, 5 )

        self.checkBox66 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox66, 1, wx.ALL, 5 )

        self.checkBox67 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox67, 1, wx.ALL, 5 )

        self.checkBox68 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox68, 1, wx.ALL, 5 )

        self.checkBox69 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox69, 1, wx.ALL, 5 )

        self.checkBox70 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox70, 1, wx.ALL, 5 )

        self.checkBox71 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox71, 1, wx.ALL, 5 )

        self.checkBox72 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox72, 1, wx.ALL, 5 )

        self.checkBox73 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer78.Add( self.checkBox73, 1, wx.ALL, 5 )


        bSizer74.Add( bSizer78, 0, wx.EXPAND, 5 )


        bSizer74.Add( ( 55, 0), 0, 0, 5 )

        bSizer79 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox74 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox74, 1, wx.ALL, 5 )

        self.checkBox75 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox75, 1, wx.ALL, 5 )

        self.checkBox76 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox76, 1, wx.ALL, 5 )

        self.checkBox77 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox77, 1, wx.ALL, 5 )

        self.checkBox78 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox78, 1, wx.ALL, 5 )

        self.checkBox79 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox79, 1, wx.ALL, 5 )

        self.checkBox80 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox80, 1, wx.ALL, 5 )

        self.checkBox81 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox81, 1, wx.ALL, 5 )

        self.checkBox82 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox82, 1, wx.ALL, 5 )

        self.checkBox83 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer79.Add( self.checkBox83, 1, wx.ALL, 5 )


        bSizer74.Add( bSizer79, 0, wx.EXPAND, 5 )


        bSizer74.Add( ( 45, 0), 0, 0, 5 )

        bSizer80 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox84 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox84, 1, wx.ALL, 5 )

        self.checkBox85 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox85, 1, wx.ALL, 5 )

        self.checkBox86 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox86, 1, wx.ALL, 5 )

        self.checkBox87 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox87, 1, wx.ALL, 5 )

        self.checkBox88 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox88, 1, wx.ALL, 5 )

        self.checkBox89 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox89, 1, wx.ALL, 5 )

        self.checkBox90 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox90, 1, wx.ALL, 5 )

        self.checkBox91 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox91, 1, wx.ALL, 5 )

        self.checkBox92 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox92, 1, wx.ALL, 5 )

        self.checkBox93 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer80.Add( self.checkBox93, 1, wx.ALL, 5 )


        bSizer74.Add( bSizer80, 0, wx.EXPAND, 5 )


        bSizer66.Add( bSizer74, 1, wx.EXPAND, 5 )


        bSizer65.Add( bSizer66, 1, wx.EXPAND, 5 )

        bSizer81 = wx.BoxSizer( wx.VERTICAL )


        bSizer81.Add( ( 0, 20), 0, 0, 5 )

        self.staticText38 = wx.StaticText( self, wx.ID_ANY, u"Final Inspection and Testing on Road Worthiness", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText38.Wrap( -1 )
        self.staticText38.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer81.Add( self.staticText38, 0, wx.ALL, 5 )

        bSizer83 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer83.Add( ( 20, 0), 0, 0, 5 )

        bSizer82 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox194 = wx.CheckBox( self, wx.ID_ANY, u"Brake Test", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox194.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer82.Add( self.checkBox194, 0, wx.ALL, 5 )

        self.checkBox195 = wx.CheckBox( self, wx.ID_ANY, u"Side Slip Test", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox195.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer82.Add( self.checkBox195, 0, wx.ALL, 5 )

        self.checkBox196 = wx.CheckBox( self, wx.ID_ANY, u"Smoke Emission Test", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox196.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer82.Add( self.checkBox196, 0, wx.ALL, 5 )


        bSizer83.Add( bSizer82, 1, wx.EXPAND, 5 )


        bSizer81.Add( bSizer83, 1, wx.EXPAND, 5 )


        bSizer65.Add( bSizer81, 1, wx.EXPAND, 5 )

        bSizer801 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer801.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button17 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button17.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer801.Add( self.button17, 0, wx.ALL, 5 )

        self.button18 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button18.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer801.Add( self.button18, 0, wx.ALL, 5 )


        bSizer65.Add( bSizer801, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer65 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.checkBox34.Bind( wx.EVT_CHECKBOX, self.Check1 )
        self.checkBox35.Bind( wx.EVT_CHECKBOX, self.Check2 )
        self.checkBox36.Bind( wx.EVT_CHECKBOX, self.Check3 )
        self.checkBox37.Bind( wx.EVT_CHECKBOX, self.Check4 )
        self.checkBox38.Bind( wx.EVT_CHECKBOX, self.Check5 )
        self.checkBox39.Bind( wx.EVT_CHECKBOX, self.Check6 )
        self.checkBox40.Bind( wx.EVT_CHECKBOX, self.Check7 )
        self.checkBox41.Bind( wx.EVT_CHECKBOX, self.Check8 )
        self.checkBox42.Bind( wx.EVT_CHECKBOX, self.Check9 )
        self.checkBox43.Bind( wx.EVT_CHECKBOX, self.Check10 )
        self.checkBox44.Bind( wx.EVT_CHECKBOX, self.Check11 )
        self.checkBox45.Bind( wx.EVT_CHECKBOX, self.Check12 )
        self.checkBox46.Bind( wx.EVT_CHECKBOX, self.Check13 )
        self.checkBox47.Bind( wx.EVT_CHECKBOX, self.Check14 )
        self.checkBox48.Bind( wx.EVT_CHECKBOX, self.Check15 )
        self.checkBox49.Bind( wx.EVT_CHECKBOX, self.Check16 )
        self.checkBox50.Bind( wx.EVT_CHECKBOX, self.Check17 )
        self.checkBox51.Bind( wx.EVT_CHECKBOX, self.Check18 )
        self.checkBox52.Bind( wx.EVT_CHECKBOX, self.Check19 )
        self.checkBox53.Bind( wx.EVT_CHECKBOX, self.Check20 )
        self.checkBox54.Bind( wx.EVT_CHECKBOX, self.Check21 )
        self.checkBox55.Bind( wx.EVT_CHECKBOX, self.Check22 )
        self.checkBox56.Bind( wx.EVT_CHECKBOX, self.Check23 )
        self.checkBox57.Bind( wx.EVT_CHECKBOX, self.Check24 )
        self.checkBox58.Bind( wx.EVT_CHECKBOX, self.Check25 )
        self.checkBox59.Bind( wx.EVT_CHECKBOX, self.Check26 )
        self.checkBox60.Bind( wx.EVT_CHECKBOX, self.Check27 )
        self.checkBox61.Bind( wx.EVT_CHECKBOX, self.Check28 )
        self.checkBox62.Bind( wx.EVT_CHECKBOX, self.Check29 )
        self.checkBox63.Bind( wx.EVT_CHECKBOX, self.Check30 )
        self.checkBox64.Bind( wx.EVT_CHECKBOX, self.Check31 )
        self.checkBox65.Bind( wx.EVT_CHECKBOX, self.Check32 )
        self.checkBox66.Bind( wx.EVT_CHECKBOX, self.Check33 )
        self.checkBox67.Bind( wx.EVT_CHECKBOX, self.Check34 )
        self.checkBox68.Bind( wx.EVT_CHECKBOX, self.Check35 )
        self.checkBox69.Bind( wx.EVT_CHECKBOX, self.Check36 )
        self.checkBox70.Bind( wx.EVT_CHECKBOX, self.Check37 )
        self.checkBox71.Bind( wx.EVT_CHECKBOX, self.Check38 )
        self.checkBox72.Bind( wx.EVT_CHECKBOX, self.Check39 )
        self.checkBox73.Bind( wx.EVT_CHECKBOX, self.Check40 )
        self.checkBox74.Bind( wx.EVT_CHECKBOX, self.Check41 )
        self.checkBox75.Bind( wx.EVT_CHECKBOX, self.Check42 )
        self.checkBox76.Bind( wx.EVT_CHECKBOX, self.Check43 )
        self.checkBox77.Bind( wx.EVT_CHECKBOX, self.Check44 )
        self.checkBox78.Bind( wx.EVT_CHECKBOX, self.Check45 )
        self.checkBox79.Bind( wx.EVT_CHECKBOX, self.Check46 )
        self.checkBox80.Bind( wx.EVT_CHECKBOX, self.Check47 )
        self.checkBox81.Bind( wx.EVT_CHECKBOX, self.Check48 )
        self.checkBox82.Bind( wx.EVT_CHECKBOX, self.Check49 )
        self.checkBox83.Bind( wx.EVT_CHECKBOX, self.Check50 )
        self.checkBox84.Bind( wx.EVT_CHECKBOX, self.Check51 )
        self.checkBox85.Bind( wx.EVT_CHECKBOX, self.Check52 )
        self.checkBox86.Bind( wx.EVT_CHECKBOX, self.Check53 )
        self.checkBox87.Bind( wx.EVT_CHECKBOX, self.Check54 )
        self.checkBox88.Bind( wx.EVT_CHECKBOX, self.Check55 )
        self.checkBox89.Bind( wx.EVT_CHECKBOX, self.Check56 )
        self.checkBox90.Bind( wx.EVT_CHECKBOX, self.Check57 )
        self.checkBox91.Bind( wx.EVT_CHECKBOX, self.Check58 )
        self.checkBox92.Bind( wx.EVT_CHECKBOX, self.Check59 )
        self.checkBox93.Bind( wx.EVT_CHECKBOX, self.Check60 )
        self.checkBox194.Bind( wx.EVT_CHECKBOX, self.Check61 )
        self.checkBox195.Bind( wx.EVT_CHECKBOX, self.Check62 )
        self.checkBox196.Bind( wx.EVT_CHECKBOX, self.Check63 )
        self.button17.Bind( wx.EVT_BUTTON, self.Reset )
        self.button18.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Check1( self, event ):
            global c2
            box29 = self.checkBox34.GetValue()

            if box29 == True:
                c2 = 1
            else:
                c2 = 0

    def Check2( self, event ):
            global d2
            box30 = self.checkBox35.GetValue()

            if box30 == True:
                d2 = 1
            else:
                d2 = 0

    def Check3( self, event ):
            global e2
            box31 = self.checkBox36.GetValue()

            if box31 == True:
                e2 = 1
            else:
                e2 = 0


    def Check4( self, event ):
            global f2
            box32 = self.checkBox37.GetValue()

            if box32 == True:
                f2 = 1
            else:
                f2 = 0

    def Check5( self, event ):
            global g2
            box33 = self.checkBox38.GetValue()

            if box33 == True:
                g2 = 1
            else:
                g2 = 0

    def Check6( self, event ):
            global h2
            box34 = self.checkBox39.GetValue()

            if box34 == True:
                h2 = 1
            else:
                h2 = 0

    def Check7( self, event ):
            global i2
            box35 = self.checkBox40.GetValue()

            if box35 == True:
                i2 = 1
            else:
                i2 = 0

    def Check8( self, event ):
            global j2
            box36 = self.checkBox41.GetValue()

            if box36 == True:
                j2 = 1
            else:
                j2 = 0

    def Check9( self, event ):
            global k2
            box37 = self.checkBox42.GetValue()

            if box37 == True:
                k2 = 1
            else:
                k2 = 0

    def Check10( self, event ):
            global l2
            box38 = self.checkBox43.GetValue()

            if box38 == True:
                l2 = 1
            else:
                l2 = 0

    def Check11( self, event ):
            global m2
            box39 = self.checkBox44.GetValue()

            if box39 == True:
                m2 = 1
            else:
                m2 = 0

    def Check12( self, event ):
            global n2
            box40 = self.checkBox45.GetValue()

            if box40 == True:
                n2 = 1
            else:
                n2 = 0

    def Check13( self, event ):
            global o2
            box41 = self.checkBox46.GetValue()

            if box41 == True:
                o2 = 1
            else:
                o2 = 0

    def Check14( self, event ):
            global p2
            box42 = self.checkBox47.GetValue()

            if box42 == True:
                p2 = 1
            else:
                p2 = 0

    def Check15( self, event ):
            global q2
            box43 = self.checkBox48.GetValue()

            if box43 == True:
                q2 = 1
            else:
                q2 = 0

    def Check16( self, event ):
            global r2
            box44 = self.checkBox48.GetValue()

            if box44 == True:
                r2 = 1
            else:
                r2 = 0

    def Check17( self, event ):
            global s2
            box45 = self.checkBox50.GetValue()

            if box45 == True:
                s2 = 1
            else:
                s2 = 0

    def Check18( self, event ):
            global t2
            box46 = self.checkBox51.GetValue()

            if box46 == True:
                t2 = 1
            else:
                t2 = 0

    def Check19( self, event ):
            global u2
            box47 = self.checkBox52.GetValue()

            if box47 == True:
                u2 = 1
            else:
                u2 = 0

    def Check20( self, event ):
            global v2
            box48 = self.checkBox53.GetValue()

            if box48 == True:
                v2 = 1
            else:
                v2 = 0

    def Check21( self, event ):
            global w2
            box49 = self.checkBox54.GetValue()

            if box49 == True:
                w2 = 1
            else:
                w2 = 0

    def Check22( self, event ):
            global x2
            box50 = self.checkBox55.GetValue()

            if box50 == True:
                x2 = 1
            else:
                x2 = 0

    def Check23( self, event ):
            global y2
            box51 = self.checkBox56.GetValue()

            if box51 == True:
                y2 = 1
            else:
                y2 = 0


    def Check24( self, event ):
            global z2
            box52 = self.checkBox57.GetValue()

            if box52 == True:
                z2 = 1

            else:
                z2 = 0


    def Check25( self, event ):
            global a3
            box53 = self.checkBox58.GetValue()

            if box53 == True:
                a3 = 1

            else:
                a3 = 0

    def Check26( self, event ):
            global b3
            box54 = self.checkBox59.GetValue()

            if box54 == True:
                b3 = 1

            else:
                b3 = 0

    def Check27( self, event ):
            global c3
            box55 = self.checkBox60.GetValue()

            if box55 == True:
                c3 = 1

            else:
                c3 = 0

    def Check28( self, event ):
            global d3
            box56 = self.checkBox61.GetValue()

            if box56 == True:
                d3 = 1

            else:
                d3 = 0

    def Check29( self, event ):
            global e3
            box57 = self.checkBox62.GetValue()

            if box57 == True:
                e3 = 1

            else:
                e3 = 0

    def Check30( self, event ):
            global f3
            box58 = self.checkBox63.GetValue()

            if box58 == True:
                f3 = 1

            else:
                f3 = 0

    def Check31( self, event ):
            global g3
            box59 = self.checkBox64.GetValue()

            if box59 == True:
                g3 = 1

            else:
                g3 = 0

    def Check32( self, event ):
            global h3
            box60 = self.checkBox65.GetValue()

            if box60 == True:
                h3 = 1

            else:
                h3 = 0

    def Check33( self, event ):
            global i3
            box61 = self.checkBox66.GetValue()

            if box61 == True:
                i3 = 1

            else:
                i3 = 0

    def Check34( self, event ):
            global j3
            box62 = self.checkBox67.GetValue()

            if box62 == True:
                j3 = 1

            else:
                j3 = 0

    def Check35( self, event ):
            global k3
            box63 = self.checkBox68.GetValue()

            if box63 == True:
                k3 = 1

            else:
                k3 = 0

    def Check36( self, event ):
            global l3
            box64 = self.checkBox69.GetValue()

            if box64 == True:
                l3 = 1

            else:
                l3 = 0

    def Check37( self, event ):
            global m3
            box65 = self.checkBox70.GetValue()

            if box65 == True:
                m3 = 1

            else:
                m3 = 0

    def Check38( self, event ):
            global n3
            box66 = self.checkBox71.GetValue()

            if box66 == True:
                n3 = 1

            else:
                n3 = 0

    def Check39( self, event ):
            global o3
            box67 = self.checkBox72.GetValue()

            if box67 == True:
                o3 = 1

            else:
                o3 = 0

    def Check40( self, event ):
            global p3
            box68 = self.checkBox73.GetValue()

            if box68 == True:
                p3 = 1

            else:
                p3 = 0

    def Check41( self, event ):
            global q3
            box69 = self.checkBox74.GetValue()

            if box69 == True:
                q3 = 1

            else:
                q3 = 0

    def Check42( self, event ):
            global r3
            box70 = self.checkBox75.GetValue()

            if box70 == True:
                r3 = 1

            else:
                r3 = 0

    def Check43( self, event ):
            global s3
            box71 = self.checkBox76.GetValue()

            if box71 == True:
                s3 = 1

            else:
                s3 = 0

    def Check44( self, event ):
            global t3
            box72 = self.checkBox77.GetValue()

            if box72 == True:
                t3 = 1

            else:
                t3 = 0

    def Check45( self, event ):
            global u3
            box73 = self.checkBox78.GetValue()

            if box73 == True:
                u3 = 1

            else:
                u3 = 0

    def Check46( self, event ):
            global v3
            box74 = self.checkBox79.GetValue()

            if box74 == True:
                v3 = 1

            else:
                v3 = 0

    def Check47( self, event ):
            global w3
            box75 = self.checkBox80.GetValue()

            if box75 == True:
                w3 = 1

            else:
                w3 = 0

    def Check48( self, event ):
            global x3
            box76 = self.checkBox81.GetValue()

            if box76 == True:
                x3 = 1

            else:
                x3 = 0

    def Check49( self, event ):
            global y3
            box77 = self.checkBox82.GetValue()

            if box77 == True:
                y3 = 1

            else:
                y3 = 0
            

    def Check50( self, event ):
            global z3
            box78 = self.checkBox83.GetValue()

            if box78 == True:
                z3 = 1

            else:
                z3 = 0

    def Check51( self, event ):
            global a4
            box79 = self.checkBox84.GetValue()

            if box79 == True:
                a4 = 1

            else:
                a4 = 0

    def Check52( self, event ):
            global b4
            box80 = self.checkBox85.GetValue()

            if box80 == True:
                b4 = 1

            else:
                b4 = 0

    def Check53( self, event ):
            global c4
            box81 = self.checkBox86.GetValue()

            if box81 == True:
                c4 = 1

            else:
                c4 = 0

    def Check54( self, event ):
            global d4
            box82 = self.checkBox87.GetValue()

            if box82 == True:
                d4 = 1

            else:
                d4 = 0

    def Check55( self, event ):
            global e4
            box83 = self.checkBox88.GetValue()
            

            if box83 == True:
                e4 = 1

            else:
                e4 = 0

    def Check56( self, event ):
            global f4
            box84 = self.checkBox89.GetValue()

            if box84 == True:
                f4 = 1

            else:
                f4 = 0

    def Check57( self, event ):
            global g4
            box85 = self.checkBox90.GetValue()

            if box85 == True:
                g4 = 1

            else:
                g4 = 0

    def Check58( self, event ):
            global h4
            box86 = self.checkBox91.GetValue()

            if box86 == True:
                h4 = 1

            else:
                h4 = 0

    def Check59( self, event ):
            global i4
            box87 = self.checkBox92.GetValue()

            if box87 == True:
                i4 = 1

            else:
                i4 = 0

    def Check60( self, event ):
            global j4
            box88 = self.checkBox93.GetValue()

            if box88 == True:
                j4 = 1

            else:
                j4 = 0

    def Check61( self, event ):
            global k4
            box89 = self.checkBox194.GetValue()

            if box89 == True:
                k4 = 1

            else:
                k4 = 0

    def Check62( self, event ):
            global l4
            box90 = self.checkBox195.GetValue()

            if box90 == True:
                l4 = 1

            else:
                l4 = 0

    def Check63( self, event ):
            global m4
            box91 = self.checkBox196.GetValue()

            if box91 == True:
                m4 = 1

            else:
                m4 = 0

    def Reset( self, event ):
            self.checkBox34.SetValue(False)
            self.checkBox35.SetValue(False)
            self.checkBox36.SetValue(False)
            self.checkBox37.SetValue(False)
            self.checkBox38.SetValue(False)
            self.checkBox39.SetValue(False)
            self.checkBox40.SetValue(False)
            self.checkBox41.SetValue(False)
            self.checkBox42.SetValue(False)
            self.checkBox43.SetValue(False)
            self.checkBox44.SetValue(False)
            self.checkBox45.SetValue(False)
            self.checkBox46.SetValue(False)
            self.checkBox47.SetValue(False)
            self.checkBox48.SetValue(False)
            self.checkBox49.SetValue(False)
            self.checkBox50.SetValue(False)
            self.checkBox51.SetValue(False)
            self.checkBox52.SetValue(False)
            self.checkBox53.SetValue(False)
            self.checkBox54.SetValue(False)
            self.checkBox55.SetValue(False)
            self.checkBox56.SetValue(False)
            self.checkBox57.SetValue(False)
            self.checkBox58.SetValue(False)
            self.checkBox59.SetValue(False)
            self.checkBox60.SetValue(False)
            self.checkBox61.SetValue(False)
            self.checkBox62.SetValue(False)
            self.checkBox63.SetValue(False)
            self.checkBox64.SetValue(False)
            self.checkBox65.SetValue(False)
            self.checkBox66.SetValue(False)
            self.checkBox67.SetValue(False)
            self.checkBox68.SetValue(False)
            self.checkBox69.SetValue(False)
            self.checkBox70.SetValue(False)
            self.checkBox71.SetValue(False)
            self.checkBox72.SetValue(False)
            self.checkBox73.SetValue(False)
            self.checkBox74.SetValue(False)
            self.checkBox75.SetValue(False)
            self.checkBox76.SetValue(False)
            self.checkBox77.SetValue(False)
            self.checkBox78.SetValue(False)
            self.checkBox79.SetValue(False)
            self.checkBox80.SetValue(False)
            self.checkBox81.SetValue(False)
            self.checkBox82.SetValue(False)
            self.checkBox83.SetValue(False)
            self.checkBox84.SetValue(False)
            self.checkBox85.SetValue(False)
            self.checkBox86.SetValue(False)
            self.checkBox87.SetValue(False)
            self.checkBox88.SetValue(False)
            self.checkBox89.SetValue(False)
            self.checkBox90.SetValue(False)
            self.checkBox91.SetValue(False)
            self.checkBox92.SetValue(False)
            self.checkBox93.SetValue(False)
            self.checkBox194.SetValue(False)
            self.checkBox195.SetValue(False)
            self.checkBox196.SetValue(False)
            
            DataReset = Reset(parent = self.button17)
            SoundReset = soundReset(parent = self.button17)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
            
            

    def Save( self, event ):
            global a29x
            global a30x
            global a31x
            global a32x
            global a33x
            global a34x
            global a35x
            global a36x
            global a37x
            global a38x
            global a39x
            global a40x
            global a41x
            global a42x
            global a43x
            global a44x
            global a45x
            global a46x
            global a47x
            global a48x
            global a49x
            global a50x
            global a51x
            global a52x
            global a53x
            global a54x
            global a55x
            global a56x
            global a57x
            global a58x
            global a59x
            global a60x
            global a61x
            global a62x
            global a63x
            global a64x
            global a65x
            global a66x
            global a67x
            global a68x
            global a69x
            global a70x
            global a71x
            global a72x
            global a73x
            global a74x
            global a75x
            global a76x
            global a77x
            global a78x
            global a79x
            global a80x
            global a81x
            global a82x
            global a83x
            global a84x
            global a85x
            global a86x
            global a87x
            global a88x
            global a89x
            global a90x
            global a91x
            
            if c2 == 1:
                a29x = 1
            else:
                a29x = 0
            
            if d2 == 1:
                a30x = 1
            else:
                a30x = 0
                
            if e2 == 1:
                a31x = 1
            else:
                a31x = 0
                
            if f2 == 1:
                a32x = 1
            else:
                a32x = 0
                
            if g2 == 1:
                a33x = 1
            else:
                a33x = 0
                
            if h2 == 1:
                a34x = 1
            else:
                a34x = 0
                
            if i2 == 1:
                a35x = 1
            else:
                a35x = 0
                
            if j2 == 1:
                a36x = 1
            else:
                a36x = 0
                
            if k2 == 1:
                a37x = 1
            else:
                a37x = 0
                
            if l2 == 1:
                a38x = 1
            else:
                a38x = 0
                
            if m2 == 1:
                a39x = 1
            else:
                a39x = 0
                
            if n2 == 1:
                a40x = 1
            else:
                a40x = 0
                
            if o2 == 1:
                a41x = 1
            else:
                a41x = 0
                
            if p2 == 1:
                a42x = 1
            else:
                a42x = 0
                
            if q2 == 1:
                a43x = 1
            else:
                a43x = 0
                
            if r2 == 1:
                a44x = 1
            else:
                a44x = 0
                
            if s2 == 1:
                a45x = 1
            else:
                a45x = 0
                
            if t2 == 1:
                a46x = 1
            else:
                a46x = 0
                
            if u2 == 1:
                a47x = 1
            else:
                a47x = 0
                
            if v2 == 1:
                a48x = 1
            else:
                a48x = 0
                
            if w2 == 1:
                a49x = 1
            else:
                a49x = 0
                
            if x2 == 1:
                a50x = 1
            else:
                a50x = 0
                
            if y2 == 1:
                a51x = 1
            else:
                a51x = 0
                
            if z2 == 1:
                a52x = 1
            else:
                a52x = 0
                
            if a3 == 1:
                a53x = 1
            else:
                a53x = 0
                
            if b3 == 1:
                a54x = 1
            else:
                a54x = 0
                
            if c3 == 1:
                a55x = 1
            else:
                a55x = 0
                
            if d3 == 1:
                a56x = 1
            else:
                a56x = 0
                
            if e3 == 1:
                a57x = 1
            else:
                a57x = 0
                
            if f3 == 1:
                a58x = 1
            else:
                a58x = 0
                
            if g3 == 1:
                a59x = 1
            else:
                a59x = 0
                
            if h3 == 1:
                a60x = 1
            else:
                a60x = 0
                
            if i3 == 1:
                a60x = 1
            else:
                a60x = 0
                
            if j3 == 1:
                a61x = 1
            else:
                a61x = 0
                
            if k3 == 1:
                a62x = 1
            else:
                a62x = 0
                
            if l3 == 1:
                a63x = 1
            else:
                a63x = 0
                
            if m3 == 1:
                a64x = 1
            else:
                a64x = 0
                
            if n3 == 1:
                a65x = 1
            else:
                a65x = 0
                
            if o3 == 1:
                a66x = 1
            else:
                a66x = 0
                
            if p3 == 1:
                a67x = 1
            else:
                a67x = 0
                
            if q3 == 1:
                a68x = 1
            else:
                a68x = 0
                
            if r3 == 1:
                a69x = 1
            else:
                a69x = 0
                
            if s3 == 1:
                a70x = 1
            else:
                a70x = 0
                
            if t3 == 1:
                a71x = 1
            else:
                a71x = 0
                
            if u3 == 1:
                a72x = 1
            else:
                a72x = 0
                
            if v3 == 1:
                a73x = 1
            else:
                a73x = 0
            
            if w3 == 1:
                a74x = 1
            else:
                a74x = 0
                
            if x3 == 1:
                a75x = 1
            else:
                a75x = 0
                
            if x3 == 1:
                a76x = 1
            else:
                a76x = 0
                
            if y3 == 1:
                a77x = 1
            else:
                a77x = 0
                
            if z3 == 1:
                a78x = 1
            else:
                a78x = 0
                
            if a4 == 1:
                a79x = 1
            else:
                a79x = 0
                
            if b4 == 1:
                a80x = 1
            else:
                a80x = 0
                
            if c4 == 1:
                a81x = 1
            else:
                a81x = 0
                
            if d4 == 1:
                a82x = 1
            else:
                a82x = 0
                
            if e4 == 1:
                a83x = 1
            else:
                a83x = 0
                
            if f4 == 1:
                a84x = 1
            else:
                a84x = 0
                
            if g4 == 1:
                a85x = 1
            else:
                a85x = 0
                
            if h4 == 1:
                a86x = 1
            else:
                a86x = 0
                
            if i4 == 1:
                a87x = 1
            else:
                a87x = 0
                
            if j4 == 1:
                a88x = 1
            else:
                a88x = 0
                
            if k4 == 1:
                a89x = 1
            else:
                a89x = 0
                
            if l4 == 1:
                a90x = 1
            else:
                a90x = 0
             
            if m4 == 1:
                a91x = 1
            else:
                a91x = 0

            SavingData = DataSaved(parent = self.button18)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button18)
            AcceptSound.Accept.Play()

###########################################################################
## Class PartD
###########################################################################

class PartD ( wx.Frame ):
    global a92x
    global a93x
    global a94x
    global a95x
    global a96x
    global a97x
    global a98x
    
    global n4
    global o4
    global p4
    global q4
    global r4
    global s4
    global t4
    
    n4 = 0
    o4 = 0
    p4 = 0
    q4 = 0
    r4 = 0
    s4 = 0
    t4 = 0
    
    a92x = 0
    a93x = 0
    a94x = 0
    a95x = 0
    a96x = 0
    a97x = 0
    a98x = 0
        
        

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Standard Compliance To\nVehicle Type Approval (VTA)\n-Safety Parts-", pos = wx.DefaultPosition, size = wx.Size( 791,320 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(791,320), wx.Size(791,320) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )

        bSizer81 = wx.BoxSizer( wx.VERTICAL )

        bSizer82 = wx.BoxSizer( wx.HORIZONTAL )

        self.staticText37 = wx.StaticText( self, wx.ID_ANY, u"Compliance to E-marking that follow condition MS/ UN", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText37.Wrap( -1 )
        self.staticText37.SetFont( wx.Font( 14, 70, 90, 92, False, "Arial" ) )

        bSizer82.Add( self.staticText37, 0, wx.ALL, 5 )


        bSizer82.Add( ( 50, 0), 0, 0, 5 )

        self.staticText38 = wx.StaticText( self, wx.ID_ANY, u"Standards", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText38.Wrap( -1 )
        self.staticText38.SetFont( wx.Font( 14, 70, 90, 92, False, "Arial" ) )

        bSizer82.Add( self.staticText38, 0, wx.ALL, 5 )


        bSizer81.Add( bSizer82, 0, wx.EXPAND, 5 )

        bSizer84 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer84.Add( ( 20, 0), 0, 0, 5 )

        bSizer83 = wx.BoxSizer( wx.VERTICAL )

        self.checkBox87 = wx.CheckBox( self, wx.ID_ANY, u"Safety Glass", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox87.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer83.Add( self.checkBox87, 0, wx.ALL, 5 )

        self.checkBox88 = wx.CheckBox( self, wx.ID_ANY, u"Brake Lamp Performance", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox88.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer83.Add( self.checkBox88, 0, wx.ALL, 5 )

        self.checkBox89 = wx.CheckBox( self, wx.ID_ANY, u"Indicator Performance", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox89.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer83.Add( self.checkBox89, 0, wx.ALL, 5 )

        self.checkBox90 = wx.CheckBox( self, wx.ID_ANY, u"Headlamp Performance", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox90.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer83.Add( self.checkBox90, 0, wx.ALL, 5 )

        self.checkBox91 = wx.CheckBox( self, wx.ID_ANY, u"Tire and Wheel", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox91.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer83.Add( self.checkBox91, 0, wx.ALL, 5 )

        self.checkBox92 = wx.CheckBox( self, wx.ID_ANY, u"Rear Marking Plates", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox92.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer83.Add( self.checkBox92, 0, wx.ALL, 5 )

        self.m_checkBox93 = wx.CheckBox( self, wx.ID_ANY, u"Front Underrun Protection", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_checkBox93.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer83.Add( self.m_checkBox93, 0, wx.ALL, 5 )


        bSizer84.Add( bSizer83, 1, wx.EXPAND, 5 )

        bSizer85 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer85.Add( ( 30, 0), 0, 0, 5 )

        bSizer86 = wx.BoxSizer( wx.VERTICAL )

        self.staticText43 = wx.StaticText( self, wx.ID_ANY, u"MS 595/UN R43", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText43.Wrap( -1 )
        self.staticText43.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.staticText43, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.staticText44 = wx.StaticText( self, wx.ID_ANY, u"MS 1776/UN R7", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText44.Wrap( -1 )
        self.staticText44.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.staticText44, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.staticText45 = wx.StaticText( self, wx.ID_ANY, u"MS 1851/UN R6", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText45.Wrap( -1 )
        self.staticText45.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.staticText45, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.staticText46 = wx.StaticText( self, wx.ID_ANY, u"UN R112", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText46.Wrap( -1 )
        self.staticText46.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.staticText46, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.staticText47 = wx.StaticText( self, wx.ID_ANY, u"UN R54", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText47.Wrap( -1 )
        self.staticText47.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.staticText47, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.staticText48 = wx.StaticText( self, wx.ID_ANY, u"UN R70", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText48.Wrap( -1 )
        self.staticText48.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.staticText48, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"UN R 93", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText49.Wrap( -1 )
        self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer86.Add( self.staticText49, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer85.Add( bSizer86, 1, wx.EXPAND, 5 )


        bSizer84.Add( bSizer85, 1, wx.EXPAND, 5 )


        bSizer81.Add( bSizer84, 1, wx.EXPAND, 5 )

        bSizer87 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer87.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button19 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button19.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer87.Add( self.button19, 0, wx.ALL, 5 )

        self.button20 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button20.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer87.Add( self.button20, 0, wx.ALL, 5 )


        bSizer81.Add( bSizer87, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer81 )
        self.Layout()

        self.Centre( wx.BOTH )
        # Connect Events
        self.checkBox87.Bind( wx.EVT_CHECKBOX, self.Check1 )
        self.checkBox88.Bind( wx.EVT_CHECKBOX, self.Check2 )
        self.checkBox89.Bind( wx.EVT_CHECKBOX, self.Check3 )
        self.checkBox90.Bind( wx.EVT_CHECKBOX, self.Check4 )
        self.checkBox91.Bind( wx.EVT_CHECKBOX, self.Check5 )
        self.checkBox92.Bind( wx.EVT_CHECKBOX, self.Check6 )
        self.m_checkBox93.Bind( wx.EVT_CHECKBOX, self.Check7 )
        self.button19.Bind( wx.EVT_BUTTON, self.Reset )
        self.button20.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Check1( self, event ):
            global n4
            box92 = self.checkBox87.GetValue()

            if box92 == True:
                n4 = 1

            else:
                n4 = 0

    def Check2( self, event ):
            global o4
            box93 = self.checkBox88.GetValue()

            if box93 == True:
                o4 = 1

            else:
                o4 = 0

    def Check3( self, event ):
            global p4
            box94 = self.checkBox89.GetValue()

            if box94 == True:
                p4 = 1

            else:
                p4 = 0

    def Check4( self, event ):
            global q4
            box95 = self.checkBox90.GetValue()

            if box95 == True:
                q4 = 1

            else:
                q4 = 0

    def Check5( self, event ):
            global r4
            box96 = self.checkBox91.GetValue()

            if box96 == True:
                r4 = 1

            else:
                r4 = 0

    def Check6( self, event ):
            global s4
            box97 = self.checkBox92.GetValue()

            if box97 == True:
                s4 = 1

            else:
                s4 = 0

    def Check7( self, event ):
            global t4
            box98 = self.m_checkBox93.GetValue()

            if box98 == True:
                t4 = 1

            else:
                t4 = 0

    def Reset( self, event ):
            self.checkBox87.SetValue(False)
            self.checkBox88.SetValue(False)
            self.checkBox89.SetValue(False)
            self.checkBox90.SetValue(False)
            self.checkBox91.SetValue(False)
            self.checkBox92.SetValue(False)
            self.m_checkBox93.SetValue(False)
            
            DataReset = Reset(parent = self.button19)
            SoundReset = soundReset(parent = self.button19)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global a92x
            global a93x
            global a94x
            global a95x
            global a96x
            global a97x
            global a98x
            
            if n4 == 1:
                a92x = 1
            else:
                a92x = 0
                
            if o4 == 1:
                a93x = 1
            else:
                a93x = 0
                
            if p4 == 1:
                a94x = 1
            else:
                a94x = 0
                
            if q4 == 1:
                a95x = 1
            else:
                a95x = 0
                
            if r4 == 1:
                a96x = 1
            else:
                a96x = 0
                
            if s4 == 1:
                a97x = 1
            else:
                a97x = 0
                
            if t4 == 1:
                a98x = 1
            else:
                a98x = 0
            
            SavingData = DataSaved(parent = self.button20)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button20)
            AcceptSound.Accept.Play()
            

###########################################################################
## Class ChecklistNew
###########################################################################
class ChecklistNew ( wx.Frame ):
    
    global aa
    global ba
    global ca
    global da
    global ea
    
    global ab
    global bb
    global cb
    global db
    global eb
    global fb
    global gb
    global hb
    global ib
    global jb
    
    global ac
    global bc
    global cc
    global dc
    global ec
    global fc
    global gc
    
    global ad
    global bd
    global cd
    
    global ae
    
    
    global a12
    global b12
    global c12
    global d12
    global e12
    global f12
    global g12
    global h12
    global i12
    global j12
    global k12
    global l12
    global m12
    global n12
    global o12
    
    global p12
    global q12
    global r12
    global s12
    global t12
    global u12
    global v12
    global w12
    global x12
    global y12
    global z12
    global a13
    global b13
    global c13
    global d13
    global e13
    global f13
    global g13
    global h13
    global i13
    global j13
    global k13
    global l13
    global m13
    global n13
    global o13
    global p13
    global q13
    global r13
    global s13
    
    global t13
    global u13
    global v13
    global w13
    global x13
    global y13
    global z13
    global a14
    global b14
    global c14
    global d14
    global e14
    global f14
    global g14
    global h14
    global i14
    global j14
    global k14
    global l14
    global m14
    global n14
    
    global o14
    global p14
    global q14
    global r14
    global s14
    global t14
    global u14
    global v14
    global w14
    
    global x14
    global y14
    global z14
    
    global mop1
    global mop2
    global mop3
    global mop4
    global mop5
    
    global mop6
    global mop7
    global mop8
    global mop9
    global mop10
    global mop11
    global mop12
    global mop13
    global mop14
    global mop15
    
    global mop16
    global mop17
    global mop18
    global mop19
    global mop20
    global mop21
    global mop22
    
    global mop23
    global mop24
    global mop25
    
    global mop26
    
    global mup1
    global mup2
    global mup3
    global mup4
    global mup5
    global mup6
    global mup7
    global mup8
    global mup9
    
    global b1x
    global b2x
    global b3x
    global b4x
    global b5x
    global b6x
    global b7x
    global b8x
    global b9x
    global b10x
    global b11x
    global b12x
    global b13x
    global b14x
    global b15x
    
    global b16x
    global b17x
    global b18x
    global b19x
    global b20x
    global b21x
    global b22x
    global b23x
    global b24x
    global b25x
    global b26x
    global b27x

    global aiy2
    global moo2
    
    aiy2 = 0

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Checklist New", pos = wx.DefaultPosition, size = wx.Size( 500,500 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(500,500), wx.Size(500,500) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer92 = wx.BoxSizer( wx.VERTICAL )


        bSizer92.Add( ( 0, 15), 0, 0, 5 )

        bSizer254 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer254.Add( ( 100, 0), 0, 0, 5 )

        self.staticText223 = wx.StaticText( self, wx.ID_ANY, u"State Date", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText223.Wrap( -1 )

        self.staticText223.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer254.Add( self.staticText223, 0, wx.ALL, 5 )

        self.datePicker2 = wx.adv.DatePickerCtrl( self, wx.ID_ANY, wx.DefaultDateTime, wx.DefaultPosition, wx.DefaultSize, wx.adv.DP_DEFAULT )
        self.datePicker2.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer254.Add( self.datePicker2, 0, wx.ALL, 5 )

        self.button77 = wx.Button( self, wx.ID_ANY, u"Set", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.button77.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer254.Add( self.button77, 0, wx.ALL, 5 )


        bSizer92.Add( bSizer254, 1, wx.EXPAND, 5 )

        bSizer93 = wx.BoxSizer( wx.VERTICAL )

        self.button26 = wx.Button( self, wx.ID_ANY, u"Clause 7.1 \nGeneral Requirement", wx.DefaultPosition, wx.Size( 160,60 ), 0 )
        self.button26.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer93.Add( self.button26, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.button27 = wx.Button( self, wx.ID_ANY, u"Clause 7.3\nRemanufacturing SOP", wx.DefaultPosition, wx.Size( 160,-1 ), 0 )
        self.button27.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer93.Add( self.button27, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.button28 = wx.Button( self, wx.ID_ANY, u"Clause 9.1 \nLabelling and packaging", wx.DefaultPosition, wx.Size( 160,-1 ), 0 )
        self.button28.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer93.Add( self.button28, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.button29 = wx.Button( self, wx.ID_ANY, u"Clause 10\nWarranty", wx.DefaultPosition, wx.Size( 160,-1 ), 0 )
        self.button29.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer93.Add( self.button29, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.button30 = wx.Button( self, wx.ID_ANY, u"Clause 11\nSupplier mark", wx.DefaultPosition, wx.Size( 160,-1 ), 0 )
        self.button30.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer93.Add( self.button30, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.button78 = wx.Button( self, wx.ID_ANY, u"Auditor\nVerification", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.button78.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer93.Add( self.button78, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer92.Add( bSizer93, 1, wx.EXPAND, 5 )

        bSizer94 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer94.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button31 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button31.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer94.Add( self.button31, 0, wx.ALL, 5 )


        bSizer92.Add( bSizer94, 0, wx.EXPAND, 5 )


        self.SetSizer( bSizer92 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.datePicker2.Bind( wx.adv.EVT_DATE_CHANGED, self.Date2 )
        self.button77.Bind( wx.EVT_BUTTON, self.Set2 )
        self.button26.Bind( wx.EVT_BUTTON, self.Clause1 )
        self.button27.Bind( wx.EVT_BUTTON, self.Clause2 )
        self.button28.Bind( wx.EVT_BUTTON, self.Clause3 )
        self.button29.Bind( wx.EVT_BUTTON, self.Clause4 )
        self.button30.Bind( wx.EVT_BUTTON, self.Clause5 )
        self.button78.Bind( wx.EVT_BUTTON, self.Auditor )
        self.button31.Bind( wx.EVT_BUTTON, self.Reset )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    # Virtual event handlers, overide them in your derived class
    def Date2( self, event ):
            global mine2
            mine2 =  event.GetDate()
            

    def Set2( self, event ):
            global aiy2
            global moo2
            
            moo2 = mine2.Format("%d/%m/%Y")
            aiy2 = 1
            
            SavingData = DataSaved(parent = self.button77)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button77)
            AcceptSound.Accept.Play()

    def Clause1( self, event ):
            #event.Skip()
            Clause1Form = Clause71(parent = self.button26)
            Clause1Form.Show()

    def Clause2( self, event ):
            #event.Skip()
            Clause2Form = Clause73(parent = self.button27)
            Clause2Form.Show()

    def Clause3( self, event ):
            #event.Skip()
            Clause3Form = Clause91a(parent = self.button28)
            Clause3Form.Show()

    def Clause4( self, event ):
            #event.Skip()
            Clause4Form = Clause10(parent = self.button29)
            Clause4Form.Show()

    def Clause5( self, event ):
            #event.Skip()
            Clause5Form = Clause11(parent = self.button30)
            Clause5Form.Show()

    def Auditor( self, event ):
            #event.Skip()
            AuditorForm = ChecklistNewDetail(parent = self.button78)
            AuditorForm.Show()

    def Reset( self, event ):
            global aa
            global ba
            global ca
            global da
            global ea
            
            global ab
            global bb
            global cb
            global db
            global eb
            global fb
            global gb
            global hb
            global ib
            global jb
            
            global ac
            global bc
            global cc
            global dc
            global ec
            global fc
            global gc
            
            global ad
            global bd
            global cd
            
            global ae
            
            
            global a12
            global b12
            global c12
            global d12
            global e12
            global f12
            global g12
            global h12
            global i12
            global j12
            global k12
            global l12
            global m12
            global n12
            global o12
            
            global p12
            global q12
            global r12
            global s12
            global t12
            global u12
            global v12
            global w12
            global x12
            global y12
            global z12
            global a13
            global b13
            global c13
            global d13
            global e13
            global f13
            global g13
            global h13
            global i13
            global j13
            global k13
            global l13
            global m13
            global n13
            global o13
            global p13
            global q13
            global r13
            global s13
            
            global t13
            global u13
            global v13
            global w13
            global x13
            global y13
            global z13
            global a14
            global b14
            global c14
            global d14
            global e14
            global f14
            global g14
            global h14
            global i14
            global j14
            global k14
            global l14
            global m14
            global n14
            
            global o14
            global p14
            global q14
            global r14
            global s14
            global t14
            global u14
            global v14
            global w14
            
            global x14
            global y14
            global z14
            
            global mop1
            global mop2
            global mop3
            global mop4
            global mop5
            
            global mop6
            global mop7
            global mop8
            global mop9
            global mop10
            global mop11
            global mop12
            global mop13
            global mop14
            global mop15
            
            global mop16
            global mop17
            global mop18
            global mop19
            global mop20
            global mop21
            global mop22
            
            global mop23
            global mop24
            global mop25
            
            global mop26
            
            global mup1
            global mup2
            global mup3
            global mup4
            global mup5
            global mup6
            global mup7
            global mup8
            global mup9
            
            
            global b1x
            global b2x
            global b3x
            global b4x
            global b5x
            global b6x
            global b7x
            global b8x
            global b9x
            global b10x
            global b11x
            global b12x
            global b13x
            global b14x
            global b15x
            
            global b16x
            global b17x
            global b18x
            global b19x
            global b20x
            global b21x
            global b22x
            global b23x
            global b24x
            global b25x
            global b26x
            global b27x



            aa = 0
            ba = 0
            ca = 0
            da = 0
            ea = 0
            
            ab = 0
            bb = 0
            cb = 0
            db = 0
            eb = 0
            fb = 0
            gb = 0
            hb = 0
            ib = 0
            jb = 0
            
            ac = 0
            bc = 0
            cc = 0
            dc = 0
            ec = 0
            fc = 0
            gc = 0
            
            ad = 0
            bd = 0
            cd = 0
            
            ae = 0
            
            
            a12 = 0
            b12 = 0
            c12 = 0
            d12 = 0
            e12 = 0
            f12 = 0
            g12 = 0
            h12 = 0
            i12 = 0
            j12 = 0
            k12 = 0
            l12 = 0
            m12 = 0
            n12 = 0
            o12 = 0
            
            p12 = 0
            q12 = 0
            r12 = 0
            s12 = 0
            t12 = 0
            u12 = 0
            v12 = 0
            w12 = 0
            x12 = 0
            y12 = 0
            z12 = 0
            a13 = 0
            b13 = 0
            c13 = 0
            d13 = 0
            e13 = 0
            f13 = 0
            g13 = 0
            h13 = 0
            i13 = 0
            j13 = 0
            k13 = 0
            l13 = 0
            m13 = 0
            n13 = 0
            o13 = 0
            p13 = 0
            q13 = 0
            r13 = 0
            s13 = 0
            
            t13 = 0
            u13 = 0
            v13 = 0
            w13 = 0
            x13 = 0
            y13 = 0
            z13 = 0
            a14 = 0
            b14 = 0
            c14 = 0
            d14 = 0
            e14 = 0
            f14 = 0
            g14 = 0
            h14 = 0
            i14 = 0
            j14 = 0
            k14 = 0
            l14 = 0
            m14 = 0
            n14 = 0
            
            o14 = 0
            p14 = 0
            q14 = 0
            r14 = 0
            s14 = 0
            t14 = 0
            u14 = 0
            v14 = 0
            w14 = 0
            
            x14 = 0
            y14 = 0
            z14 = 0
            
            
            b1x = ""
            b2x = ""
            b3x = ""
            b4x = ""
            b5x = ""
            b6x = ""
            b7x = ""
            b8x = ""
            b9x = ""
            b10x = ""
            b11x = ""
            b12x = ""
            b13x = ""
            b14x = ""
            b15x = ""
            
            b16x = ""
            b17x = ""
            b18x = ""
            b19x = ""
            b20x = ""
            b21x = ""
            b22x = ""
            b23x = ""
            b24x = ""
            b25x = ""
            b26x = ""
            b27x = ""

            
            mop1 = ""
            mop2 = ""
            mop3 = ""
            mop4 = ""
            mop5 = ""
            
            mop6 = ""
            mop7 = ""
            mop8 = ""
            mop9 = ""
            mop10 = ""
            mop11 = ""
            mop12 = ""
            mop13 = ""
            mop14 = ""
            mop15 = ""
        
            mop16 = ""
            mop17 = ""
            mop18 = ""
            mop19 = ""
            mop20 = ""
            mop21 = ""
            mop22 = ""
            
            mop23 = ""
            mop24 = ""
            mop25 = ""
            
            mop26 = ""
            
            mup1 = ""
            mup2 = ""
            mup3 = ""
            mup4 = ""
            mup5 = ""
            mup6 = ""
            mup7 = ""
            mup8 = ""
            mup9 = ""
            
            DataReset = Reset(parent = self.button31)
            SoundReset = soundReset(parent = self.button31)
            
            DataReset.Show()
            SoundReset.Resetto.Play()


###########################################################################
## Class ChecklistNewDetail
###########################################################################

class ChecklistNewDetail ( wx.Frame ):
    global a11
    global b11
    global c11
    global d11
    global e11
    global f11
    global g11
    global h11
    global i11
    
    global mup1
    global mup2
    global mup3
    global mup4
    global mup5
    global mup6
    global mup7
    global mup8
    global mup9
    
    a11 = 0
    b11 = 0
    c11 = 0
    d11 = 0
    e11 = 0
    f11 = 0
    g11 = 0
    h11 = 0
    i11 = 0
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Checklist New Detail", pos = wx.DefaultPosition, size = wx.Size( 1187,225 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(1187,225), wx.Size(1187,225) )

        bSizer307 = wx.BoxSizer( wx.VERTICAL )

        bSizer308 = wx.BoxSizer( wx.HORIZONTAL )

        self.staticText275 = wx.StaticText( self, wx.ID_ANY, u"Description", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText275.Wrap( -1 )

        self.staticText275.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial" ) )

        bSizer308.Add( self.staticText275, 0, wx.ALL, 5 )


        bSizer308.Add( ( 130, 0), 0, wx.EXPAND, 5 )

        self.staticText276 = wx.StaticText( self, wx.ID_ANY, u"Auditor 1", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText276.Wrap( -1 )

        self.staticText276.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial" ) )

        bSizer308.Add( self.staticText276, 0, wx.ALL, 5 )


        bSizer308.Add( ( 250, 0), 0, 0, 5 )

        self.staticText277 = wx.StaticText( self, wx.ID_ANY, u"Auditor 2", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText277.Wrap( -1 )

        self.staticText277.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial" ) )

        bSizer308.Add( self.staticText277, 0, wx.ALL, 5 )


        bSizer308.Add( ( 270, 0), 0, wx.EXPAND, 5 )

        self.staticText278 = wx.StaticText( self, wx.ID_ANY, u"Auditor 3", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText278.Wrap( -1 )

        self.staticText278.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial" ) )

        bSizer308.Add( self.staticText278, 0, wx.ALL, 5 )


        bSizer307.Add( bSizer308, 0, wx.EXPAND, 5 )

        bSizer309 = wx.BoxSizer( wx.HORIZONTAL )

        bSizer310 = wx.BoxSizer( wx.VERTICAL )

        self.staticText279 = wx.StaticText( self, wx.ID_ANY, u"Name", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText279.Wrap( -1 )

        self.staticText279.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer310.Add( self.staticText279, 0, wx.ALL, 5 )


        bSizer310.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText280 = wx.StaticText( self, wx.ID_ANY, u"Position", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText280.Wrap( -1 )

        self.staticText280.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer310.Add( self.staticText280, 0, wx.ALL, 5 )


        bSizer310.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText281 = wx.StaticText( self, wx.ID_ANY, u"Organisation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText281.Wrap( -1 )

        self.staticText281.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer310.Add( self.staticText281, 0, wx.ALL, 5 )


        bSizer309.Add( bSizer310, 0, wx.EXPAND, 5 )


        bSizer309.Add( ( 30, 0), 0, wx.EXPAND, 5 )

        bSizer311 = wx.BoxSizer( wx.VERTICAL )

        self.textCtrl80 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer311.Add( self.textCtrl80, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer311.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.textCtrl81 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer311.Add( self.textCtrl81, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer311.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.textCtrl82 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer311.Add( self.textCtrl82, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer309.Add( bSizer311, 1, wx.EXPAND, 5 )

        bSizer312 = wx.BoxSizer( wx.VERTICAL )

        self.textCtrl84 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer312.Add( self.textCtrl84, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer312.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.textCtrl85 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer312.Add( self.textCtrl85, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer312.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.textCtrl86 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer312.Add( self.textCtrl86, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer309.Add( bSizer312, 1, wx.EXPAND, 5 )

        bSizer313 = wx.BoxSizer( wx.VERTICAL )

        self.textCtrl88 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer313.Add( self.textCtrl88, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer313.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.textCtrl89 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer313.Add( self.textCtrl89, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer313.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.textCtrl90 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer313.Add( self.textCtrl90, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer309.Add( bSizer313, 1, wx.EXPAND, 5 )


        bSizer307.Add( bSizer309, 1, wx.EXPAND, 5 )

        bSizer314 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer314.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button79 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size(-1,50), 0 )
        self.button79.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer314.Add( self.button79, 0, wx.ALL, 5 )

        self.button80 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size(-1,50), 0 )
        self.button80.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer314.Add( self.button80, 0, wx.ALL, 5 )


        bSizer307.Add( bSizer314, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer307 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.textCtrl80.Bind( wx.EVT_TEXT, self.auditor1name )
        self.textCtrl81.Bind( wx.EVT_TEXT, self.auditor1position )
        self.textCtrl82.Bind( wx.EVT_TEXT, self.auditor1organisation )
        self.textCtrl84.Bind( wx.EVT_TEXT, self.auditor2name )
        self.textCtrl85.Bind( wx.EVT_TEXT, self.auditor2position )
        self.textCtrl86.Bind( wx.EVT_TEXT, self.auditor2organisation )
        self.textCtrl88.Bind( wx.EVT_TEXT, self.auditor3name )
        self.textCtrl89.Bind( wx.EVT_TEXT, self.auditor3position )
        self.textCtrl90.Bind( wx.EVT_TEXT, self.auditor3organisation )
        self.button79.Bind( wx.EVT_BUTTON, self.Reset )
        self.button80.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def auditor1name( self, event ):
            global poach1
            global a11
            poach1 = self.textCtrl80.GetValue()
            a11 = 1

    def auditor1position( self, event ):
            global poach2
            global b11
            poach2 = self.textCtrl81.GetValue()
            b11 = 1

    def auditor1organisation( self, event ):
            global poach3
            global c11
            poach3 = self.textCtrl82.GetValue()
            c11 = 1

    def auditor2name( self, event ):
            global poach4
            global d11
            poach4 = self.textCtrl84.GetValue()
            d11 = 1

    def auditor2position( self, event ):
            global poach5
            global e11
            poach5 = self.textCtrl85.GetValue()
            e11 = 1

    def auditor2organisation( self, event ):
            global poach6
            global f11
            poach6 = self.textCtrl86.GetValue()
            f11 = 1

    def auditor3name( self, event ):
            global poach7
            global g11
            poach7 = self.textCtrl88.GetValue()
            g11 = 1

    def auditor3position( self, event ):
            global poach8
            global h11
            poach8 = self.textCtrl89.GetValue()
            h11 = 1

    def auditor3organisation( self, event ):
            global poach9
            global i11
            poach9 = self.textCtrl90.GetValue()
            i11 = 1

    def Reset( self, event ):
            self.textCtrl80.SetValue("")
            self.textCtrl81.SetValue("")
            self.textCtrl82.SetValue("")
            self.textCtrl84.SetValue("")
            self.textCtrl85.SetValue("")
            self.textCtrl86.SetValue("")
            self.textCtrl88.SetValue("")
            self.textCtrl89.SetValue("")
            self.textCtrl90.SetValue("")
            
            DataReset = Reset(parent = self.button79)
            SoundReset = soundReset(parent = self.button79)
            
            DataReset.Show()
            SoundReset.Resetto.Play()


    def Save( self, event ):

            global mup1
            global mup2
            global mup3
            global mup4
            global mup5
            global mup6
            global mup7
            global mup8
            global mup9

            if a11 == 1:
                mup1 = poach1
            else:
                mup1 = ""
                
            if b11 == 1:
                mup2 = poach2
            else:
                mup2 = ""
                
            if c11 == 1:
                mup3 = poach3
            else:
                mup3 = ""
                
            if d11 == 1:
                mup4 = poach4
            else:
                mup4 = ""
                
            if e11 == 1:    
                mup5 = poach5
            else:
                mup5 = ""
            
            if f11 == 1:
                mup6 = poach6
            else:
                mup6 = ""
                
            if g11 == 1:    
                mup7 = poach7
            else:
                mup7 = ""
             
            if h11 == 1:
                mup8 = poach8
            else:
                mup8 = ""
                
            if i11 == 1:
                mup9 = poach9
            else:
                mup9 = ""
                
            SavingData = DataSaved(parent = self.button80)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button80)
            AcceptSound.Accept.Play()
                

###########################################################################
## Class Clause71
###########################################################################
class Clause71 ( wx.Frame ):
    global aa
    global ba
    global ca
    global da
    global ea
    
    
    global a12
    global b12
    global c12
    global d12
    global e12
    global f12
    global g12
    global h12
    global i12
    global j12
    global k12
    global l12
    global m12
    global n12
    global o12
    
    global mop1
    global mop2
    global mop3
    global mop4
    global mop5
    
    global b1x
    global b2x
    global b3x
    global b4x
    global b5x
    global b6x

    
    aa = 0
    ba = 0
    ca = 0
    da = 0
    ea = 0
    
    a12 = 0
    b12 = 0
    c12 = 0
    d12 = 0
    e12 = 0
    f12 = 0
    g12 = 0
    h12 = 0
    i12 = 0
    j12 = 0
    k12 = 0
    l12 = 0
    m12 = 0
    n12 = 0
    o12 = 0
    
    b1x = ""
    b2x = ""
    b3x = ""
    b4x = ""
    b5x = ""
    b6x = ""
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Clause 7.1", pos = wx.DefaultPosition, size = wx.Size( 1500,591 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
        
        self.SetSizeHintsSz( wx.Size(1500,591), wx.Size(1500,591) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
        
        bSizer95 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer117 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText64.Wrap( -1 )
        self.staticText64.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 370, 0), 0, 0, 5 )
        
        self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText65.Wrap( -1 )
        self.staticText65.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 545, 0), 0, 0, 5 )
        
        self.staticText270 = wx.StaticText( self, wx.ID_ANY, u"Remark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText270.Wrap( -1 )
        self.staticText270.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText270, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer117, 1, wx.EXPAND, 5 )
        
        bSizer116 = wx.BoxSizer( wx.HORIZONTAL )
        
        bSizer104 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer96 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.1", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText49.Wrap( -1 )
        self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"Handling remanufacturing of the\nused parts and components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText103.Wrap( -1 )
        self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 90, 0), 1, 0, 5 )
        
        self.staticText50 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing\nof the used parts and components shall \nbe to  demonstrate that  it has the legal \nright to transfer their ownership to \nanother party.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText50.Wrap( -1 )
        self.staticText50.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText50, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer96, 0, 0, 5 )
        
        bSizer97 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.2", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText51.Wrap( -1 )
        self.staticText51.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText51, 0, wx.ALL, 5 )
        
        
        bSizer97.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText105 = wx.StaticText( self, wx.ID_ANY, u"Provide specified information \nof remanufactured parts\nand components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText105.Wrap( -1 )
        self.staticText105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText105, 0, wx.ALL, 5 )
        
        
        bSizer97.Add( ( 105, 0), 0, 0, 5 )
        
        self.staticText52 = wx.StaticText( self, wx.ID_ANY, u"The company shall be able to provide\nspecified information on the condition\nof the remanufactured parts and\ncomponents supplied.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText52.Wrap( -1 )
        self.staticText52.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText52, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer97, 1, wx.EXPAND, 5 )
        
        bSizer98 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText57 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.3", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText57.Wrap( -1 )
        self.staticText57.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText57, 0, wx.ALL, 5 )
        
        
        bSizer98.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText106 = wx.StaticText( self, wx.ID_ANY, u"Manpower certification ", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText106.Wrap( -1 )
        self.staticText106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText106, 0, wx.ALL, 5 )
        
        
        bSizer98.Add( ( 147, 0), 0, 0, 5 )
        
        self.staticText58 = wx.StaticText( self, wx.ID_ANY, u"Person dismantling, cleaning, examining,\nremediating, re-assembling, testing and \nhandling of the core and remanufactured\nparts or components shall be of \ntechnical specialist.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText58.Wrap( -1 )
        self.staticText58.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText58, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer98, 0, wx.EXPAND, 5 )
        
        bSizer100 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText59 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.4", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText59.Wrap( -1 )
        self.staticText59.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText59, 0, wx.ALL, 5 )
        
        
        bSizer100.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText107 = wx.StaticText( self, wx.ID_ANY, u"Tool and Equipment", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText107.Wrap( -1 )
        self.staticText107.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText107, 0, wx.ALL, 5 )
        
        
        bSizer100.Add( ( 165, 0), 0, 0, 5 )
        
        self.staticText60 = wx.StaticText( self, wx.ID_ANY, u"The facilities of the company involved\nin the remanufacturing process shall have\nappropriate tools and equipment.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText60.Wrap( -1 )
        self.staticText60.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText60, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer100, 1, wx.EXPAND, 5 )
        
        bSizer101 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText61 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.1.5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText61.Wrap( -1 )
        self.staticText61.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer101.Add( self.staticText61, 0, wx.ALL, 5 )
        
        
        bSizer101.Add( ( 60, 0), 0, 0, 5 )
        
        self.m_staticText108 = wx.StaticText( self, wx.ID_ANY, u"Statutory and regulatory\nrequirements and \nindustry best practices", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText108.Wrap( -1 )
        self.m_staticText108.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer101.Add( self.m_staticText108, 0, wx.ALL, 5 )
        
        
        bSizer101.Add( ( 140, 0), 0, 0, 5 )
        
        self.staticText62 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing shall be performed in     \naccordance to the relevant statutory\nand regulatory requirements and\nindustry best practices.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText62.Wrap( -1 )
        self.staticText62.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer101.Add( self.staticText62, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer101, 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer104, 1, wx.EXPAND, 5 )
        
        bSizer105 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer106 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox93 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox93.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox93, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox94 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox94.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox94, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox95 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox95.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox95, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        
        bSizer105.Add( bSizer106, 1, wx.EXPAND, 5 )
        
        bSizer107 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox96 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox96.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox96, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox97 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox97.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox97, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox98 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox98.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox98, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        
        bSizer105.Add( bSizer107, 1, wx.EXPAND, 5 )
        
        bSizer108 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox99 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox99.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox99, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox100 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox100.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox100, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox101 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox101.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox101, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        
        bSizer105.Add( bSizer108, 1, wx.EXPAND, 5 )
        
        bSizer114 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox102 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox102.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox102, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox103 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox103, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox104 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox104.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox104, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        
        bSizer105.Add( bSizer114, 1, wx.EXPAND, 5 )
        
        bSizer115 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox105 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox105, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox106 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox106, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        self.checkBox107 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox107.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox107, 0, wx.ALL|wx.ALIGN_CENTER_VERTICAL, 5 )
        
        
        bSizer105.Add( bSizer115, 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer105, 1, wx.EXPAND, 5 )
        
        bSizer301 = wx.BoxSizer( wx.VERTICAL )
        
        
        bSizer301.Add( ( 0, 30), 0, 0, 5 )
        
        self.textCtrl54 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer301.Add( self.textCtrl54, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer301.Add( ( 0, 50), 0, 0, 5 )
        
        self.textCtrl55 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer301.Add( self.textCtrl55, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer301.Add( ( 0, 55), 0, 0, 5 )
        
        self.textCtrl56 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer301.Add( self.textCtrl56, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer301.Add( ( 0, 50), 0, 0, 5 )
        
        self.textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer301.Add( self.textCtrl57, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer301.Add( ( 0, 50), 0, 0, 5 )
        
        self.textCtrl58 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer301.Add( self.textCtrl58, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer301.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer301, 1, wx.EXPAND, 5 )
        
        
        bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )
        
        bSizer103 = wx.BoxSizer( wx.HORIZONTAL )
        
        
        bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button33, 0, wx.ALL, 5 )
        
        self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button34, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )
        
        
        self.SetSizer( bSizer95 )
        self.Layout()
        
        self.Centre( wx.BOTH )
        
        # Connect Events
        self.checkBox93.Bind( wx.EVT_CHECKBOX, self.check1 )
        self.checkBox94.Bind( wx.EVT_CHECKBOX, self.check2 )
        self.checkBox95.Bind( wx.EVT_CHECKBOX, self.check3 )
        self.checkBox96.Bind( wx.EVT_CHECKBOX, self.check4 )
        self.checkBox97.Bind( wx.EVT_CHECKBOX, self.check5 )
        self.checkBox98.Bind( wx.EVT_CHECKBOX, self.check6 )
        self.checkBox99.Bind( wx.EVT_CHECKBOX, self.check7 )
        self.checkBox100.Bind( wx.EVT_CHECKBOX, self.check8 )
        self.checkBox101.Bind( wx.EVT_CHECKBOX, self.check9 )
        self.checkBox102.Bind( wx.EVT_CHECKBOX, self.check10 )
        self.checkBox103.Bind( wx.EVT_CHECKBOX, self.check11 )
        self.checkBox104.Bind( wx.EVT_CHECKBOX, self.check12 )
        self.checkBox105.Bind( wx.EVT_CHECKBOX, self.check13 )
        self.checkBox106.Bind( wx.EVT_CHECKBOX, self.check14 )
        self.checkBox107.Bind( wx.EVT_CHECKBOX, self.check15 )
        self.textCtrl54.Bind( wx.EVT_TEXT, self.Remark1 )
        self.textCtrl55.Bind( wx.EVT_TEXT, self.Remark2 )
        self.textCtrl56.Bind( wx.EVT_TEXT, self.Remark3 )
        self.textCtrl57.Bind( wx.EVT_TEXT, self.Remark4 )
        self.textCtrl58.Bind( wx.EVT_TEXT, self.Remark5 )
        self.button33.Bind( wx.EVT_BUTTON, self.Reset )
        self.button34.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
        pass
    
    
    # Virtual event handlers, overide them in your derived class
    def check1( self, event ):
            global a12
            box99 = self.checkBox93.GetValue()
            if box99 == True:
                a12 = 1
            else:
                a12 = 0

    def check2( self, event ):
            global b12
            box100 = self.checkBox94.GetValue()
            if box100 == True:
                b12 = 1
            else:
                b12 = 0
            

    def check3( self, event ):
            global c12
            box101 = self.checkBox95.GetValue()
            if box101 == True:
                c12 = 1
            else:
                c12 = 0

    def check4( self, event ):
            global d12
            box102 = self.checkBox96.GetValue()
            if box102 == True:
                d12 = 1
            else:
                d12 = 0

    def check5( self, event ):
            global e12
            box103 = self.checkBox97.GetValue()
            if box103 == True:
                e12 = 1
            else:
                e12 = 0

    def check6( self, event ):
            global f12
            box104 = self.checkBox98.GetValue()
            if box104 == True:
                f12 = 1
            else:
                f12 = 0

    def check7( self, event ):
            global g12
            box105 = self.checkBox99.GetValue()
            if box105 == True:
                g12 = 1
            else:
                g12 = 0

    def check8( self, event ):
            global h12
            box106 = self.checkBox100.GetValue()
            if box106 == True:
                h12 = 1
            else:
                h12 = 0

    def check9( self, event ):
            global i12
            box107 = self.checkBox101.GetValue()
            if box107 == True:
                i12 = 1
            else:
                i12 = 0

    def check10( self, event ):
            global j12
            box108 = self.checkBox102.GetValue()
            if box108 == True:
                j12 = 1
            else:
                j12 = 0

    def check11( self, event ):
            global k12
            box109 = self.checkBox103.GetValue()
            if box109 == True:
                k12 = 1
            else:
                k12 = 0

    def check12( self, event ):
            global l12
            box110 = self.checkBox104.GetValue()
            if box110 == True:
                l12 = 1
            else:
                l12 = 0

    def check13( self, event ):
            global m12
            box111 = self.checkBox105.GetValue()

            if box111 == True:
                m12 = 1
            else:
                m12 = 0

    def check14( self, event ):
            global n12
            box112 = self.checkBox106.GetValue()

            if box112 == True:
                n12 = 1
            else:
                n12 = 0

    def check15( self, event ):
            global o12
            box113 = self.checkBox107.GetValue()

            if box113 == True:
                o12 = 1
            else:
                o12 = 0
    
    def Remark1( self, event ):
            global pear1
            global aa
            pear1 = self.textCtrl54.GetValue()
            aa = 1
    
    def Remark2( self, event ):
            global pear2
            global ba
            pear2 = self.textCtrl55.GetValue()
            ba = 1
    
    def Remark3( self, event ):
            global pear3
            global ca
            pear3 = self.textCtrl56.GetValue()
            ca = 1
    
    def Remark4( self, event ):
            global pear4
            global da
            pear4 = self.textCtrl57.GetValue()
            da = 1
    
    def Remark5( self, event ):
            global pear5
            global ea
            pear5 = self.textCtrl58.GetValue()
            ea = 1
    
    def Reset( self, event ):
            self.checkBox93.SetValue(False)
            self.checkBox94.SetValue(False)
            self.checkBox95.SetValue(False)
            self.checkBox96.SetValue(False)
            self.checkBox97.SetValue(False)
            self.checkBox98.SetValue(False)
            self.checkBox99.SetValue(False)
            self.checkBox100.SetValue(False)
            self.checkBox101.SetValue(False)
            self.checkBox102.SetValue(False)
            self.checkBox103.SetValue(False)
            self.checkBox104.SetValue(False)
            self.checkBox105.SetValue(False)
            self.checkBox106.SetValue(False)
            self.checkBox107.SetValue(False)
            self.textCtrl54.SetValue("")
            self.textCtrl55.SetValue("")
            self.textCtrl56.SetValue("")
            self.textCtrl57.SetValue("")
            self.textCtrl58.SetValue("")
            
            DataReset = Reset(parent = self.button33)
            SoundReset = soundReset(parent = self.button33)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
    
    def Save( self, event ):
            
            global mop1
            global mop2
            global mop3
            global mop4
            global mop5
            
            global b1x
            global b2x
            global b3x
            global b4x
            global b5x
            global b6x
            
            if aa == 1:
                mop1 = pear1
            else:
                mop1 = ""
            
            if ba == 1:
                mop2 = pear2
            else:
                mop2 = ""
                
            if ca == 1:
                mop3 = pear3
            else:
                mop3 = ""
                
            if da == 1:
                mop4 = pear4
            else:
                mop4 = ""
                
            if ea == 1:
                mop5 = pear5
            else:
                mop5 = ""

            
            if a12 == 1:
                b1x = "Comply"
            else:
                pass
                
            if b12 == 1:
                b1x = "Observation"
            else:
                pass
            
            if c12 == 1:
                b1x = "Not Comply"
            else:
                pass
                
            if d12 == 1:
                b2x = "Comply"
            else:
                pass
                
            if e12 == 1:
                b2x = "Observation"
            else:
                pass
                
            if f12 == 1:
                b2x = "Not Comply"
            else:
                pass
                
            if g12 == 1:
                b3x = "Comply"
            else:
                pass
                
            if h12 == 1:
                b3x = "Observation"
            else:
                pass
                
            if i12 == 1:
                b3x = "Not Comply"
            else:
                pass
                
            if j12 == 1:
                b4x = "Comply"
            else:
                pass
                
            if k12 == 1:
                b4x = "Observation"
            else:
                pass
                
            if l12 == 1:
                b4x = "Not Comply"
            else:
                pass
                
            if m12 == 1:
                b5x = "Comply"
            else:
                pass
                
            if n12 == 1:
                b6x = "Observation"
            else:
                pass
                
            if o12 == 1:
                b6x = "Not Comply"
            else:
                pass
                
            SavingData = DataSaved(parent = self.button34)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button34)
            AcceptSound.Accept.Play()

###########################################################################
## Class Clause73
###########################################################################
class Clause73 ( wx.Frame ):
    
    global ab
    global bb
    global cb
    global db
    global eb
    global fb
    global gb
    global hb
    global ib
    global jb
    
    global p12
    global q12
    global r12
    global s12
    global t12
    global u12
    global v12
    global w12
    global x12
    global y12
    global z12
    global a13
    global b13
    global c13
    global d13
    global e13
    global f13
    global g13
    global h13
    global i13
    global j13
    global k13
    global l13
    global m13
    global n13
    global o13
    global p13
    global q13
    global r13
    global s13
    
    global mop6
    global mop7
    global mop8
    global mop9
    global mop10
    global mop11
    global mop12
    global mop13
    global mop14
    global mop15
    
    global b7x
    global b8x
    global b9x
    global b10x
    global b11x
    global b12x
    global b13x
    global b14x
    global b15x
    global b16x
    
    p12 = 0
    q12 = 0
    r12 = 0
    s12 = 0
    t12 = 0
    u12 = 0
    v12 = 0
    w12 = 0
    x12 = 0
    y12 = 0
    z12 = 0
    a13 = 0
    b13 = 0
    c13 = 0
    d13 = 0
    e13 = 0
    f13 = 0
    g13 = 0
    h13 = 0
    i13 = 0
    j13 = 0
    k13 = 0
    l13 = 0
    m13 = 0
    n13 = 0
    o13 = 0
    p13 = 0
    q13 = 0
    r13 = 0
    s13 = 0
    
    ab = 0
    bb = 0
    cb = 0
    db = 0
    eb = 0
    fb = 0
    gb = 0
    hb = 0
    ib = 0
    jb = 0
    
    b7x = ""
    b8x = ""
    b9x = ""
    b10x = ""
    b11x = ""
    b12x = ""
    b13x = ""
    b14x = ""
    b15x = ""
    b16x = ""
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Clause 7.3", pos = wx.DefaultPosition, size = wx.Size( 1500,983 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
        
        self.SetSizeHintsSz( wx.Size(1500,983), wx.Size(1500,983) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
        
        bSizer95 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer117 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText64.Wrap( -1 )
        self.staticText64.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 360, 0), 0, 0, 5 )
        
        self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText65.Wrap( -1 )
        self.staticText65.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 578, 0), 0, 0, 5 )
        
        self.m_staticText271 = wx.StaticText( self, wx.ID_ANY, u"Remark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText271.Wrap( -1 )
        self.m_staticText271.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.m_staticText271, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer117, 1, wx.EXPAND, 5 )
        
        bSizer116 = wx.BoxSizer( wx.HORIZONTAL )
        
        bSizer104 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer96 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText49.Wrap( -1 )
        self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 78, 0), 0, 0, 5 )
        
        self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"The remanufacturing process\ninvolves:", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText103.Wrap( -1 )
        self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 90, 0), 1, 0, 5 )
        
        self.staticText50 = wx.StaticText( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText50.Wrap( -1 )
        bSizer96.Add( self.staticText50, 1, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer104.Add( bSizer96, 0, 0, 5 )
        
        bSizer97 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(a)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText51.Wrap( -1 )
        self.staticText51.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText51, 0, wx.ALL, 5 )
        
        
        bSizer97.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText105 = wx.StaticText( self, wx.ID_ANY, u"Core management, core \nsorting, dismantling", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText105.Wrap( -1 )
        self.staticText105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText105, 0, wx.ALL, 5 )
        
        
        bSizer97.Add( ( 104, 0), 0, 0, 5 )
        
        self.staticText52 = wx.StaticText( self, wx.ID_ANY, u"The core management process is fully \nimplemented in terms of traceability\nrecord of the parts and components.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText52.Wrap( -1 )
        self.staticText52.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText52, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer97, 1, wx.EXPAND, 5 )
        
        bSizer98 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText57 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(b)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText57.Wrap( -1 )
        self.staticText57.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText57, 0, wx.ALL, 5 )
        
        
        bSizer98.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText106 = wx.StaticText( self, wx.ID_ANY, u"Cleaning of all internal \nand external components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText106.Wrap( -1 )
        self.staticText106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText106, 0, wx.ALL, 5 )
        
        
        bSizer98.Add( ( 104, 0), 0, 0, 5 )
        
        self.staticText58 = wx.StaticText( self, wx.ID_ANY, u"Cleaning is required for used\nparts and components to ensure\nthat they are free from any impurities\nand hazardous materials.\n", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText58.Wrap( -1 )
        self.staticText58.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText58, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer98, 0, wx.EXPAND, 5 )
        
        bSizer100 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText59 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(c)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText59.Wrap( -1 )
        self.staticText59.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText59, 0, wx.ALL, 5 )
        
        
        bSizer100.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText107 = wx.StaticText( self, wx.ID_ANY, u"Replacement of parts and\ncomponent", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText107.Wrap( -1 )
        self.staticText107.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText107, 0, wx.ALL, 5 )
        
        
        bSizer100.Add( ( 100, 0), 0, 0, 5 )
        
        self.staticText60 = wx.StaticText( self, wx.ID_ANY, u"All missing parts, restoration of all\nimpaired, defective or substantially\nworn parts to a sound condition or\nreplacement thereof", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText60.Wrap( -1 )
        self.staticText60.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText60, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer100, 1, wx.EXPAND, 5 )
        
        bSizer101 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText61 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(d)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText61.Wrap( -1 )
        self.staticText61.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer101.Add( self.staticText61, 0, wx.ALL, 5 )
        
        
        bSizer101.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText108 = wx.StaticText( self, wx.ID_ANY, u"Reworking and machining ", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText108.Wrap( -1 )
        self.staticText108.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer101.Add( self.staticText108, 0, wx.ALL, 5 )
        
        
        bSizer101.Add( ( 95, 0), 0, 0, 5 )
        
        self.staticText62 = wx.StaticText( self, wx.ID_ANY, u"Performing such other operations as\nare necessary to put the\npart in original\nworking condition or better", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText62.Wrap( -1 )
        self.staticText62.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer101.Add( self.staticText62, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer101, 1, wx.EXPAND, 5 )
        
        bSizer218 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText160 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(e)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText160.Wrap( -1 )
        self.staticText160.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer218.Add( self.staticText160, 0, wx.ALL, 5 )
        
        
        bSizer218.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText162 = wx.StaticText( self, wx.ID_ANY, u"Component assembly\n", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText162.Wrap( -1 )
        self.staticText162.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer218.Add( self.staticText162, 0, wx.ALL, 5 )
        
        
        bSizer218.Add( ( 125, 0), 0, 0, 5 )
        
        self.staticText161 = wx.StaticText( self, wx.ID_ANY, u"Assembly done according to\nmanufacturer specification\n", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText161.Wrap( -1 )
        self.staticText161.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer218.Add( self.staticText161, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer104.Add( bSizer218, 1, wx.EXPAND, 5 )
        
        bSizer1011 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText611 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.1(f) \nand\nClause 7.3.6", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText611.Wrap( -1 )
        self.staticText611.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1011.Add( self.staticText611, 0, wx.ALL, 5 )
        
        
        bSizer1011.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText1081 = wx.StaticText( self, wx.ID_ANY, u"Final testing of each\nremanufactured part\n", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText1081.Wrap( -1 )
        self.staticText1081.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1011.Add( self.staticText1081, 0, wx.ALL, 5 )
        
        
        bSizer1011.Add( ( 135, 0), 0, 0, 5 )
        
        self.staticText621 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing shall be performed in     \naccordance to the relevant statutory\nand regulatory requirements and\nindustry best practices.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText621.Wrap( -1 )
        self.staticText621.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1011.Add( self.staticText621, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer1011, 1, wx.EXPAND, 5 )
        
        bSizer2181 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText1601 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.2", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText1601.Wrap( -1 )
        self.staticText1601.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer2181.Add( self.staticText1601, 0, wx.ALL, 5 )
        
        
        bSizer2181.Add( ( 75, 0), 0, 0, 5 )
        
        self.staticText1621 = wx.StaticText( self, wx.ID_ANY, u"Identification mark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText1621.Wrap( -1 )
        self.staticText1621.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer2181.Add( self.staticText1621, 0, wx.ALL, 5 )
        
        
        bSizer2181.Add( ( 153, 0), 0, 0, 5 )
        
        self.staticText1611 = wx.StaticText( self, wx.ID_ANY, u"Any remanufactured parts and components that\nare regulated by relevant authorities shall be\naffixed with an identification mark. This\nidentification mark shall identify the source\nof the used parts and components. A detailed\nrecord of the identification mark shall be maintained", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText1611.Wrap( -1 )
        self.staticText1611.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer2181.Add( self.staticText1611, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer2181, 1, wx.EXPAND, 5 )
        
        bSizer971 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText511 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.3", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText511.Wrap( -1 )
        self.staticText511.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer971.Add( self.staticText511, 0, wx.ALL, 5 )
        
        
        bSizer971.Add( ( 75, 0), 0, 0, 5 )
        
        self.staticText1051 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing permanent\nmark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText1051.Wrap( -1 )
        self.staticText1051.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer971.Add( self.staticText1051, 0, wx.ALL, 5 )
        
        
        bSizer971.Add( ( 85, 0), 0, 0, 5 )
        
        self.staticText521 = wx.StaticText( self, wx.ID_ANY, u"All remanufactured part shall\nhave a non-removable and\npermanent marking to identify\nit as a remanufactured product", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText521.Wrap( -1 )
        self.staticText521.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer971.Add( self.staticText521, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer971, 1, wx.EXPAND, 5 )
        
        bSizer972 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText512 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.4", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText512.Wrap( -1 )
        self.staticText512.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer972.Add( self.staticText512, 0, wx.ALL, 5 )
        
        
        bSizer972.Add( ( 73, 0), 0, 0, 5 )
        
        self.staticText1052 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing records", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText1052.Wrap( -1 )
        self.staticText1052.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer972.Add( self.staticText1052, 0, wx.ALL, 5 )
        
        
        bSizer972.Add( ( 107, 0), 0, 0, 5 )
        
        self.staticText522 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing process shall be conducted\nin accordance to an established industrialised\nprocess, guided by industry best practices\nand shall be fully documented and recorded", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText522.Wrap( -1 )
        self.staticText522.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer972.Add( self.staticText522, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer972, 1, wx.EXPAND, 5 )
        
        bSizer973 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText513 = wx.StaticText( self, wx.ID_ANY, u"Clause 7.3.5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText513.Wrap( -1 )
        self.staticText513.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer973.Add( self.staticText513, 0, wx.ALL, 5 )
        
        
        bSizer973.Add( ( 75, 0), 0, 0, 5 )
        
        self.staticText1053 = wx.StaticText( self, wx.ID_ANY, u"Authorised from person\nin charge", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText1053.Wrap( -1 )
        self.staticText1053.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer973.Add( self.staticText1053, 0, wx.ALL, 5 )
        
        
        bSizer973.Add( ( 117, 0), 0, 0, 5 )
        
        self.staticText523 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing of the\nused parts and components shall be to \ndemonstrate that  it has the legal right to\ntransfer their ownership to another party", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText523.Wrap( -1 )
        self.staticText523.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer973.Add( self.staticText523, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer973, 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer104, 1, wx.EXPAND, 5 )
        
        bSizer105 = wx.BoxSizer( wx.VERTICAL )
        
        
        bSizer105.Add( ( 0, 50), 0, 0, 5 )
        
        bSizer106 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox108 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox108.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox108, 0, wx.ALL, 5 )
        
        self.checkBox109 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox109.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox109, 0, wx.ALL, 5 )
        
        self.checkBox110 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox110.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox110, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer106, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 60), 0, 0, 5 )
        
        bSizer107 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox111 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox111.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox111, 0, wx.ALL, 5 )
        
        self.checkBox112 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox112.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox112, 0, wx.ALL, 5 )
        
        self.checkBox113 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox113.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox113, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer107, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 60), 0, 0, 5 )
        
        bSizer108 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox114 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox114.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox114, 0, wx.ALL, 5 )
        
        self.checkBox115 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox115.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox115, 0, wx.ALL, 5 )
        
        self.checkBox116 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox116.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox116, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer108, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 50), 0, 0, 5 )
        
        bSizer114 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox117 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox117.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox117, 0, wx.ALL, 5 )
        
        self.checkBox118 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox118.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox118, 0, wx.ALL, 5 )
        
        self.checkBox119 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox119.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox119, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer114, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 50), 0, 0, 5 )
        
        bSizer115 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox120 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox120.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox120, 0, wx.ALL, 5 )
        
        self.checkBox121 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox121.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox121, 0, wx.ALL, 5 )
        
        self.checkBox122 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox122.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox122, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer115, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 50), 0, 0, 5 )
        
        bSizer1061 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox123 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox123.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1061.Add( self.checkBox123, 0, wx.ALL, 5 )
        
        self.checkBox124 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox124.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1061.Add( self.checkBox124, 0, wx.ALL, 5 )
        
        self.checkBox125 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox125.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1061.Add( self.checkBox125, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer1061, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 60), 0, 0, 5 )
        
        bSizer1062 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox126 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox126.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1062.Add( self.checkBox126, 0, wx.ALL, 5 )
        
        self.checkBox127 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox127.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1062.Add( self.checkBox127, 0, wx.ALL, 5 )
        
        self.checkBox128 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox128.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1062.Add( self.checkBox128, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer1062, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 80), 0, 0, 5 )
        
        bSizer1063 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox129 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox129.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1063.Add( self.checkBox129, 0, wx.ALL, 5 )
        
        self.checkBox130 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox130.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1063.Add( self.checkBox130, 0, wx.ALL, 5 )
        
        self.checkBox131 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox131.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1063.Add( self.checkBox131, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer1063, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 50), 0, 0, 5 )
        
        bSizer1064 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox132 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox132.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1064.Add( self.checkBox132, 0, wx.ALL, 5 )
        
        self.checkBox133 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox133.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1064.Add( self.checkBox133, 0, wx.ALL, 5 )
        
        self.checkBox134 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox134.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1064.Add( self.checkBox134, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer1064, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 60), 0, 0, 5 )
        
        bSizer1065 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox135 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox135.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1065.Add( self.checkBox135, 0, wx.ALL, 5 )
        
        self.checkBox136 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox136.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1065.Add( self.checkBox136, 0, wx.ALL, 5 )
        
        self.checkBox137 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox137.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1065.Add( self.checkBox137, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer1065, 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer105, 1, 0, 5 )
        
        bSizer302 = wx.BoxSizer( wx.VERTICAL )
        
        
        bSizer302.Add( ( 0, 50), 0, wx.EXPAND, 5 )
        
        self.textCtrl59 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl59, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 55), 0, wx.EXPAND, 5 )
        
        self.textCtrl60 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl60, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 55), 0, wx.EXPAND, 5 )
        
        self.textCtrl61 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl61, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 45), 0, wx.EXPAND, 5 )
        
        self.textCtrl62 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl62, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 45), 0, wx.EXPAND, 5 )
        
        self.textCtrl63 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl63, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 40), 0, wx.EXPAND, 5 )
        
        self.textCtrl64 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl64, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 50), 0, wx.EXPAND, 5 )
        
        self.textCtrl65 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl65, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 80), 0, wx.EXPAND, 5 )
        
        self.textCtrl66 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl66, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 40), 0, wx.EXPAND, 5 )
        
        self.textCtrl67 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl67, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 55), 0, wx.EXPAND, 5 )
        
        self.textCtrl68 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer302.Add( self.textCtrl68, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer302.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer302, 1, wx.EXPAND, 5 )
        
        
        bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )
        
        bSizer103 = wx.BoxSizer( wx.HORIZONTAL )
        
        
        bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button33, 0, wx.ALL, 5 )
        
        self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button34, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )
        
        
        self.SetSizer( bSizer95 )
        self.Layout()
        
        self.Centre( wx.BOTH )
        
        # Connect Events
        self.checkBox108.Bind( wx.EVT_CHECKBOX, self.check1 )
        self.checkBox109.Bind( wx.EVT_CHECKBOX, self.check2 )
        self.checkBox110.Bind( wx.EVT_CHECKBOX, self.check3 )
        self.checkBox111.Bind( wx.EVT_CHECKBOX, self.check4 )
        self.checkBox112.Bind( wx.EVT_CHECKBOX, self.check5 )
        self.checkBox113.Bind( wx.EVT_CHECKBOX, self.check6 )
        self.checkBox114.Bind( wx.EVT_CHECKBOX, self.check7 )
        self.checkBox115.Bind( wx.EVT_CHECKBOX, self.check8 )
        self.checkBox116.Bind( wx.EVT_CHECKBOX, self.check9 )
        self.checkBox117.Bind( wx.EVT_CHECKBOX, self.check10 )
        self.checkBox118.Bind( wx.EVT_CHECKBOX, self.check11 )
        self.checkBox119.Bind( wx.EVT_CHECKBOX, self.check12 )
        self.checkBox120.Bind( wx.EVT_CHECKBOX, self.check13 )
        self.checkBox121.Bind( wx.EVT_CHECKBOX, self.check14 )
        self.checkBox122.Bind( wx.EVT_CHECKBOX, self.check15 )
        self.checkBox123.Bind( wx.EVT_CHECKBOX, self.check16 )
        self.checkBox124.Bind( wx.EVT_CHECKBOX, self.check17 )
        self.checkBox125.Bind( wx.EVT_CHECKBOX, self.check18 )
        self.checkBox126.Bind( wx.EVT_CHECKBOX, self.check19 )
        self.checkBox127.Bind( wx.EVT_CHECKBOX, self.check20 )
        self.checkBox128.Bind( wx.EVT_CHECKBOX, self.check21 )
        self.checkBox129.Bind( wx.EVT_CHECKBOX, self.check22 )
        self.checkBox130.Bind( wx.EVT_CHECKBOX, self.check23 )
        self.checkBox131.Bind( wx.EVT_CHECKBOX, self.check24 )
        self.checkBox132.Bind( wx.EVT_CHECKBOX, self.check25 )
        self.checkBox133.Bind( wx.EVT_CHECKBOX, self.check26 )
        self.checkBox134.Bind( wx.EVT_CHECKBOX, self.check27 )
        self.checkBox135.Bind( wx.EVT_CHECKBOX, self.check28 )
        self.checkBox136.Bind( wx.EVT_CHECKBOX, self.check29 )
        self.checkBox137.Bind( wx.EVT_CHECKBOX, self.check30 )
        self.textCtrl59.Bind( wx.EVT_TEXT, self.Remark1 )
        self.textCtrl60.Bind( wx.EVT_TEXT, self.Remark2 )
        self.textCtrl61.Bind( wx.EVT_TEXT, self.Remark3 )
        self.textCtrl62.Bind( wx.EVT_TEXT, self.Remark4 )
        self.textCtrl63.Bind( wx.EVT_TEXT, self.Remark5 )
        self.textCtrl64.Bind( wx.EVT_TEXT, self.Remark6 )
        self.textCtrl65.Bind( wx.EVT_TEXT, self.Remark7 )
        self.textCtrl66.Bind( wx.EVT_TEXT, self.Remark8 )
        self.textCtrl67.Bind( wx.EVT_TEXT, self.Remark9 )
        self.textCtrl68.Bind( wx.EVT_TEXT, self.Remark10 )
        self.button33.Bind( wx.EVT_BUTTON, self.Reset )
        self.button34.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
        pass
    
    
    # Virtual event handlers, overide them in your derived class
    def check1( self, event ):
            global p12
            box114 = self.checkBox108.GetValue()

            if box114 == True:
                p12 = 1
            else:
                p12 = 0

    def check2( self, event ):
            global q12
            box115 = self.checkBox109.GetValue()

            if box115 == True:
                q12 = 1
            else:
                q12 = 0

    def check3( self, event ):
            global r12
            box116 = self.checkBox110.GetValue()

            if box116 == True:
                r12 = 1
            else:
                r12 = 0

    def check4( self, event ):
            global s12
            box117 = self.checkBox111.GetValue()

            if box117 == True:
                s12 = 1
            else:
                s12 = 0

    def check5( self, event ):
            global t12
            box118 = self.checkBox112.GetValue()

            if box118 == True:
                t12 = 1
            else:
                t12 = 0

    def check6( self, event ):
            global u12
            box119 = self.checkBox113.GetValue()

            if box119 == True:
                u12 = 1
            else:
                u12 = 0

    def check7( self, event ):
            global v12
            box120 = self.checkBox114.GetValue()

            if box120 == True:
                v12 = 1
            else:
                v12 = 0

    def check8( self, event ):
            global w12
            box121 = self.checkBox115.GetValue()

            if box121 == True:
                w12 = 1
            else:
                w12 = 0

    def check9( self, event ):
            global x12
            box122 = self.checkBox116.GetValue()

            if box122 == True:
                x12 = 1
            else:
                x12 = 0

    def check10( self, event ):
            global y12
            box123 = self.checkBox117.GetValue()

            if box123 == True:
                y12 = 1
            else:
                y12 = 0

    def check11( self, event ):
            global z12
            box124 = self.checkBox118.GetValue()

            if box124 == True:
                z12 = 1
            else:
                z12 = 0

    def check12( self, event ):
            global a13
            box125 = self.checkBox119.GetValue()

            if box125 == True:
                a13 = 1
            else:
                a13 = 0

    def check13( self, event ):
            global b13
            box126 = self.checkBox120.GetValue()

            if box126 == True:
                b13 = 1
            else:
                b13 = 0

    def check14( self, event ):
            global c13
            box127 = self.checkBox121.GetValue()
            
            if box127 == True:
                c13 = 1
            else:
                c13 = 0

    def check15( self, event ):
            global d13
            box128 = self.checkBox122.GetValue()
            
            if box128 == True:
                d13 = 1
            else:
                d13 = 0

    def check16( self, event ):
            global e13
            box129 = self.checkBox123.GetValue()
            
            if box129 == True:
                e13 = 1
            else:
                e13 = 0

    def check17( self, event ):
            global f13
            box130 = self.checkBox124.GetValue()
            
            if box130 == True:
                f13 = 1
            else:
                f13 = 1

    def check18( self, event ):
            global g13
            box131 = self.checkBox125.GetValue()
            
            if box131 == True:
                g13 = 1
            else:
                g13 = 0

    def check19( self, event ):
            global h13
            box132 = self.checkBox126.GetValue()
            
            if box132 == True:
                h13 = 1
            else:
                h13 = 0

    def check20( self, event ):
            global i13
            box133 = self.checkBox127.GetValue()
            
            if box133 == True:
                i13 = 1
            else:
                i13 = 0

    def check21( self, event ):
            global j13
            box134 = self.checkBox128.GetValue()
            
            if box134 == True:
                j13 = 1
            else:
                j13 = 0

    def check22( self, event ):
            global k13
            box135 = self.checkBox129.GetValue()
            
            if box135 == True:
                k13 = 1
            else:
                k13 = 0

    def check23( self, event ):
            global l13
            box136 = self.checkBox130.GetValue()
            
            if box136 == True:
                l13 = 1
            else:
                l13 = 0

    def check24( self, event ):
            global m13
            box137 = self.checkBox131.GetValue()
            
            if box137 == True:
                m13 = 1
            else:
                m13 = 0

    def check25( self, event ):
            global n13
            box138 = self.checkBox132.GetValue()
            
            if box138 == True:
                n13 = 1
            else:
                n13 = 0

    def check26( self, event ):
            global o13
            box139 = self.checkBox133.GetValue()
            
            if box139 == True:
                o13 = 1
            else:
                o13 = 0

    def check27( self, event ):
            global p13
            box140 = self.checkBox134.GetValue()
            
            
            if box140 == True:
                p13 = 1
            else:
                p13 = 0

    def check28( self, event ):
            global q13
            box141 = self.checkBox135.GetValue()
            
            if box141 == True:
                q13 = 1
            else:
                q13 = 0

    def check29( self, event ):
            global r13
            box142 = self.checkBox136.GetValue()
            
            if box142 == True:
                r13 = 1
            else:
                r13 = 0

    def check30( self, event ):
            global s13
            box143 = self.checkBox137.GetValue()
            
            if box143 == True:
                s13 = 1
            else:
                s13 = 0
    
    def Remark1( self, event ):
            global pear6
            global ab
            pear6 = self.textCtrl59.GetValue()
            ab = 1
            
    def Remark2( self, event ):
            global pear7
            global bb
            pear7 = self.textCtrl60.GetValue()
            bb = 1
    
    def Remark3( self, event ):
            global pear8
            global cb
            pear8 = self.textCtrl61.GetValue()
            cb = 1
            
    def Remark4( self, event ):
            global pear9
            global db
            pear9 = self.textCtrl62.GetValue()
            db = 1
    
    def Remark5( self, event ):
            global pear10
            global eb
            pear10 = self.textCtrl63.GetValue()
            eb = 1
    
    def Remark6( self, event ):
            global pear11
            global fb
            pear11 = self.textCtrl64.GetValue()
            fb = 1
    
    def Remark7( self, event ):
            global pear12
            global gb
            pear12 = self.textCtrl65.GetValue()
            gb = 1
    
    def Remark8( self, event ):
            global pear13
            global hb
            pear13 = self.textCtrl66.GetValue()
            hb = 1
            
    def Remark9( self, event ):
            global pear14
            global ib
            pear14 = self.textCtrl67.GetValue()
            ib = 1
    
    def Remark10( self, event ):
            global pear15
            global jb
            pear15 = self.textCtrl68.GetValue()
            jb = 1
    
    def Reset( self, event ):
            self.checkBox108.SetValue(False)
            self.checkBox109.SetValue(False)
            self.checkBox110.SetValue(False)
            self.checkBox111.SetValue(False)
            self.checkBox112.SetValue(False)
            self.checkBox113.SetValue(False)
            self.checkBox114.SetValue(False)
            self.checkBox115.SetValue(False)
            self.checkBox116.SetValue(False)
            self.checkBox117.SetValue(False)
            self.checkBox118.SetValue(False)
            self.checkBox119.SetValue(False)
            self.checkBox120.SetValue(False)
            self.checkBox121.SetValue(False)
            self.checkBox122.SetValue(False)
            self.checkBox123.SetValue(False)
            self.checkBox124.SetValue(False)
            self.checkBox125.SetValue(False)
            self.checkBox126.SetValue(False)
            self.checkBox127.SetValue(False)
            self.checkBox128.SetValue(False)
            self.checkBox129.SetValue(False)
            self.checkBox130.SetValue(False)
            self.checkBox131.SetValue(False)
            self.checkBox132.SetValue(False)
            self.checkBox133.SetValue(False)
            self.checkBox134.SetValue(False)
            self.checkBox135.SetValue(False)
            self.checkBox136.SetValue(False)
            self.checkBox137.SetValue(False)
            
            self.textCtrl59.SetValue("")
            self.textCtrl60.SetValue("")
            self.textCtrl61.SetValue("")
            self.textCtrl62.SetValue("")
            self.textCtrl63.SetValue("")
            self.textCtrl64.SetValue("")
            self.textCtrl65.SetValue("")
            self.textCtrl66.SetValue("")
            self.textCtrl67.SetValue("")
            self.textCtrl68.SetValue("")
            
            DataReset = Reset(parent = self.button33)
            SoundReset = soundReset(parent = self.butto33)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
                        

    def Save( self, event ):
            global mop6
            global mop7
            global mop8
            global mop9
            global mop10
            global mop11
            global mop12
            global mop13
            global mop14
            global mop15
            
            global b7x
            global b8x
            global b9x
            global b10x
            global b11x
            global b12x
            global b13x
            global b14x
            global b15x
            global b16x

            
            if ab == 1:
                mop6 = pear6
            else:
                mop6 = ""
                
            if bb == 1:
                mop7 = pear7
            else:
                mop7 = ""
                
            if cb == 1:
                mop8 = pear8
            else:
                mop8 = ""
                
            if db == 1:    
                mop9 = pear9
            else:
                mop9 = ""
                
            if eb == 1:
                mop10 = pear10
            else:
                mop10 = ""
                
            if fb == 1:
                mop11 = pear11
            else:
                mop11 = ""
                
            if gb == 1:
                mop12 = pear12
            else:
                mop12 = ""
            
            if hb == 1:
                mop13 = pear13
            else:
                mop13 = ""
                
            if ib == 1:
                mop14 = pear14
            else:
                mop14 = ""
                
            if jb == 1:
                mop15 = pear15
            else:
                mop15 = ""
            
   
            
            
            if p12 == 1:
                b7x = "Comply"
            else:
                pass
                
            if q12 == 1:
                b7x = "Observation"
            else:
                pass
                
            if r12 == 1:
                b7x = "Not Comply"
            else:
                pass
                
            if s12 == 1:
                b8x = "Comply"
            else:
                pass
                
            if t12 == 1:
                b8x = "Observation"
            else:
                pass
                
            if u12 == 1:
                b8x = "Not Comply"
            else:
                pass
                
            if v12 == 1:
                b9x = "Comply"
            else:
                pass
                
            if w12 == 1:
                b9x = "Observation"
            else:
                pass
                
            if x12 == 1:
                b9x = "Not Comply"
            else:
                pass
                
            if y12 == 1:
                b10x = "Comply"
            else:
                pass
                
            if z12 == 1:
                b10x = "Observation"
            else:
                pass
                
            if a13 == 1:
                b10x = "Not Comply"
            else:
                pass
                
            if b13 == 1:
                b11x = "Comply"
            else:
                pass
                
            if c13 == 1:
                b11x = "Observation"
            else:
                pass
                
            if d13 == 1:
                b11x = "Not Comply"
            else:
                pass
                
            if e13 == 1:
                b12x = "Comply"
            else:
                pass
                
            if f13 == 1:
                b12x = "Observation"
            else:
                pass
                
            if g13 == 1:
                b12x = "Not Comply"
            else:
                pass
                
            if h13 == 1:
                b13x = "Comply"
            else:
                pass
                
            if i13 == 1:
                b13x = "Observation"
            else:
                pass
                
            if j13 == 1:
                b13x = "Not Comply"
            else:
                pass
                
            if k13 == 1:
                b14x = "Comply"
            else:
                pass
                
            if l13 == 1:
                b14x = "Observation"
            else:
                pass
                
            if m13 == 1:
                b14x = "Not Comply"
            else:
                pass
                
            if n13 == 1:
                b15x = "Comply"
            else:
                pass
                
            if o13 == 1:
                b15x = "Observation"
            else:
                pass
                
            if p13 == 1:
                b15x = "Not Comply"
            else:
                pass
                
            if q13 == 1:
                b16x = "Comply"
            else:
                pass
                
            if r13 == 1:
                b16x = "Observation"
            else:
                pass
                
            if s13 == 1:
                b16x = "Not Comply"
            else:
                pass
                
            SavingData = DataSaved(parent = self.button34)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button34)
            AcceptSound.Accept.Play()

###########################################################################
## Class Clause91a
###########################################################################
class Clause91a ( wx.Frame ):
    
    global ac
    global bc
    global cc
    global dc
    global ec
    global fc
    global gc
    
    global t13
    global u13
    global v13
    global w13
    global x13
    global y13
    global z13
    global a14
    global b14
    global c14
    global d14
    global e14
    global f14
    global g14
    global h14
    global i14
    global j14
    global k14
    global l14
    global m14
    global n14
    
    global mop16
    global mop17
    global mop18
    global mop19
    global mop20
    global mop21
    global mop22
    
    global b17x
    global b18x
    global b19x
    global b20x
    global b21x
    global b22x
    global b23x
    
    t13 = 0
    u13 = 0
    v13 = 0
    w13 = 0
    x13 = 0
    y13 = 0
    z13 = 0
    a14 = 0
    b14 = 0
    c14 = 0
    d14 = 0
    e14 = 0
    f14 = 0
    g14 = 0
    h14 = 0
    i14 = 0
    j14 = 0
    k14 = 0
    l14 = 0
    m14 = 0
    n14 = 0
    
    ac = 0
    bc = 0
    cc = 0
    dc = 0
    ec = 0
    fc = 0
    gc = 0
    
    b17x = ""
    b18x = ""
    b19x = ""
    b20x = ""
    b21x = ""
    b22x = ""
    b23x = ""
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Clause 9.1", pos = wx.DefaultPosition, size = wx.Size( 1500,600 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
        
        self.SetSizeHintsSz( wx.Size(1500,600), wx.Size(1500,600) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
        
        bSizer95 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer117 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.1", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText64.Wrap( -1 )
        self.staticText64.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 330, 0), 0, 0, 5 )
        
        self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText65.Wrap( -1 )
        self.staticText65.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 385, 0), 0, 0, 5 )
        
        self.staticText272 = wx.StaticText( self, wx.ID_ANY, u"Remark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText272.Wrap( -1 )
        self.staticText272.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText272, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer117, 0, wx.EXPAND, 5 )
        
        
        bSizer95.Add( ( 0, 20), 0, 0, 5 )
        
        bSizer116 = wx.BoxSizer( wx.HORIZONTAL )
        
        bSizer104 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer96 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText49.Wrap( -1 )
        self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 75, 0), 0, 0, 5 )
        
        self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"Labelling shall include the following\ninformation:\n", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText103.Wrap( -1 )
        self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 90, 0), 1, 0, 5 )
        
        self.staticText50 = wx.StaticText( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText50.Wrap( -1 )
        bSizer96.Add( self.staticText50, 1, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer104.Add( bSizer96, 0, 0, 5 )
        
        bSizer97 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(a)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText51.Wrap( -1 )
        self.staticText51.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText51, 0, wx.ALL, 5 )
        
        
        bSizer97.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText105 = wx.StaticText( self, wx.ID_ANY, u"Company name", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText105.Wrap( -1 )
        self.staticText105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText105, 0, wx.ALL, 5 )
        
        
        bSizer97.Add( ( 100, 0), 0, 0, 5 )
        
        
        bSizer104.Add( bSizer97, 1, wx.EXPAND, 5 )
        
        bSizer98 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText57 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(b)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText57.Wrap( -1 )
        self.staticText57.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText57, 0, wx.ALL, 5 )
        
        
        bSizer98.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText106 = wx.StaticText( self, wx.ID_ANY, u"Parts identification\nnumber, if applicable", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText106.Wrap( -1 )
        self.staticText106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText106, 0, wx.ALL, 5 )
        
        
        bSizer98.Add( ( 95, 0), 0, 0, 5 )
        
        
        bSizer104.Add( bSizer98, 1, wx.EXPAND, 5 )
        
        bSizer100 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText59 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(c)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText59.Wrap( -1 )
        self.staticText59.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText59, 0, wx.ALL, 5 )
        
        
        bSizer100.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText107 = wx.StaticText( self, wx.ID_ANY, u"Make and model", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText107.Wrap( -1 )
        self.staticText107.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText107, 0, wx.ALL, 5 )
        
        
        bSizer100.Add( ( 140, 0), 0, 0, 5 )
        
        self.staticText60 = wx.StaticText( self, wx.ID_ANY, u"Products shall be properly\npackaged, labelled and\nidentified as reused,\nrepaired or remanufactured", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText60.Wrap( -1 )
        self.staticText60.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer100.Add( self.staticText60, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer100, 1, wx.EXPAND, 5 )
        
        bSizer101 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText61 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(d)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText61.Wrap( -1 )
        self.staticText61.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer101.Add( self.staticText61, 0, wx.ALL, 5 )
        
        
        bSizer101.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText108 = wx.StaticText( self, wx.ID_ANY, u"Engine capacity,\nif applicable", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText108.Wrap( -1 )
        self.staticText108.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer101.Add( self.staticText108, 0, wx.ALL, 5 )
        
        
        bSizer101.Add( ( 85, 0), 0, 0, 5 )
        
        
        bSizer104.Add( bSizer101, 1, wx.EXPAND, 5 )
        
        bSizer218 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText160 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(e)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText160.Wrap( -1 )
        self.staticText160.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer218.Add( self.staticText160, 0, wx.ALL, 5 )
        
        
        bSizer218.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText162 = wx.StaticText( self, wx.ID_ANY, u"Classifications according to\nindustry best practices e.g.\nreuse, remanufacturing or repair", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText162.Wrap( -1 )
        self.staticText162.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer218.Add( self.staticText162, 0, wx.ALL, 5 )
        
        
        bSizer218.Add( ( 110, 0), 0, 0, 5 )
        
        
        bSizer104.Add( bSizer218, 1, wx.EXPAND, 5 )
        
        bSizer1011 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText611 = wx.StaticText( self, wx.ID_ANY, u"Clause 9.2(f)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText611.Wrap( -1 )
        self.staticText611.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1011.Add( self.staticText611, 0, wx.ALL, 5 )
        
        
        bSizer1011.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText1081 = wx.StaticText( self, wx.ID_ANY, u"Recovered from local\nvehicles or importe", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText1081.Wrap( -1 )
        self.staticText1081.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1011.Add( self.staticText1081, 0, wx.ALL, 5 )
        
        
        bSizer1011.Add( ( 115, 0), 0, 0, 5 )
        
        
        bSizer104.Add( bSizer1011, 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer104, 1, 0, 5 )
        
        bSizer105 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer106 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox138 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox138.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox138, 0, wx.ALL, 5 )
        
        self.checkBox139 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox139.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox139, 0, wx.ALL, 5 )
        
        self.checkBox140 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox140.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox140, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer106, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 30), 0, 0, 5 )
        
        bSizer107 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox141 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox141.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox141, 0, wx.ALL, 5 )
        
        self.checkBox142 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox142.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox142, 0, wx.ALL, 5 )
        
        self.checkBox143 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox143.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox143, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer107, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 60), 0, 0, 5 )
        
        bSizer108 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox144 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox144.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox144, 0, wx.ALL, 5 )
        
        self.checkBox145 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox145.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox145, 0, wx.ALL, 5 )
        
        self.checkBox146 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox146.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox146, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer108, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 40), 0, 0, 5 )
        
        bSizer114 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox147 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox147.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox147, 0, wx.ALL, 5 )
        
        self.checkBox148 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox148.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox148, 0, wx.ALL, 5 )
        
        self.checkBox149 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox149.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer114.Add( self.checkBox149, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer114, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 50), 0, 0, 5 )
        
        bSizer115 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox150 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox150.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox150, 0, wx.ALL, 5 )
        
        self.checkBox151 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox151.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox151, 0, wx.ALL, 5 )
        
        self.checkBox152 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox152.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer115.Add( self.checkBox152, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer115, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 40), 0, 0, 5 )
        
        bSizer1061 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox153 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox153.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1061.Add( self.checkBox153, 0, wx.ALL, 5 )
        
        self.checkBox154 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox154.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1061.Add( self.checkBox154, 0, wx.ALL, 5 )
        
        self.checkBox155 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox155.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1061.Add( self.checkBox155, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer1061, 1, wx.EXPAND, 5 )
        
        
        bSizer105.Add( ( 0, 40), 0, 0, 5 )
        
        bSizer1062 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox156 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox156.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1062.Add( self.checkBox156, 0, wx.ALL, 5 )
        
        self.checkBox157 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox157.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1062.Add( self.checkBox157, 0, wx.ALL, 5 )
        
        self.checkBox158 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox158.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer1062.Add( self.checkBox158, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer1062, 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer105, 0, 0, 5 )
        
        bSizer303 = wx.BoxSizer( wx.VERTICAL )
        
        
        bSizer303.Add( ( 0, 0), 0, wx.EXPAND, 5 )
        
        self.textCtrl69 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer303.Add( self.textCtrl69, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer303.Add( ( 0, 20), 0, wx.EXPAND, 5 )
        
        self.textCtrl70 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer303.Add( self.textCtrl70, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer303.Add( ( 0, 55), 0, wx.EXPAND, 5 )
        
        self.textCtrl71 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer303.Add( self.textCtrl71, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer303.Add( ( 0, 35), 0, wx.EXPAND, 5 )
        
        self.textCtrl72 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer303.Add( self.textCtrl72, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer303.Add( ( 0, 45), 0, wx.EXPAND, 5 )
        
        self.textCtrl73 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer303.Add( self.textCtrl73, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer303.Add( ( 0, 35), 0, wx.EXPAND, 5 )
        
        self.textCtrl74 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer303.Add( self.textCtrl74, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer303.Add( ( 0, 35), 0, wx.EXPAND, 5 )
        
        self.textCtrl75 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer303.Add( self.textCtrl75, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer303, 1, wx.EXPAND, 5 )
        
        
        bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )
        
        bSizer103 = wx.BoxSizer( wx.HORIZONTAL )
        
        
        bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button33, 0, wx.ALL, 5 )
        
        self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button34, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )
        
        
        self.SetSizer( bSizer95 )
        self.Layout()
        
        self.Centre( wx.BOTH )
        
        # Connect Events
        self.checkBox138.Bind( wx.EVT_CHECKBOX, self.check1 )
        self.checkBox139.Bind( wx.EVT_CHECKBOX, self.check2 )
        self.checkBox140.Bind( wx.EVT_CHECKBOX, self.check3 )
        self.checkBox141.Bind( wx.EVT_CHECKBOX, self.check4 )
        self.checkBox142.Bind( wx.EVT_CHECKBOX, self.check5 )
        self.checkBox143.Bind( wx.EVT_CHECKBOX, self.check6 )
        self.checkBox144.Bind( wx.EVT_CHECKBOX, self.check7 )
        self.checkBox145.Bind( wx.EVT_CHECKBOX, self.check8 )
        self.checkBox146.Bind( wx.EVT_CHECKBOX, self.check9 )
        self.checkBox147.Bind( wx.EVT_CHECKBOX, self.check10 )
        self.checkBox148.Bind( wx.EVT_CHECKBOX, self.check11 )
        self.checkBox149.Bind( wx.EVT_CHECKBOX, self.check12 )
        self.checkBox150.Bind( wx.EVT_CHECKBOX, self.check13 )
        self.checkBox151.Bind( wx.EVT_CHECKBOX, self.check14 )
        self.checkBox152.Bind( wx.EVT_CHECKBOX, self.check15 )
        self.checkBox153.Bind( wx.EVT_CHECKBOX, self.check16 )
        self.checkBox154.Bind( wx.EVT_CHECKBOX, self.check17 )
        self.checkBox155.Bind( wx.EVT_CHECKBOX, self.check18 )
        self.checkBox156.Bind( wx.EVT_CHECKBOX, self.check19 )
        self.checkBox157.Bind( wx.EVT_CHECKBOX, self.check20 )
        self.checkBox158.Bind( wx.EVT_CHECKBOX, self.check21 )
        self.textCtrl69.Bind( wx.EVT_TEXT, self.Remark1 )
        self.textCtrl70.Bind( wx.EVT_TEXT, self.Remark2 )
        self.textCtrl71.Bind( wx.EVT_TEXT, self.Remark3 )
        self.textCtrl72.Bind( wx.EVT_TEXT, self.Remark4 )
        self.textCtrl73.Bind( wx.EVT_TEXT, self.Remark5 )
        self.textCtrl74.Bind( wx.EVT_TEXT, self.Remark6 )
        self.textCtrl75.Bind( wx.EVT_TEXT, self.Remark7 )
        self.button33.Bind( wx.EVT_BUTTON, self.Reset )
        self.button34.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
        pass
    
    
    # Virtual event handlers, overide them in your derived class
    def check1( self, event ):
            global t13
            box144 = self.checkBox138.GetValue()

            if box144 == True:
                t13 = 1
            else:
                t13 = 0

    def check2( self, event ):
            global u13
            box145 = self.checkBox139.GetValue()

            if box145 == True:
                u13 = 1
            else:
                u13 = 0

    def check3( self, event ):
            global v13
            box146 = self.checkBox140.GetValue()

            if box146 == True:
                v13 = 1
            else:
                v13 = 0

    def check4( self, event ):
            global w13
            box147 = self.checkBox141.GetValue()

            if box147 == True:
                w13 = 1
            else:
                w13 = 0

    def check5( self, event ):
            global x13
            box148 = self.checkBox142.GetValue()

            if box148 == True:
                x13 = 1
            else:
                x13 = 0

    def check6( self, event ):
            global y13
            box149 = self.checkBox143.GetValue()

            if box149 == True:
                y13 = 1
            else:
                y13 = 0

    def check7( self, event ):
            global z13
            box150 = self.checkBox144.GetValue()

            if box150 == True:
                z13 = 1
            else:
                z13 = 0

    def check8( self, event ):
            global a14
            box151 = self.checkBox145.GetValue()

            if box151 == True:
                a14 = 1
            else:
                a14 = 0

    def check9( self, event ):
            global b14
            box152 = self.checkBox146.GetValue()

            if box152 == True:
                b14 = 1
            else:
                b14 = 0

    def check10( self, event ):
            global c14
            box153 = self.checkBox147.GetValue()

            if box153 == True:
                c14 = 1
            else:
                c14 = 0

    def check11( self, event ):
            global d14
            box154 = self.checkBox148.GetValue()

            if box154 == True:
                d14 = 1
            else:
                d14 = 0

    def check12( self, event ):
            global e14
            box155 = self.checkBox149.GetValue()

            if box155 == True:
                e14 = 1
            else:
                e14 = 0

    def check13( self, event ):
            global f14
            box156 = self.checkBox150.GetValue()

            if box156 == True:
                f14 = 1
            else:
                f14 = 0

    def check14( self, event ):
            global g14
            box157 = self.checkBox151.GetValue()

            if box157 == True:
                g14 = 1
            else:
                g14 = 0

    def check15( self, event ):
            global h14
            box158 = self.checkBox152.GetValue()

            if box158 == True:
                h14 = 1
            else:
                h14 = 0

    def check16( self, event ):
            global i14
            box159 = self.checkBox153.GetValue()

            if box159 == True:
                i14 = 1
            else:
                i14 = 0

    def check17( self, event ):
            global j14
            box160 = self.checkBox154.GetValue()

            if box160 == True:
                j14 = 1
            else:
                j14 = 0

    def check18( self, event ):
            global k14
            box161 = self.checkBox155.GetValue()

            if box161 == True:
                k14 = 1
            else:
                k14 = 0

    def check19( self, event ):
            global l14
            box162 = self.checkBox156.GetValue()

            if box162 == True:
                l14 = 1
            else:
                l14 = 0

    def check20( self, event ):
            global m14
            box163 = self.checkBox157.GetValue()

            if box163 == True:
                m14 = 1
            else:
                m14 = 0

    def check21( self, event ):
            global n14
            box164 = self.checkBox158.GetValue()

            if box164 == True:
                n14 = 1
            else:
                n14 = 0
    
    def Remark1( self, event ):
            global pear16
            global ac
            pear16 = self.textCtrl69.GetValue()
            ac = 1
    
    def Remark2( self, event ):
            global pear17
            global bc
            pear17 = self.textCtrl70.GetValue()
            bc = 1
    
    def Remark3( self, event ):
            global pear18
            global cc
            pear18 = self.textCtrl71.GetValue()
            cc = 1
    
    def Remark4( self, event ):
            global pear19
            global dc
            pear19 = self.textCtrl72.GetValue()
            dc = 1
    
    def Remark5( self, event ):
            global pear20
            global ec
            pear20 = self.textCtrl73.GetValue()
            ec = 1
    
    def Remark6( self, event ):
            global pear21
            global fc
            pear21 = self.textCtrl74.GetValue()
            fc = 1
    
    def Remark7( self, event ):
            global pear22
            global gc
            pear22 = self.textCtrl75.GetValue()
            gc = 1
    
    def Reset( self, event ):
            self.checkBox138.SetValue(False)
            self.checkBox139.SetValue(False)
            self.checkBox140.SetValue(False)
            self.checkBox141.SetValue(False)
            self.checkBox142.SetValue(False)
            self.checkBox143.SetValue(False)
            self.checkBox144.SetValue(False)
            self.checkBox145.SetValue(False)
            self.checkBox146.SetValue(False)
            self.checkBox147.SetValue(False)
            self.checkBox148.SetValue(False)
            self.checkBox149.SetValue(False)
            self.checkBox150.SetValue(False)
            self.checkBox151.SetValue(False)
            self.checkBox152.SetValue(False)
            self.checkBox153.SetValue(False)
            self.checkBox154.SetValue(False)
            self.checkBox155.SetValue(False)
            self.checkBox156.SetValue(False)
            self.checkBox157.SetValue(False)
            self.checkBox158.SetValue(False)
            self.textCtrl69.SetValue("")
            self.textCtrl70.SetValue("")
            self.textCtrl71.SetValue("")
            self.textCtrl72.SetValue("")
            self.textCtrl73.Setvalue("")
            self.textCtrl74.SetValue("")
            self.textCtrl75.SetValue("")
            
            DataReset = Reset(parent = self.button33)
            SoundReset = soundReset(parent = self.button33)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
            

    def Save( self, event ):
            global mop16
            global mop17
            global mop18
            global mop19
            global mop20
            global mop21
            global mop22
            
            global b17x
            global b18x
            global b19x
            global b20x
            global b21x
            global b22x
            global b23x
            
            if ac == 1:
                mop16 = pear16
            else:
                mop16 = ""
                
            if bc == 1:
                mop17 = pear17
            else:
                mop17 = ""
                
            if cc == 1:
                mop18 = pear18
            else:
                mop18 = ""
                
            if dc == 1:
                mop19 = pear19
            else:
                mop19 = ""
                
            if ec == 1:
                mop20 = pear20
            else:
                mop20 = ""
                
            if fc == 1:
                mop21 = pear21
            else:
                mop21 = ""
                
            if gc == 1:
                mop22 = pear22
            else:
                mop22 = ""
            
            
            if t13 == 1:
                b17x = "Comply"
            else:
                pass
            
            if u13 == 1:
                b17x = "Observation"
            else:
                pass
                
            if v13 == 1:
                b17x = "Not Comply"
            else:
                pass
                
            if w13 == 1:
                b18x = "Comply"
            else:
                pass
                
            if x13 == 1:
                b18x = "Observation"
            else:
                pass
                
            if y13 == 1:
                b18x = "Not Comply"
            else:
                pass
                
            if z13 == 1:
                b19x = "Comply"
            else:
                pass
                
            if a14 == 1:
                b19x = "Observation"
            else:
                pass
                
            if b14 == 1:
                b19x = "Not Comply"
            else:
                pass
                
            if c14 == 1:
                b20x = "Comply"
            else:
                pass
                
            if d14 == 1:
                b20x = "Observation"
            else:
                pass
                
            if e14 == 1:
                b20x = "Not Comply"
            else:
                pass
                
            if f14 == 1:
                b21x = "Comply"
            else:
                pass
                
            if g14 == 1:
                b21x = "Observation"
            else:
                pass
                
            if h14 == 1:
                b21x = "Not Comply"
            else:
                pass
                
            if i14 == 1:
                b22x = "Comply"
            else:
                pass
                
            if j14 == 1:
                b22x = "Observation"
            else:
                pass
                
            if k14 == 1:
                b22x = "Not Comply"
            else:
                pass
                
            if l14 == 1:
                b23x = "Comply"
            else:
                pass
                
            if m14 == 1:
                b23x = "Observation"
            else:
                pass
                
            if n14 == 1:
                b23x = "Not Comply"
            else:
                pass

            SavingData = DataSaved(parent = self.button34)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button34)
            AcceptSound.Accept.Play()
            
            
###########################################################################
## Class Clause10
###########################################################################
class Clause10 ( wx.Frame ):
    global ad
    global bd
    global cd
    
    global o14
    global p14
    global q14
    global r14
    global s14
    global t14
    global u14
    global v14
    global w14
    
    global mop23
    global mop24
    global mop25
    
    global b24x
    global b25x
    global b26x
    
    o14 = 0
    p14 = 0
    q14 = 0
    r14 = 0
    s14 = 0
    t14 = 0
    u14 = 0
    v14 = 0
    w14 = 0
    
    ad = 0
    bd = 0
    cd = 0
    
    b24x = ""
    b25x = ""
    b26x = ""
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Clause 10", pos = wx.DefaultPosition, size = wx.Size( 1500,426 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
        
        self.SetSizeHintsSz( wx.Size(1500,426), wx.Size(1500,426) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
        
        bSizer95 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer117 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 10", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText64.Wrap( -1 )
        self.staticText64.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 375, 0), 0, 0, 5 )
        
        self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText65.Wrap( -1 )
        self.staticText65.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 540, 0), 0, wx.EXPAND, 5 )
        
        self.staticText273 = wx.StaticText( self, wx.ID_ANY, u"Remark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText273.Wrap( -1 )
        self.staticText273.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText273, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer117, 0, wx.EXPAND, 5 )
        
        bSizer116 = wx.BoxSizer( wx.HORIZONTAL )
        
        bSizer104 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer96 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 10(a)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText49.Wrap( -1 )
        self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"Handling remanufacturing of the\nused parts and components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText103.Wrap( -1 )
        self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 90, 0), 1, 0, 5 )
        
        self.staticText50 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing\nof the used parts and components shall \nbe to  demonstrate that  it has the legal \nright to transfer their ownership to \nanother party", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText50.Wrap( -1 )
        self.staticText50.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText50, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer96, 0, 0, 5 )
        
        bSizer97 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText51 = wx.StaticText( self, wx.ID_ANY, u"Clause 10(b)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText51.Wrap( -1 )
        self.staticText51.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText51, 0, wx.ALL, 5 )
        
        
        bSizer97.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText105 = wx.StaticText( self, wx.ID_ANY, u"Provide specified information \nof remanufactured parts\nand components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText105.Wrap( -1 )
        self.staticText105.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText105, 0, wx.ALL, 5 )
        
        
        bSizer97.Add( ( 105, 0), 0, 0, 5 )
        
        self.staticText52 = wx.StaticText( self, wx.ID_ANY, u"The company handling remanufacturing\nof the used parts and components shall \nbe to  demonstrate that  it has the legal \nright to transfer their ownership to \nanother party", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText52.Wrap( -1 )
        self.staticText52.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer97.Add( self.staticText52, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer97, 1, wx.EXPAND, 5 )
        
        bSizer98 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText57 = wx.StaticText( self, wx.ID_ANY, u"Clause 10(c)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText57.Wrap( -1 )
        self.staticText57.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText57, 0, wx.ALL, 5 )
        
        
        bSizer98.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText106 = wx.StaticText( self, wx.ID_ANY, u"Manpower certification ", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText106.Wrap( -1 )
        self.staticText106.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText106, 0, wx.ALL, 5 )
        
        
        bSizer98.Add( ( 147, 0), 0, 0, 5 )
        
        self.staticText58 = wx.StaticText( self, wx.ID_ANY, u"Person dismantling, cleaning, examining,\nremediating, re-assembling, testing and \nhandling of the core and remanufactured\nparts or components shall be of \ntechnical specialist", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText58.Wrap( -1 )
        self.staticText58.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer98.Add( self.staticText58, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer98, 0, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer104, 1, wx.EXPAND, 5 )
        
        bSizer105 = wx.BoxSizer( wx.VERTICAL )
        
        
        bSizer105.Add( ( 0, 20), 0, 0, 5 )
        
        bSizer106 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox159 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox159.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox159, 0, wx.ALL, 5 )
        
        self.checkBox160 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox160.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox160, 0, wx.ALL, 5 )
        
        self.checkBox161 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox161.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox161, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer106, 0, wx.ALIGN_CENTER_HORIZONTAL, 5 )
        
        
        bSizer105.Add( ( 0, 50), 0, 0, 5 )
        
        bSizer107 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox162 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox162.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox162, 0, wx.ALL, 5 )
        
        self.checkBox163 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox163.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox163, 0, wx.ALL, 5 )
        
        self.checkBox164 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox164.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer107.Add( self.checkBox164, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer107, 0, wx.ALIGN_CENTER_HORIZONTAL, 5 )
        
        
        bSizer105.Add( ( 0, 100), 0, 0, 5 )
        
        bSizer108 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox165 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox165.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox165, 0, wx.ALL, 5 )
        
        self.checkBox166 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox166.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox166, 0, wx.ALL, 5 )
        
        self.checkBox167 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox167.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer108.Add( self.checkBox167, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer108, 0, wx.ALIGN_CENTER_HORIZONTAL, 5 )
        
        
        bSizer116.Add( bSizer105, 1, 0, 5 )
        
        bSizer305 = wx.BoxSizer( wx.VERTICAL )
        
        
        bSizer305.Add( ( 0, 20), 0, wx.EXPAND, 5 )
        
        self.textCtrl76 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer305.Add( self.textCtrl76, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer305.Add( ( 0, 45), 0, wx.EXPAND, 5 )
        
        self.textCtrl77 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer305.Add( self.textCtrl77, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer305.Add( ( 0, 95), 0, wx.EXPAND, 5 )
        
        self.textCtrl78 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer305.Add( self.textCtrl78, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer305.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer305, 1, wx.EXPAND, 5 )
        
        
        bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )
        
        bSizer103 = wx.BoxSizer( wx.HORIZONTAL )
        
        
        bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button33, 0, wx.ALL, 5 )
        
        self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button34, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )
        
        
        self.SetSizer( bSizer95 )
        self.Layout()
        
        self.Centre( wx.BOTH )
        
        # Connect Events
        self.checkBox159.Bind( wx.EVT_CHECKBOX, self.check1 )
        self.checkBox160.Bind( wx.EVT_CHECKBOX, self.check2 )
        self.checkBox161.Bind( wx.EVT_CHECKBOX, self.check3 )
        self.checkBox162.Bind( wx.EVT_CHECKBOX, self.check4 )
        self.checkBox163.Bind( wx.EVT_CHECKBOX, self.check5 )
        self.checkBox164.Bind( wx.EVT_CHECKBOX, self.check6 )
        self.checkBox165.Bind( wx.EVT_CHECKBOX, self.check7 )
        self.checkBox166.Bind( wx.EVT_CHECKBOX, self.check8 )
        self.checkBox167.Bind( wx.EVT_CHECKBOX, self.check9 )
        self.textCtrl76.Bind( wx.EVT_TEXT, self.Remark1 )
        self.textCtrl77.Bind( wx.EVT_TEXT, self.Remark2 )
        self.textCtrl78.Bind( wx.EVT_TEXT, self.Remark3 )
        self.button33.Bind( wx.EVT_BUTTON, self.Reset )
        self.button34.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
        pass
    
    
    # Virtual event handlers, overide them in your derived class
    def check1( self, event ):
            global o14
            box165 = self.checkBox159.GetValue()

            if box165 == True:
                o14 = 1
            else:
                o14 = 0

    def check2( self, event ):
            global p14
            box166 = self.checkBox160.GetValue()

            if box166 == True:
                p14 = 1
            else:
                p14 = 0

    def check3( self, event ):
            global q14
            box167 = self.checkBox161.GetValue()

            if box167 == True:
                q14 = 1
            else:
                q14 = 0

    def check4( self, event ):
            global r14
            box168 = self.checkBox162.GetValue()

            if box168 == True:
                r14 = 1
            else:
                r14 = 0

    def check5( self, event ):
            global s14           
            box169 = self.checkBox163.GetValue()

            if box169 == True:
                s14 = 1
            else:
                s14 = 0

    def check6( self, event ):
            global t14
            box170 = self.checkBox164.GetValue()

            if box170 == True:
                t14 = 1
            else:
                t14 = 0

    def check7( self, event ):
            global u14
            box171 = self.checkBox165.GetValue()

            if box171 == True:
                u14 = 1
            else:
                u14 = 0


    def check8( self, event ):
            global v14
            box172 = self.checkBox166.GetValue()

            if box172 == True:
                v14 = 1
            else:
                v14 = 0

    def check9( self, event ):
            global w14
            box173 = self.checkBox167.GetValue()

            if box173 == True:
                w14 = 1
            else:
                w14 = 0
    
    def Remark1( self, event ):
            global pear23
            global ad
            pear23 = self.textCtrl76.GetValue()
            ad = 1
    
    def Remark2( self, event ):
            global pear24
            global bd
            pear24 = self.textCtrl77.GetValue()
            bd = 1
    
    def Remark3( self, event ):
            global pear25
            global cd
            pear25 = self.textCtrl78.GetValue()
            cd = 1
    
    def Reset( self, event ):
            self.checkBox159.SetValue(False)
            self.checkBox160.SetValue(False)
            self.checkBox161.SetValue(False)
            self.checkBox162.SetValue(False)
            self.checkBox163.SetValue(False)
            self.checkBox164.SetValue(False)
            self.checkBox165.SetValue(False)
            self.checkBox166.SetValue(False)
            self.checkBox167.SetValue(False)
            self.textCtrl76.SetValue("")
            self.textCtrl77.SetValue("")
            self.textCtrl78.SetValue("")
            
            DataReset = Reset(parent = self.button33)
            SoundReset = soundReset(parent = self.button33)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
        
            global mop23
            global mop24
            global mop25
            
            global b24x
            global b25x
            global b26x
                        
            if ad == 1:
                mop23 = pear23
            else:
                mop23 = ""
                
            if bd == 1:
                mop24 = pear24
            else:
                mop24 = ""
                
            if cd == 1:
                mop25 = pear25
            else:
                mop25 = ""
            
            
            if o14 == 1:
                b24x = "Comply"
            else:
                pass
                
            if p14 == 1:
                b24x = "Observation"
            else:
                pass
                
            if q14 == 1:
                b24x = "Not Comply"
            else:
                pass
                
            if r14 == 1:
                b25x = "Comply"
            else:
                pass
                
            if s14 == 1:
                b25x = "Observation"
            else:
                pass
                
            if t14 == 1:
                b25x = "Not Comply"
            else:
                pass
                
            if u14 == 1:
                b26x = "Comply"
            else:
                pass
                
            if v14 == 1:
                b26x = "Observation"
            else:
                pass
                
            if w14 == 1:
                b26x = "Not Comply"
            else:
                pass
    
            SavingData = DataSaved(parent = self.button34)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button34)
            AcceptSound.Accept.Play()

###########################################################################
## Class Clause11
###########################################################################
class Clause11 ( wx.Frame ):
    global ae
    
    global x14
    global y14
    global z14
    
    global b27x
    
    global mop26
    
    x14 = 0
    y14 = 0
    z14 = 0
    
    ae = 0
    
    b27x = ""

    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Clause 11", pos = wx.DefaultPosition, size = wx.Size( 1300,264 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )
        
        self.SetSizeHintsSz( wx.Size(1300,264), wx.Size(1300,264) )
        self.SetBackgroundColour( wx.Colour( 165, 218, 182 ) )
        
        bSizer95 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer117 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText64 = wx.StaticText( self, wx.ID_ANY, u"Clause 11", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText64.Wrap( -1 )
        self.staticText64.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText64, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 200, 0), 0, 0, 5 )
        
        self.staticText65 = wx.StaticText( self, wx.ID_ANY, u"Interpretation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText65.Wrap( -1 )
        self.staticText65.SetFont( wx.Font( 12, 70, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText65, 0, wx.ALL, 5 )
        
        
        bSizer117.Add( ( 510, 0), 0, wx.EXPAND, 5 )
        
        self.staticText274 = wx.StaticText( self, wx.ID_ANY, u"Remark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText274.Wrap( -1 )
        self.staticText274.SetFont( wx.Font( 12, 74, 90, 92, False, "Arial" ) )
        
        bSizer117.Add( self.staticText274, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer117, 0, wx.EXPAND, 5 )
        
        bSizer116 = wx.BoxSizer( wx.HORIZONTAL )
        
        bSizer104 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer96 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.staticText49 = wx.StaticText( self, wx.ID_ANY, u"Clause 11", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText49.Wrap( -1 )
        self.staticText49.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText49, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 60, 0), 0, 0, 5 )
        
        self.staticText103 = wx.StaticText( self, wx.ID_ANY, u"Supplier mark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText103.Wrap( -1 )
        self.staticText103.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText103, 0, wx.ALL, 5 )
        
        
        bSizer96.Add( ( 50, 0), 1, 0, 5 )
        
        self.staticText50 = wx.StaticText( self, wx.ID_ANY, u"Each product shall be marked with\nits supplier mark and conforms to\nthe requirements of this standard\nupon of sales. Supplier mark shall\nbe affixed to the product.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText50.Wrap( -1 )
        self.staticText50.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer96.Add( self.staticText50, 0, wx.ALL, 5 )
        
        
        bSizer104.Add( bSizer96, 0, 0, 5 )
        
        
        bSizer116.Add( bSizer104, 1, wx.EXPAND, 5 )
        
        bSizer105 = wx.BoxSizer( wx.VERTICAL )
        
        bSizer106 = wx.BoxSizer( wx.HORIZONTAL )
        
        self.checkBox168 = wx.CheckBox( self, wx.ID_ANY, u"Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox168.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox168, 0, wx.ALL, 5 )
        
        self.checkBox169 = wx.CheckBox( self, wx.ID_ANY, u"Observation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox169.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox169, 0, wx.ALL, 5 )
        
        self.checkBox170 = wx.CheckBox( self, wx.ID_ANY, u"Not Comply", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.checkBox170.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )
        
        bSizer106.Add( self.checkBox170, 0, wx.ALL, 5 )
        
        
        bSizer105.Add( bSizer106, 0, wx.ALIGN_CENTER_HORIZONTAL, 5 )
        
        
        bSizer116.Add( bSizer105, 1, 0, 5 )
        
        bSizer306 = wx.BoxSizer( wx.VERTICAL )
        
        self.textCtrl81 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer306.Add( self.textCtrl81, 0, wx.ALL|wx.EXPAND, 5 )
        
        
        bSizer116.Add( bSizer306, 1, wx.EXPAND, 5 )
        
        
        bSizer95.Add( bSizer116, 1, wx.EXPAND, 5 )
        
        bSizer103 = wx.BoxSizer( wx.HORIZONTAL )
        
        
        bSizer103.Add( ( 0, 0), 1, wx.EXPAND, 5 )
        
        self.button33 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button33.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button33, 0, wx.ALL, 5 )
        
        self.button34 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button34.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )
        
        bSizer103.Add( self.button34, 0, wx.ALL, 5 )
        
        
        bSizer95.Add( bSizer103, 0, wx.EXPAND, 5 )
        
        
        self.SetSizer( bSizer95 )
        self.Layout()
        
        self.Centre( wx.BOTH )
        
        # Connect Events
        self.checkBox168.Bind( wx.EVT_CHECKBOX, self.check1 )
        self.checkBox169.Bind( wx.EVT_CHECKBOX, self.check2 )
        self.checkBox170.Bind( wx.EVT_CHECKBOX, self.check3 )
        self.textCtrl81.Bind( wx.EVT_TEXT, self.Remark1 )
        self.button33.Bind( wx.EVT_BUTTON, self.Reset )
        self.button34.Bind( wx.EVT_BUTTON, self.Save )
    
    def __del__( self ):
        pass
    
    
    # Virtual event handlers, overide them in your derived class
    def check1( self, event ):
            global x14
            box174 = self.checkBox168.GetValue()

            if box174 == True:
                x14 = 1
            else:
                x14 = 0

    def check2( self, event ):
            global y14
            box175 = self.checkBox169.GetValue()

            if box175 == True:
                y14 = 1
            else:
                y14 = 0

    def check3( self, event ):
            global z14
            box176 = self.checkBox170.GetValue()

            if box176 == True:
                z14 = 1
            else:
                z14 = 0
    
    def Remark1( self, event ):
            global pear26
            global ae
            pear26 = self.textCtrl81.GetValue()
            ae = 1
    
    def Reset( self, event ):
            self.checkBox168.SetValue(False)
            self.checkBox169.SetValue(False)
            self.checkBox170.SetValue(False)
            self.textCtrl81.SetValue("")
            
            DataReset = Reset(parent = self.button33)
            SoundReset = soundReset(parent = self.button33)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global mop26
            
            global b27x

            
            if ae ==1:
                mop26 = pear26
            else:
                mop26 = ""
            
            if x14 == 1:
                b27x = "Comply"
            else:
                pass
                
            if y14 == 1:
                b27x = "Observation"
            else:
                pass
                
            if z14 == 1:
                b27x = "Not Comply"
            else:
                pass
                
            SavingData = DataSaved(parent = self.button34)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button34)
            AcceptSound.Accept.Play()
    

###########################################################################
## Class AuditScore
###########################################################################
class AuditScore ( wx.Frame ):
    global word1
    global word2
    global word3
    global word4
    global word5
    global word6
    global word7
    global word8
    global word9
    global total1
    
    global word10
    global word11
    global word12
    global word13
    global word14
    global word15
    global word16
    global word17
    global word18
    global word19
    global word20
    global total2
    
    global word21
    global total3
    
    global word22
    global word23
    global total4
    
    global word24
    global word25
    global word26
    global total5
    
    global word27
    global total6
    
    global word28
    global word29
    global total7
    
    global word30
    global word31
    global word32
    global word33
    global total8
    
    global word34
    global word35
    global word36
    global word37
    global total9
    
    global word38
    global total10
    
    global a15
    global b15
    global c15
    global d15
    global e15
    global f15
    global g15
    global h15
    global i15
    
    global j15
    global k15
    global l15
    global m15
    global n15
    global o15
    global p15
    global q15
    global r15
    global s15
    global t15
    
    global u15

    global v15
    global w15

    global x15
    global y15
    global z15

    global a16
   
    global b16
    global c16
    
    global d16
    global e16
    global f16
    global g16

    global h16
    global i16
    global j16
    global k16

    global l16
    
    global aiy1
    global moo3
    
    aiy1 = 0
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Audit Score", pos = wx.DefaultPosition, size = wx.Size( 491,831 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(491,831), wx.Size(491,831) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

        bSizer244 = wx.BoxSizer( wx.VERTICAL )

        bSizer255 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer255.Add( ( 120, 0), 0, 0, 5 )

        self.staticText224 = wx.StaticText( self, wx.ID_ANY, u"State Date", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText224.Wrap( -1 )

        self.staticText224.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer255.Add( self.staticText224, 0, wx.ALL, 5 )

        self.datePicker3 = wx.adv.DatePickerCtrl( self, wx.ID_ANY, wx.DefaultDateTime, wx.DefaultPosition, wx.DefaultSize, wx.adv.DP_DEFAULT )
        self.datePicker3.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer255.Add( self.datePicker3, 0, wx.ALL, 5 )

        self.button75 = wx.Button( self, wx.ID_ANY, u"Set", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer255.Add( self.button75, 0, wx.ALL, 5 )
        self.button75.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer244.Add( bSizer255, 1, wx.EXPAND, 5 )

        bSizer245 = wx.BoxSizer( wx.VERTICAL )

        self.m_button48 = wx.Button( self, wx.ID_ANY, u"Section 7.1 - GENERAL REQUIREMENT", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button48.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button48, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button49 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (a) - CORE MANAGEMENT\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\n", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button49.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button49, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button50 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (b) - CLEANING OF ALL INTERNAL AND EXTERNAL COMPONENTS\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t\n", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button50.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button50, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button51 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (c) - REPLACEMENT AND/OR\nRESTORATION OF COMPONENTS", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button51.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button51, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button52 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (d) - MACHINING, INSPECTION AND TESTING", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button52.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button52, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button53 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (e) - COMPONENT ASSEMBLY", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button53.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button53, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button54 = wx.Button( self, wx.ID_ANY, u"Section 7.3.1 (f) - FINAL INSPECTION", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button54.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button54, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button55 = wx.Button( self, wx.ID_ANY, u"Section 9.1 - LABELLING AND PACKAGING\t", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button55.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button55, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button56 = wx.Button( self, wx.ID_ANY, u"Section 10 - WARRANTY", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button56.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button56, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )

        self.m_button57 = wx.Button( self, wx.ID_ANY, u"Section 11 - SUPPLIER MARK", wx.DefaultPosition, wx.Size( -1,60 ), 0 )
        self.m_button57.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer245.Add( self.m_button57, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


        bSizer244.Add( bSizer245, 1, wx.EXPAND, 5 )

        bSizer246 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer246.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.m_button58 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.m_button58.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer246.Add( self.m_button58, 0, wx.ALL, 5 )


        bSizer244.Add( bSizer246, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer244 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.datePicker3.Bind( wx.adv.EVT_DATE_CHANGED, self.Date3 )
        self.button75.Bind( wx.EVT_BUTTON, self.SetDate )
        self.m_button48.Bind( wx.EVT_BUTTON, self.Section1 )
        self.m_button49.Bind( wx.EVT_BUTTON, self.Section2 )
        self.m_button50.Bind( wx.EVT_BUTTON, self.Section3 )
        self.m_button51.Bind( wx.EVT_BUTTON, self.Section4 )
        self.m_button52.Bind( wx.EVT_BUTTON, self.Section5 )
        self.m_button53.Bind( wx.EVT_BUTTON, self.Section6 )
        self.m_button54.Bind( wx.EVT_BUTTON, self.Section7 )
        self.m_button55.Bind( wx.EVT_BUTTON, self.Section8 )
        self.m_button56.Bind( wx.EVT_BUTTON, self.Section9 )
        self.m_button57.Bind( wx.EVT_BUTTON, self.Section10 )
        self.m_button58.Bind( wx.EVT_BUTTON, self.Reset )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Date3( self, event ):
            global mine3
            mine3 =  event.GetDate()
            
            
    def SetDate( self, event ):
            global aiy1
            global moo3
            moo3 = mine3.Format("%d/%m/%Y")
            aiy1 = 1
            
            SavingData = DataSaved(parent = self.button75)
            SavingData.Show()
            
            AcceptSound = soundAccept(parent = self.button75)
            AcceptSound.Accept.Play()
            

    def Section1( self, event ):
            #event.Skip()
            Section1Form = Section71(parent=self.m_button48)
            Section1Form.Show()

    def Section2( self, event ):
            #event.Skip()
            Section2Form = Section731a(parent=self.m_button49)
            Section2Form.Show()

    def Section3( self, event ):
            #event.Skip()
            Section3Form = Section731b(parent=self.m_button50)
            Section3Form.Show()

    def Section4( self, event ):
            #event.Skip()
            Section4Form = Seksyen731c(parent=self.m_button51)
            Section4Form.Show()

    def Section5( self, event ):
            #event.Skip()
            Section5Form = Section731d(parent=self.m_button52)
            Section5Form.Show()

    def Section6( self, event ):
            #event.Skip()
            Section6Form = Section731e(parent=self.m_button53)
            Section6Form.Show()

    def Section7( self, event ):
            #event.Skip()
            Section7Form = Section731f(parent = self.m_button54)
            Section7Form.Show()

    def Section8( self, event ):
            #event.Skip()
            Section8Form = Section91(parent = self.m_button55)
            Section8Form.Show()

    def Section9( self, event ):
            #event.Skip()
            Section9Form = Section10(parent = self.m_button56)
            Section9Form.Show()

    def Section10( self, event ):
            #event.Skip()
            Section10Form = Section11(parent = self.m_button57)
            Section10Form.Show()

    def Reset( self, event ):
            global word1
            global word2
            global word3
            global word4
            global word5
            global word6
            global word7
            global word8
            global word9
            global total1
            
            global word10
            global word11
            global word12
            global word13
            global word14
            global word15
            global word16
            global word17
            global word18
            global word19
            global word20
            global total2
            
            global word21
            global total3
            
            global word22
            global word23
            global total4
            
            global word24
            global word25
            global word26
            global total5
            
            global word27
            global total6
            
            global word28
            global word29
            global total7
            
            global word30
            global word31
            global word32
            global word33
            global total8
            
            global word34
            global word35
            global word36
            global word37
            global total9
            
            global word38
            global total10
            
            global a15
            global b15
            global c15
            global d15
            global e15
            global f15
            global g15
            global h15
            global i15
            
            global j15
            global k15
            global l15
            global m15
            global n15
            global o15
            global p15
            global q15
            global r15
            global s15
            global t15
            
            global u15

            global v15
            global w15

            global x15
            global y15
            global z15

            global a16
   
            global b16
            global c16
            
            global d16
            global e16
            global f16
            global g16

            global h16
            global i16
            global j16
            global k16
    
            global l16

            
            a15 = 0
            b15 = 0
            c15 = 0
            d15 = 0
            e15 = 0
            f15 = 0
            g15 = 0
            h15 = 0
            i15 = 0
            
            j15 = 0
            k15 = 0
            l15 = 0
            m15 = 0
            n15 = 0
            o15 = 0
            p15 = 0
            q15 = 0
            r15 = 0
            s15 = 0
            t15 = 0
            
            u15 = 0
            
            v15 = 0
            w15 = 0
            
            x15 = 0
            y15 = 0
            z15 = 0
            
            a16 = 0
            
            b16 = 0
            c16 = 0
            
            d16 = 0
            e16 = 0
            f16 = 0
            g16 = 0
            
            h16 = 0
            i16 = 0
            j16 = 0
            k16 = 0
            
            l16 = 0
            
            
            word1 = 0
            word2 = 0
            word3 = 0
            word4 = 0
            word5 = 0
            word6 = 0
            word7 = 0
            word8 = 0
            word9 = 0
            total1 = 0
            
            word10 = 0
            word11 = 0
            word12 = 0
            word13 = 0
            word14 = 0
            word15 = 0
            word16 = 0
            word17 = 0
            word18 = 0
            word19 = 0
            word20 = 0
            total2 = 0
            
            word21 = 0
            total3 = 0
            
            word22 = 0
            word23 = 0
            total4 = 0
            
            word24 = 0
            word25 = 0
            word26 = 0
            total5 = 0
            
            word27 = 0
            total6 = 0
            
            word28 = 0
            word29 = 0
            total7 = 0
            
            word30 = 0
            word31 = 0
            word32 = 0
            word33 = 0
            total8 = 0
            
            word34 = 0
            word35 = 0
            word36 = 0
            word37 = 0
            total9 = 0
            
            word38 = 0
            total10 = 0
            
            DataReset = Reset(parent = self.m_button58)
            SoundReset = soundReset(parent = self.m_button58)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
            
            
            


###########################################################################
## Class Section71
###########################################################################

class Section71 ( wx.Frame ):
    global word1
    global word2
    global word3
    global word4
    global word5
    global word6
    global word7
    global word8
    global word9
    global total1
    
    global a15
    global b15
    global c15
    global d15
    global e15
    global f15
    global g15
    global h15
    global i15
    
    a15 = 0
    b15 = 0
    c15 = 0
    d15 = 0
    e15 = 0
    f15 = 0
    g15 = 0
    h15 = 0
    i15 = 0
    
    word1 = 0
    word2 = 0
    word3 = 0
    word4 = 0
    word5 = 0
    word6 = 0
    word7 = 0
    word8 = 0
    word9 = 0
    total1 = 0
    
    

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"General Requirement", pos = wx.DefaultPosition, size = wx.Size( 1151,750 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(1151,750), wx.Size(1151,750) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Internal Audit and Management Review", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 50), 0, 0, 5 )

        self.staticText213 = wx.StaticText( self, wx.ID_ANY, u"II. Safety, Health and Environmental Regulation Compliance", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText213.Wrap( -1 )
        self.staticText213.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText213, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 50), 0, 0, 5 )

        self.staticText214 = wx.StaticText( self, wx.ID_ANY, u"III. Competency Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText214.Wrap( -1 )
        self.staticText214.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText214, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 50), 0, 0, 5 )

        self.staticText215 = wx.StaticText( self, wx.ID_ANY, u"IV. Facilities Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText215.Wrap( -1 )
        self.staticText215.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText215, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 45), 0, 0, 5 )

        self.staticText216 = wx.StaticText( self, wx.ID_ANY, u"V. Environmental Quality Act 1974.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText216.Wrap( -1 )
        self.staticText216.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText216, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 35), 0, 0, 5 )

        self.staticText217 = wx.StaticText( self, wx.ID_ANY, u"VI. Standard Operating Procedure", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText217.Wrap( -1 )
        self.staticText217.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText217, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 45), 0, 0, 5 )

        self.staticText218 = wx.StaticText( self, wx.ID_ANY, u"VII. Calibration Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText218.Wrap( -1 )
        self.staticText218.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText218, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 40), 0, 0, 5 )

        self.staticText219 = wx.StaticText( self, wx.ID_ANY, u"VIII. Tools and Equipments For Assembly", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText219.Wrap( -1 )
        self.staticText219.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText219, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 40), 0, 0, 5 )

        self.staticText220 = wx.StaticText( self, wx.ID_ANY, u"IX. Documentation", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText220.Wrap( -1 )
        self.staticText220.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText220, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 1, wx.EXPAND, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Conduct for Management review for the 4R2S implementation performance. \n4R2S Management review record and conduct 4R2S internal self audit.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText222 = wx.StaticText( self, wx.ID_ANY, u"Machining of remanufacturing automotive parts and component \nshall be done in compliance with the relevant safety, health and \nenvironmental regulations. Parts that cannot be reused must be replaced.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText222.Wrap( -1 )
        self.staticText222.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText222, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText223 = wx.StaticText( self, wx.ID_ANY, u"Documented, establish the personnel competency training requirement.\nIdentify any personnel under training certification from SKM.\nRemanufacturing process shall be conducted in accordance to industry best practices.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText223.Wrap( -1 )
        self.staticText223.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText223, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText224 = wx.StaticText( self, wx.ID_ANY, u"The facilities shall have a dedicated area/section in their compound for each\nof the remanufacturing operation. The area should be in compliance with the\nrelevant safety, health and environmental regulations.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText224.Wrap( -1 )
        self.staticText224.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText224, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText225 = wx.StaticText( self, wx.ID_ANY, u"Storage, labelling and disposal of hazardous wastes shall be managed in\naccordance with the Environmental Quality Act 1974.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText225.Wrap( -1 )
        self.staticText225.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText225, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText226 = wx.StaticText( self, wx.ID_ANY, u"Remanufacturing processes required as good as new per Standard\nOperating Procedure/ OEM manual.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText226.Wrap( -1 )
        self.staticText226.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText226, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText227 = wx.StaticText( self, wx.ID_ANY, u"Tools and equipment used for inspection and testing and machines used for\nremanufacturing process must be calibrated and the calibration reports\nare able to be presented.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText227.Wrap( -1 )
        self.staticText227.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText227, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText228 = wx.StaticText( self, wx.ID_ANY, u"Personnel shall be equipped with necessary tools and equipments to\nperform the assembly.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText228.Wrap( -1 )
        self.staticText228.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText228, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText229 = wx.StaticText( self, wx.ID_ANY, u"All inspection and testing results shall be fully documented, recorded and\nretained for at least 7 years upon completion of sales.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText229.Wrap( -1 )
        self.staticText229.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText229, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer249, 1, wx.EXPAND, 5 )

        bSizer185 = wx.BoxSizer( wx.VERTICAL )

        bSizer258 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer258.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText228 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText228.Wrap( -1 )
        bSizer258.Add( self.staticText228, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer258, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 40), 0, 0, 5 )

        bSizer259 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl7 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer259.Add( self.textCtrl7, 0, wx.ALL, 5 )

        self.staticText229 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText229.Wrap( -1 )
        bSizer259.Add( self.staticText229, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer259, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 40), 0, 0, 5 )

        bSizer260 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl8 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer260.Add( self.textCtrl8, 0, wx.ALL, 5 )

        self.staticText230 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText230.Wrap( -1 )
        bSizer260.Add( self.staticText230, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer260, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 40), 0, 0, 5 )

        bSizer261 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl9 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer261.Add( self.textCtrl9, 0, wx.ALL, 5 )

        self.m_staticText231 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText231.Wrap( -1 )
        bSizer261.Add( self.m_staticText231, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer261, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 45), 0, 0, 5 )

        bSizer262 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl10 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer262.Add( self.textCtrl10, 0, wx.ALL, 5 )

        self.staticText232 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText232.Wrap( -1 )
        bSizer262.Add( self.staticText232, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer262, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 25), 0, 0, 5 )

        bSizer263 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl11 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer263.Add( self.textCtrl11, 0, wx.ALL, 5 )

        self.staticText233 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText233.Wrap( -1 )
        bSizer263.Add( self.staticText233, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer263, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 30), 0, 0, 5 )

        bSizer264 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl12 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer264.Add( self.textCtrl12, 0, wx.ALL, 5 )

        self.staticText234 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText234.Wrap( -1 )
        bSizer264.Add( self.staticText234, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer264, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 45), 0, 0, 5 )

        bSizer265 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl13 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer265.Add( self.textCtrl13, 0, wx.ALL, 5 )

        self.staticText235 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText235.Wrap( -1 )
        bSizer265.Add( self.staticText235, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer265, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 25), 0, 0, 5 )

        bSizer266 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl14 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString , wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer266.Add( self.textCtrl14, 0, wx.ALL, 5 )

        self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText236.Wrap( -1 )
        bSizer266.Add( self.staticText236, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer266, 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer185, 1, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 1, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.textCtrl7.Bind( wx.EVT_TEXT, self.Score2 )
        self.textCtrl8.Bind( wx.EVT_TEXT, self.Score3 )
        self.textCtrl9.Bind( wx.EVT_TEXT, self.Score4 )
        self.textCtrl10.Bind( wx.EVT_TEXT, self.Score5 )
        self.textCtrl11.Bind( wx.EVT_TEXT, self.Score6 )
        self.textCtrl12.Bind( wx.EVT_TEXT, self.Score7 )
        self.textCtrl13.Bind( wx.EVT_TEXT, self.Score8 )
        self.textCtrl14.Bind( wx.EVT_TEXT, self.Score9 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )
    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global conv1
            global get1
            global a15
            get1 = self.textCtrl6.GetValue()
            conv1 = int(get1)
            
            a15 = 1

    def Score2( self, event ):
            global conv2
            global get2
            global b15
            get2 = self.textCtrl7.GetValue()
            conv2 = int(get2)
            
            b15 = 1

    def Score3( self, event ):
            global conv3
            global get3
            global c15
            get3 = self.textCtrl8.GetValue()
            conv3 = int(get3)
            
            c15 = 1

    def Score4( self, event ):
            global conv4
            global get4
            global d15
            get4 = self.textCtrl9.GetValue()
            conv4 = int(get4)
            
            d15 = 1

    def Score5( self, event ):
            global conv5
            global get5
            global e15
            get5 = self.textCtrl10.GetValue()
            conv5 = int(get5)
            
            e15 = 1

    def Score6( self, event ):
            global conv6
            global get6
            global f15
            get6 = self.textCtrl11.GetValue()
            conv6 = int(get6)
            
            f15 = 1

    def Score7( self, event ):
            global conv7
            global get7
            global g15
            get7 = self.textCtrl12.GetValue()
            conv7 = int(get7)
            
            g15 = 1

    def Score8( self, event ):
            global conv8
            global get8
            global h15
            get8 = self.textCtrl13.GetValue()
            conv8 = int(get8)
            
            h15 = 1

    def Score9( self, event ):
            global conv9
            global get9
            global i15
            get9 = self.textCtrl14.GetValue()
            conv9 = int(get9)
            
            i15 =1

    def Reset( self, event ):
        self.textCtrl6.SetValue("")
        self.textCtrl7.SetValue("")
        self.textCtrl8.SetValue("")
        self.textCtrl9.SetValue("")
        self.textCtrl10.SetValue("")
        self.textCtrl11.SetValue("")
        self.textCtrl12.SetValue("")
        self.textCtrl13.SetValue("")
        self.textCtrl14.SetValue("")
        
        DataReset = Reset(parent = self.button55)
        SoundReset = soundReset(parent = self.button55)
            
        DataReset.Show()
        SoundReset.Resetto.Play()
        

    def Save( self, event ):
            global word1
            global word2
            global word3
            global word4
            global word5
            global word6
            global word7
            global word8
            global word9
            global total1
            
            sum1 = [conv1,conv2,conv3,conv4,conv5,conv6,conv7,conv8,conv9]
            total1 = sum(sum1)

            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
             
            if conv1 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
               
            elif conv2 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
               
            elif conv3 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
               
            elif conv4 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
               
            elif conv5 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
                       
            elif conv6 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
               
            elif conv7 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
               
            elif conv8 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
               
            elif conv9 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
                
              
            else:
            
                word1 = get1
                word2 = get2
                word3 = get3
                word4 = get4
                word5 = get5
                word6 = get6
                word7 = get7
                word8 = get8
                word9 = get9
                
                SavingScore.Show()
                AcceptSound.Accept.Play()
            

###########################################################################
## Class Section731a
###########################################################################

class Section731a ( wx.Frame ):
    global word10
    global word11
    global word12
    global word13
    global word14
    global word15
    global word16
    global word17
    global word18
    global word19
    global word20
    global total2
    
    global j15
    global k15
    global l15
    global m15
    global n15
    global o15
    global p15
    global q15
    global r15
    global s15
    global t15
    
    j15 = 0
    k15 = 0
    l15 = 0
    m15 = 0
    n15 = 0
    o15 = 0
    p15 = 0
    q15 = 0
    r15 = 0
    s15 = 0
    t15 = 0
    
    word10 = 0
    word11 = 0
    word12 = 0
    word13 = 0
    word14 = 0
    word15 = 0
    word16 = 0
    word17 = 0
    word18 = 0
    word19 = 0
    word20 = 0
    total2 = 0

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Core Management", pos = wx.DefaultPosition, size = wx.Size( 857,980 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(857,980), wx.Size(857,980) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )

        bSizer196 = wx.BoxSizer( wx.HORIZONTAL )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Data Retention", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 30), 0, 0, 5 )

        self.staticText213 = wx.StaticText( self, wx.ID_ANY, u"II. Documented Information", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText213.Wrap( -1 )
        self.staticText213.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText213, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 25), 0, 0, 5 )

        self.staticText214 = wx.StaticText( self, wx.ID_ANY, u"III. Parts and Components Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText214.Wrap( -1 )
        self.staticText214.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText214, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 20), 0, 0, 5 )

        self.staticText215 = wx.StaticText( self, wx.ID_ANY, u"IV. Core Sorting", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText215.Wrap( -1 )
        self.staticText215.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText215, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 25), 0, 0, 5 )

        self.staticText216 = wx.StaticText( self, wx.ID_ANY, u"V. Disassembly Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText216.Wrap( -1 )
        self.staticText216.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText216, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 30), 0, 0, 5 )

        self.staticText217 = wx.StaticText( self, wx.ID_ANY, u"VI. Disassembly Selection", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText217.Wrap( -1 )
        self.staticText217.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText217, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 35), 0, 0, 5 )

        self.staticText218 = wx.StaticText( self, wx.ID_ANY, u"VII. Physical Examination", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText218.Wrap( -1 )
        self.staticText218.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText218, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 260), 0, 0, 5 )

        self.staticText219 = wx.StaticText( self, wx.ID_ANY, u"VIII. Test Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText219.Wrap( -1 )
        self.staticText219.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText219, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 125), 0, 0, 5 )

        self.staticText220 = wx.StaticText( self, wx.ID_ANY, u"IX.  Component for Recycle", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText220.Wrap( -1 )
        self.staticText220.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText220, 0, wx.ALL, 5 )


        bSizer196.Add( bSizer248, 1, wx.EXPAND, 5 )


        bSizer196.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Essential information such as and not limited to serial number\nand part name are recorded to ensure traceability of components.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 10), 0, 0, 5 )

        self.staticText222 = wx.StaticText( self, wx.ID_ANY, u"Document such as quality manual, procedure, SOP or form and\netc are well maintained.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText222.Wrap( -1 )
        self.staticText222.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText222, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 10), 0, 0, 5 )

        self.staticText223 = wx.StaticText( self, wx.ID_ANY, u"Parts and components which are not allowed as used parts and\ncomponents under the relevant regulation shall not be repaired.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText223.Wrap( -1 )
        self.staticText223.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText223, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 10), 0, 0, 5 )

        self.staticText224 = wx.StaticText( self, wx.ID_ANY, u"Cores are sorted to relevant categories and are tagged\naccordingly to reuse the core or recycle the core.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText224.Wrap( -1 )
        self.staticText224.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText224, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 10), 0, 0, 5 )

        self.staticText225 = wx.StaticText( self, wx.ID_ANY, u"Disassembly is required for used parts and components to\ndetermine its usability.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText225.Wrap( -1 )
        self.staticText225.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText225, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 10), 0, 0, 5 )

        self.staticText226 = wx.StaticText( self, wx.ID_ANY, u"Consideration shall be taken with regards to the type of\ndisassembly method, tools and machine selected for\ndisassembly processes.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText226.Wrap( -1 )
        self.staticText226.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText226, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 10), 0, 0, 5 )

        self.staticText190 = wx.StaticText( self, wx.ID_ANY, u"a. Regulated by relevant authorities with an identification mark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText190.Wrap( -1 )
        self.staticText190.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

        bSizer249.Add( self.staticText190, 0, wx.TOP|wx.RIGHT|wx.LEFT, 5 )

        self.staticText227 = wx.StaticText( self, wx.ID_ANY, u"Any reused parts and components that are regulated by relevant\nauthorities shall be affixed with an identification mark. This\nidentification mark shall identify the source of the used parts and\ncomponents. A detailed record of the identification mark shall be\nmaintained.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText227.Wrap( -1 )
        self.staticText227.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText227, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 5), 0, 0, 5 )

        self.staticText191 = wx.StaticText( self, wx.ID_ANY, u"b. Part and component", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText191.Wrap( -1 )
        self.staticText191.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

        bSizer249.Add( self.staticText191, 0, wx.TOP|wx.RIGHT|wx.LEFT, 5 )

        self.staticText228 = wx.StaticText( self, wx.ID_ANY, u"The serial number for the parts shall be recorded. If the partsor \ncomponents serial numbers cannot be read or do not match\nexisting available records for that vehicle. The safety parts and\ncomponents shall not be sold as used, shall be recycled.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText228.Wrap( -1 )
        self.staticText228.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText228, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 5), 0, 0, 5 )

        self.staticText192 = wx.StaticText( self, wx.ID_ANY, u"c. Part visual inspection", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText192.Wrap( -1 )
        self.staticText192.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

        bSizer249.Add( self.staticText192, 0, wx.TOP|wx.RIGHT|wx.LEFT, 5 )

        self.staticText229 = wx.StaticText( self, wx.ID_ANY, u"Visual inspection of parts and components shall be undertaken to \nlook for any critical or physical damages.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText229.Wrap( -1 )
        self.staticText229.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText229, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 10), 0, 0, 5 )

        self.staticText193 = wx.StaticText( self, wx.ID_ANY, u"a. Test requirement -Functional test", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText193.Wrap( -1 )
        self.staticText193.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

        bSizer249.Add( self.staticText193, 0, 0, 5 )

        self.staticText188 = wx.StaticText( self, wx.ID_ANY, u"General condition of the used parts and components shall be\nassessed and recorded in order to determine its usability\ncondition. The used parts and components shall be subjected to\na functional/bench test in order to ensure that the parts are fit to\nbe reused. Proof of functionality shall be recorded, where\nnecessary. (Include visual and dimensional check, Fluids\ninspection)", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText188.Wrap( -1 )
        self.staticText188.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText188, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 10), 0, 0, 5 )

        self.staticText194 = wx.StaticText( self, wx.ID_ANY, u"a. Depolluted consideration", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText194.Wrap( -1 )
        self.staticText194.SetFont( wx.Font( 10, 74, 90, 92, False, "Arial" ) )

        bSizer249.Add( self.staticText194, 0, wx.TOP|wx.RIGHT|wx.LEFT, 5 )

        self.staticText189 = wx.StaticText( self, wx.ID_ANY, u"Consideration shall be taken to ensure that the products of the\nrecyclable materials are depolluted prior to dismantling, handling\nand segregation.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText189.Wrap( -1 )
        self.staticText189.SetFont( wx.Font( 10, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText189, 0, wx.ALL, 5 )


        bSizer196.Add( bSizer249, 1, wx.EXPAND, 5 )


        bSizer196.Add( ( 20, 0), 0, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.VERTICAL )

        bSizer267 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer267.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText237.Wrap( -1 )
        bSizer267.Add( self.staticText237, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer267, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 15), 0, 0, 5 )

        bSizer268 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl7 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer268.Add( self.textCtrl7, 0, wx.ALL, 5 )

        self.staticText238 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText238.Wrap( -1 )
        bSizer268.Add( self.staticText238, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer268, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 20), 0, 0, 5 )

        bSizer269 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl8 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer269.Add( self.textCtrl8, 0, wx.ALL, 5 )

        self.staticText239 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText239.Wrap( -1 )
        bSizer269.Add( self.staticText239, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer269, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 15), 0, 0, 5 )

        bSizer270 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl9 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer270.Add( self.textCtrl9, 0, wx.ALL, 5 )

        self.staticText240 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText240.Wrap( -1 )
        bSizer270.Add( self.staticText240, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer270, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 15), 0, 0, 5 )

        bSizer271 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl10 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer271.Add( self.textCtrl10, 0, wx.ALL, 5 )

        self.staticText241 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText241.Wrap( -1 )
        bSizer271.Add( self.staticText241, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer271, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 15), 0, 0, 5 )

        bSizer272 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl11 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer272.Add( self.textCtrl11, 0, wx.ALL, 5 )

        self.staticText242 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText242.Wrap( -1 )
        bSizer272.Add( self.staticText242, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer272, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 30), 0, 0, 5 )

        bSizer273 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl12 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer273.Add( self.textCtrl12, 0, wx.ALL, 5 )

        self.staticText243 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText243.Wrap( -1 )
        bSizer273.Add( self.staticText243, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer273, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 75), 0, 0, 5 )

        bSizer274 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl13 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer274.Add( self.textCtrl13, 0, wx.ALL, 5 )

        self.staticText244 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText244.Wrap( -1 )
        bSizer274.Add( self.staticText244, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer274, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 60), 0, 0, 5 )

        bSizer275 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl14 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer275.Add( self.textCtrl14, 0, wx.ALL, 5 )

        self.staticText245 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText245.Wrap( -1 )
        bSizer275.Add( self.staticText245, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer275, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 30), 0, 0, 5 )

        bSizer276 = wx.BoxSizer( wx.HORIZONTAL )

        self.m_textCtrl33 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer276.Add( self.m_textCtrl33, 0, wx.ALL, 5 )

        self.staticText246 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText246.Wrap( -1 )
        bSizer276.Add( self.staticText246, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer276, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 100), 0, 0, 5 )

        bSizer277 = wx.BoxSizer( wx.HORIZONTAL )

        self.m_textCtrl34 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer277.Add( self.m_textCtrl34, 0, wx.ALL, 5 )

        self.staticText247 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText247.Wrap( -1 )
        bSizer277.Add( self.staticText247, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer277, 1, wx.EXPAND, 5 )


        bSizer196.Add( bSizer185, 1, wx.EXPAND, 5 )


        bSizer247.Add( bSizer196, 1, wx.EXPAND, 5 )

        bSizer197 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer197.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button53 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button53.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer197.Add( self.button53, 0, wx.ALL, 5 )

        self.button54 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button54.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer197.Add( self.button54, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer197, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.textCtrl7.Bind( wx.EVT_TEXT, self.Score2 )
        self.textCtrl8.Bind( wx.EVT_TEXT, self.Score3 )
        self.textCtrl9.Bind( wx.EVT_TEXT, self.Score4 )
        self.textCtrl10.Bind( wx.EVT_TEXT, self.Score5 )
        self.textCtrl11.Bind( wx.EVT_TEXT, self.Score6 )
        self.textCtrl12.Bind( wx.EVT_TEXT, self.Score7 )
        self.textCtrl13.Bind( wx.EVT_TEXT, self.Score8 )
        self.textCtrl14.Bind( wx.EVT_TEXT, self.Score9 )
        self.m_textCtrl33.Bind( wx.EVT_TEXT, self.Score10 )
        self.m_textCtrl34.Bind( wx.EVT_TEXT, self.Score11 )
        self.button53.Bind( wx.EVT_BUTTON, self.Reset )
        self.button54.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get10
            global conv10
            global j15
            get10 = self.textCtrl6.GetValue()
            conv10 = int(get10)
            
            j15 = 1

    def Score2( self, event ):
            global get11
            global conv11
            global k15
            get11 = self.textCtrl7.GetValue()
            conv11 = int(get11)
            
            k15 = 1

    def Score3( self, event ):
            global get12
            global conv12
            global l15
            get12 = self.textCtrl8.GetValue()
            conv12 = int(get12)
            
            l15 = 1

    def Score4( self, event ):
            global get13
            global conv13
            global m15
            get13 = self.textCtrl9.GetValue()
            conv13 = int(get13)
            
            m15 = 1

    def Score5( self, event ):
            global get14
            global conv14
            global n15
            get14 = self.textCtrl10.GetValue()
            conv14 = int(get14)
            
            n15 = 1

    def Score6( self, event ):
            global get15
            global conv15
            global o15
            get15 = self.textCtrl11.GetValue()
            conv15 = int(get15)
            
            o15 = 1

    def Score7( self, event ):
            global get16
            global conv16
            global p15
            get16 = self.textCtrl12.GetValue()
            conv16 = int(get16)
            
            p15 = 1

    def Score8( self, event ):
            global get17
            global conv17
            global q15
            get17 = self.textCtrl13.GetValue()
            conv17 = int(get17)
            
            q15 = 1

    def Score9( self, event ):
            global get18
            global conv18
            global r15
            get18 = self.textCtrl14.GetValue()
            conv18 = int(get18)
            
            r15 = 1

    def Score10( self, event ):
            global get19
            global conv19
            global s15
            get19 = self.m_textCtrl33.GetValue()
            conv19 = int(get19)
            
            s15 = 1

    def Score11( self, event ):
            global get20
            global conv20
            global t15
            get20 = self.m_textCtrl34.GetValue()
            conv20 = int(get20)
            
            t15 = 1

    def Reset( self, event ):
            self.textCtrl7.SetValue("")
            self.textCtrl8.SetValue("")
            self.textCtrl9.SetValue("")
            self.textCtrl10.SetValue("")
            self.textCtrl11.SetValue("")
            self.textCtrl12.SetValue("")
            self.textCtrl13.SetValue("")
            self.textCtrl14.SetValue("")
            self.m_textCtrl33.SetValue("")
            self.m_textCtrl34.SetValue("")
            
            DataReset = Reset(parent = self.button53)
            SoundReset = soundReset(parent = self.button53)
            
            DataReset.Show()
            SoundReset.Resetto.Play()


    def Save( self, event ):
            global word10
            global word11
            global word12
            global word13
            global word14
            global word15
            global word16
            global word17
            global word18
            global word19
            global word20
            global total2
            
            sum2 = [conv10,conv11,conv12,conv13,conv14,conv15,conv16,conv17,conv18,conv19,conv20]
            
            total2 = sum(sum2)
            
            ErrorScore = Error1(parent = self.button54)
            SavingScore = ScoreSaved(parent = self.button54)
            AcceptSound = soundAccept(parent = self.button54)
            ErrorSound = soundError(parent = self.button54)
            
            if conv10 > 5 :
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv11 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv12 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv13 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv14 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv15 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv16 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv17 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv18 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv19 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv20 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
                
            else:

                word10 = get10
                word11 = get11
                word12 = get12
                word13 = get13
                word14 = get14
                word15 = get15
                word16 = get16
                word17 = get17
                word18 = get18
                word19 = get19
                word20 = get20
                
                SavingScore.Show()
                AcceptSound.Accept.Play()

###########################################################################
## Class Section731b
###########################################################################

class Section731b ( wx.Frame ):
    global word21
    global total3
    
    global u15
    
    u15 = 0
    
    word21 = 0
    total3 = 0

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Cleaning Of All Internal And External Components", pos = wx.DefaultPosition, size = wx.Size( 779,172 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(779,172), wx.Size(779,172) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 10, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Cleaning Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Internal and external components that will undergo\nremanufacturing process shall be cleaned.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer249, 1, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer185.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText248 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText248.Wrap( -1 )
        bSizer185.Add( self.staticText248, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 10), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )


        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get21
            global conv21
            global u15
            get21 = self.textCtrl6.GetValue()
            conv21 = int(get21)
            
            u15 = 1

    def Reset( self, event ):
            self.textCtrl6.SetValue("")
            
            DataReset = Reset(parent = self.button55)
            SoundReset = soundReset(parent = self.button55)
            
            DataReset.Show()
            SoundReset.Resetto.Play()
        

    def Save( self, event ):
            global word21
            global total3
            
            sum3 = [conv21]
            
            total3 = sum(sum3)
            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
        
            if conv21 > 5:
                
                ErrorScore.Show()
                ErrorSound.Error.Play()
                
            else:
                
                word21 = get21
                SavingScore.Show()
                AcceptSound.Accept.Play()



###########################################################################
## Class Seksyen731c
###########################################################################

class Seksyen731c ( wx.Frame ):
    global word22
    global word23
    global total4
    
    global v15
    global w15
    
    v15 = 0
    w15 = 0
    
    word22 = 0
    word23 = 0
    total4 = 0

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Replacement And/Or Restoration of Components", pos = wx.DefaultPosition, size = wx.Size( 758,310 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(758,310), wx.Size(758,310) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 10, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Replacement of Components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 65), 0, 0, 5 )

        self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Restoration of Components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText236.Wrap( -1 )
        self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Parts and components which are not allowed\nas used parts and components under the \nrelevant regulation(brake lining, battery,\ntires) and/or missing shall be replaced.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"Malfunction parts and components that is fit\nfor restoring within approved tolerance\naccording to established industrial\nspecifications shall be repaired.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText237.Wrap( -1 )
        self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText237, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer249, 1, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.VERTICAL )

        bSizer278 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer278.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText249 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText249.Wrap( -1 )
        bSizer278.Add( self.staticText249, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer278, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 60), 0, 0, 5 )

        bSizer279 = wx.BoxSizer( wx.HORIZONTAL )

        self.m_textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer279.Add( self.m_textCtrl57, 0, wx.ALL, 5 )

        self.staticText250 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText250.Wrap( -1 )
        bSizer279.Add( self.staticText250, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer279, 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 10), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )


        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.m_textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get22
            global conv22
            global v15
            get22 = self.textCtrl6.GetValue()
            conv22 = int(get22)
            
            v15 = 1

    def Score2( self, event ):
            global get23
            global conv23
            global w15
            get23 = self.m_textCtrl57.GetValue()
            conv23 = int(get23)
            
            w15 = 1

    def Reset( self, event ):
            self.textCtrl6.SetValue("")
            self.m_textCtrl57.SetValue("")
            
            DataReset = Reset(parent = self.button55)
            SoundReset = soundReset(parent = self.button55)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global word22
            global word23
            global total4
            
            sum4 = [conv22,conv23]
            
            total4 = sum(sum4)
            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
            
            if conv22 > 5 :
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv23 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
                
            else:

                word22 = get22
                word23 = get23
                
                SavingScore.Show()
                AcceptSound.Accept.Play()


###########################################################################
## Class Section731d
###########################################################################

class Section731d ( wx.Frame ):
    global word24
    global word25
    global word26
    global total5
    
    global x15
    global y15
    global z15
    
    x15 = 0
    y15 = 0
    z15 = 0
    
    word24 = 0
    word25 = 0
    word26 = 0
    total5 = 0
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Machining, Inspection and Testing", pos = wx.DefaultPosition, size = wx.Size( 770,360 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(770,360), wx.Size(770,360) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 10, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Replacement of Components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 50), 0, 0, 5 )

        self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Restoration of Components", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText236.Wrap( -1 )
        self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 50), 0, 0, 5 )

        self.staticText246 = wx.StaticText( self, wx.ID_ANY, u"III. Tools and Equipments", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText246.Wrap( -1 )
        self.staticText246.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText246, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"All remanufacturing components are required to undergo\ninspection and testing before assembly to ensure\nall the internals are in working condition.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"Reworking, machining or performing such other operations\nas are necessary to put the part in original working condition\nor better.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText237.Wrap( -1 )
        self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText237, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

        self.staticText247 = wx.StaticText( self, wx.ID_ANY, u"Personnel shall be equipped with necessary PPE to perform\npainting process as per relevant Standard Operating Procedure (SOP).", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText247.Wrap( -1 )
        self.staticText247.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText247, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer249, 1, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.VERTICAL )

        bSizer280 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer280.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText251 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText251.Wrap( -1 )
        bSizer280.Add( self.staticText251, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer280, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 40), 0, 0, 5 )

        bSizer281 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer281.Add( self.textCtrl57, 0, wx.ALL, 5 )

        self.staticText252 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText252.Wrap( -1 )
        bSizer281.Add( self.staticText252, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer281, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 40), 0, 0, 5 )

        bSizer282 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl62 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer282.Add( self.textCtrl62, 0, wx.ALL, 5 )

        self.staticText253 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText253.Wrap( -1 )
        bSizer282.Add( self.staticText253, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer282, 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 10), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )


        bSizer199.Add( ( 0, 30), 0, 0, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )


        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
        self.textCtrl62.Bind( wx.EVT_TEXT, self.Score3 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get24
            global conv24
            global x15
            get24 = self.textCtrl6.GetValue()
            conv24 = int(get24)
            
            x15 = 1

    def Score2( self, event ):
            global get25
            global conv25
            global y15
            get25 = self.textCtrl57.GetValue()
            conv25 = int(get25)
            
            y15 = 1

    def Score3( self, event ):
            global get26
            global conv26
            global z15
            get26 = self.textCtrl62.GetValue()
            conv26 = int(get26)
            
            z15 = 1

    def Reset( self, event ):
            self.textCtrl6.SetValue("")
            self.textCtrl57.SetValue("")
            self.textCtrl62.SetValue("")
            
            DataReset = Reset(parent = self.button55)
            SoundReset = soundReset(parent = self.button55)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global word24
            global word25
            global word26
            global total5
            
            sum5 = [conv24,conv25,conv26]
            
            total5 = sum(sum5)
            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
            
            if conv24 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv25 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv26 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
                
            else:    

                word24 = get24
                word25 = get25
                word26 = get26
                
                SavingScore.Show()
                AcceptSound.Accept.Play()


###########################################################################
## Class Section731e
###########################################################################

class Section731e ( wx.Frame ):
    global word27
    global total6
    
    global a16
    
    a16 = 0
    
    word27 = 0
    total6 = 0
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Component Assembly", pos = wx.DefaultPosition, size = wx.Size( 785,180 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(785,180), wx.Size(785,180) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 10, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Assembly Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Assembly process was conducted in accordance to standard\noperating procedure.\t", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer249, 1, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer185.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText254 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText254.Wrap( -1 )
        bSizer185.Add( self.staticText254, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 10), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get27
            global conv27
            global a16
            get27 = self.textCtrl6.GetValue()
            conv27 = int(get27)
            
            a16 = 1

    def Reset( self, event ):
            self.textCtrl6.SetValue("")
            
            DataReset = Reset(parent = self.button55)
            SoundReset = soundReset(parent = self.button55)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global word27
            global total6
            
            sum6 = [conv27]
            
            total6 = sum(sum6)
            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
            
            if conv27 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            
            else:

                word27 = get27
                SavingScore.Show()
                AcceptSound.Accept.Play()



###########################################################################
## Class Section731f
###########################################################################

class Section731f ( wx.Frame ):
    global word28
    global word29
    global total7
    
    global b16
    global c16
    
    b16 = 0
    c16 = 0
    
    word28 = 0
    word29 = 0
    total7 = 0
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Final Inspection", pos = wx.DefaultPosition, size = wx.Size( 785,322 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(785,322), wx.Size(785,322) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 10, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Inspection Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 110), 0, 0, 5 )

        self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Inspection and Testing Information", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText236.Wrap( -1 )
        self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Components shall undergo final inspection that\ninclude testing. Components that comply\nrelevant testing standards shall be recorded with\nQC inspection, provided the product conforms\nto the requirements of the standard with inspection label.\nInspection label may be affixed to the product or\npackaging, where appropriate.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"Components were tested to its nature of working condition\naccording to relevant standard information.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText237.Wrap( -1 )
        self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText237, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer249, 1, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.VERTICAL )

        bSizer283 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer283.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText255 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText255.Wrap( -1 )
        bSizer283.Add( self.staticText255, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer283, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        bSizer284 = wx.BoxSizer( wx.HORIZONTAL )

        self.m_textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer284.Add( self.m_textCtrl57, 0, wx.ALL, 5 )

        self.staticText256 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText256.Wrap( -1 )
        bSizer284.Add( self.staticText256, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer284, 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 10), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.m_textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get28
            global conv28
            global b16
            get28 = self.textCtrl6.GetValue()
            conv28 = int(get28)
            
            b16 = 1

    def Score2( self, event ):
            global get29
            global conv29
            global c16
            get29 = self.m_textCtrl57.GetValue()
            conv29 = int(get29)
            
            c16 = 1
            

    def Reset( self, event ):
            self.textCtrl6.SetValue("")
            self.textCtrl57.SetValue("")
            
            DataReset = Reset(parent = self.button55)
            SoundReset = soundReset(parent = self.button55)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global word28
            global word29
            global total7
            
            sum7 = [conv28,conv29]
            
            total7 = sum(sum7)
            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
            
            if conv28 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv29 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
                
            else:

                word28 = get28
                word29 = get29
                
                SavingScore.Show()
                AcceptSound.Accept.Play()



###########################################################################
## Class Section91
###########################################################################

class Section91 ( wx.Frame ):
    global word30
    global word31
    global word32
    global word33
    global total8
    
    global d16
    global e16
    global f16
    global g16
    
    d16 = 0
    e16 = 0
    f16 = 0
    g16 = 0
    
    word30 = 0
    word31 = 0
    word32 = 0
    word33 = 0
    total8 = 0
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"labelling And Packaging", pos = wx.DefaultPosition, size = wx.Size( 839,547 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(839,547), wx.Size(839,547) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 10, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Labelling and Identification", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 35), 0, 0, 5 )

        self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Labelling Information", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText236.Wrap( -1 )
        self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 130), 0, 0, 5 )

        self.staticText246 = wx.StaticText( self, wx.ID_ANY, u"III. Packaging Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText246.Wrap( -1 )
        self.staticText246.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText246, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 53), 0, 0, 5 )

        self.staticText272 = wx.StaticText( self, wx.ID_ANY, u"IV. Packaging information", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText272.Wrap( -1 )
        self.staticText272.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText272, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Products shall have a permanent label to be identified\nas reused, repaired or remanufactured parts or components.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"\"Labelling shall include, but not limited to, the following information:\n\na) company name\nb) parts identification, if applicable\nc) make and model\nd) engine capacity, if applicable\ne) classifications according to industry best practices (reuse, remanufacturing or repair)\nf) recovered from local vehicles or imported.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText237.Wrap( -1 )
        self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText237, 0, wx.RIGHT|wx.LEFT, 5 )


        bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

        self.staticText247 = wx.StaticText( self, wx.ID_ANY, u"Types of packaging used shall be able to protect the reused,\nrepaired and remanufactured parts from damage during\nlogistics and transportations.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText247.Wrap( -1 )
        self.staticText247.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText247, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

        self.staticText273 = wx.StaticText( self, wx.ID_ANY, u"Packaging shall include, but not limited to, the following information:\n\na) name of product\nb) manufacturer's details\nc) type of warranty covered\t\t\n", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText273.Wrap( -1 )
        self.staticText273.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText273, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer249, 1, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.VERTICAL )

        bSizer285 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer285.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.m_staticText257 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText257.Wrap( -1 )
        bSizer285.Add( self.m_staticText257, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer285, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 25), 0, 0, 5 )

        bSizer286 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer286.Add( self.textCtrl57, 0, wx.ALL, 5 )

        self.staticText258 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText258.Wrap( -1 )
        bSizer286.Add( self.staticText258, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer286, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 130), 0, 0, 5 )

        bSizer287 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl62 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer287.Add( self.textCtrl62, 0, wx.ALL, 5 )

        self.staticText259 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText259.Wrap( -1 )
        bSizer287.Add( self.staticText259, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer287, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 50), 0, 0, 5 )

        bSizer288 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl75 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer288.Add( self.textCtrl75, 0, wx.ALL, 5 )

        self.staticText260 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText260.Wrap( -1 )
        bSizer288.Add( self.staticText260, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer288, 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 10), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )


        bSizer199.Add( ( 0, 30), 0, 0, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
        self.textCtrl62.Bind( wx.EVT_TEXT, self.Score3 )
        self.textCtrl75.Bind( wx.EVT_TEXT, self.Score4 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get30
            global conv30
            global d16
            get30 = self.textCtrl6.GetValue()
            conv30 = int(get30)
            
            d16 = 1


    def Score2( self, event ):
            global get31
            global conv31
            global e16
            get31 = self.textCtrl57.GetValue()
            conv31 = int(get31)
            
            e16 = 1

    def Score3( self, event ):
            global get32
            global conv32
            global f16
            get32 = self.textCtrl62.GetValue()
            conv32 = int(get32)
            
            f16 = 1

    def Score4( self, event ):
            global get33
            global conv33
            global g16
            get33 = self.textCtrl75.GetValue()
            conv33 = int(get33)
            
            g16 = 1

    def Reset( self, event ):
            self.textCtrl6.SetValue("")
            self.textCtrl57.SetValue("")
            self.textCtrl62.SetValue("")
            self.textCtrl75.SetValue("")
            
            DataReset = Reset(parent = self.button55)
            SoundReset = soundReset(parent = self.button55)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global word30
            global word31
            global word32
            global word33
            global total8
            
            sum8 = [conv30,conv31,conv32,conv33]
            
            total8 = sum(sum8)
            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
            
            if conv30 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv31 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv32 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv33 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
                
            else:
                word30 = get30
                word31 = get31
                word32 = get32
                word33 = get33
                
                SavingScore.Show()
                AcceptSound.Accept.Play()


###########################################################################
## Class Section10
###########################################################################

class Section10 ( wx.Frame ):
    global word34
    global word35
    global word36
    global word37
    global total9
    
    global h16
    global i16
    global j16
    global k16
    
    h16 = 0
    i16 = 0
    j16 = 0
    k16 = 0
    
    word34 = 0
    word35 = 0
    word36 = 0
    word37 = 0
    total9 = 0
    
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Warranty", pos = wx.DefaultPosition, size = wx.Size( 935,478 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(935,478), wx.Size(935,478) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 10, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Warranty Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 35), 0, 0, 5 )

        self.staticText236 = wx.StaticText( self, wx.ID_ANY, u"II. Warranty Information", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText236.Wrap( -1 )
        self.staticText236.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText236, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 95), 0, 0, 5 )

        self.staticText246 = wx.StaticText( self, wx.ID_ANY, u"III. Record of Warranty Claim", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText246.Wrap( -1 )
        self.staticText246.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText246, 0, wx.ALL, 5 )


        bSizer248.Add( ( 0, 36), 0, 0, 5 )

        self.staticText272 = wx.StaticText( self, wx.ID_ANY, u"IV. Record Information", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText272.Wrap( -1 )
        self.staticText272.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText272, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Warranty shall be provided for reused parts, repaired parts and\nremanufacturing parts.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 0, 0, 5 )

        self.staticText237 = wx.StaticText( self, wx.ID_ANY, u"Remanufactured parts and components - not less than 90 days\nfrom the date of purchase\nRepaired parts and components - not less than 45 days from the date of purchase\nReused parts and components - not less than 30 days from the date of purchase\nWarranty shall be in writing and assured by the warranty\nprovider.\" cal vehicles or imported.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText237.Wrap( -1 )
        self.staticText237.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText237, 0, wx.RIGHT|wx.LEFT, 5 )


        bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

        self.staticText247 = wx.StaticText( self, wx.ID_ANY, u"Record of warranty claim is kept and retained for 7 years for\nreferences.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText247.Wrap( -1 )
        self.staticText247.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText247, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 20), 1, wx.EXPAND, 5 )

        self.staticText273 = wx.StaticText( self, wx.ID_ANY, u"Records shall include, but not limited to, the following information:\n\na) customer's particular\nb) component type\nc) component serial number", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText273.Wrap( -1 )
        self.staticText273.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText273, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer249, 1, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.VERTICAL )

        bSizer289 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer289.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText261 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText261.Wrap( -1 )
        bSizer289.Add( self.staticText261, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer289, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 25), 0, 0, 5 )

        bSizer290 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl57 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer290.Add( self.textCtrl57, 0, wx.ALL, 5 )

        self.staticText262 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText262.Wrap( -1 )
        bSizer290.Add( self.staticText262, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer290, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 95), 0, 0, 5 )

        bSizer291 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl62 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer291.Add( self.textCtrl62, 0, wx.ALL, 5 )

        self.staticText263 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText263.Wrap( -1 )
        bSizer291.Add( self.staticText263, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer291, 1, wx.EXPAND, 5 )


        bSizer185.Add( ( 0, 30), 0, 0, 5 )

        bSizer292 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl75 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer292.Add( self.textCtrl75, 0, wx.ALL, 5 )

        self.staticText264 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText264.Wrap( -1 )
        bSizer292.Add( self.staticText264, 0, wx.ALL, 5 )


        bSizer185.Add( bSizer292, 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 10), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )


        bSizer199.Add( ( 0, 30), 0, 0, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.textCtrl57.Bind( wx.EVT_TEXT, self.Score2 )
        self.textCtrl62.Bind( wx.EVT_TEXT, self.Score3 )
        self.textCtrl75.Bind( wx.EVT_TEXT, self.Score4 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get34
            global conv34
            global h16
            get34 = self.textCtrl6.GetValue()
            conv34 = int(get34)
            
            h16 = 1

    def Score2( self, event ):
            global get35
            global conv35
            global i16
            get35 = self.textCtrl57.GetValue()
            conv35 = int(get35)
            
            i16 = 1

    def Score3( self, event ):
            global get36
            global conv36
            global j16
            get36 = self.textCtrl62.GetValue()
            conv36 = int(get36)
            
            j16 = 1

    def Score4( self, event ):
            global get37
            global conv37
            global k16
            get37 = self.textCtrl75.GetValue()
            conv37 = int(get37)
            
            k16 = 1
            

    def Reset( self, event ):
            self.textCtrl6.SetValue("")
            self.textCtrl57.SetValue("")
            self.textCtrl62.SetValue("")
            self.textCtrl75.SetValue("")
            
            DataReset = Reset(parent = self.button55)
            SoundReset = soundReset(parent = self.button55)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global word34
            global word35
            global word36
            global word37
            global total9
            
            sum9 = [conv34,conv35,conv36,conv37]
            
            total9 = sum(sum9)
            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
            
            if conv34 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv35 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv36 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            elif conv37 > 5:
                ErrorScore.Show()
                ErrorSound.Error.Play()
            
            else:

                word34 = get34
                word35 = get35
                word36 = get36
                word37 = get37
                
                SavingScore.Show()
                AcceptSound.Accept.Play()



###########################################################################
## Class Section11
###########################################################################

class Section11 ( wx.Frame ):
    global word38
    
    global l16
    
    global total10
    
    l16 = 0
    
    word38 = 0
    
    total10 = 0
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Supplier Mark", pos = wx.DefaultPosition, size = wx.Size( 782,172 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHintsSz( wx.Size(782,172), wx.Size(782,172) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVEBORDER ) )

        bSizer247 = wx.BoxSizer( wx.VERTICAL )


        bSizer247.Add( ( 0, 20), 0, 0, 5 )

        bSizer198 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer198.Add( ( 10, 0), 0, 0, 5 )

        bSizer248 = wx.BoxSizer( wx.VERTICAL )

        self.staticText212 = wx.StaticText( self, wx.ID_ANY, u"I. Supplier Mark", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText212.Wrap( -1 )
        self.staticText212.SetFont( wx.Font( 11, 74, 90, 92, False, "Arial" ) )

        bSizer248.Add( self.staticText212, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer248, 0, wx.EXPAND, 5 )


        bSizer198.Add( ( 20, 0), 0, 0, 5 )

        bSizer249 = wx.BoxSizer( wx.VERTICAL )

        self.staticText221 = wx.StaticText( self, wx.ID_ANY, u"Each product be marked with its certification mark based on company standard. \nSupplier mark be affixed to the product or packaging, where appropriate.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText221.Wrap( -1 )
        self.staticText221.SetFont( wx.Font( 11, 74, 90, 90, False, "Arial" ) )

        bSizer249.Add( self.staticText221, 0, wx.ALL, 5 )


        bSizer249.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer198.Add( bSizer249, 1, 0, 5 )

        bSizer185 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl6 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( 60,-1 ), 0 )
        bSizer185.Add( self.textCtrl6, 0, wx.ALL, 5 )

        self.staticText265 = wx.StaticText( self, wx.ID_ANY, u"/ 5", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText265.Wrap( -1 )
        bSizer185.Add( self.staticText265, 0, wx.ALL, 5 )


        bSizer198.Add( bSizer185, 0, wx.EXPAND, 5 )


        bSizer247.Add( bSizer198, 0, wx.EXPAND, 5 )


        bSizer247.Add( ( 0, 10), 0, 0, 5 )

        bSizer199 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer199.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button55 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button55.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button55, 0, wx.ALL, 5 )

        self.button56 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button56.SetFont( wx.Font( 14, 74, 90, 90, False, "Arial" ) )

        bSizer199.Add( self.button56, 0, wx.ALL, 5 )


        bSizer247.Add( bSizer199, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer247 )
        self.Layout()

        self.Centre( wx.BOTH )


        # Connect Events
        self.textCtrl6.Bind( wx.EVT_TEXT, self.Score1 )
        self.button55.Bind( wx.EVT_BUTTON, self.Reset )
        self.button56.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Score1( self, event ):
            global get38
            global conv38
            global l16
            get38 = self.textCtrl6.GetValue()
            conv38 = int(get38)
            
            l16 = 1

    def Reset( self, event ):
            self.textCtrl6.SetValue("")
            
            DataReset = Reset(parent = self.button55)
            SoundReset = soundReset(parent = self.button55)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global word38
            global total10
            
            sum10 = [conv38]
            
            total10 = sum(sum10)
            
            ErrorScore = Error1(parent = self.button56)
            SavingScore = ScoreSaved(parent = self.button56)
            AcceptSound = soundAccept(parent = self.button56)
            ErrorSound = soundError(parent = self.button56)
            
            if conv38 > 5 :
                
                ErrorScore.Show()
                ErrorSound.Error.Play()
                
            else:
                word38 = get38
                SavingScore.Show()
                AcceptSound.Accept.Play()



###########################################################################
## Class ExecutiveSummary
###########################################################################
class ExecutiveSummary ( wx.Frame ):
    global u4
    global v4
    global w4
    global x4
    global y4
    global z4
    global a5
    global b5
    global c5
    global d5
    global e5
    global f5
    
    global moo4
    
    u4 = 0
    v4 = 0
    w4 = 0
    x4 = 0
    y4 = 0
    z4 = 0
    a5 = 0
    b5 = 0
    c5 = 0
    d5 = 0
    e5 = 0
    f5 = 0
    
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 623,460 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

        bSizer304 = wx.BoxSizer( wx.VERTICAL )

        bSizer256 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer256.Add( ( 180, 0), 0, 0, 5 )

        self.staticText225 = wx.StaticText( self, wx.ID_ANY, u"State Date", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText225.Wrap( -1 )

        self.staticText225.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer256.Add( self.staticText225, 0, wx.ALL, 5 )

        self.datePicker4 = wx.adv.DatePickerCtrl( self, wx.ID_ANY, wx.DefaultDateTime, wx.DefaultPosition, wx.DefaultSize, wx.adv.DP_DEFAULT )
        self.datePicker4.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer256.Add( self.datePicker4, 0, wx.ALL, 5 )

        self.button89 = wx.Button( self, wx.ID_ANY, u"Set", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer256.Add( self.button89, 0, wx.ALL, 5 )


        bSizer304.Add( bSizer256, 1, wx.EXPAND, 5 )

        bSizer314 = wx.BoxSizer( wx.HORIZONTAL )

        self.staticText300 = wx.StaticText( self, wx.ID_ANY, u"Requirement", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText300.Wrap( -1 )

        self.staticText300.SetFont( wx.Font( 12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial" ) )

        bSizer314.Add( self.staticText300, 0, wx.ALL, 5 )


        bSizer314.Add( ( 195, 0), 0, 0, 5 )

        self.staticText301 = wx.StaticText( self, wx.ID_ANY, u"Available", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText301.Wrap( -1 )

        self.staticText301.SetFont( wx.Font( 12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial" ) )

        bSizer314.Add( self.staticText301, 0, wx.ALL, 5 )


        bSizer314.Add( ( 10, 0), 0, wx.EXPAND, 5 )

        self.staticText302 = wx.StaticText( self, wx.ID_ANY, u"In Process", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText302.Wrap( -1 )

        self.staticText302.SetFont( wx.Font( 12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial" ) )

        bSizer314.Add( self.staticText302, 0, wx.ALL, 5 )


        bSizer314.Add( ( 5, 0), 0, wx.EXPAND, 5 )

        self.staticText303 = wx.StaticText( self, wx.ID_ANY, u"Not Available", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText303.Wrap( -1 )

        self.staticText303.SetFont( wx.Font( 12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial" ) )

        bSizer314.Add( self.staticText303, 0, wx.ALL, 5 )


        bSizer314.Add( ( 40, 0), 0, wx.EXPAND, 5 )


        bSizer304.Add( bSizer314, 0, wx.EXPAND, 5 )

        bSizer305 = wx.BoxSizer( wx.HORIZONTAL )

        bSizer306 = wx.BoxSizer( wx.VERTICAL )

        self.staticText294 = wx.StaticText( self, wx.ID_ANY, u"ISO 9001: 2015", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText294.Wrap( -1 )

        self.staticText294.SetFont( wx.Font( 11, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer306.Add( self.staticText294, 1, wx.ALL|wx.EXPAND, 5 )


        bSizer306.Add( ( 0, 20), 1, wx.EXPAND, 5 )

        self.staticText295 = wx.StaticText( self, wx.ID_ANY, u"Improvement in Remanufacturing Activities", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText295.Wrap( -1 )

        self.staticText295.SetFont( wx.Font( 11, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer306.Add( self.staticText295, 1, wx.ALL|wx.EXPAND, 5 )


        bSizer306.Add( ( 0, 20), 1, wx.EXPAND, 5 )

        self.staticText296 = wx.StaticText( self, wx.ID_ANY, u"VTA Compliance", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText296.Wrap( -1 )

        self.staticText296.SetFont( wx.Font( 11, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer306.Add( self.staticText296, 1, wx.ALL|wx.EXPAND, 5 )


        bSizer306.Add( ( 0, 23), 1, wx.EXPAND, 5 )

        self.staticText297 = wx.StaticText( self, wx.ID_ANY, u"4R2S Training", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText297.Wrap( -1 )

        self.staticText297.SetFont( wx.Font( 11, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer306.Add( self.staticText297, 1, wx.ALL|wx.EXPAND, 5 )


        bSizer306.Add( ( 0, 23), 1, wx.EXPAND, 5 )

        self.staticText298 = wx.StaticText( self, wx.ID_ANY, u"Score", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText298.Wrap( -1 )

        self.staticText298.SetFont( wx.Font( 11, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer306.Add( self.staticText298, 1, wx.ALL|wx.EXPAND, 5 )


        bSizer305.Add( bSizer306, 1, wx.EXPAND, 5 )


        bSizer305.Add( ( 50, 0), 0, wx.EXPAND, 5 )

        bSizer307 = wx.BoxSizer( wx.VERTICAL )

        bSizer309 = wx.BoxSizer( wx.HORIZONTAL )

        self.checkBox171 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer309.Add( self.checkBox171, 0, wx.ALL, 5 )


        bSizer309.Add( ( 60, 0), 0, 0, 5 )

        self.checkBox172 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer309.Add( self.checkBox172, 0, wx.ALL, 5 )


        bSizer309.Add( ( 70, 0), 0, 0, 5 )

        self.checkBox173 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer309.Add( self.checkBox173, 0, wx.ALL, 5 )


        bSizer307.Add( bSizer309, 0, wx.EXPAND, 5 )


        bSizer307.Add( ( 0, 33), 0, 0, 5 )

        bSizer310 = wx.BoxSizer( wx.HORIZONTAL )

        self.checkBox174 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer310.Add( self.checkBox174, 0, wx.ALL, 5 )


        bSizer310.Add( ( 60, 0), 0, 0, 5 )

        self.checkBox175 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer310.Add( self.checkBox175, 0, wx.ALL, 5 )


        bSizer310.Add( ( 70, 0), 0, 0, 5 )

        self.checkBox176 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer310.Add( self.checkBox176, 0, wx.ALL, 5 )


        bSizer307.Add( bSizer310, 0, wx.EXPAND, 5 )


        bSizer307.Add( ( 0, 30), 0, 0, 5 )

        bSizer311 = wx.BoxSizer( wx.HORIZONTAL )

        self.checkBox177 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer311.Add( self.checkBox177, 0, wx.ALL, 5 )


        bSizer311.Add( ( 60, 0), 0, 0, 5 )

        self.checkBox178 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer311.Add( self.checkBox178, 0, wx.ALL, 5 )


        bSizer311.Add( ( 70, 0), 0, 0, 5 )

        self.checkBox179 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer311.Add( self.checkBox179, 0, wx.ALL, 5 )


        bSizer307.Add( bSizer311, 0, wx.EXPAND, 5 )


        bSizer307.Add( ( 0, 30), 0, 0, 5 )

        bSizer312 = wx.BoxSizer( wx.HORIZONTAL )

        self.checkBox180 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer312.Add( self.checkBox180, 0, wx.ALL, 5 )


        bSizer312.Add( ( 60, 0), 0, 0, 5 )

        self.checkBox181 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer312.Add( self.checkBox181, 0, wx.ALL, 5 )


        bSizer312.Add( ( 70, 0), 0, 0, 5 )

        self.checkBox182 = wx.CheckBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer312.Add( self.checkBox182, 0, wx.ALL, 5 )


        bSizer307.Add( bSizer312, 0, wx.EXPAND, 5 )


        bSizer307.Add( ( 0, 25), 0, 0, 5 )

        bSizer313 = wx.BoxSizer( wx.HORIZONTAL )

        self.textCtrl91 = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer313.Add( self.textCtrl91, 0, wx.ALL, 5 )


        bSizer307.Add( bSizer313, 0, 0, 5 )


        bSizer305.Add( bSizer307, 1, wx.EXPAND, 5 )


        bSizer304.Add( bSizer305, 0, wx.EXPAND, 5 )

        bSizer315 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer315.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button89 = wx.Button( self, wx.ID_ANY, u"Reset", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button89.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer315.Add( self.button89, 0, wx.ALL, 5 )

        self.button90 = wx.Button( self, wx.ID_ANY, u"Save", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button90.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer315.Add( self.button90, 0, wx.ALL, 5 )


        bSizer304.Add( bSizer315, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer304 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.datePicker4.Bind( wx.adv.EVT_DATE_CHANGED, self.Date4 )
        self.button89.Bind( wx.EVT_BUTTON, self.DateSet )
        self.checkBox171.Bind( wx.EVT_CHECKBOX, self.check1 )
        self.checkBox172.Bind( wx.EVT_CHECKBOX, self.check2 )
        self.checkBox173.Bind( wx.EVT_CHECKBOX, self.check3 )
        self.checkBox174.Bind( wx.EVT_CHECKBOX, self.check4 )
        self.checkBox175.Bind( wx.EVT_CHECKBOX, self.check5 )
        self.checkBox176.Bind( wx.EVT_CHECKBOX, self.check6 )
        self.checkBox177.Bind( wx.EVT_CHECKBOX, self.check7 )
        self.checkBox178.Bind( wx.EVT_CHECKBOX, self.check8 )
        self.checkBox179.Bind( wx.EVT_CHECKBOX, self.check9 )
        self.checkBox180.Bind( wx.EVT_CHECKBOX, self.check10 )
        self.checkBox181.Bind( wx.EVT_CHECKBOX, self.check11 )
        self.checkBox182.Bind( wx.EVT_CHECKBOX, self.check12 )
        self.textCtrl91.Bind( wx.EVT_TEXT, self.score1 )
        self.button89.Bind( wx.EVT_BUTTON, self.Reset )
        self.button90.Bind( wx.EVT_BUTTON, self.Save )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Date4( self, event ):
            global mine4
            mine4 = event.GetDate()

    def DateSet( self, event ):
            global moo4
            moo4 = mine4.Format("%d/%m/%Y")

    def check1( self, event ):
            global u4
            box177 = self.checkBox171.GetValue()

            if box177 == True:
                u4 = 1
                #put177.value = "/"
            else:
                u4 = 0

    def check2( self, event ):
            global v4
            box178 = self.checkBox172.GetValue()

            if box178 == True:
                v4 = 1
                #put178.value = "O"
            else:
                v4 = 0

    def check3( self, event ):
            global w4
            box179 = self.checkBox173.GetValue()

            if box179 == True:
                w4 = 1
                #put179.value = "X"
                
            else:
                w4 = 0

    def check4( self, event ):
            global x4
            box180 = self.checkBox180.GetValue()

            if box180 == True:
                x4 = 1
                #put180.value = "/"
            else:
                x4 = 0

    def check5( self, event ):
            global y4
            box181 = self.checkBox175.GetValue()

            if box181 == True:
                y4 = 1
                #put181.value = "O"
            else:
                y4 = 0

    def check6( self, event ):
            global z4
            box182 = self.checkBox176.GetValue()

            if box182 == True:
                z4 = 1
                #put182.value = "X"
            else:
                z4 = 0

    def check7( self, event ):
            global a5
            box183 = self.checkBox177.GetValue()

            if box183 == True:
                #input183.value = "/"
                a5 = 1
            else:
                a5 = 0

    def check8( self, event ):
            global b5
            box184 = self.checkBox178.GetValue()

            if box184 == True:
                b5 = 1
                #put184.value = "O"
            else:
                b5 = 0

    def check9( self, event ):
            global c5
            box185 = self.checkBox179.GetValue()

            if box185 == True:
                c5 = 1
                #put185.value = "X"
            else:
                c5 = 0

    def check10( self, event ):
            global d5
            box186 = self.checkBox180.GetValue()

            if box186 == True:
                #put186.value = "/"
                d5 = 1
            else:
                d5 = 0

    def check11( self, event ):
            global e5
            box187 = self.checkBox181.GetValue()

            if box187 == True:
                #put187.value = "O"
                e5 = 1
            else:
                e5 = 0

    def check12( self, event ):
            global f5
            box188 = self.checkBox182.GetValue()

            if box188 == True:
                #put188.value = "X"
                f5 = 1
            else:
                f5 = 1

    def score1( self, event ):
            global get39
            get39 = self.textCtrl91.GetValue()

    def Reset( self, event ):
            self.checkBox171.SetValue(False)
            self.checkBox172.SetValue(False)
            self.checkBox173.SetValue(False)
            self.checkBox174.SetValue(False)
            self.checkBox175.SetValue(False)
            self.checkBox176.SetValue(False)
            self.checkBox177.SetValue(False)
            self.checkBox178.SetValue(False)
            self.checkBox179.SetValue(False)
            self.checkBox180.SetValue(False)
            self.checkBox181.SetValue(False)
            self.checkBox182.SetValue(False)
            self.textCtrl91.SetValue("")
            
            DataReset = Reset(parent = self.button89)
            SoundReset = soundReset(parent = self.button89)
            
            DataReset.Show()
            SoundReset.Resetto.Play()

    def Save( self, event ):
            global word39
            
            global b79x
            global b80x
            global b81x
            global b82x
            global b83x
            global b84x
            global b85x
            global b86x
            global b87x
            global b88x
            global b89x
            global b90x
            global b91x
            
            word39 = int(get39)
            
            if u4 == 1:
                b79x = 1
            else:
                b79x = 0
                
            if v4 == 1:
                b80x = 1
            else:
                b80x = 0
                
            if w4 == 1:
                b81x = 1
            else:
                b81x = 0
                
            if x4 == 1:
                b82x = 1
            else:
                b82x = 0
                
            if y4 == 1:
                b83x = 1
            else:
                b83x = 0
                
            if z4 == 1:
                b85x = 1
            else:
                b85x = 0
                
            if a5 == 1:
                b86x = 1
            else:
                b86x = 0
                
            if b5 == 1:
                b87x = 1
            else:
                b87x = 0
                
            if c5 == 1:
                b88x = 1
            else:
                b88x = 0
                
            if d5 == 1:
                b89x = 1
            else:
                b89x = 0
            
            if e5 == 1:
                b90x = 1
            else:
                b90x = 0
                
            if f5 == 1:
                b91x = 1
            else:
                b91x = 0
                
            SavingScore = ScoreSaved(parent = self.button90)
            AcceptSound = soundAccept(parent = self.button90)
            
            SavingScore.Show()
            AcceptSound.Accept.Play()
    
                
###########################################################################
## Class Generate
###########################################################################

class Generate ( wx.Frame ):
    global FileName1
    global FileName2
    
    global PracTotal
    global PercentageTotal
    global TrueTotal
    global Gradex
    
    
    momo = ""
    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Generate File", pos = wx.DefaultPosition, size = wx.Size( 500,270 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(500,270), wx.Size(500,270) )
        self.SetForegroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_BTNTEXT ) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

        bSizer314 = wx.BoxSizer( wx.VERTICAL )


        bSizer314.Add( ( 0, 50), 0, wx.EXPAND, 5 )

        bSizer315 = wx.BoxSizer( wx.VERTICAL )

        self.staticText282 = wx.StaticText( self, wx.ID_ANY, u"Please Select A File", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText282.Wrap( -1 )

        self.staticText282.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial Black" ) )

        bSizer315.Add( self.staticText282, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.dirPicker3 = wx.DirPickerCtrl( self, wx.ID_ANY, wx.EmptyString, u"Select a folder", wx.DefaultPosition, wx.Size( -1,50 ), wx.DIRP_DEFAULT_STYLE )
        self.dirPicker3.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer315.Add( self.dirPicker3, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL|wx.EXPAND, 5 )


        bSizer315.Add( ( 0, 25), 0, wx.EXPAND, 5 )

        bSizer316 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer316.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button81 = wx.Button( self, wx.ID_ANY, u"Cancel", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button81.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer316.Add( self.button81, 0, wx.ALL, 5 )

        self.button82 = wx.Button( self, wx.ID_ANY, u"Generate", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button82.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer316.Add( self.button82, 0, wx.ALL, 5 )


        bSizer315.Add( bSizer316, 1, wx.EXPAND, 5 )


        bSizer314.Add( bSizer315, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer314 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button81.Bind( wx.EVT_BUTTON, self.Cancel )
        self.button82.Bind( wx.EVT_BUTTON, self.Generate )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Cancel( self, event ):
        #event.Skip()
        self.Destroy()

    def Generate( self, event ):
        global FileName1
        global FileName2
        
        global PracTotal
        global PercentageTotal
        global TrueTotal
        global Gradex
        

#############################################################################
########## Audit Score Detail Company and Scoring ###########################
#############################################################################

        #For Company Name
        name1 = LoadWorkbook(parent=self.button82)
        mname1 = name1.wb.get_sheet_by_name('Cover Page')
        pick1 = mname1['E13']
        if a == 1:
            pick1.value = "{}".format(map1)
        else:
            pick1.value = ""
                
        #For Address
        mname2 = name1.wb.get_sheet_by_name('Cover Page')
        pick2 = mname2['E18']
        if b == 1:
            pick2.value = "{}".format(map2)
        else:
            pick2.value = ""
        
        #For Address1
        mname3 = name1.wb.get_sheet_by_name('Cover Page')
        pick3 = mname3['E19']
        if c == 1:
            pick3.value = "{}".format(map3)
        else:
            pick3.value = ""
        
        #For Address2
        mname4 = name1.wb.get_sheet_by_name('Cover Page')
        pick4 = mname4['E20']
        if d == 1:
            pick4.value = "{}".format(map4)
        else:
            pick4.value = ""
            
        #For Address3
        mname5 = name1.wb.get_sheet_by_name('Cover Page')
        pick5 = mname5['E21']
        if e == 1:
            pick5.value = "{}".format(map5)
        else:
            pick5.value = ""
            
        #For Address4
        mname6 = name1.wb.get_sheet_by_name('Cover Page')
        pick6 = mname6['E22']
        if f == 1:
            pick6.value = "{}".format(map6)
        else:
            pick6.value = ""
            
        #For Address5
        mname7 = name1.wb.get_sheet_by_name('Cover Page')
        pick7 = mname7['E23']
        if g == 1:
            pick7.value = "{}".format(map7)
        else:
            pick7.value = ""
            
        #For Auditee
        mname8 = name1.wb.get_sheet_by_name('Cover Page')
        pick8 = mname8['K9']
        if h == 1:
            pick8.value = "{}".format(map8)
        else:
            pick8.value = ""
            
        #For Contact1
        mname9 = name1.wb.get_sheet_by_name('Cover Page')
        pick9 = mname9['I16']
        if i == 1:
            pick9.value = "{}".format(map9)
        else:
            pick9.value = ""
            
        #For email1
        mname10 = name1.wb.get_sheet_by_name('Cover Page')
        pick10 = mname10['I19']
        if j == 1:
            pick10.value = "{}".format(map10)
        else:
            pick10.value = ""
            
        #For email2
        mname11 = name1.wb.get_sheet_by_name('Cover Page')
        pick11 = mname11['I20']
        if k == 1:
            pick11.value = "{}".format(map11)
        else:
            pick11.value = ""
            
        #For email3
        mname12 = name1.wb.get_sheet_by_name('Cover Page')
        pick12 = mname12['I21']
        if l == 1:
            pick12.value = "{}".format(map12)
        else:
            pick12.value = ""
            
        #For Audit
        mname13 = name1.wb.get_sheet_by_name('Cover Page')
        pick13 = mname13['E11']
        if m == 1:
            pick13.value = "{}".format(map13)
        else:
            pick13.value = ""
            
        #For Plant
        mname14 = name1.wb.get_sheet_by_name('Cover Page')
        pick14 = mname14['E16']
        if n == 1:
            pick14.value = "{}".format(map14)
        else:
            pick14.value = ""
        
#############################################################################
########## Checklist New  ###################################################
#############################################################################
        #For auditor1name
        nexus1 = LoadWorkbook2(parent=self.button82)
        mnome1 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill1 = mnome1['D80']
        if a11 == 1:
            pill1.value = "{}".format(mup1)
        else:
            pill1.value = ""
            
        #For Checklist New Date
        mdateChecklistNew = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        checklistnewdate1 = mdateChecklistNew['D86']
        checklistnewdate2 = mdateChecklistNew['E86']
        checklistnewdate3 = mdateChecklistNew['G86']
        
        if aiy2 == 1:
            checklistnewdate1.value = "{}".format(moo2)
            checklistnewdate2.value = "{}".format(moo2)
            checklistnewdate3.value = "{}".format(moo2)
        else:
            checklistnewdate1.value = ""
            checklistnewdate2.value = ""
            checklistnewdate3.value = ""
        
        
        
        #For auditor1position
        mpoisson1 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill2 = mpoisson1['D82']
        if b11 == 1:
            pill2.value = "{}".format(mup2)
        else:
            pill2.value = ""
            
        #For auditor1organisation
        morgan1 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill3 = morgan1['D84']
        if c11 == 1:
            pill3.value = "{}".format(mup3)
        else:
            pill3.value = ""
            
        #For auditor2name
        mnome2 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill4 = mnome2['E80']
        if d11 == 1:
            pill4.value = "{}".format(mup4)
        else:
            pill4.value = ""
            
        #For auditor2position
        mpoisson2 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill5 = mpoisson2['E82']
        if e11 == 1:
            pill5.value = "{}".format(mup5)
        else:
            pill5.value = ""
            
        #For auditor2organisation
        morgan2 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill6 = morgan2['E84']
        if f11 == 1:
            pill6.value = "{}".format(mup6)
        else:
            pill6.value = ""
            
        #For auditor3name
        mnome3 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill7 = mnome3['G80']
        if g11 == 1:
            pill7.value = "{}".format(mup7)
        else:
            pill7.value = ""
        
        #For auditor3position
        mpoisson3 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill8 = mpoisson3['G82']
        if h11 == 1:
            pill8.value = "{}".format(mup8)
        else:
            pill8.value = ""
            
        #For auditor3organisation
        morgan3 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        pill9 = morgan3['G84']
        if i11 == 1:
            pill9.value = "{}".format(mup9)
        else:
            pill9.value = ""
            
        #Clause 7.1 Check 1
        mtick99 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put99 = mtick99['E11']
        if a12 == 1:
            put99.value = "/"
        else:
            put99.value = ""
        
        #Clause 7.1 Check 2
        mtick100 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put100 = mtick100['F11']
        if b12 == 1:
            put100.value = "/"
        else:
            put100.value = ""
        
        #Clause 7.1 Check 3
        mtick101 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put101 = mtick101['G11']
        if c12 == 1:
            put101.value = "/"
        else:
            put101.value = ""
            
        #Clause 7.1 Check 4
        mtick102 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put102 = mtick102['E13']
        if d12 == 1:
            put102.value = "/"
        else:
            put102.value = ""
            
        #Clasue 7.1 Check 5
        mtick103 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put103 = mtick103['F13']
        if e12 == 1:
            put103.value = "/"
        else:
            put103.value = ""
            
        #Clause 7.1 Check 6
        mtick104 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put104 = mtick104['G13']
        if f12 == 1:
            put104.value = "/"
        else:
            put104.value = ""
            
        #Clause 7.1 Check 7
        mtick105 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put105 = mtick105['E15']
        if g12 == 1:
            put105.value = "/"
        else:
            put105.value = ""
            
        #Clause 7.1 Check 8
        mtick106 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put106 = mtick106['F15']
        if h12 == 1:
            put106.value = "/"
        else:
            put106.value = ""
            
        #Clause 7.1 Check 9
        mtick107 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put107 = mtick107['G15']
        if i12 == 1:
            put107.value = "/"
        else:
            put107.value = ""
            
        #Clause 7.1 Check 10
        mtick108 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put108 = mtick108['E17']
        if j12 == 1:
            put108.value = "/"
        else:
            put108.value = ""
            
        #Clause 7.1 Check 11
        mtick109 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put109 = mtick109['F17']
        if k12 == 1:
            put109.value = "/"
        else:
            put109.value = ""
        
        #Clause 7.1 Check 12
        mtick110 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put110 = mtick110['G17']
        if l12 == 1:
            put110.value = "/"
        else:
            put110.value = ""
            
        #Clause 7.1 Check 13
        mtick111 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put111 = mtick111['E19']
        if m12 == 1:
            put111.value = "/"
        else:
            put111.value = ""
            
        #Clause 7.1 Check 14
        mtick112 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put112 = mtick112['F19']
        if n12 == 1:
            put112.value = "/"
        else:
            put112.value = ""
            
        #Clause 7.1 Check 15
        mtick113 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put113 = mtick113['G19']
        if o12 == 1:
            put113.value = "/"
        else:
            put113.value = ""
        
        #Clause 7.1 Remark1
        mcomment1 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark1 = mcomment1['H11']
        if aa == 1:
            mark1.value = "{}".format(mop1)
        else:
            mark1.value = ""
            
        #Clause 7.1 Remark2
        mcomment2 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark2 = mcomment2['H13']
        if ba == 1:
            mark2.value = "{}".format(mop2)
        else:
            mark2.value = ""
            
        #Clause 7.1 Remark3
        mcomment3 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark3 = mcomment3['H15']
        if ca == 1:
            mark3.value = "{}".format(mop3)
        else:
            mark3.value = ""
            
        #Clause 7.1 Remark4
        mcomment4 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark4 = mcomment4['H17']
        if da == 1:
            mark4.value = "{}".format(mop4)
        else:
            mark4.value = ""
            
        #Clause 7.1 Remark5
        mcomment5 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark5 = mcomment5['H19']
        if ea == 1:
            mark5.value = "{}".format(mop5)
        else:
            mark5.value = ""
        
        #Clause 7.3 Check1
        mtick114 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put114 = mtick114['E21']
        if p12 == 1:
            put114.value = "/"
        else:
            put114.value = ""
            
        #Clause 7.3 Check2
        mtick115 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put115 = mtick115['F21']
        if q12 == 1:
            put115.value = "/"
        else:
            put115.value = ""
        
        #Clause 7.3 Check3
        mtick116 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put116 = mtick116['G21']
        if r12 == 1:
            put116.value = "/"
        else:
            put116.value = ""
            
        #Clause 7.3 Check4
        mtick117 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put117 = mtick117['E24']
        if s12 == 1:
            put117.value = "/"
        else:
            put117.value = ""
            
        #Clause 7.3 Check5
        mtick118 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put118 = mtick118['F24']
        if t12 == 1:
            put118.value = "/"
        else:
            put118.value = ""
            
        #Clause 7.3 Check6
        mtick119 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put119 = mtick119['G24']
        if u12 == 1:
            put119.value = "/"
        else:
            put119.value = ""
            
        #Clause 7.3 Check7
        mtick120 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put120 = mtick120['E26']
        if v12 == 1:
            put120.value = "/"
        else:
            put120.value = ""
            
        #Clause 7.3 Check8
        mtick121 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put121 = mtick121['F26']
        if w12 == 1:
            put121.value = "/"
        else:
            put121.value = ""
            
        #Clause 7.3 Check9
        mtick122 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put122 = mtick122['G26']
        if x12 == 1:
            put122.value = "/"
        else:
            put122.value = ""
            
        #Clause 7.3 Check10
        mtick123 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put123 = mtick123['E28']
        if y12 == 1:
            put123.value = "/"
        else:
            put123.value = ""
            
        #Clause 7.3 Check11
        mtick124 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put124 = mtick124['F28']
        if z12 == 1:
            put124.value = "/"
        else:
            put124.value = ""
        
        #Clause 7.3 Check12
        mtick125 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put125 = mtick125['G28']
        if a13 == 1:
            put125.value = "/"
        else:
            put125.value = ""
        
        #Clause 7.3 Check13
        mtick126 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put126 = mtick126['E30']
        if b13 == 1:
            put126.value = "/"
        else:
            put126.value = ""
        
        #Clause 7.3 Check14
        mtick127 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put127 = mtick127['F30']
        if c13 == 1:
            put127.value = "/"
        else:
            put127.value = ""
            
        #Clause 7.3 Check15
        mtick128 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put128 = mtick128['G30']
        if d13 == 1:
            put128.value = "/"
        else:
            put128.value = ""
        
        #Clause 7.3 Check16
        mtick129 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put129 = mtick129['E31']
        if e13 == 1:
            put129.value = "/"
        else:
            put129.value = ""
            
        #Clause 7.3 Check17
        mtick130 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put130 = mtick130['F31']
        if f13 == 1:
            put130.value = "/"
        else:
            put130.value = ""
            
        #Clause 7.3 Check18
        mtick131 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put131 = mtick131['G31']
        if g13 == 1:
            put131.value = "/"
        else:
            put131.value = ""
            
        #Clause 7.3 Check19
        mtick132 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put132 = mtick132['E32']
        if h13 == 1:
            put132.value = "/"
        else:
            put132.value = ""
            
        #Clause 7.3 Check20
        mtick133 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put133 = mtick133['F32']
        if i13 == 1:
            put133.value = "/"
        else:
            put133.value = ""
            
        #Clause 7.3 Check21
        mtick134 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put134 = mtick134['G32']
        if j13 == 1:
            put134.value = "/"
        else:
            put134.value = ""
        
        #Clause 7.3 Check22
        mtick135 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put135 = mtick135['E35']
        if k13 == 1:
            put135.value = "/"
        else:
            put135.value = ""
        
        #Clause 7.3 Check23
        mtick136 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put136 = mtick136['F35']
        if l13 == 1:
            put136.value = "/"
        else:
            put136.value = ""
        
        #Clause 7.3 Check24
        mtick137 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put137 = mtick137['G35']
        if m13 == 1:
            put137.value = "/"
        else:
            put137.value = ""
        
        #Clause 7.3 Check25
        mtick138 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put138 = mtick138['E37']
        if n13 == 1:
            put138.value = "/"
        else:
            put138.value = ""
        
        #Clause 7.3 Check26
        mtick139 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put139 = mtick139['F37']
        if o13 == 1:
            put139.value = "/"
        else:
            put139.value = ""
        
        #Clause 7.3 Check27
        mtick140 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put140 = mtick140['G37']
        if p13 == 1:
            put140.value = "/"
        else:
            put140.value = ""
        
        #Clause 7.3 Check28
        mtick141 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put141 = mtick141['E39']
        if q13 == 1:
            put141.value = "/"
        else:
            put141.value = ""
            
        #Clause 7.3 Check29
        mtick142 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put142 = mtick142['F39']
        if r13 == 1:
            put142.value = "/"
        else:
            put142.value = ""
            
        #Clause 7.3 Check30
        mtick143 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put143 = mtick143['G39']
        if s13 == 1:
            put143.value = "/"
        else:
            put143.value = ""
        
        #Clause 7.3 Remark1
        mcomment6 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark6 = mcomment6['H21']
        if ab == 1:
            mark6.value = "{}".format(mop6)
        else:
            mark6.value = ""
            
        #Clause 7.3 Remark2
        mcomment7 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark7 = mcomment7['H24']
        if bb == 1:
            mark7.value = "{}".format(mop7)
        else:
            mark7.value = ""
            
        #Clause 7.3 Remark3
        mcomment8 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark8 = mcomment8['H26']
        if cb == 1:
            mark8.value = "{}".format(mop8)
        else:
            mark8.value = ""
            
        #Clause 7.3 Remark4
        mcomment9 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark9 = mcomment9['H28']
        if db == 1:
            mark9.value = "{}".format(mop9)
        else:
            mark9.value = ""
            
        #Clause 7.3 Remark5
        mcomment10 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark10 = mcomment10['H30']
        if eb == 1:
            mark10.value = "{}".format(mop10)
        else:
            mark10.value = ""
            
        #Clause 7.3 Remark6
        mcomment11 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark11 = mcomment11['H31']
        if fb == 1:
            mark11.value = "{}".format(mop11)
        else:
            mark11.value = ""
        
        #Clause 7.3 Remark7
        mcomment12 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark12 = mcomment12['H32']
        if gb == 1:
            mark12.value = "{}".format(mop12)
        else:
            mark12.value = ""
        
        #Clause 7.3 Remark8
        mcomment13 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark13 = mcomment13['H35']
        if hb == 1:
            mark13.value = "{}".format(mop13)
        else:
            mark13.value = ""
            
        #Clause 7.3 Remark9
        mcomment14 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark14 = mcomment14['H37']
        if ib == 1:
            mark14.value = "{}".format(mop14)
        else:
            mark14.value = ""            
        
        #Clause 7.3 Remark10
        mcomment15 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark15 = mcomment15['H39']
        if jb == 1:
            mark15.value = "{}".format(mop15)
        else:
            mark15.value = ""
            
        #Clause 9.1 Check1
        mtick144 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put144 = mtick144['E42']
        if t13 == 1:
            put144.value = "/"
        else:
            put144.value = ""
            
        #Clause 9.1 Check2
        mtick145 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put145 = mtick145['F42']
        if u13 == 1:
            put145.value = "/"
        else:
            put145.value = ""
        
        #Clause 9.1 Check3
        mtick146 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put146 = mtick146['G42']
        if v13 == 1:
            put146.value = "/"
        else:
            put146.value = ""
        
        #Clause 9.1 Check4
        mtick147 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put147 = mtick147['E42']
        if w13 == 1:
            put147.value = "/"
        else:
            put147.value = ""
            
        #Clause 9.1 Check5
        mtick148 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put148 = mtick148['F42']
        if x13 == 1:
            put148.value = "/"
        else:
            put148.value  = ""
            
        #Clause 9.1 Check6
        mtick149 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put149 = mtick149['G42']
        if y13 == 1:
            put149.value = "/"
        else:
            put149.value = ""
        
        #Clause 9.1 Check7
        mtick150 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put150 = mtick150['E45']
        if z13 == 1:
            put150.value = "/"
        else:
            put150.value = ""
        
        #Clause 9.1 Check8
        mtick151 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put151 = mtick151['F45']
        if a14 == 1:
            put151.value = "/"
        else:
            put151.value = ""
        
        #Clause 9.1 Check9
        mtick152 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put152 = mtick152['G45']
        if b14 == 1:
            put152.value = "/"
        else:
            put152.value = ""
            
        #Clause 9.1 Check10
        mtick153 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put153 = mtick153['E46']
        if c14 == 1:
            put153.value = "/"
        else:
            put153.value = ""
        
        #Clause 9.1 Check11
        mtick154 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put154 = mtick154['F46']
        if d14 == 1:
            put154.value = "/"
        else:
            put154.value = ""
        
        #Clause 9.1 Check12
        mtick155 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put155 = mtick155['G46']
        if e14 == 1:
            put155.value = "/"
        else:
            put155.value = ""
        
        #Clause 9.1 Check13
        mtick156 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put156 = mtick156['E47']
        if f14 == 1:
            put156.value = "/"
        else:
            put156.value = ""
        
        #Clause 9.1 Check14
        mtick157 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put157 = mtick157['F47']
        if g14 == 1:
            put157.value = "/"
        else:
            put157.value = ""
        
        #Clause 9.1 Check15
        mtick158 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put158 = mtick158['G47']
        if h14 == 1:
            put158.value = "/"
        else:
            put158.value = ""
        
        #Clause 9.1 Check16
        mtick159 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put159 = mtick159['E48']
        if i14 == 1:
            put159.value = "/"
        else:
            put159.value = ""
        
        #Clause 9.1 Check17
        mtick160 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put160 = mtick160['F48']
        if j14 == 1:
            put160.value = "/"
        else:
            put160.value = ""
        
        #Clause 9.1 Check18
        mtick161 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put161 = mtick161['G48']
        if k14 == 1:
            put161.value = "/"
        else:
            put161.value = ""
        
        #Clause 9.1 Check19
        mtick162 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put162 = mtick162['E49']
        if l14 == 1:
            put162.value = "/"
        else:
            put162.value = ""
        
        #Clause 9.1 Check20
        mtick163 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put163 = mtick163['F49']
        if m14 == 1:
            put163.value = "/"
        else:
            put163.value = ""
        
        #Clause 9.1 Check21
        mtick164 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put164 = mtick164['G49']
        if n14 == 1:
            put164.value = "/"
        else:
            put164.value = ""
            
        #Clause 9.1 Remark1
        mcomment16 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark16 = mcomment16['H42']
        if ac == 1:
            mark16.value = "{}".format(mop16)
        else:
            mark16.value = ""
        
        #Clause 9.1 Remark2
        mcomment17 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark17 = mcomment17['H45']
        if bc == 1:
            mark17.value = "{}".format(mop17)
        else:
            mark17.value = ""
        
        #Clause 9.1 Remark3
        mcomment18 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark18 = mcomment18['H46']
        if cc == 1:
            mark18.value = "{}".format(mop18)
        else:
            mark18.value = ""
        
        #Clause 9.1 Remark4
        mcomment19 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark19 = mcomment19['H47']
        if dc == 1:
            mark19.value = "{}".format(mop19)
        else:
            mark19.value = ""
        
        #Clause 9.1 Remark5
        mcomment20 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark20 = mcomment20['H47']
        if ec == 1:
            mark20.value = "{}".format(mop20)
        else:
            mark20.value = ""
        
        #Clause 9.1 Remark6
        mcomment21 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark21 = mcomment21['H48']
        if fc == 1:
            mark21.value = "{}".format(mop21)
        else:
            mark21.value = ""
        
        #Clause 9.1 Remark7
        mcomment22 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark22 = mcomment22['H49']
        if gc == 1:
            mark22.value = "{}".format(mop22)
        else:
            mark22.value = ""
        
        #Clause 10 Check1
        mtick165 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put165 = mtick165['E50']
        if o14 == 1:
            put165.value = "/"
        else:
            put165.value = ""
        
        #Clause 10 Check2
        mtick166 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put166 = mtick166['F50']
        if p14 == 1:
            put166.value = "/"
        else:
            put166.value = ""
        
        #Clause 10 Check3
        mtick167 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put167 = mtick167['G50']
        if q14 == 1:
            put167.value = "/"
        else:
            put167.value = ""
        
        #Clause 10 Check4
        mtick168 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put168 = mtick168['E53']
        if r14 == 1:
            put168.value = "/"
        else:
            put168.value = ""
        
        #Clause 10 Check5
        mtick169 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put169 = mtick169['F53']
        if s14 == 1:
            put169.value = "/"
        else:
            put169.value = ""
        
        #Clause 10 Check6
        mtick170 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put170 = mtick170['G53']
        if t14 == 1:
            put170.value = "/"
        else:
            put170.value = ""
        
        #Clause 10 Check7
        mtick171 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put171 = mtick171['E55']
        if u14 == 1:
            put171.value = "/"
        else:
            put171.value = ""
        
        #Clause 10 Check8
        mtick172 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put172 = mtick172['F55']
        if v14 == 1:
            put172.value = "/"
        else:
            put172.value = ""
        
        #Clause 10 Check9
        mtick173 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put173 = mtick173['G55']
        if w14 == 1:
            put173.value = "/"
        else:
            put173.value = ""
        
        #Clause 10 Remark1
        mcomment23 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark23 = mcomment23['H50']
        if ad == 1:
            mark23.value = "{}".format(mop23)
        else:
            mark23.value = ""
        
        #Clause 10 Remark2
        mcomment24 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark24 = mcomment24['H53']
        if bd == 1:
            mark24.value = "{}".format(mop24)
        else:
            mark24.value = ""
        
        #Clause 10 Remark3
        mcomment25 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark25 = mcomment25['H55']
        if cd == 1:
            mark25.value = "{}".format(mop25)
        else:
            mark25.value = ""
            
        #Clause 11 Check1
        mtick174 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put174 = mtick174['E57']
        if x14 == 1:
            put174.value = "/"
        else:
            put174.value = ""
        
        #Clause Check2
        mtick175 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put175 = mtick175['F57']
        if y14 == 1:
            put175.value = "/"
        else:
            put175.value = ""
        
        #Clause Check3
        mtick176 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        put176 = mtick176['G57']
        if z14 == 1:
            put176.value = "/"
        else:
            put176.value = ""
        
        #Clause Remark1
        mcomment26 = nexus1.wb2.get_sheet_by_name('Audit Checklist')
        mark26 = mcomment26['H57']
        if ae == 1:
            mark26.value = "{}".format(mop26)
        else:
            mark26.value = ""
            
#############################################################################
########## Audit Score  #####################################################
#############################################################################
        #Audit Score Date
        mdateaudit = name1.wb.get_sheet_by_name('Cover Page')
        auditdate1 = mdateaudit['E9']
        auditdate2 = mdateaudit['D48']
        auditdate3 = mdateaudit['I48']
        if aiy1 == 1:
            auditdate1.value = "{}".format(moo3)
            auditdate2.value = "{}".format(moo3)
            auditdate3.value = "{}".format(moo3)
        else:
            auditdate1.value = ""
            auditdate2.value = ""
            auditdate3.value = ""
            
        #Section 7.1 Score1
        mscore1 = name1.wb.get_sheet_by_name('Sec 7.1')
        input1 = mscore1['C8']
        if a15 == 1:
            input1.value = int("{}".format(word1))
        else:
            input1.value = "0"
        
        #Section 7.1 Score2
        mscore2 = name1.wb.get_sheet_by_name('Sec 7.1')
        input2 = mscore2['C12']
        if b15 == 1:
            input2.value = int("{}".format(word2))
        else:
            input2.value = "0"
            
        #Section 7.1 Score3
        mscore3 = name1.wb.get_sheet_by_name('Sec 7.1')
        input3 = mscore3['C16']
        if c15 == 1:
            input3.value = int("{}".format(word3))
        else:
            input3.value = "0"
        
        #Section 7.1 Score4
        mscore4 = name1.wb.get_sheet_by_name('Sec 7.1')
        input4 = mscore4['C21']
        if d15 == 1:
            input4.value = int("{}".format(word4))
        else:
            input4.value = "0"
        
        #Section 7.1 Score5
        mscore5 = name1.wb.get_sheet_by_name('Sec 7.1')
        input5 = mscore5['C25']
        if e15 == 1:
            input5.value = int("{}".format(word5))
        else:
            input5.value = "0"
        
        #Section 7.1 Score6
        mscore6 = name1.wb.get_sheet_by_name('Sec 7.1')
        input6 = mscore6['C29']
        if f15 == 1:
            input6.value = int("{}".format(word6))
        else:
            input6.value = "0"
        
        #Section 7.1 Score7
        mscore7 = name1.wb.get_sheet_by_name('Sec 7.1')
        input7 = mscore7['C33']
        if g15 == 1:
            input7.value = int("{}".format(word7))
        else:
            input7.value = "0"
        
        #Section 7.1 Score8
        mscore8 = name1.wb.get_sheet_by_name('Sec 7.1')
        input8 = mscore8['C37']
        if h15 == 1:
            input8.value = int("{}".format(word8))
        else:
            input8.value = "0"
        
        #Section 7.1 Score9
        mscore9 = name1.wb.get_sheet_by_name('Sec 7.1')
        input9 = mscore9['C41']
        if i15 == 1:
            input9.value = int("{}".format(word9))
        else:
            input9.value = "0"
        
        
        #Section 7.3.1(a) Score1
        mscore10 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input10 = mscore10['C9']
        if j15 == 1:
            input10.value = int("{}".format(word10))
        else:
            input10.value = "0"
        
        #Section 7.3.1(a) Score2
        mscore11 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input11 = mscore11['C13']
        if k15 == 1:
            input11.value = int("{}".format(word11))
        else:
            input11.value = "0"
            
        #Section 7.3.1(a) Score3
        mscore12 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input12 = mscore12['C17']
        if l15 == 1:
            input12.value = int("{}".format(word12))
        else:
            input12.value = "0"
        
        #Section 7.3.1(a) Score4
        mscore13 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input13 = mscore13['C21']
        if m15 == 1:
            input13.value = int("{}".format(word13))
        else:
            input13.value = "0"
        
        #Section 7.3.1(a) Score5
        mscore14 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input14 = mscore14['C25']
        if n15 == 1:
            input14.value = int("{}".format(word14))
        else:
            input14.value = "0"
        
        #Section 7.3.1(a) Score6
        mscore15 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input15 = mscore15['C29']
        if o15 == 1:
            input15.value = int("{}".format(word15))
        else:
            input15.value = "0"
        
        #Section 7.3.1(a) Score7
        mscore16 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input16 = mscore16['C34']
        if p15 == 1:
            input16.value = int("{}".format(word16))
        else:
            input16.value = "0"
        
        #Section 7.3.1(a) Score8
        mscore17 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input17 = mscore17['C38']
        if q15 == 1:
            input17.value = int("{}".format(word17))
        else:
            input17.value = "0"
        
        #Section 7.3.1(a) Score9
        mscore18 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input18 = mscore18['C42']
        if r15 == 1:
            input18.value = int("{}".format(word18))
        else:
            input18.value = "0"
        
        #Section 7.3.1(a) Score10
        mscore19 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input19 = mscore19['C47']
        if s15 == 1:
            input19.value = int("{}".format(word19))
        else:
            input19.value = "0"
            
        #Section 7.3.1(a) Score11
        mscore20 = name1.wb.get_sheet_by_name('Sec 7.3.1 (a)')
        input20 = mscore20['C52']
        if t15 == 1:
            input20.value = int("{}".format(word20))
        else:
            input20.value = "0"
            
            
        #Section 7.3.1(b) Score1
        mscore21 = name1.wb.get_sheet_by_name('Sec 7.3.1 (b)')
        input21 = mscore21['C8']
        if u15 == 1:
            input21.value = int("{}".format(word21))
        else:
            input21.value = "0"
            
        
        #Section 7.3.1(c) Score1
        mscore22 = name1.wb.get_sheet_by_name('Sec 7.3.1 (c)')
        input22 = mscore22['C8']
        if v15 == 1:
            input22.value = int("{}".format(word22))
        else:
            input22.value = "0"
        
        #Section 7.3.1(c) Score2
        mscore23 = name1.wb.get_sheet_by_name('Sec 7.3.1 (c)')
        input23 = mscore23['C12']
        if w15 == 1:
            input23.value = int("{}".format(word23))
        else:
            input23.value = "0"
            
        
        #Section 7.3.1(d) Score1
        mscore24 = name1.wb.get_sheet_by_name('Sec 7.3.1 (d)')
        input24 = mscore24['C7']
        if x15 == 1:
            input24.value = int("{}".format(word24))
        else:
            input24.value = "0"
        
        #Section 7.3.1(d) Score2
        mscore25 = name1.wb.get_sheet_by_name('Sec 7.3.1 (d)')
        input25 = mscore25['C11']
        if y15 == 1:
            input25.value = int("{}".format(word25))
        else:
            input25.value = "0"
        
        #Section 7.3.1(d) Score3
        mscore26 = name1.wb.get_sheet_by_name('Sec 7.3.1 (d)')
        input26 = mscore26['C15']
        if z15 == 1:
            input26.value = int("{}".format(word26))
        else:
            input26.value = "0"
            
            
        #Section 7.3.1(e) Score1
        mscore27 = name1.wb.get_sheet_by_name('Sec 7.3.1 (e)')
        input27 = mscore27['C8']
        if a16 == 1:
            input27.value = int("{}".format(word27))
        else:
            input27.value = "0"
        
        #Section 7.3.1(f) Score1 
        mscore28 = name1.wb.get_sheet_by_name('Sec 7.3.1 (f)')
        input28 = mscore28['C8']
        if b16 == 1:
            input28.value = int("{}".format(word28))
        else:
            input28.value = "0"
        
        #Section 7.3.1(f) Score2
        mscore29 = name1.wb.get_sheet_by_name('Sec 7.3.1 (f)')
        input29 = mscore29['C12']
        if c16 == 1:
            input29.value = int("{}".format(word29))
        else:
            input29.value = "0"
        
        #Section 9.1 Score1
        mscore30 = name1.wb.get_sheet_by_name('Sec 9.1')
        input30 = mscore30['C8']
        if d16 == 1:
            input30.value = int("{}".format(word30))
        else:
            input30.value = "0"
            
        #Section 9.1 Score2
        mscore31 = name1.wb.get_sheet_by_name('Sec 9.1')
        input31 = mscore31['C12']
        if e16 == 1:
            input31.value = int("{}".format(word31))
        else:
            input31.value = "0"
        
        #Section 9.1 Score3
        mscore32 = name1.wb.get_sheet_by_name('Sec 9.1')
        input32 = mscore32['C16']
        if f16 == 1:
            input32.value = int("{}".format(word32))
        else:
            input32.value = "0"
            
        #Section 9.1 Score4
        mscore33 = name1.wb.get_sheet_by_name('Sec 9.1')
        input33 = mscore33['C20']
        if g16 == 1:
            input33.value = int("{}".format(word33))
        else:
            input33.value = "0"
                     
        #Section 10 Score1
        mscore34 = name1.wb.get_sheet_by_name('Sec 10')
        input34 = mscore34['C8'] 
        if h16 == 1:
            input34.value = int("{}".format(word34))
        else:
            input34.value = "0"
        
        #Section 10 Score2
        mscore35 = name1.wb.get_sheet_by_name('Sec 10')
        input35 = mscore35['C12']
        if i16 == 1:
            input35.value = int("{}".format(word35))
        else:
            input35.value = "0"
        
        #Section 10 Score3
        mscore36 = name1.wb.get_sheet_by_name('Sec 10')
        input36 = mscore36['C16']
        if j16 == 1:
            input36.value = int("{}".format(word36))
        else:
            input36.value = "0"
        
        #Section 10 Score4
        mscore37 = name1.wb.get_sheet_by_name('Sec 10')
        input37 = mscore37['C20']
        if k16 == 1:
            input37.value = int("{}".format(word37))
        else:
            input37.value = "0"
        
        
        #Section 11 Score 1
        mscore38 = name1.wb.get_sheet_by_name('Sec 11')
        input38 = mscore38['C8']
        if l16 == 1:
            input38.value = int("{}".format(word38))
        else:
            input38.value = "0"
        

#############################################################################
########## Grade Calculation  ###############################################
#############################################################################            
        TrueTotal = 190
        
        SumX = [total1,total2,total3,total4,total5,total6,total7,total8,total9,total10]
        PracTotal = sum(SumX)
        
        PercentageTotal = int((PracTotal/TrueTotal)*100)
        
        
        #Result Grade
        Nice = name1.wb.get_sheet_by_name('Cover Page')
        Grade = Nice['I34']
        if PercentageTotal < 40:
            Grade.value = "D"
            Gradex = "D"
        elif (PercentageTotal >= 40 and PercentageTotal <= 59):
            Grade.value = "C"
            Gradex = "C"
        elif (PercentageTotal >= 60 and PercentageTotal <= 79):
            Grade.value = "B"
            Gradex = "B"
        elif PercentageTotal > 80:
            Grade.value = "A"
            Gradex = "A"
        else:
            Grade.value = "Error"
            
#############################################################################
########## Saving File  #####################################################
#############################################################################     
        momo = self.dirPicker3.GetPath()
        name1.wb.save(momo + '/Audit Score for {}.xlsx'.format(map1) )
        nexus1.wb2.save(momo + '/Checklist for {}.xlsx'.format(map1) )
        
        FileName1 = momo + '/Audit Score for {}.xlsx'.format(map1) 
        FileName2 = momo + '/Checklist for {}.xlsx'.format(map1)
        
        GenerateFile = GenerateMessage(parent = self.button82)
        AcceptSound = soundAccept(parent = self.button82)
        GenerateFile.Show()
        AcceptSound.Accept.Play()
        
###########################################################################
## Server
###########################################################################
class Server ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Generate File", pos = wx.DefaultPosition, size = wx.Size( 500,255 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )
        self.SetForegroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_BTNTEXT ) )
        self.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_INACTIVECAPTION ) )

        bSizer314 = wx.BoxSizer( wx.VERTICAL )


        bSizer314.Add( ( 0, 50), 0, wx.EXPAND, 5 )

        bSizer315 = wx.BoxSizer( wx.VERTICAL )

        self.staticText282 = wx.StaticText( self, wx.ID_ANY, u"Please Insert the Connection String", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText282.Wrap( -1 )

        self.staticText282.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Arial Black" ) )

        bSizer315.Add( self.staticText282, 0, wx.ALL|wx.EXPAND, 5 )

        self.m_textCtrl89 = wx.TextCtrl( self, wx.ID_ANY, "mongodb://localhost:27017/?readPreference=primary&appname=MongoDB%20Compass&ssl=false", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer315.Add( self.m_textCtrl89, 0, wx.ALL|wx.EXPAND, 5 )


        bSizer315.Add( ( 0, 25), 0, wx.EXPAND, 5 )

        bSizer316 = wx.BoxSizer( wx.HORIZONTAL )


        bSizer316.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.button81 = wx.Button( self, wx.ID_ANY, u"Cancel", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button81.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer316.Add( self.button81, 0, wx.ALL, 5 )

        self.button82 = wx.Button( self, wx.ID_ANY, u"Upload Data", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button82.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer316.Add( self.button82, 0, wx.ALL, 5 )


        bSizer315.Add( bSizer316, 1, wx.EXPAND, 5 )


        bSizer314.Add( bSizer315, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer314 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.m_textCtrl89.Bind( wx.EVT_TEXT, self.finding )
        self.button81.Bind( wx.EVT_BUTTON, self.Cancel )
        self.button82.Bind( wx.EVT_BUTTON, self.Datasent )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def finding( self, event ):
            global servtext
            servtext = self.m_textCtrl89.GetValue()
    
    def Cancel( self, event ):
            #event.Skip()
            self.Destroy()

    def Datasent( self, event ):
            global Connex
            Connex = Connecting(parent = self.button82)
            Connex.Show()
            
            try:
                Great = DataUploaded(self)
                client = pymongo.MongoClient(f"{servtext}",serverSelectionTimeoutMS = 2000)
                client.server_info()
                Connex.Destroy()
                Great.Show()
                
                
            
            except:
                fail = Connectionfail(self)
                fail.Show()
                Connex.Destroy()
                
            mydb = client ["Remanufacture"]
            
            col1 = mydb ["CompanyDetail"]
            col2 = mydb ["AuditReportScore"]
            col3 = mydb ["ChecklistNew"]
            
            comdet = {"Audit Date":moo3,"Audit Standard":map13,"Auditee Name": map8,"Company": map1,"Plant/Location": map14,"Address":map2+map3+map4+map5+map6+map7,"Phone":map9,"E-Mail":mailelement}
            audirepsco = {"Audit Date": moo3 ,"Company":map1,"Score":PracTotal,"Full Score":TrueTotal,"Quality %": PercentageTotal,"Ranking":Gradex}
            chklst = {"Clause 71":b1x,"Clause 711":b2x,"Clause 712":b3x,"Clause 713":b4x,"Clause 714":b5x,"Clause 715":b6x,"Clause 731(a)":b7x,
                      "Clause 731(b)":b8x,"Clause 731(c)":b9x,"Clause 731(d)":b10x,"Clause 731(e)":b11x,"Clause 731(f) & 736":b12x,"Clause 732":b13x,
                      "Clause 733":b14x,"Clause 734":b15x,"Clause 735":b16x,"Clause 91":b17x,"Clause 92(a)":b18x,"Clause 92(b)":b19x,"Clause 92(c)":b20x,
                      "Clause 92(d)":b21x,"Clause 92(e)":b22x,"Clause 92(f)":b23x,"Clause 10(a)":b24x,"Clause 10(b)":b25x,"Clause 10(c)":b26x,"Clause 11":b27x}
            
            col1.insert_one(comdet)
            col2.insert_one(audirepsco)
            col3.insert_one(chklst)

###########################################################################
## Class Connecting
###########################################################################
class Connecting ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 200,200 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )
        self.SetBackgroundColour( wx.Colour( 241, 201, 250 ) )

        bSizer316 = wx.BoxSizer( wx.VERTICAL )

        bSizer326 = wx.BoxSizer( wx.VERTICAL )


        bSizer326.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText284 = wx.StaticText( self, wx.ID_ANY, u"Conecting......\n           :3", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText284.Wrap( -1 )

        self.staticText284.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Candara" ) )

        bSizer326.Add( self.staticText284, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer326.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer316.Add( bSizer326, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer316 )
        self.Layout()

        self.Centre( wx.BOTH )

    def __del__( self ):
        pass

###########################################################################
## Class Connection Fail
###########################################################################
class Connectionfail ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 400,350 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.DefaultSize, wx.DefaultSize )
        self.SetBackgroundColour( wx.Colour( 241, 201, 250 ) )

        bSizer316 = wx.BoxSizer( wx.VERTICAL )

        self.bitmap10 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"Pusheenerrorsmall.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer316.Add( self.bitmap10, 0, wx.ALL, 5 )

        bSizer326 = wx.BoxSizer( wx.VERTICAL )


        bSizer326.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText284 = wx.StaticText( self, wx.ID_ANY, u"Connection Error\n               :(", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText284.Wrap( -1 )

        self.staticText284.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Candara" ) )

        bSizer326.Add( self.staticText284, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer326.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer316.Add( bSizer326, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer316 )
        self.Layout()

        self.Centre( wx.BOTH )

    def __del__( self ):
        pass


###########################################################################
## Class Error1
###########################################################################
class Error1 ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = u"Error", pos = wx.DefaultPosition, size = wx.Size( 400,350 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(400,350), wx.Size(400,350) )
        self.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )
        self.SetBackgroundColour( wx.Colour( 241, 201, 250 ) )

        bSizer317 = wx.BoxSizer( wx.VERTICAL )

        self.m_bitmap2 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"Capoo What (Small).png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer317.Add( self.m_bitmap2, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        bSizer318 = wx.BoxSizer( wx.VERTICAL )

        self.staticText283 = wx.StaticText( self, wx.ID_ANY, u"THERE IS AN EXCEED AMOUNT ON THE SCORE\n         INPUT. PLEASE RE-ENTER THE SCORE.", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText283.Wrap( -1 )

        self.staticText283.SetFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer318.Add( self.staticText283, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.button83 = wx.Button( self, wx.ID_ANY, u"OK", wx.DefaultPosition, wx.Size( -1,50 ), 0 )
        self.button83.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial" ) )

        bSizer318.Add( self.button83, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer317.Add( bSizer318, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer317 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button83.Bind( wx.EVT_BUTTON, self.Close )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Close( self, event ):
            self.Destroy()
          
###########################################################################
## Class DataSaved
###########################################################################
class DataSaved ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 400,350 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(400,350), wx.Size(400,350) )
        self.SetBackgroundColour( wx.Colour( 241, 201, 250 ) )

        bSizer316 = wx.BoxSizer( wx.VERTICAL )

        self.bitmap3 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"200x200bb.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer316.Add( self.bitmap3, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        bSizer322 = wx.BoxSizer( wx.VERTICAL )


        bSizer322.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText284 = wx.StaticText( self, wx.ID_ANY, u"THE DATA HAS BEEN SAVED\n\t   :D", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText284.Wrap( -1 )

        self.staticText284.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Candara" ) )

        bSizer322.Add( self.staticText284, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.button84 = wx.Button( self, wx.ID_ANY, u"OK", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer322.Add( self.button84, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer322.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer316.Add( bSizer322, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer316 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button84.Bind( wx.EVT_BUTTON, self.Close )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Close( self, event ):
            self.Destroy()

###########################################################################
## Class ScoreSaved
###########################################################################
class ScoreSaved ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 400,350 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(400,350), wx.Size(400,350) )
        self.SetBackgroundColour( wx.Colour( 241, 201, 250 ) )

        bSizer316 = wx.BoxSizer( wx.VERTICAL )

        self.bitmap4 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"Capoo OK.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer316.Add( self.bitmap4, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        bSizer323 = wx.BoxSizer( wx.VERTICAL )


        bSizer323.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText284 = wx.StaticText( self, wx.ID_ANY, u"THE SCORE HAD BEEN SAVED \n\t    :D", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText284.Wrap( -1 )

        self.staticText284.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Candara" ) )

        bSizer323.Add( self.staticText284, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.button85 = wx.Button( self, wx.ID_ANY, u"OK", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer323.Add( self.button85, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer323.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer316.Add( bSizer323, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer316 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button85.Bind( wx.EVT_BUTTON, self.Close )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Close( self, event ):
            self.Destroy()

###########################################################################
## Class DataUploaded
###########################################################################
class DataUploaded ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 400,350 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(400,350), wx.Size(400,350) )
        self.SetBackgroundColour( wx.Colour( 241, 201, 250 ) )

        bSizer316 = wx.BoxSizer( wx.VERTICAL )

        self.bitmap5 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"Capoo handsome.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer316.Add( self.bitmap5, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        bSizer324 = wx.BoxSizer( wx.VERTICAL )


        bSizer324.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText284 = wx.StaticText( self, wx.ID_ANY, u"THE DATA HAD BEEN UPLOADED\n\t         :D", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText284.Wrap( -1 )

        self.staticText284.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Candara" ) )

        bSizer324.Add( self.staticText284, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.button86 = wx.Button( self, wx.ID_ANY, u"OK", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer324.Add( self.button86, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer324.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer316.Add( bSizer324, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer316 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button86.Bind( wx.EVT_BUTTON, self.Close )
        self.Bind( wx.EVT_ACTIVATE, self.sender)

    def __del__( self ):
        pass

    
    def sender (self):
        pass
        
    # Virtual event handlers, overide them in your derived class
    def Close( self, event ):
            self.Destroy()
            
###########################################################################
## Class Reset
###########################################################################
class Reset ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 400,380 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(400,380), wx.Size(400,380) )
        self.SetBackgroundColour( wx.Colour( 241, 201, 250 ) )

        bSizer316 = wx.BoxSizer( wx.VERTICAL )

        self.bitmap6 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"Capoo Shock.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer316.Add( self.bitmap6, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        bSizer325 = wx.BoxSizer( wx.VERTICAL )


        bSizer325.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText284 = wx.StaticText( self, wx.ID_ANY, u"RESET SUCCESSFUL\n                  :O", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText284.Wrap( -1 )

        self.staticText284.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Candara" ) )

        bSizer325.Add( self.staticText284, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.button86 = wx.Button( self, wx.ID_ANY, u"OK", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer325.Add( self.button86, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer325.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer316.Add( bSizer325, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer316 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button86.Bind( wx.EVT_BUTTON, self.Close )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Close( self, event ):
        self.Destroy()

###########################################################################
## Class GenerateMessage
###########################################################################

class GenerateMessage ( wx.Frame ):

    def __init__( self, parent ):
        wx.Frame.__init__ ( self, parent, id = wx.ID_ANY, title = wx.EmptyString, pos = wx.DefaultPosition, size = wx.Size( 400,380 ), style = wx.DEFAULT_FRAME_STYLE|wx.TAB_TRAVERSAL )

        self.SetSizeHints( wx.Size(400,380), wx.Size(400,380) )
        self.SetBackgroundColour( wx.Colour( 241, 201, 250 ) )

        bSizer316 = wx.BoxSizer( wx.VERTICAL )

        self.bitmap7 = wx.StaticBitmap( self, wx.ID_ANY, wx.Bitmap( u"sticker.png", wx.BITMAP_TYPE_ANY ), wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer316.Add( self.bitmap7, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        bSizer326 = wx.BoxSizer( wx.VERTICAL )


        bSizer326.Add( ( 0, 0), 1, wx.EXPAND, 5 )

        self.staticText284 = wx.StaticText( self, wx.ID_ANY, u"GENERATE COMPLETE\n                    :3", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.staticText284.Wrap( -1 )

        self.staticText284.SetFont( wx.Font( 14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD, False, "Candara" ) )

        bSizer326.Add( self.staticText284, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )

        self.button86 = wx.Button( self, wx.ID_ANY, u"OK", wx.DefaultPosition, wx.DefaultSize, 0 )
        bSizer326.Add( self.button86, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )


        bSizer326.Add( ( 0, 0), 1, wx.EXPAND, 5 )


        bSizer316.Add( bSizer326, 1, wx.EXPAND, 5 )


        self.SetSizer( bSizer316 )
        self.Layout()

        self.Centre( wx.BOTH )

        # Connect Events
        self.button86.Bind( wx.EVT_BUTTON, self.Close )

    def __del__( self ):
        pass


    # Virtual event handlers, overide them in your derived class
    def Close( self, event ):
        self.Destroy()
        
###########################################################################
## Class Sound Error
###########################################################################
class soundError(object):
    def __init__(self,parent):
        self.Error = wx.adv.Sound("Quack.wav")
        
###########################################################################
## Class Sound Granted
###########################################################################
class soundAccept(object):
    def __init__(self,parent):
        self.Accept = wx.adv.Sound("Accept Sound.wav")

###########################################################################
## Class Sound Reset
###########################################################################
class soundReset(object):
    def __init__(self,parent):
        self.Resetto = wx.adv.Sound("nier_automata_mail.wav")

class MyApp(wx.App):
    def OnInit(self):
        myframe = MainSystem(None)
        myframe.Show(True)
        return True


if __name__ == '__main__':
    app = MyApp()
    app.MainLoop()